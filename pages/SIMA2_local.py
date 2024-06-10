import streamlit as st
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import math
import shutil
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import mysql.connector

# Importo el módulo sima_funciones.py
import sima_funciones

st.markdown("# SIMA_local")
#st.sidebar.markdown("# SIMA")

estaciones = ['Invierno', 'Verano']
estacion = st.radio("Seleccioná la estación de cultivo", estaciones)
st.write("La opción elegida es: ", estacion)

# Bloque 1: Arrastrar archivo y leerlo
archivo_excel = st.file_uploader("Arrastra el archivo acá", type=["xlsx", "xls"])

if archivo_excel is not None:
    progress_bar = st.progress(0)
    
    for porcentaje_completado in range(100):
        time.sleep(0.1)  # Ajusta el tiempo de espera según el tamaño del archivo
        progress_bar.progress(porcentaje_completado + 1)
    st.success("Archivo subido correctamente")

    is_clicked = st.button("PROCESAR")
    if is_clicked:
        
        @st.cache_data(experimental_allow_widgets=True)
        def convertir_a_sima(archivo_excel,estacion):

            # Conexión a BBDD MySQL 8 DAT
            host='66.97.33.201'
            password='Dat##2023Tabla$'
            port=33306
            user='datcrea_tablas'
            database='datcrea_tablas'

            conn = create_engine("mysql+mysqlconnector://{user}:{password}@{host}:{port}/{database}".format(host=host,user=user,password=password,database=database,port=port),connect_args={'auth_plugin': 'mysql_native_password'}).connect()
           

            try:
                # 1) Lectura del archivo de SIMA
                sima = pd.read_excel(archivo_excel)
            
                # Copia del df original
                df = sima.copy()

                # Importacion de funciones en orden pre-establecido
                
                #1
                sima_funciones.filtrar_campana(df, conn)
                #2
                sima_funciones.formato_fecha(df,conn)
                #3
                sima_funciones.validar_codigo_genetica(df, conn)
                #4
                sima_funciones.validar_cultivos(df,conn)
                #5
                sima_funciones.validar_genetica(df,conn)
                #6
                sima_funciones.validar_semillero(df,conn)
                #7
                sima_funciones.validar_genetica_semillero(df, conn)
                #8
                sima_funciones.fertilizante_int(df)
                #9
                sima_funciones.validar_id_senasa_fertilizantes(df,conn)
                #10
                sima_funciones.validar_nombre_fertilizante(df, conn)
                #11
                sima_funciones.validar_fertilizante(codigo,conn)
                #12
                sima_funciones.validar_dosis_productos(df,conn)
                #13
                sima_funciones.validar_momentos(df,conn)
                #14
                sima_funciones.calcular_densidad(df, conn)
                #15
                sima_funciones.reindex_columns(df, conn)
                #16
                sima_funciones.otro_cuic_campo_lote(df, conn)
                #17
                sima_funciones.validar_localidad(df,conn)
                #18
                sima_funciones.validar_departamento(df,conn)
                #19
                sima_funciones.validar_provincia(df, conn)
                #20
                sima_funciones.validar_tenencia(df,conn)
                #21
                sima_funciones.validar_destino(df,conn)
                #22
                sima_funciones.validar_cultivo_antecesor(df,conn)
                #23
                sima_funciones.validar_calidad_ambiente(df,conn)

                #24
                #sima_funciones.validar_subgrupo_cultivo(df,conn)
                sima_funciones.validar_subgrupo(df,conn)
                
                #25
                sima_funciones.validar_antecesores(df,conn)
                #26
                sima_funciones.validar_biotecnologia(row)
                #27
                sima_funciones.validar_fertilizacion(df,conn)
                #28
                sima_funciones.replace_ceros(df)
                #29
                sima_funciones.validar_forma_fertilizacion(df,conn)
                #30
                sima_funciones.validar_sistema_riego(df,conn)
                #31
                sima_funciones.influencia_napa(df,conn)


                #sima_funciones.validar_herbicidas(df,conn)
                #sima_funciones.validar_fungicida(df, conn)
                
                              
                
                # 3.8) Reemplazar el Producto de acuerdo a lo que dice el Código

                
                reg_fert_query = """SELECT * FROM datcrea_tablas.registro_fertilizantes WHERE id_senasa IS NOT NULL"""
                registros_fertilizantes = pd.read_sql(reg_fert_query, con=conn)
                registros_fertilizantes['id_senasa'] = registros_fertilizantes['id_senasa'].astype(int)

                # Crear un diccionario para almacenar los resultados
                id_senasa_dict = {}


                for index, row in registros_fertilizantes.iterrows():
                    fertilizante = row['fertilizante']
                    id_senasa = row['id_senasa']
                    if fertilizante in id_senasa_dict:
                        id_senasa_dict[fertilizante].append(id_senasa)
                    else:
                        id_senasa_dict[fertilizante] = [id_senasa]
                
                
                
                
                
            except Exception as e:
                st.error(f'Ocurrió un error: {e}')
                return None
            
            finally:
                conn.close()




            # COPIA DE ARCHIVOS EN DAT
            df.dropna(subset=['Campo'], inplace=True)
            df = df[df['Campo']!= '']
            
            
            
            
            # Subset del DF con la campaña 23-24 INVIERNO
            
            cultivos_invierno_query = """SELECT cultivo FROM datcrea_tablas.cultivos WHERE ciclo = 'i'""" # I MAYÚSCULA
            df = pd.read_sql(cultivos_invierno_query, conn)
            cultivos_invierno = df['cultivo'].tolist()


            cultivos_verano_query = """SELECT cultivo FROM datcrea_tablas.cultivos WHERE ciclo = 'v'""" # V MAYÚSCULA
            df = pd.read_sql(cultivos_verano_query, conn)
            cultivos_verano = df['cultivo'].tolist()

            cultivos_perennes_query = """SELECT cultivo FROM datcrea_tablas.cultivos WHERE ciclo = 'p'""" # P MAYÚSCULA
            df = pd.read_sql(cultivos_perennes_query, conn)
            cultivos_perennes = df['cultivo'].tolist()
            
            if estacion == 'Invierno':
                                          
                # Filtro el DataFrame

                invierno = df[df['Cultivo'].isin(cultivos_invierno)]

                seccion = 200  # Máx registros por planilla xls
                registros_campana = len(invierno.index)  # cuenta el total de registros
                registros_wb = math.ceil(registros_campana / seccion)  # cociente para saber cantidad de worksheets

                # Armo listado de subset de 200 (máx) registros cada uno
                dataframes = []

                for i in range(registros_wb):
                    inicio = i * seccion
                    fin = (i + 1) * seccion
                    seccion_df = invierno.iloc[inicio:fin]
                    dataframes.append(seccion_df)

                # Ruta del archivo original
                ruta_original = 'C:/Users/EPolacco/Documents/pruebas_local/DAT-Cultivos-de-invierno-2023-24.xlsx'

                # Cargar el archivo original
                wb_original = load_workbook(filename=ruta_original)

                # Obtener la hoja original
                planilla_original = wb_original['Planilla ']

                columns_to_exclude = {'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'U', 'W', 'Y', 'Z',
                                    'AF', 'AG', 'AK', 'AN', 'CB', 'CK', 'CN', 'CM', 'CO', 'CP', 'CQ',
                                    'CR', 'DU', 'DV', 'DW', 'IA','IB','IC','ID','IE','IF','IG','IH',
                                    'II','IJ','IK','IL','IM','IN','IO'}

                # Iterar sobre los df y guardar en archivos xls separados
                for i, seccion_df in enumerate(dataframes, 1):
                    # Crear un workbook para cada sección
                    if i == 1:
                        wb = wb_original  # Utilizar el archivo original para el primer df
                        sheet = planilla_original
                    else:

                        # Crear una copia del archivo original para los df subsiguientes
                        ruta_copia = f'C:/Users/EPolacco/Documents/pruebas_local/DAT-Cultivos-de-invierno-2023-24_SECCION_{i}.xlsx'
                        shutil.copy(ruta_original, ruta_copia)

                        # Cargar la copia
                        wb = load_workbook(filename=ruta_copia)
                        sheet = wb['Planilla ']
                        sheet.title = f'Planilla '

                    # Borrar datos existentes en la hoja, después de la fila 14 porque sino se pisan los datos
                    for row in sheet.iter_rows(min_row=15, max_row=sheet.max_row, min_col=6, max_col=sheet.max_column):
                        for cell in row:
                            if cell.column_letter not in columns_to_exclude:
                                cell.value = None

                    # Obtener las filas de datos del df
                    rows = dataframe_to_rows(seccion_df, index=False, header=False)

                    # Copiar en worksheet a partir de la fila 15 y columna 7
                    for r_idx, row in enumerate(rows, 15):
                        for c_idx, value in enumerate(row, 6):
                            if not pd.isna(value):
                                sheet.cell(row=r_idx, column=c_idx, value=value)

                    # Guardar el workbook, si no es el primer df
                    if i != 1:
                        wb.save(ruta_copia)

                    # Pedir al usuario que ingrese el nombre del archivo
                    nombre_archivo = st.text_input(f"Ingrese el nombre para el archivo {i}:", f"parte {i}")

                    # Guardar workbook con el primer df
                    ruta_archivo = f'C:/Users/EPolacco/Documents/pruebas_local/{nombre_archivo}.xlsx'
                    wb.save(ruta_archivo)

                    # Añadir botón de descarga
                    st.download_button(f"Descargar {nombre_archivo}", 
                                    open(ruta_archivo, 'rb').read(), 
                                    file_name=f"{nombre_archivo}.xlsx", 
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            else:
                # Filtro el DataFrame

                verano = df[df['Cultivo'].isin(cultivos_verano)]

                seccion = 200  # Máx registros por planilla xls
                registros_campana = len(verano.index)  # cuenta el total de registros
                registros_wb = math.ceil(registros_campana / seccion)  # cociente para saber cantidad de worksheets

                # Armo listado de subset de 200 (máx) registros cada uno
                dataframes = []

                for i in range(registros_wb):
                    inicio = i * seccion
                    fin = (i + 1) * seccion
                    seccion_df = verano.iloc[inicio:fin]
                    dataframes.append(seccion_df)

                # Ruta del archivo original
                ruta_original = 'C:/Users/EPolacco/Documents/pruebas_local/DAT-Cultivos-de-verano-2023-24.xlsx'

                # Cargar el archivo original
                wb_original = load_workbook(filename=ruta_original)

                # Obtener la hoja original
                planilla_original = wb_original['Planilla ']

                # Columnas excluídas porque tienen fórmulas y validaciones.
                columns_to_exclude = {'I', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'U', 'W', 'Y', 'Z', 'AC',
                      'AF', 'AG', 'AH', 'AK', 'AO', 'CC', 'CS', 'CU', 'CV', 'CW', 'CX',
                      'CY', 'CZ', 'DI', 'DJ', 'DK','DL', 'DM', 'DN','DO','DP','DQ','DR', 
                      'DS','DT','DU', 'DV', 'DW', 'DY', 'DZ', 'EA',
                      'IQ','IR','IS','IT','IU','IV','IW','IX','IY','IZ', 'JA', 'JB', 'JC',
                      'JD', 'JE'}

                # Iterar sobre los df y guardar en archivos xls separados
                for i, seccion_df in enumerate(dataframes, 1):
                    # Crear un workbook para cada sección
                    if i == 1:
                        wb = wb_original  # Utilizar el archivo original para el primer df
                        sheet = planilla_original
                    else:

                        # Crear una copia del archivo original para los df subsiguientes
                        ruta_copia = f'C:/Users/EPolacco/Documents/pruebas_local/DAT-Cultivos-de-verano-2023-24_SECCION_{i}.xlsx'
                        shutil.copy(ruta_original, ruta_copia)

                        # Cargar la copia
                        wb = load_workbook(filename=ruta_copia)
                        sheet = wb['Planilla ']
                        sheet.title = f'Planilla '

                    # Borrar datos existentes en la hoja, después de la fila 14 porque sino se pisan los datos
                    for row in sheet.iter_rows(min_row=15, max_row=sheet.max_row, min_col=6, max_col=sheet.max_column):
                        for cell in row:
                            if cell.column_letter not in columns_to_exclude:
                                cell.value = None

                    # Obtener las filas de datos del df
                    rows = dataframe_to_rows(seccion_df, index=False, header=False)

                    # Copiar en worksheet a partir de la fila 15 y columna 7
                    for r_idx, row in enumerate(rows, 15):
                        for c_idx, value in enumerate(row, 6):
                            if not pd.isna(value):
                                sheet.cell(row=r_idx, column=c_idx, value=value)

                    # Guardar el workbook, si no es el primer df
                    if i != 1:
                        wb.save(ruta_copia)

                    # Pedir al usuario que ingrese el nombre del archivo
                    nombre_archivo = st.text_input(f"Ingrese el nombre para el archivo {i}:", f"parte {i}")

                    # Guardar workbook con el primer df
                    ruta_archivo = f'C:/Users/EPolacco/Documents/pruebas_local/{nombre_archivo}.xlsx'
                    wb.save(ruta_archivo)

                    # Añadir botón de descarga
                    st.download_button(f"Descargar {nombre_archivo}", 
                                    open(ruta_archivo, 'rb').read(), 
                                    file_name=f"{nombre_archivo}.xlsx", 
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        convertir_a_sima(archivo_excel)
        st.success('Proceso terminado')