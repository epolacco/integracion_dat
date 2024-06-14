import streamlit as st
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import math
import shutil
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import mysql.connector

# Importo el módulo sima_funciones.py

#import sima_funciones

st.markdown("# SIMA")
#st.sidebar.markdown("# SIMA")

estaciones = ['Invierno', 'Verano']
estacion = st.radio("Seleccioná la estación de cultivo", estaciones)
st.write("La opción elegida es: ", estacion)
campana_filtro = st.selectbox("Seleccioná campaña a procesar",['2023-2024'])
# Bloque 1: Arrastrar archivo y leerlo
archivo_excel = st.file_uploader("Arrastra el archivo acá", type=["xlsx", "xls"])

if archivo_excel is not None:
    progress_bar = st.progress(0)
    
    for porcentaje_completado in range(100):
        time.sleep(0.1)  # Ajusta el tiempo de espera según el tamaño del archivo
        progress_bar.progress(porcentaje_completado + 1)
    st.success("Archivo subido correctamente")

    is_clicked = st.button("PROCESAR")
    if is_clicked:
        
        @st.cache_data(experimental_allow_widgets=True)
        def convertir_a_sima(archivo_excel,estacion, campana_filtro):
            if estacion == "Verano":
                           
                # Importo librerías
                from openpyxl import load_workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                import pandas as pd
                import numpy as np
                from datetime import datetime
                import math
                import shutil

                # Conexión a la base
                from sqlalchemy import create_engine
                import mysql.connector
                
                host='66.97.33.201'
                password='Dat##2023Tabla$'
                port=33306
                user='datcrea_tablas'
                database='datcrea_tablas'
                
                conn = create_engine("mysql+mysqlconnector://{user}:{password}@{host}:{port}/{database}".format(host=host,user=user,password=password,database=database,port=port),connect_args={'auth_plugin': 'mysql_native_password'})
                
            
                #ruta_completa = directorio + archivo
                #df = pd.read_excel(ruta_completa)
                #campana_filtro = '2023-2024'
                df = pd.read_excel(archivo_excel)
                cuic = df['CUIC'][0]
                
                def filtrar_campana(df, campana_filtro):
                    
                    filtro = df['Ciclo'] == campana_filtro
                    df = df[filtro]
                    return df

                df = filtrar_campana(df, campana_filtro)

                
                def formato_fecha(df):
                    df['FechaSiembra'] = df['FechaSiembra'].str[:10] #Elimino horas en fechas
                    df['FechaSiembra'] = pd.to_datetime(df['FechaSiembra'])
                    df['FechaSiembra'] = df['FechaSiembra'].dt.strftime('%d/%m/%Y')    
                    return df
                    
                df = formato_fecha(df)


                # 3.2) Validar código genética #FUNCIONA
                # Genética
                # Consultar los nro_registro desde la base de datos
                def validar_codigo_genetica(df, conn):
                    query = "SELECT nro_registro FROM datcrea_tablas.materiales WHERE nro_registro IS NOT NULL"
                    df_materiales = pd.read_sql(query, conn)
                    materiales = df_materiales['nro_registro'].astype(int).tolist()
                    df['Codigo_Genetica'] = df['Codigo_Genetica'].fillna(0)
                    df['Codigo_Genetica'] = df['Codigo_Genetica'].astype('Int64')
                    df['Codigo_Genetica'] = ['' if x not in materiales else x for x in df['Codigo_Genetica']]
                    return df

                df = validar_codigo_genetica(df, conn)



                ## 3.3) Validación Cultivo #FUNCIONA
                # Cultivos
                def validar_cultivos(df, conn):
                    cultivos_query = """SELECT cultivo FROM datcrea_tablas.cultivos"""
                    cultivos_df = pd.read_sql(cultivos_query,con=conn)
                    cultivos = cultivos_df['cultivo'].tolist()
                    df['Cultivo'] = df['Cultivo'].replace(['Soja - De segunda', 'Mani'], ['Soja', 'Maní'])
                    df['Cultivo'] = ['' if x not in cultivos else x for x in df['Cultivo']]
                    return df
                
                df = validar_cultivos(df,conn)

                ## 3.4) Validación Genética
                
                # Cultivos
                def validar_genetica(df,conn): #FUNCIONA
                    
                    genetica_query = """SELECT material FROM datcrea_tablas.materiales"""
                    genetica_df = pd.read_sql(genetica_query,con=conn)
                    
                    genetica = genetica_df['material'].tolist()
                    df['Genetica'] = ['' if x not in genetica else x for x in df['Genetica']]
                    return df
                    
                df = validar_genetica(df,conn)

                ## 3.9) Validación Semillero #FUNCIONA
                def validar_semillero(df, conn): 
                    
                    semillero_query = """SELECT semillero FROM datcrea_tablas.semilleros"""
                    semillero = pd.read_sql(semillero_query,con=conn)
                    semillero = semillero['semillero'].tolist()
                    df['Semillero'] = ['' if x not in semillero else x for x in df['Semillero']]
                    return df
                
                df = validar_semillero(df, conn)

                def validar_genetica_semillero(df, conn): #FUNCIONA
                    # VALIDA GENÉTICA Y SEMILLERO EN BASE A CÓDIGO GENÉTICO Y CULTIVO
                    asignaciones = [
                        (('Girasol', '18615'), ('Nidera.Girasol', '106 CL HO')),
                        (('Trigo', '9993'), ('Aca.Trigo', '1801F')),
                        (('Trigo', '11270'), ('Aca.Trigo', '1901F')),
                        (('Soja', '19320'), ('Stine.Soja', '25EB32')),
                        (('Soja', '19321'), ('Stine.Soja', '29EB02')),
                        (('Maíz', '16570'), ('Nuseed.Maíz', '3790 RR2-CL')),
                        (('Soja','19324'), ('Stine.Soja','38EB03')),
                        (('Soja','19348'), ('Nidera.Soja','4031 E NS')),
                        (('Soja','19326'), ('Stine.Soja','40EB20')),
                        (('Soja','17688'), ('MacroSeed.Soja','40MS01 E STS')),
                        (('Soja','20271'), ('Stine.Soja','45EB52')),
                        (('Soja','20277'), ('Stine.Soja','47EA32')),
                        (('Soja','15994'), ('MacroSeed.Soja','47MS01 STS')),
                        (('Soja','19331'), ('Stine.Soja','48EB20')),
                        (('Soja','18457'), ('MacroSeed.Soja','48MS01 E')),
                        (('Soja','15267'), ('LaTijereta.Soja','4914 IPRO')),
                        (('Soja','20958'), ('Nidera.Soja','4941 E STS NS')),
                        (('Soja','16661'), ('MacroSeed.Soja','50MS01 STS')),
                        (('Soja','20069'), ('MacroSeed.Soja','51MS01 E')),
                        (('Soja','17113'), ('MacroSeed.Soja','53MS01 IPRO')),
                        (('Soja','17640'), ('MacroSeed.Soja','53MS02 STS')),
                        (('Soja','16545'), ('MacroSeed.Soja','59MS01 IPRO STS')),
                        (('Soja','15479'), ('MacroSeed.Soja','60MS01 STS')),
                        (('Soja','19100'), ('MacroSeed.Soja','61MS01 STS')),
                        (('Maíz','20857'), ('Nidera.Maíz','6223 CE')),
                        (('Soja','20857'), ('Nidera.Soja','6223 CE')),
                        (('Soja','16717'), ('MacroSeed.Soja','62MS01 IPRO STS')),
                        (('Soja','20099'), ('MacroSeed.Soja','62MS02 E')),
                        (('Soja','20175'), ('MacroSeed.Soja','63MS01 CE')),
                        (('Soja','20194'), ('Credenz.Soja','6522')),
                        (('Soja','17639'), ('MacroSeed.Soja','66MS01')),
                        (('Soja','20171'), ('MacroSeed.Soja','68MS01 CE')),
                        (('Soja','16669'), ('MacroSeed.Soja','78MS01 IPRO')),
                        (('Soja','6915'), ('Nidera.Soja','A 3302 RG')),
                        (('Soja','9986'), ('Nidera.Soja','A 3731 RG')),
                        (('Soja','6165'), ('Nidera.Soja','A 3901 RG')),
                        (('Soja','9217'), ('Nidera.Soja','A 4209 RG')),
                        (('Soja','7246'), ('Nidera.Soja','A 4303 RG')),
                        (('Soja','8272'), ('Nidera.Soja','A 4505 RG')),
                        (('Soja','9214'), ('Nidera.Soja','A 4553 RG')),
                        (('Soja','8756'), ('Nidera.Soja','A 4613 RG')),
                        (('Soja','10082'), ('Nidera.Soja','A 5009 RG')),
                        (('Soja','5829'), ('Nidera.Soja','A 5409 RG')),
                        (('Soja','6742'), ('Nidera.Soja','A 5503')),
                        (('Soja','11072'), ('Nidera.Soja','A 5509 RG')),
                        (('Soja','11073'), ('Nidera.Soja','A 5909 RG')),
                        (('Soja','8759'), ('Nidera.Soja','A 6126 RG')),
                        (('Soja','8273'), ('Nidera.Soja','A 6411 RG')),
                        (('Soja','10083'), ('Nidera.Soja','A 6509 RG')),
                        (('Soja','6739'), ('Nidera.Soja','A 7636 RG')),
                        (('Soja','5828'), ('Nidera.Soja','A 8000 RG')),
                        (('Colza','11426'), ('HighTech.Colza','ABILITY')),
                        (('Trigo','10021'), ('Aca.Trigo','ACA 201')),
                        (('Trigo','11267'), ('Aca.Trigo','ACA 202')),
                        (('Girasol','10219'), ('ACA.Girasol','ACA 203')),
                        (('Trigo','7632'), ('Aca.Trigo','ACA 302')),
                        (('Trigo','7633'), ('Aca.Trigo','ACA 303')),
                        (('Trigo','15747'), ('Aca.Trigo','ACA 303 plus')),
                        (('Trigo','8470'), ('Aca.Trigo','ACA 304')),
                        (('Trigo','20497'), ('Aca.Trigo','ACA 308')),
                        (('Trigo','9530'), ('Aca.Trigo','ACA 315')),
                        (('Trigo','11271'), ('Aca.Trigo','ACA 320')),
                        (('Girasol','15388'), ('ACA.Girasol','ACA 350')),
                        (('Trigo','14471'), ('Aca.Trigo','ACA 360')),
                        (('Soja','7203'), ('ACA.Soja','ACA 360 GR')),
                        (('Trigo','21453'), ('Aca.Trigo','ACA 363')),
                        (('Trigo','21451'), ('Aca.Trigo','ACA 364')),
                        (('Soja','7674'), ('ACA.Soja','ACA 420 GR')),
                        (('Soja','8395'), ('ACA.Soja','ACA 460 GR')),
                        (('Maíz','13902'), ('ACA.Maíz','ACA 470')),
                        (('Soja','7202'), ('ACA.Soja','ACA 470 GR')),
                        (('Maíz','14836'), ('ACA.Maíz','ACA 474')),
                        (('Maíz','14087'), ('ACA.Maíz','ACA 480')),
                        (('Soja','7697'), ('ACA.Soja','ACA 480 GR')),
                        (('Soja','3937'), ('ACA.Soja','ACA 490')),
                        (('Maíz','16320'), ('ACA.Maíz','ACA 493')),
                        (('Soja','16543'), ('ACA.Soja','ACA 4949 IPRO')),
                        (('Soja','15547'), ('ACA.Soja','ACA 5020 IPRO')),
                        (('Maíz','14089'), ('ACA.Maíz','ACA 514 Flint')),
                        (('Maíz','14088'), ('ACA.Maíz','ACA 530 Flint')),
                        (('Soja','7204'), ('ACA.Soja','ACA 530 GR')),
                        (('Sorgo','14103'), ('ACA.Sorgo','ACA 548')),
                        (('Sorgo','1846'), ('ACA.Sorgo','ACA 550')),
                        (('Sorgo','5975'), ('ACA.Sorgo','ACA 558')),
                        (('Soja','3940'), ('ACA.Soja','ACA 560')),
                        (('Sorgo','10448'), ('ACA.Sorgo','ACA 561')),
                        (('Sorgo','17570'), ('ACA.Sorgo','ACA 563')),
                        (('Soja','7698'), ('ACA.Soja','ACA 570 GR')),
                        (('Maíz','16019'), ('ACA.Maíz','ACA 580 Flint')),
                        (('Soja','15265'), ('ACA.Soja','ACA 5814 IPRO')),
                        (('Soja','15394'), ('ACA.Soja','ACA 5825 IPRO')),
                        (('Trigo','8067'), ('Aca.Trigo','ACA 601')),
                        (('Trigo','14472'), ('Aca.Trigo','ACA 602')),
                        (('Trigo','19477'), ('Aca.Trigo','ACA 604')),
                        (('Alfalfa','9455'), ('ACA.Alfalfa','ACA 605')),
                        (('Maíz','16020'), ('ACA.Maíz','ACA 622 Flint')),
                        (('Soja','14613'), ('ACA.Soja','ACA 6513 IPRO')),
                        (('Soja','20178'), ('ACA.Soja','ACA 70A70 CE')),
                        (('Soja','16023'), ('ACA.Soja','ACA 7890 IPRO')),
                        (('Trigo','7979'), ('Aca.Trigo','ACA 801')),
                        (('Soja','14949'), ('ACA.Soja','ACA 8080 IPRO')),
                        (('Girasol','9307'), ('ACA.Girasol','ACA 861')),
                        (('Girasol','15884'), ('ACA.Girasol','ACA 869')),
                        (('Girasol','17526'), ('ACA.Girasol','ACA 870')),
                        (('Girasol','2330'), ('ACA.Girasol','ACA 884')),
                        (('Girasol','11822'), ('ACA.Girasol','ACA 887')),
                        (('Girasol','15885'), ('ACA.Girasol','ACA 889')),
                        (('Trigo','9520'), ('Aca.Trigo','ACA 901')),
                        (('Alfalfa','10740'), ('ACA.Alfalfa','ACA 903')),
                        (('Trigo','10740'), ('Aca.Trigo','ACA 903')),
                        (('Trigo','10739'), ('Aca.Trigo','ACA 905PA')),
                        (('Trigo','12340'), ('Aca.Trigo','ACA 906')),
                        (('Trigo','14473'), ('Aca.Trigo','ACA 908')),
                        (('Trigo','15733'), ('Aca.Trigo','ACA 909')),
                        (('Trigo','19479'), ('Aca.Trigo','ACA 917')),
                        (('Maíz','16139'), ('ACA.Maíz','ACA VG 48')),
                        (('Alfalfa','2974'), ('LosPrados.Alfalfa','ACONCAGUA')),
                        (('Sorgo','12653'), ('AdSur.Sorgo','AD-91 SUCROL')),
                        (('Sorgo','15370'), ('Advanta.Sorgo','ADV 1100')),
                        (('Sorgo','17874'), ('Advanta.Sorgo','ADV 1114')),
                        (('Sorgo','12726'), ('Advanta.Sorgo','ADV 114')),
                        (('Sorgo','16870'), ('Advanta.Sorgo','ADV 1250 IG')),
                        (('Sorgo','16415'), ('Advanta.Sorgo','ADV 1302')),
                        (('Sorgo','16414'), ('Advanta.Sorgo','ADV 1305')),
                        (('Sorgo','16972'), ('Advanta.Sorgo','ADV 1350 IG')),
                        (('Sorgo','14804'), ('Advanta.Sorgo','ADV 2010')),
                        (('Maíz','19860'), ('Advanta.Maíz','ADV 23.0')),
                        (('Sorgo','18582'), ('Advanta.Sorgo','ADV 2450 IG')),
                        (('Sorgo','14138'), ('Advanta.Sorgo','ADV 2499')),
                        (('Sorgo','17702'), ('Advanta.Sorgo','ADV 2701')),
                        (('Sorgo','14137'), ('Advanta.Sorgo','ADV 2800')),
                        (('Girasol','14857'), ('Advanta.Girasol','ADV 5200')),
                        (('Girasol','17932'), ('Advanta.Girasol','ADV 5304')),
                        (('Girasol','21663'), ('Advanta.Girasol','ADV 5310')),
                        (('Girasol','20601'), ('Advanta.Girasol','ADV 5407')),
                        (('Girasol','16183'), ('Advanta.Girasol','ADV 5500')),
                        (('Girasol','19063'), ('Advanta.Girasol','ADV 5566')),
                        (('Maíz','14793'), ('Advanta.Maíz','ADV 8101')),
                        (('Maíz','18158'), ('Advanta.Maíz','ADV 8560 T')),
                        (('Maíz','19859'), ('Advanta.Maíz','ADV 8570 T')),
                        (('Sorgo','14396'), ('Agseed.Sorgo','AG 1816')),
                        (('Sorgo','14393'), ('Agseed.Sorgo','AG 1817')),
                        (('Maíz','16719'), ('Agseed.Maíz','AG 7004')),
                        (('Maíz','11799'), ('Agseed.Maíz','AG 8000')),
                        (('Maíz','9577'), ('Agseed.Maíz','AG 9005')),
                        (('Maíz','17727'), ('Agseed.Maíz','AG 9300')),
                        (('Maíz','18407'), ('Agseed.Maíz','AG 9400')),
                        (('Sorgo','6874'), ('Agseed.Sorgo','AG SILO 200')),
                        (('Trigo','13123'), ('Buck.Trigo','AGP 127')),
                        (('Trigo','11565'), ('Buck.Trigo','AGP FAST')),
                        (('Girasol','11685'), ('Advanta.Girasol','AGUARA 6')),
                        (('Trigo','22406'), ('Bioseminis.Trigo','AGUARIBAY')),
                        (('Trigo','21629'), ('Buck.Trigo','AIMARÁ')),
                        (('Cebada','4675'), ('INTA.Cebada','ALICIA INTA ')),
                        (('Soja','8276'), ('ALM.Soja','ALM 3530')),
                        (('Soja','10565'), ('ALM.Soja','ALM 3830')),
                        (('Soja','8078'), ('ALM.Soja','ALM 4650')),
                        (('Soja','9850'), ('ALM.Soja','ALM 4930')),
                        (('Soja','8787'), ('Sursem.Soja','ANDREA 60')),
                        (('Soja','8785'), ('Sursem.Soja','ANDREA 63')),
                        (('Soja','8342'), ('Sursem.Soja','ANTA81')),
                        (('Maíz','16964'), ('AlumniSeed.Maíz','AP2505')),
                        (('Maíz','19373'), ('AlumniSeed.Maíz','AP2509')),
                        (('Maíz','16965'), ('AlumniSeed.Maíz','AP4512')),
                        (('Maíz','12343'), ('AlumniSeed.Maíz','AP6002')),
                        (('Maíz','16963'), ('AlumniSeed.Maíz','AP6005')),
                        (('Maíz','18612'), ('AlumniSeed.Maíz','AP8204')),
                        (('Maíz','18613'), ('AlumniSeed.Maíz','AP8205')),
                        (('Trigo','22343'), ('Bioseminis.Trigo','ARAZÁ')),
                        (('Trigo','21523'), ('Bioseminis.Trigo','ARCE')),
                        (('Maíz','17593'), ('Argenetics.Maíz','ARG 7712 BT RR')),
                        (('Maíz','19480'), ('Argenetics.Maíz','ARG 7716 BT RR')),
                        (('Maíz','15925'), ('Argenetics.Maíz','ARG 7730 BT')),
                        (('Maíz','15924'), ('Argenetics.Maíz','ARG 7732 BT CL')),
                        (('Maíz','17449'), ('Argenetics.Maíz','ARG 7742 FCL')),
                        (('Maíz','17535'), ('Argenetics.Maíz','ARG 8410 BT')),
                        (('Maíz','17450'), ('Argenetics.Maíz','ARG 8500 T')),
                        (('Maíz','17452'), ('Argenetics.Maíz','ARG 8800 T')),
                        (('Maíz','17454'), ('Argenetics.Maíz','ARG 8900 W')),
                        (('Maíz','15416'), ('Argenetics.Maíz','ARGENPOP 141')),
                        (('Maíz','15420'), ('Argenetics.Maíz','ARGENPOP 142')),
                        (('Sorgo','13412'), ('Argenetics.Sorgo','ARGENSOR 110 T')),
                        (('Sorgo','11806'), ('Argenetics.Sorgo','ARGENSOR 121')),
                        (('Sorgo','14106'), ('Argenetics.Sorgo','ARGENSOR 125 B')),
                        (('Sorgo','13614'), ('Argenetics.Sorgo','ARGENSOR 130 T')),
                        (('Sorgo','15428'), ('Argenetics.Sorgo','ARGENSOR 134 T')),
                        (('Sorgo','11805'), ('Argenetics.Sorgo','ARGENSOR 151 DP')),
                        (('Sorgo','14783'), ('Argenetics.Sorgo','ARGENSOR 155 DP')),
                        (('Maíz','13859'), ('Arvales.Maíz','ARV 2155')),
                        (('Maíz','12510'), ('Arvales.Maíz','ARV 2180')),
                        (('Maíz','12498'), ('Arvales.Maíz','ARV 2194')),
                        (('Maíz','12499'), ('Arvales.Maíz','ARV 2310')),
                        (('Sorgo','14782'), ('Arvales.Sorgo','ARV 300')),
                        (('Sorgo','13300'), ('Arvales.Sorgo','ARV 390')),
                        (('Soja','12013'), ('ASP.Soja','AS 3601')),
                        (('Soja','12400'), ('ASP.Soja','AS 3911')),
                        (('Soja','9857'), ('ASP.Soja','AS 4201')),
                        (('Soja','12037'), ('ASP.Soja','AS 4402')),
                        (('Soja','16421'), ('ASP.Soja','AS 4731')),
                        (('Soja','9764'), ('ASP.Soja','AS 4801')),
                        (('Soja','13543'), ('ASP.Soja','AS 4931')),
                        (('Soja','10573'), ('ASP.Soja','AS 5308i')),
                        (('Maní','14799'), ('INTA.Maní','ASEM 400 INTA')),
                        (('Maní','13703'), ('INTA.Maní','ASEM VICTOR INTA')),
                        (('Alfalfa','2394'), ('BayaCasal.Alfalfa','AURORA')),
                        (('Soja','16521'), ('ASGROW.Soja','AW 3806 IPRO')),
                        (('Soja','18422'), ('ASGROW.Soja','AW 3920 IPRO')),
                        (('Soja','18427'), ('ASGROW.Soja','AW 4320 IPRO')),
                        (('Soja','16408'), ('ASGROW.Soja','AW 4326 IPRO')),
                        (('Soja','18435'), ('ASGROW.Soja','AW 4610 IPRO')),
                        (('Soja','16420'), ('ASGROW.Soja','AW 4736 IPRO')),
                        (('Soja','16885'), ('ASGROW.Soja','AW 4927 IPRO')),
                        (('Soja','18441'), ('ASGROW.Soja','AW 5021 IPRO')),
                        (('Soja','15320'), ('ASGROW.Soja','AW 5714 IPRO')),
                        (('Soja','15318'), ('ASGROW.Soja','AW 5815 IPRO')),
                        (('Soja','18437'), ('ASGROW.Soja','AW 5920 IPRO')),
                        (('Soja','14645'), ('ASGROW.Soja','AW 6211 IPRO')),
                        (('Soja','18425'), ('ASGROW.Soja','AW 6320 IPRO')),
                        (('Soja','17787'), ('ASGROW.Soja','AW 7720 IPRO')),
                        (('Maíz','15325'), ('Nidera.Maíz','AX 7761')),
                        (('Maíz','16558'), ('Nidera.Maíz','AX 7784')),
                        (('Maíz','15326'), ('Nidera.Maíz','AX 7822')),
                        (('Maíz','15862'), ('Nidera.Maíz','AX 7918')),
                        (('Maíz','15862'), ('Nidera.Maíz','AX 7919')),
                        (('Maíz','16557'), ('Nidera.Maíz','AX 8010')),
                        (('Maíz','10269'), ('Nidera.Maíz','AX 852')),
                        (('Maíz','12677'), ('Nidera.Maíz','AX 887')),
                        (('Soja','8588'), ('Sursem.Soja','AYELEN22')),
                        (('Soja','7592'), ('Sursem.Soja','AZUL35')),
                        (('Colza','15253'), ('Limagrain.Colza','Albatros')),
                        (('Trigo','18190'), ('DonMario.Trigo','Alerce')),
                        (('Alfalfa','19212'), ('Gentos.Alfalfa','Alfalfa Latte 920')),
                        (('Alfalfa','13853'), ('Gentos.Alfalfa','Alfalfa Nobel 620')),
                        (('Alfalfa','15109'), ('Gentos.Alfalfa','Alfalfa Nobel 720')),
                        (('Trigo','15699'), ('DonMario.Trigo','Algarrobo')),
                        (('Trigo','14576'), ('LG.Trigo','Alhambra')),
                        (('Cebada','18704'), ('CyMQuilmes.Cebada','Alhue')),
                        (('Soja','12135'), ('INTA.Soja','Alim 5.09')),
                        (('Alfalfa','19944'), ('Forratec.Alfalfa','AlphaTec 621')),
                        (('Alfalfa','19945'), ('Forratec.Alfalfa','AlphaTec 821')),
                        (('Alfalfa','19947'), ('Forratec.Alfalfa','AlphaTec 921')),
                        (('Cebada','13129'), ('MalteriaPampa.Cebada','Andreia')),
                        (('Trigo','12406'), ('DonMario.Trigo','Arex')),
                        (('Girasol','10945'), ('Argenetics.Girasol','Argensol 20')),
                        (('Girasol','17453'), ('Argenetics.Girasol','Argensol 20 Max')),
                        (('Girasol','13316'), ('Argenetics.Girasol','Argensol 40')),
                        (('Girasol','18034'), ('Argenetics.Girasol','Argensol 54 AO')),
                        (('Girasol','17448'), ('Argenetics.Girasol','Argensol 72 CL')),
                        (('Girasol','10945'), ('Argenetics.Girasol','Argensol 76')),
                        (('Trigo','17150'), ('LG.Trigo','Arlask')),
                        (('Trigo','19732'), ('DonMario.Trigo','Aromo')),
                        (('Trigo','5378'), ('Buck.Trigo','Arriero')),
                        (('Alfalfa','18394'), ('Barenbrug.Alfalfa','Artemis')),
                        (('Trigo','19379'), ('LG.Trigo','Aryal')),
                        (('Trigo','10823'), ('DonMario.Trigo','Atlax')),
                        (('Trigo','17458'), ('DonMario.Trigo','Audaz')),
                        (('Poroto','6992'), ('INTA.Leales.Poroto','Azabache INTA')),
                        (('Maíz','19835'), ('Basso.Maíz','BAS 2206')),
                        (('Maíz','12736'), ('Basso.Maíz','BAS 5802')),
                        (('Maíz','16271'), ('Basso.Maíz','BAS 5803')),
                        (('Maíz','12645'), ('Basso.Maíz','BAS 6101')),
                        (('Maíz','12737'), ('Basso.Maíz','BAS 6102')),
                        (('Maíz','16270'), ('Basso.Maíz','BAS 6104')),
                        (('Trigo','15764'), ('Bioceres.Trigo','BASILIO')),
                        (('Trigo','21706'), ('Bioseminis.Trigo','BERMEJO HB4')),
                        (('Trigo','16330'), ('Bioceres.Trigo','BIO 1008')),
                        (('Soja','12860'), ('Bioceres.Soja','BIO 3.5')),
                        (('Soja','11769'), ('Bioceres.Soja','BIO 3.8')),
                        (('Soja','13547'), ('Bioceres.Soja','BIO 3.9')),
                        (('Soja','10574'), ('Bioceres.Soja','BIO 4.2')),
                        (('Soja','12465'), ('Bioceres.Soja','BIO 4.6')),
                        (('Soja','11302'), ('Bioceres.Soja','BIO 4.8')),
                        (('Soja','13541'), ('Bioceres.Soja','BIO 4.9')),
                        (('Soja','12862'), ('Bioceres.Soja','BIO 5.4')),
                        (('Colza','10066'), ('Nuseed.Colza','BIOAUREO 2386')),
                        (('Colza','10067'), ('Nuseed.Colza','BIOAUREO 2486')),
                        (('Soja','15568'), ('Bioceres.Soja','BIOCERES 3.41')),
                        (('Soja','15967'), ('Bioceres.Soja','BIOCERES 4.11')),
                        (('Soja','15968'), ('Bioceres.Soja','BIOCERES 4.51')),
                        (('Soja','18412'), ('Bioceres.Soja','BIOCERES 4.52')),
                        (('Soja','15966'), ('Bioceres.Soja','BIOCERES 4.91')),
                        (('Soja','15969'), ('Bioceres.Soja','BIOCERES 5.11')),
                        (('Soja','14894'), ('Bioceres.Soja','BIOCERES 5.21')),
                        (('Soja','16769'), ('Bioceres.Soja','BIOCERES 5.92')),
                        (('Colza','9054'), ('Nuseed.Colza','BIOLZA 440')),
                        (('Soja','9920'), ('Buck.Soja','BK42')),
                        (('Soja','18588'), ('BullMark.Soja','BK44P41')),
                        (('Soja','21001'), ('BullMark.Soja','BK44P41 STS')),
                        (('Maíz','13253'), ('Forratec.Maíz','BMR 126')),
                        (('Girasol','20671'), ('Brevant.Girasol','BRV 3304')),
                        (('Soja','21502'), ('Brevant.Soja','BRV 53722 SE')),
                        (('Soja','17857'), ('Brevant.Soja','BRV 54321E')),
                        (('Soja','17855'), ('Brevant.Soja','BRV 54621SE')),
                        (('Soja','19140'), ('Brevant.Soja','BRV 55021SE')),
                        (('Soja','17855'), ('Brevant.Soja','BRV 55621SE')),
                        (('Soja','20211'), ('Brevant.Soja','BRV 56123SCE')),
                        (('Soja','20088'), ('Brevant.Soja','BRV 56222 E')),
                        (('Maíz','19942'), ('Brevant.Maíz','BRV 8380')),
                        (('Maíz','20927'), ('Brevant.Maíz','BRV 8421')),
                        (('Maíz','20112'), ('Brevant.Maíz','BRV 8472')),
                        (('Cebada','21359'), ('Buck.Cebada','BUCK 316')),
                        (('Trigo','9107'), ('Buck.Trigo','BUCK 75 Aniversario')),
                        (('Trigo','20475'), ('Buck.Trigo','BUCK BRAVÍO CL2')),
                        (('Trigo','6246'), ('Nidera.Trigo','Baguette 10')),
                        (('Trigo','6246'), ('Nidera.Trigo','Baguette 11')),
                        (('Trigo','6246'), ('Nidera.Trigo','Baguette 13')),
                        (('Trigo','10602'), ('Nidera.Trigo','Baguette 17')),
                        (('Trigo','10603'), ('Nidera.Trigo','Baguette 18')),
                        (('Trigo','8996'), ('Nidera.Trigo','Baguette 19')),
                        (('Trigo','7815'), ('Nidera.Trigo','Baguette 21')),
                        (('Trigo','10152'), ('Nidera.Trigo','Baguette 30')),
                        (('Trigo','10451'), ('Nidera.Trigo','Baguette 31')),
                        (('Trigo','17153'), ('Nidera.Trigo','Baguette 450')),
                        (('Trigo','13642'), ('Nidera.Trigo','Baguette 501')),
                        (('Trigo','21533'), ('Nidera.Trigo','Baguette 525')),
                        (('Trigo','17979'), ('Nidera.Trigo','Baguette 550')),
                        (('Trigo','13028'), ('Nidera.Trigo','Baguette 560 CL')),
                        (('Trigo','13130'), ('Nidera.Trigo','Baguette 601')),
                        (('Trigo','6246'), ('Nidera.Trigo','Baguette 610')),
                        (('Trigo','17147'), ('Nidera.Trigo','Baguette 620')),
                        (('Trigo','17146'), ('Nidera.Trigo','Baguette 680')),
                        (('Trigo','13642'), ('Nidera.Trigo','Baguette 701')),
                        (('Trigo','16313'), ('Nidera.Trigo','Baguette 750')),
                        (('Trigo','13511'), ('Nidera.Trigo','Baguette 802')),
                        (('Trigo','20447'), ('Nidera.Trigo','Baguette 820')),
                        (('Trigo','9808'), ('Nidera.Trigo','Baguette 9')),
                        (('Trigo','22466'), ('Klein.Trigo','Ballesta')),
                        (('Arveja','15956'), ('Bioseminis.Arveja','Balltrap')),
                        (('Trigo','9942'), ('Buck.Trigo','Baqueano')),
                        (('Alfalfa','5384'), ('Barenbrug.Alfalfa','Baralfa 85')),
                        (('Sorgo','15923'), ('Barenbrug.Sorgo','Bardoble')),
                        (('Cebada','6360'), ('Cargill.Cebada','Barke')),
                        (('Sorgo','15412'), ('Barenbrug.Sorgo','Barkilos')),
                        (('Sorgo','15413'), ('Barenbrug.Sorgo','Barluz')),
                        (('Sorgo','15411'), ('Barenbrug.Sorgo','Barplus')),
                        (('Trigo','15202'), ('Buck.Trigo','Bellaco')),
                        (('Poroto','1761'), ('EEAOC.Poroto','Blanco - TUC 122')),
                        (('Poroto','1760'), ('EEAOC.Poroto','Blanco - TUC 27')),
                        (('Poroto','1762'), ('EEAOC.Poroto','Blanco - TUC 56')),
                        (('Arveja','13345'), ('Inv.Suipachense.Arveja','Bluestar')),
                        (('Trigo','21897'), ('RAGT.Trigo','Borsalino')),
                        (('Trigo','6981'), ('Buck.Trigo','Brasil')),
                        (('Girasol','15435'), ('Buck.Girasol','Buck 355')),
                        (('Girasol','17864'), ('Buck.Girasol','Buck 363')),
                        (('Cebada','563'), ('INTA.Cebada','CALCU INTA ')),
                        (('Cebada','3733'), ('INTA.Cebada','CARLA INTA - MP ')),
                        (('Cebada','755'), ('INTA.Cebada','CAÑUMIL INTA ')),
                        (('Trigo','14470'), ('Aca.Trigo','CEDRO')),
                        (('Sorgo','1622'), ('Agseed.Sorgo','CENTELLA')),
                        (('Sorgo','14574'), ('Agseed.Sorgo','CENTELLA PLUS')),
                        (('Girasol','11698'), ('Advanta.Girasol','CF 101')),
                        (('Girasol','12529'), ('Advanta.Girasol','CF 202')),
                        (('Girasol','8423'), ('Advanta.Girasol','CF 27')),
                        (('Girasol','9312'), ('Advanta.Girasol','CF 31')),
                        (('Trigo','10749'), ('Inia.Trigo','CH 12507')),
                        (('Soja','8235'), ('AGD.Soja','CHAMPAQUI 5.40')),
                        (('Soja','8604'), ('AGD.Soja','CHAMPAQUI 5.80')),
                        (('Soja','8603'), ('AGD.Soja','CHAMPAQUI 5.90')),
                        (('Colza','19722'), ('DSV.Colza','CHIP CL')),
                        (('Trigo','13062'), ('Aca.Trigo','CIPRES')),
                        (('Colza','18066'), ('DSV.Colza','CLARUS')),
                        (('Trigo','18778'), ('Buck.Trigo','COLIHUE')),
                        (('Alfalfa','13329'), ('LosPrados.Alfalfa','CONSTANZA')),
                        (('Soja','9874'), ('AGD.Soja','CQ 4.55')),
                        (('Soja','9865'), ('AGD.Soja','CQ 4.90')),
                        (('Cebada','8689'), ('INTA.Cebada','CRESPA FCA ')),
                        (('Trigo','19048'), ('Buck.Trigo','CUARZO')),
                        (('Alfalfa','10185'), ('CalWestSeeds.Alfalfa','CW 194')),
                        (('Alfalfa','16036'), ('CalWestSeeds.Alfalfa','CW 197')),
                        (('Alfalfa','13339'), ('CalWestSeeds.Alfalfa','CW 660')),
                        (('Alfalfa','14806'), ('CalWestSeeds.Alfalfa','CW 809')),
                        (('Alfalfa','14805'), ('CalWestSeeds.Alfalfa','CW Premium')),
                        (('Soja','19675'), ('Credenz.Soja','CZ 3621 STS')),
                        (('Soja','15041'), ('Credenz.Soja','CZ 4306 B')),
                        (('Soja','18303'), ('Credenz.Soja','CZ 4721 STS')),
                        (('Soja','17084'), ('Credenz.Soja','CZ 4908 IPRO')),
                        (('Soja','17242'), ('Credenz.Soja','CZ 5407 IPRO')),
                        (('Soja','17108'), ('Credenz.Soja','CZ 5907 IPRO')),
                        (('Soja','16604'), ('Credenz.Soja','CZ 6806 IPRO')),
                        (('Soja','14963'), ('Credenz.Soja','CZ 7905 IPRO')),
                        (('Girasol','9615'), ('CriaderoElCencerro.Girasol','Cacique')),
                        (('Trigo','9615'), ('Klein.Trigo','Cacique')),
                        (('Girasol','13842'), ('CriaderoElCencerro.Girasol','Cacique 312')),
                        (('Girasol','19172'), ('CriaderoElCencerro.Girasol','Cacique 320')),
                        (('Cebada','14474'), ('DonMario.Cebada','Cambium')),
                        (('Trigo','14474'), ('DonMario.Trigo','Cambium')),
                        (('Centeno','8695'), ('INTA.Centeno','Camilo')),
                        (('Alfalfa','11003'), ('CriaderoElCencerro.Alfalfa','Candela')),
                        (('Trigo','8501'), ('Klein.Trigo','Capricornio')),
                        (('Alfalfa','9753'), ('Biscayart.Alfalfa','Carabela')),
                        (('Trigo','10050'), ('Klein.Trigo','Carpincho')),
                        (('Trigo','9061'), ('Klein.Trigo','Castor')),
                        (('Trigo','20929'), ('DonMario.Trigo','Catalpa')),
                        (('Trigo','5423'), ('Buck.Trigo','Caudillo')),
                        (('Alfalfa','10886'), ('Biscayart.Alfalfa','Cautiva II')),
                        (('Alfalfa','14294'), ('Biscayart.Alfalfa','Cautiva III')),
                        (('Trigo','15365'), ('DonMario.Trigo','Ceibo')),
                        (('Trigo','9042'), ('Buck.Trigo','Chacarero')),
                        (('Trigo','7746'), ('Klein.Trigo','Chaja')),
                        (('Cebada','17521'), ('Cargill.Cebada','Charles')),
                        (('Trigo','2038'), ('Buck.Trigo','Charrua')),
                        (('Cebada','17572'), ('Cargill.Cebada','Cheers')),
                        (('Trigo','18693'), ('Klein.Trigo','Cien años')),
                        (('Trigo','16333'), ('Buck.Trigo','Claraz')),
                        (('Arveja','1392'), ('INTA.Arveja','Cobri')),
                        (('Poroto','7404'), ('EEAOC.Poroto','Cranberry - TUC 241')),
                        (('Trigo','8601'), ('DonMario.Trigo','Cronox')),
                        (('Trigo','18373'), ('Buck.Trigo','Cumelen')),
                        (('Arveja','15963'), ('Bioseminis.Arveja','Curling')),
                        (('Soja','7269'), ('AgriSeed.Soja','DALIA 390')),
                        (('Soja','7663'), ('AgriSeed.Soja','DALIA 455')),
                        (('Soja','12976'), ('AgriSeed.Soja','DALIA 490')),
                        (('Soja','7664'), ('AgriSeed.Soja','DALIA 500')),
                        (('Soja','8811'), ('AgriSeed.Soja','DALIA 550')),
                        (('Soja','14838'), ('AgriSeed.Soja','DALIA 610')),
                        (('Soja','10581'), ('AgriSeed.Soja','DALIA 620')),
                        (('Soja','10583'), ('AgriSeed.Soja','DALIA 680')),
                        (('Soja','9431'), ('AgriSeed.Soja','DALIA 700')),
                        (('Soja','9432'), ('AgriSeed.Soja','DALIA 740')),
                        (('Soja','9444'), ('AgriSeed.Soja','DALIA 750')),
                        (('Soja','10584'), ('AgriSeed.Soja','DALIA 780')),
                        (('Colza','17133'), ('Nuseed.Colza','DIAMOND')),
                        (('Colza','13528'), ('HighTech.Colza','DIMENSION')),
                        (('Girasol','11926'), ('Syngenta.Girasol','DK 4065')),
                        (('Sorgo','14383'), ('Dekalb.Sorgo','DK 53')),
                        (('Sorgo','7624'), ('Dekalb.Sorgo','DK 61T')),
                        (('Sorgo','13640'), ('Dekalb.Sorgo','DK 64T')),
                        (('Maíz','14516'), ('Dekalb.Maíz','DK 66-10')),
                        (('Maíz','16385'), ('Dekalb.Maíz','DK 69-10')),
                        (('Maíz','13231'), ('Dekalb.Maíz','DK 692')),
                        (('Maíz','9457'), ('Dekalb.Maíz','DK 70-10')),
                        (('Maíz','18769'), ('Dekalb.Maíz','DK 70-20')),
                        (('Maíz','21630'), ('Dekalb.Maíz','DK 72-08')),
                        (('Maíz','14755'), ('Dekalb.Maíz','DK 72-10')),
                        (('Maíz','14802'), ('Dekalb.Maíz','DK 72-50')),
                        (('Maíz','21007'), ('Dekalb.Maíz','DK 72-70')),
                        (('Maíz','22202'), ('Dekalb.Maíz','DK 72-72')),
                        (('Maíz','22135'), ('Dekalb.Maíz','DK 73-03')),
                        (('Maíz','15854'), ('Dekalb.Maíz','DK 73-10')),
                        (('Maíz','19457'), ('Dekalb.Maíz','DK 73-20')),
                        (('Maíz','21606'), ('Dekalb.Maíz','DK 73-30')),
                        (('Maíz','22661'), ('Dekalb.Maíz','DK 74-47')),
                        (('Maíz','21639'), ('Dekalb.Maíz','DK 77-02')),
                        (('Maíz','17233'), ('Dekalb.Maíz','DK 77-10')),
                        (('Maíz','17302'), ('Dekalb.Maíz','DK 78-20')),
                        (('Maíz','15837'), ('Dekalb.Maíz','DK 79-10')),
                        (('Maíz','10976'), ('DonMario.Maíz','DM 2747')),
                        (('Maíz','10187'), ('DonMario.Maíz','DM 2753')),
                        (('Maíz','10176'), ('DonMario.Maíz','DM 2765')),
                        (('Soja','21557'), ('DonMario.Soja','DM 33E22 SE')),
                        (('Soja','20002'), ('DonMario.Soja','DM 33R22')),
                        (('Soja','15267'), ('DonMario.Soja','DM 4014 IPRO')),
                        (('Soja','22560'), ('DonMario.Soja','DM 40E23 SE')),
                        (('Soja','16637'), ('DonMario.Soja','DM 46R18 STS')),
                        (('Soja','17048'), ('DonMario.Soja','DM 46i17 IPRO')),
                        (('Soja','19679'), ('DonMario.Soja','DM 46i20 IPRO STS')),
                        (('Soja','22571'), ('DonMario.Soja','DM 47E23')),
                        (('Soja','6079'), ('DonMario.Soja','DM 4800')),
                        (('Soja','14309'), ('DonMario.Soja','DM 4915 IPRO')),
                        (('Soja','20068'), ('DonMario.Soja','DM 50E22 SE')),
                        (('Soja','17632'), ('DonMario.Soja','DM 52R19')),
                        (('Soja','15937'), ('DonMario.Soja','DM 53i53 IPRO')),
                        (('Soja','18322'), ('DonMario.Soja','DM 55R20 STS')),
                        (('Soja','20105'), ('DonMario.Soja','DM 60K60')),
                        (('Soja','20105'), ('DonMario.Soja','DM 60K60 SCE')),
                        (('Soja','17107'), ('DonMario.Soja','DM 60i62 IPRO')),
                        (('Soja','14228'), ('DonMario.Soja','DM 6262 RSF IPRO')),
                        (('Soja','16480'), ('DonMario.Soja','DM 62R63 STS')),
                        (('Soja','16210'), ('DonMario.Soja','DM 63i64 Garra IPRO STS')),
                        (('Soja','20090'), ('DonMario.Soja','DM 64E64 SE')),
                        (('Soja','21081'), ('DonMario.Soja','DM 64K64 SCE')),
                        (('Soja','17658'), ('DonMario.Soja','DM 66R69 STS')),
                        (('Soja','18349'), ('DonMario.Soja','DM 68K68 SCE')),
                        (('Soja','18349'), ('DonMario.Soja','DM 68K68 STS')),
                        (('Soja','20172'), ('DonMario.Soja','DM 75K75 CE')),
                        (('Soja','16670'), ('DonMario.Soja','DM 75i75 IPRO')),
                        (('Soja','16023'), ('DonMario.Soja','DM 7870 IPRO')),
                        (('Soja','19401'), ('DonMario.Soja','DM 80K80 SCE')),
                        (('Alfalfa','9745'), ('BayaCasal.Alfalfa','DON ENRIQUE')),
                        (('Algodón','15182'), ('Gensus.Algodón','DP 1238')),
                        (('Algodón','8402'), ('Gensus.Algodón','DP 402')),
                        (('Maíz','15676'), ('Forratec.Maíz','DUO 24')),
                        (('Maíz','15793'), ('Forratec.Maíz','DUO 28')),
                        (('Maíz','17271'), ('Forratec.Maíz','DUO 30')),
                        (('Maíz','15230'), ('Forratec.Maíz','DUO 575')),
                        (('Cebada','15653'), ('MalteriaOriental.Cebada','Danielle')),
                        (('Trigo','6816'), ('Klein.Trigo','Delfin')),
                        (('Trigo','17282'), ('Buck.Trigo','Destello')),
                        (('Centeno','8209'), ('INTA.Centeno','Don Alberto')),
                        (('Alfalfa','15565'), ('AdSur.Alfalfa','Don Carlos')),
                        (('Centeno','921'), ('INTA.Centeno','Don Enrique')),
                        (('Trigo','921'), ('Klein.Trigo','Don Enrique')),
                        (('Centeno','12213'), ('INTA.Centeno','Don Ewald')),
                        (('Centeno','3400'), ('INTA.Centeno','Don Guillermo')),
                        (('Centeno','3233'), ('INTA.Centeno','Don Lisandro')),
                        (('Centeno','8209'), ('INTA.Centeno','Don Norberto')),
                        (('Alfalfa','15557'), ('AdSur.Alfalfa','Don Ramon')),
                        (('Tricepiro','3278'), ('INTA.Tricepiro','Don Rene')),
                        (('Soja','11512'), ('Nidera.Soja','EBC 4900 RG')),
                        (('Alfalfa','9752'), ('BayaCasal.Alfalfa','EBC 90')),
                        (('Alfalfa','18018'), ('BayaCasal.Alfalfa','EBC 909 MAX')),
                        (('Maní','18383'), ('CriaderoElCarmen.Maní','EC 214')),
                        (('Maíz','14846'), ('Produsem.Maíz','EG 808 ')),
                        (('Maíz','18356'), ('Produsem.Maíz','EG 809')),
                        (('Girasol','14074'), ('LGseeds.Girasol','ES SHERPA')),
                        (('Arveja','1628'), ('Inia.Arveja','ElRemate')),
                        (('Poroto','12786'), ('INTA.Leales.Poroto','Escarlata INTA')),
                        (('Trigo','6199'), ('Klein.Trigo','Escorpion')),
                        (('Trigo','6817'), ('Klein.Trigo','Escudo')),
                        (('Trigo','6753'), ('Buck.Trigo','Esmeralda')),
                        (('Trigo','4318'), ('Klein.Trigo','Estrella')),
                        (('Cebada','13324'), ('Nidera.Cebada','Explorer')),
                        (('Sorgo','13284'), ('Biscayart.Sorgo','Expreso 131 GR')),
                        (('Sorgo','13285'), ('Biscayart.Sorgo','Expreso 636')),
                        (('Trigo','22465'), ('Klein.Trigo','Extremo')),
                        (('Sorgo','4428'), ('GAPP.Sorgo','F - 700')),
                        (('Sorgo','11010'), ('Forratec.Sorgo','F 1200')),
                        (('Sorgo','13174'), ('Forratec.Sorgo','F 1300 FS')),
                        (('Sorgo','14789'), ('Forratec.Sorgo','F 1307')),
                        (('Sorgo','11186'), ('Forratec.Sorgo','F 1400')),
                        (('Sorgo','14790'), ('Forratec.Sorgo','F 1405')),
                        (('Sorgo','14791'), ('Forratec.Sorgo','F 1497')),
                        (('Sorgo','14757'), ('Forratec.Sorgo','F 2490')),
                        (('Sorgo','14758'), ('Forratec.Sorgo','F 3590')),
                        (('Sorgo','15341'), ('Forratec.Sorgo','F 750 Plus')),
                        (('Trigo','19535'), ('Klein.Trigo','FAVORITO II')),
                        (('Colza','8335'), ('Nuseed.Colza','FILIAL UOFA')),
                        (('Trigo','12979'), ('Agseed.Trigo','FLORIPAN 100')),
                        (('Trigo','13026'), ('Agseed.Trigo','FLORIPAN 200')),
                        (('Trigo','13027'), ('Agseed.Trigo','FLORIPAN 300')),
                        (('Trigo','13027'), ('Agseed.Trigo','FLORIPAN 301')),
                        (('Soja','9416'), ('Credenz.Soja','FN 3.90')),
                        (('Soja','12176'), ('Credenz.Soja','FN 365 AP')),
                        (('Soja','8260'), ('Credenz.Soja','FN 4.10')),
                        (('Soja','11246'), ('Credenz.Soja','FN 4.50')),
                        (('Soja','13573'), ('Credenz.Soja','FN 4.95')),
                        (('Soja','11247'), ('Credenz.Soja','FN 5.20')),
                        (('Soja','12182'), ('Credenz.Soja','FN 5.25')),
                        (('Soja','13569'), ('Credenz.Soja','FN 5.55')),
                        (('Soja','11248'), ('Credenz.Soja','FN 5.75')),
                        (('Soja','12179'), ('Credenz.Soja','FN 6.25')),
                        (('Soja','13563'), ('Credenz.Soja','FN 6.55')),
                        (('Alfalfa','13833'), ('LosPrados.Alfalfa','FRANCESCA')),
                        (('Trigo','20358'), ('Aca.Trigo','FRESNO')),
                        (('Maíz','18696'), ('Forratec.Maíz','FT 2122')),
                        (('Maíz','17556'), ('Forratec.Maíz','FT 4180')),
                        (('Maíz','19608'), ('Forratec.Maíz','FT 4212')),
                        (('Arveja','6873'), ('Bioseminis.Arveja','Facon')),
                        (('Trigo','6083'), ('Buck.Trigo','Farol')),
                        (('Cebada','17586'), ('ACA.Cebada','Fatima')),
                        (('Centeno','8881'), ('INTA.Centeno','Fausto')),
                        (('Garbanzo','13061'), ('Inta.Garbanzo','Felipe')),
                        (('Trigo','13775'), ('Klein.Trigo','Flamenco')),
                        (('Trigo','8092'), ('Klein.Trigo','Flecha')),
                        (('Trigo','20345'), ('Buck.Trigo','Fulgor')),
                        (('Trigo','13998'), ('DonMario.Trigo','Fuste')),
                        (('Sorgo','19862'), ('Gentos.Sorgo','G2.85 BMR')),
                        (('Alfalfa','12939'), ('PGG.Alfalfa','G686')),
                        (('Alfalfa','12938'), ('PGG.Alfalfa','G969')),
                        (('Sorgo','8306'), ('GAPP.Sorgo','GAPP 202')),
                        (('Trigo','15271'), ('Aca.Trigo','GARDELL')),
                        (('Sorgo','14633'), ('GAPP.Sorgo','GG 409')),
                        (('Sorgo','13851'), ('GAPP.Sorgo','GP 307')),
                        (('Sorgo','12485'), ('GAPP.Sorgo','GP 702 BMR')),
                        (('Sorgo','14519'), ('GAPP.Sorgo','GPF 203 BMR')),
                        (('Poroto','6967'), ('INTA.Leales.Poroto','Gateado INTA')),
                        (('Trigo','8502'), ('Klein.Trigo','Gavilan')),
                        (('Sorgo','11019'), ('Gentos.Sorgo','Gentos 125')),
                        (('Sorgo','15470'), ('Gentos.Sorgo','Gentos 130 AA')),
                        (('Sorgo','14776'), ('Gentos.Sorgo','Gentos 135 BMR')),
                        (('Sorgo','11020'), ('Gentos.Sorgo','Gentos 75 BMR')),
                        (('Sorgo','11021'), ('Gentos.Sorgo','Gentos 95 BMR')),
                        (('Trigo','17944'), ('Bioceres.Trigo','Gingko')),
                        (('Trigo','12413'), ('Klein.Trigo','Gladiador')),
                        (('Trigo','10015'), ('Buck.Trigo','Glutino')),
                        (('Maní','7907'), ('CriaderoElCarmen.Maní','Granoleico')),
                        (('Trigo','6685'), ('Buck.Trigo','Guapo')),
                        (('Algodón','18241'), ('Gensus.Algodón','Guaraní INTA')),
                        (('Maní','14642'), ('CriaderoElCarmen.Maní','Guasu (Virginia)')),
                        (('Trigo','7355'), ('Buck.Trigo','Guatimozin')),
                        (('Trigo','17945'), ('Bioceres.Trigo','Guayabo')),
                        (('Algodón','7537'), ('Gensus.Algodón','Guazuncho 2000')),
                        (('Algodón','2126'), ('Gensus.Algodón','Guazuncho 4 INTA')),
                        (('Trigo','10051'), ('Klein.Trigo','Guerrero')),
                        (('Trigo','19553'), ('Klein.Trigo','Géminis')),
                        (('Maíz','9072'), ('DonMario.Maíz','H 2740')),
                        (('Alfalfa','5670'), ('BayaCasal.Alfalfa','HAYGRAZAER')),
                        (('Soja','12761'), ('Horus.Soja','HO 3890')),
                        (('Soja','14242'), ('Horus.Soja','HO 3998')),
                        (('Soja','16418'), ('Horus.Soja','HO 4119 IPRO')),
                        (('Soja','11987'), ('Horus.Soja','HO 4880')),
                        (('Soja','16419'), ('Horus.Soja','HO 4919 IPRO')),
                        (('Soja','13470'), ('Horus.Soja','HO 5010')),
                        (('Soja','14969'), ('Horus.Soja','HO 5310 IPRO')),
                        (('Soja','14258'), ('Horus.Soja','HO 5910')),
                        (('Soja','15910'), ('Horus.Soja','HO 6110 IPRO')),
                        (('Soja','17274'), ('Horus.Soja','HO 6620 IPRO STS')),
                        (('Soja','14326'), ('Horus.Soja','HO 6997 IPRO')),
                        (('Soja','15917'), ('Horus.Soja','HO 7510 IPRO')),
                        (('Soja','16666'), ('Horus.Soja','HO59136 IPRO')),
                        (('Colza','10479'), ('HighTech.Colza','HORNET')),
                        (('Soja','17648'), ('Horus.Soja','HO 74134 IPRO STS')),
                        (('Cebada','14590'), ('INTA.Cebada','HUILEN INTA ')),
                        (('Cebada','777'), ('INTA.Cebada','HUITRU INTA ')),
                        (('Colza','11534'), ('ADVANTA.Colza','HYOLA 433')),
                        (('Colza','14646'), ('ADVANTA.Colza','HYOLA 575 CL')),
                        (('Colza','9102'), ('ADVANTA.Colza','HYOLA 61')),
                        (('Colza','11533'), ('ADVANTA.Colza','HYOLA 76')),
                        (('Colza','15761'), ('ADVANTA.Colza','HYOLA 830 CC')),
                        (('Colza','12567'), ('ADVANTA.Colza','HYOLA 971 CL')),
                        (('Cebada','13713'), ('SABMiller.Cebada','Henrike')),
                        (('Trigo','17283'), ('DonMario.Trigo','Ho Atuel')),
                        (('Trigo','20492'), ('Illinois.Trigo','Hornero')),
                        (('Trigo','3853'), ('Prointa.Trigo','Huenpan')),
                        (('Trigo','17219'), ('Klein.Trigo','Huracan')),
                        (('Maíz','9648'), ('Illinois.Maíz','I 550')),
                        (('Colza','14919'), ('HighTech.Colza','INSPIRATION')),
                        (('Cebada','14591'), ('INTA.Cebada','INTA 7302 ')),
                        (('Trigo','21709'), ('Bioseminis.Trigo','IRUYA HB4')),
                        (('Soja','21503'), ('Illinois.Soja','IS 38.2 SE')),
                        (('Soja','10518'), ('Illinois.Soja','IS 3808')),
                        (('Soja','12766'), ('Illinois.Soja','IS 3909')),
                        (('Soja','12463'), ('Illinois.Soja','IS 4510')),
                        (('Soja','19118'), ('Illinois.Soja','IS 46.1 SE')),
                        (('Soja','21545'), ('Illinois.Soja','IS 46.2 RR/STS')),
                        (('Soja','11741'), ('Illinois.Soja','IS 4777')),
                        (('Soja','21504'), ('Illinois.Soja','IS 48.2 E')),
                        (('Soja','18329'), ('Illinois.Soja','IS 52.0 RR/STS')),
                        (('Soja','10515'), ('Illinois.Soja','IS 5250i')),
                        (('Soja','19147'), ('Illinois.Soja','IS 60.1 SE')),
                        (('Soja','18318'), ('Illinois.Soja','IS 62.1 IPRO STS')),
                        (('Soja','19399'), ('Illinois.Soja','IS-69.2 CE')),
                        (('Cebada','15238'), ('INTA.Cebada','IVANKA INTA ')),
                        (('Girasol','21711'), ('Basf.Girasol','InSun 211B22')),
                        (('Trigo','18026'), ('Bioseminis.Trigo','JACARANDÁ')),
                        (('Trigo','7766'), ('Klein.Trigo','Jabalí')),
                        (('Cebada','15146'), ('Cargill.Cebada','Jennifer')),
                        (('Cebada','15146'), ('SABMiller.Cebada','Jennifer')),
                        (('Trigo','21953'), ('LosGrobo.Trigo','Juramento')),
                        (('Soja','16869'), ('Kumagro.Soja','K 3717 STS')),
                        (('Soja','15175'), ('Kumagro.Soja','K 4001 STS')),
                        (('Soja','16868'), ('Kumagro.Soja','K 4017 STS')),
                        (('Soja','16114'), ('Kumagro.Soja','K 4616 STS')),
                        (('Soja','15657'), ('Kumagro.Soja','K 5102 STS')),
                        (('Soja','16839'), ('Kumagro.Soja','K 6000')),
                        (('Soja','15654'), ('Kumagro.Soja','K 6501 STS')),
                        (('Soja','15656'), ('Kumagro.Soja','K 7102')),
                        (('Maíz','21749'), ('KWS.Maíz','K19-120')),
                        (('Maíz','19891'), ('KWS.Maíz','K9606')),
                        (('Maíz','12876'), ('KWS.Maíz','KM 1301')),
                        (('Maíz','9792'), ('KWS.Maíz','KM 3601')),
                        (('Maíz','15951'), ('KWS.Maíz','KM 3720')),
                        (('Maíz','15501'), ('KWS.Maíz','KM 3800')),
                        (('Maíz','17447'), ('KWS.Maíz','KM 3916 GLS')),
                        (('Maíz','18217'), ('KWS.Maíz','KM 3927')),
                        (('Maíz','14431'), ('KWS.Maíz','KM 4020')),
                        (('Maíz','15949'), ('KWS.Maíz','KM 4200')),
                        (('Maíz','22843'), ('KWS.Maíz','KM 4216')),
                        (('Maíz','12826'), ('KWS.Maíz','KM 4321')),
                        (('Maíz','13447'), ('KWS.Maíz','KM 4360 AS')),
                        (('Maíz','17696'), ('KWS.Maíz','KM 4480')),
                        (('Maíz','14890'), ('KWS.Maíz','KM 4500')),
                        (('Maíz','18969'), ('KWS.Maíz','KM 4580')),
                        (('Sorgo','14707'), ('KWS.Sorgo','KSGR 28')),
                        (('Sorgo','15116'), ('KWS.Sorgo','KSGR 42')),
                        (('Maíz','21750'), ('KWS.Maíz','KWS 13-160')),
                        (('Maíz','21745'), ('KWS.Maíz','KWS 605')),
                        (('Girasol','10950'), ('KWS.Girasol','KWSol 362')),
                        (('Girasol','14709'), ('KWS.Girasol','KWSol 480')),
                        (('Alfalfa','7212'), ('Biscayart.Alfalfa','Key II')),
                        (('Garbanzo','13060'), ('Inta.Garbanzo','Kiara')),
                        (('Arveja','18631'), ('Limagrain.Arveja','Kingfisher')),
                        (('Sorgo','8715'), ('Pemán.Sorgo','Kuntur INTA-Pemán')),
                        (('Cebada','697'), ('INTA.Cebada','LA PLATA BORDEBA FA ')),
                        (('Cebada','631'), ('INTA.Cebada','LA PLATA CAFPTA ')),
                        (('Trigo','22344'), ('Bioseminis.Trigo','LAUREL')),
                        (('Soja','12759'), ('MacroSeed.Soja','LDC 3.7')),
                        (('Soja','11151'), ('MacroSeed.Soja','LDC 3.8  STS')),
                        (('Soja','13546'), ('MacroSeed.Soja','LDC 5.3')),
                        (('Soja','11149'), ('MacroSeed.Soja','LDC 5.9 STS')),
                        (('Soja','14271'), ('MacroSeed.Soja','LDC 8.5')),
                        (('Trigo','7454'), ('Inia.Trigo','LE 2210 (INIA TIJERETA) (TCL)')),
                        (('Trigo','1322'), ('Inia.Trigo','LE 2249 (INIA CHURRINCHE) (TCI)')),
                        (('Trigo','10751'), ('Sursem.Trigo','LE 2330')),
                        (('Trigo','1922'), ('Inia.Trigo','LE 2331 (INIA DON ALBERTO)')),
                        (('Trigo','22588'), ('Inia.Trigo','LE 2333 (INIA CARPINTERO)')),
                        (('Soja','10316'), ('Sursem.Soja','LEO7800')),
                        (('Maíz','22307'), ('LGseeds.Maíz','LG 30.600')),
                        (('Maíz','19602'), ('LGseeds.Maíz','LG 30.680')),
                        (('Maíz','15866'), ('LGseeds.Maíz','LG 30.860')),
                        (('Girasol','18102'), ('LGseeds.Girasol','LG 50.750')),
                        (('Girasol','19006'), ('LGseeds.Girasol','LG 50.760')),
                        (('Girasol','17566'), ('LGseeds.Girasol','LG 5626 HO')),
                        (('Girasol','15345'), ('LGseeds.Girasol','LG 5710')),
                        (('Cebada','21419'), ('Limagrain.Cebada','LG ANDANTE')),
                        (('Trigo','21689'), ('LG.Trigo','LG BAYO')),
                        (('Cebada','22246'), ('Limagrain.Cebada','LG BELCANTO')),
                        (('Trigo','21373'), ('LG.Trigo','LG MORO')),
                        (('Trigo','21802'), ('LG.Trigo','LG PICAZO')),
                        (('Cebada','21089'), ('Limagrain.Cebada','LG Zodiac')),
                        (('Maíz','13897'), ('LGseeds.Maíz','LGSA 30.850')),
                        (('Maíz','18110'), ('LGseeds.Maíz','LGSA 30775')),
                        (('Colza','10480'), ('HighTech.Colza','LILIAN')),
                        (('Maíz','13193'), ('LaTijereta.Maíz','LT 611')),
                        (('Maíz','12479'), ('LaTijereta.Maíz','LT 617')),
                        (('Maíz','10936'), ('LaTijereta.Maíz','LT 618')),
                        (('Maíz','9093'), ('LaTijereta.Maíz','LT 620')),
                        (('Maíz','13996'), ('LaTijereta.Maíz','LT 621')),
                        (('Maíz','11692'), ('LaTijereta.Maíz','LT 622')),
                        (('Maíz','13835'), ('LaTijereta.Maíz','LT 623')),
                        (('Maíz','11766'), ('LaTijereta.Maíz','LT 624')),
                        (('Maíz','9709'), ('LaTijereta.Maíz','LT 625')),
                        (('Maíz','12853'), ('LaTijereta.Maíz','LT 626')),
                        (('Maíz','10935'), ('LaTijereta.Maíz','LT 632')),
                        (('Maíz','19372'), ('LaTijereta.Maíz','LT 718')),
                        (('Maíz','15688'), ('LaTijereta.Maíz','LT 719')),
                        (('Maíz','21344'), ('LaTijereta.Maíz','LT 720')),
                        (('Maíz','19489'), ('LaTijereta.Maíz','LT 721')),
                        (('Maíz','15858'), ('LaTijereta.Maíz','LT 722')),
                        (('Maíz','22306'), ('LaTijereta.Maíz','LT 723')),
                        (('Maíz','21771'), ('LaTijereta.Maíz','LT 725')),
                        (('Maíz','15540'), ('LaTijereta.Maíz','LT 780')),
                        (('Maíz','13331'), ('LaTijereta.Maíz','LT 790')),
                        (('Maíz','17232'), ('LaTijereta.Maíz','LT 795')),
                        (('Maíz','16051'), ('LaTijereta.Maíz','LT 800')),
                        (('Trigo','16413'), ('Klein.Trigo','Lanza')),
                        (('Trigo','14467'), ('LG.Trigo','Lapacho')),
                        (('Trigo','14467'), ('Sursem.Trigo','Lapacho (Nogal 111)')),
                        (('Trigo','22344'), ('Bioceres.Trigo','Laurel')),
                        (('Poroto','10035'), ('INTA.Leales.Poroto','Leales 10 INTA')),
                        (('Poroto','10035'), ('INTA.Leales.Poroto','Leales 14 INTA')),
                        (('Poroto','11108'), ('INTA.Leales.Poroto','Leales 15 INTA')),
                        (('Poroto','10036'), ('INTA.Leales.Poroto','Leales 17 INTA')),
                        (('Poroto','10037'), ('INTA.Leales.Poroto','Leales 22 INTA')),
                        (('Poroto','11109'), ('INTA.Leales.Poroto','Leales 24 INTA')),
                        (('Poroto','10037'), ('INTA.Leales.Poroto','Leales 26 INTA')),
                        (('Poroto','14401'), ('INTA.Leales.Poroto','Leales B30 INTA')),
                        (('Poroto','14402'), ('INTA.Leales.Poroto','Leales B40 INTA')),
                        (('Poroto','14403'), ('INTA.Leales.Poroto','Leales C1 INTA')),
                        (('Poroto','14404'), ('INTA.Leales.Poroto','Leales CR5 INTA')),
                        (('Poroto','14405'), ('INTA.Leales.Poroto','Leales R4 INTA')),
                        (('Trigo','14486'), ('DonMario.Trigo','Lenga')),
                        (('Trigo','13055'), ('DonMario.Trigo','Lenox')),
                        (('Trigo','22463'), ('Klein.Trigo','Leyenda')),
                        (('Trigo','13778'), ('Klein.Trigo','Liebre')),
                        (('Trigo','20676'), ('LosGrobo.Trigo','Limay')),
                        (('Sorgo','14711'), ('LaTijereta.Sorgo','Litio')),
                        (('Trigo','13380'), ('DonMario.Trigo','Lyon')),
                        (('Soja','14994'), ('LaTijereta.Soja','M 5410 IPRO')),
                        (('Maíz','15381'), ('RedSurcos.Maíz','M5890 BT')),
                        (('Soja','14713'), ('ASGROW.Soja','M6210 IPRO')),
                        (('Soja','14714'), ('ASGROW.Soja','M6410 IPRO')),
                        (('Maní','17238'), ('Maniagro.Maní','MA-02')),
                        (('Maní','17237'), ('Maniagro.Maní','MA-121')),
                        (('Maní','17240'), ('Maniagro.Maní','MA-757')),
                        (('Maní','17234'), ('Maniagro.Maní','MA-767')),
                        (('Maní','17235'), ('Maniagro.Maní','MA-88')),
                        (('Maní','17236'), ('Maniagro.Maní','MA-90')),
                        (('Colza','16397'), ('INTA.Colza','MACACHA INTA')),
                        (('Cebada','606'), ('INTA.Cebada','MAGNIF 102 INTA ')),
                        (('Cebada','607'), ('INTA.Cebada','MAGNIF 105 INTA ')),
                        (('Sorgo','8781'), ('Argenetics.Sorgo','MALON')),
                        (('Colza','18065'), ('DSV.Colza','MARATHON')),
                        (('Soja','8788'), ('Sursem.Soja','MARIA 50')),
                        (('Vicia','22134'), ('Zinma.Vicia','MASSA')),
                        (('Alfalfa','5732'), ('LosPrados.Alfalfa','MAYACO')),
                        (('Cebada','7970'), ('INTA.Cebada','MELIPAL INTA ')),
                        (('Soja','8339'), ('Sursem.Soja','MERCEDES 76')),
                        (('Girasol','14594'), ('Brevant.Girasol','MG 303GP')),
                        (('Girasol','13281'), ('Brevant.Girasol','MG 305')),
                        (('Girasol','13920'), ('Brevant.Girasol','MG 360')),
                        (('Sorgo','14701'), ('Nord.Sorgo','MGS 76')),
                        (('Sorgo','14702'), ('Nord.Sorgo','MGS 85')),
                        (('Soja','15050'), ('INTA.Soja','MJ42 STS')),
                        (('Cebada','8688'), ('INTA.Cebada','MOROCHA FCA ')),
                        (('Cebada','11440'), ('MalteriaPampa.Cebada','MP 1012')),
                        (('Cebada','9609'), ('MalteriaPampa.Cebada','MP 1109')),
                        (('Cebada','13014'), ('MalteriaPampa.Cebada','MP 2122')),
                        (('Cebada','7880'), ('MalteriaPampa.Cebada','MP 546')),
                        (('Cebada','7880'), ('MalteriaPampa.Cebada','MP 546 ')),
                        (('Sorgo','15371'), ('Brevant.Sorgo','MS 100')),
                        (('Sorgo','9689'), ('Brevant.Sorgo','MS 102')),
                        (('Sorgo','12772'), ('Brevant.Sorgo','MS 105')),
                        (('Sorgo','15276'), ('Brevant.Sorgo','MS 106')),
                        (('Sorgo','7890'), ('Brevant.Sorgo','MS 108')),
                        (('Sorgo','9244'), ('Brevant.Sorgo','MS 109')),
                        (('Soja','14868'), ('MacroSeed.Soja','MS 4.0 IPRO')),
                        (('Soja','14881'), ('MacroSeed.Soja','MS 4.9 IPRO')),
                        (('Soja','14961'), ('MacroSeed.Soja','MS 6.3 IPRO')),
                        (('Soja','15914'), ('MacroSeed.Soja','MS 6.9 IPRO')),
                        (('Soja','14327'), ('MacroSeed.Soja','MS 7.4 IPRO')),
                        (('Trigo','16336'), ('MacroSeed.Trigo','MS INTA 116')),
                        (('Trigo','18706'), ('MacroSeed.Trigo','MS INTA 119')),
                        (('Trigo','16328'), ('MacroSeed.Trigo','MS INTA 415')),
                        (('Trigo','16335'), ('MacroSeed.Trigo','MS INTA 416')),
                        (('Trigo','21350'), ('MacroSeed.Trigo','MS INTA 521')),
                        (('Trigo','13715'), ('MacroSeed.Trigo','MS INTA 615')),
                        (('Trigo','15746'), ('MacroSeed.Trigo','MS INTA 815')),
                        (('Maíz','17518'), ('Stine.Maíz','MST 120-19')),
                        (('Trigo','9523'), ('Buck.Trigo','Malevo')),
                        (('Trigo','11566'), ('Buck.Trigo','Mangrullo')),
                        (('Trigo','7279'), ('Klein.Trigo','Martillo')),
                        (('Trigo','7744'), ('Buck.Trigo','Mataco')),
                        (('Sorgo','10972'), ('Tobin.Sorgo','Matrero')),
                        (('Arveja','14615'), ('AFA.Arveja','Meadow')),
                        (('Trigo','17140'), ('Klein.Trigo','Mercurio')),
                        (('Trigo','10658'), ('Buck.Trigo','Meteoro')),
                        (('Cebada','18905'), ('Produsem.Cebada','Militza INTA')),
                        (('Trigo','17218'), ('Klein.Trigo','Minerva')),
                        (('Cebada','17571'), ('CyMQuilmes.Cebada','Montoya')),
                        (('Girasol','5861'), ('Nidera.Girasol','N Aromo105')),
                        (('Girasol','11259'), ('Nidera.Girasol','N Aromo11')),
                        (('Girasol','8707'), ('Nidera.Girasol','N Paraiso102')),
                        (('Soja','8337'), ('Sursem.Soja','N49R')),
                        (('Cebada','15936'), ('INTA.Cebada','NELIDA INTA ')),
                        (('Soja','22613'), ('NeoGen.Soja','NEO 35S23 SE')),
                        (('Soja','21501'), ('NeoGen.Soja','NEO 40S22 SE')),
                        (('Soja','20003'), ('NeoGen.Soja','NEO 45S22 RR STS')),
                        (('Soja','20043'), ('NeoGen.Soja','NEO 46S22 SE')),
                        (('Alfalfa','6668'), ('LosPrados.Alfalfa','NEVADA')),
                        (('Maíz','16352'), ('Brevant.Maíz','NEXT 20.6')),
                        (('Maíz','16353'), ('Brevant.Maíz','NEXT 22.6')),
                        (('Maíz','17583'), ('Brevant.Maíz','NEXT 25.8')),
                        (('Maíz','10494'), ('Syngenta.Maíz','NK 135')),
                        (('Soja','9426'), ('Syngenta.Soja','NK 32-00')),
                        (('Soja','8885'), ('Syngenta.Soja','NK 34-00')),
                        (('Soja','10455'), ('Syngenta.Soja','NK 35-00')),
                        (('Soja','9333'), ('Syngenta.Soja','NK 37-00')),
                        (('Soja','11075'), ('Syngenta.Soja','NK 38-00')),
                        (('Soja','11196'), ('Syngenta.Soja','NK 39-00')),
                        (('Girasol','20445'), ('Syngenta.Girasol','NK 3969')),
                        (('Soja','9429'), ('Syngenta.Soja','NK 43-00')),
                        (('Soja','8888'), ('Syngenta.Soja','NK 47-00')),
                        (('Soja','10454'), ('Syngenta.Soja','NK 48-00')),
                        (('Soja','18227'), ('Syngenta.Soja','NK 51X22 IPRO STS')),
                        (('Soja','19674'), ('Syngenta.Soja','NK 52x21 STS')),
                        (('Soja','18218'), ('Syngenta.Soja','NK 60x21 IPRO STS')),
                        (('Maíz','20484'), ('Syngenta.Maíz','NK 800')),
                        (('Maíz','20486'), ('Syngenta.Maíz','NK 842')),
                        (('Maíz','8590'), ('Syngenta.Maíz','NK 870')),
                        (('Maíz','10137'), ('Syngenta.Maíz','NK 880')),
                        (('Maíz','19483'), ('Syngenta.Maíz','NK 885')),
                        (('Maíz','19482'), ('Syngenta.Maíz','NK 890')),
                        (('Maíz','12332'), ('Syngenta.Maíz','NK 900')),
                        (('Maíz','12331'), ('Syngenta.Maíz','NK 907')),
                        (('Maíz','10195'), ('Syngenta.Maíz','NK 910')),
                        (('Maíz','8129'), ('Syngenta.Maíz','NK 940')),
                        (('Soja','8889'), ('Syngenta.Soja','NK COKER 6.6')),
                        (('Soja','7654'), ('Syngenta.Soja','NK COKER 6.8 RR')),
                        (('Soja','7659'), ('Syngenta.Soja','NK COKER 7.5 R')),
                        (('Soja','9883'), ('Syngenta.Soja','NK COKER 8.0')),
                        (('Soja','7219'), ('Syngenta.Soja','NK MIREYA 4.2')),
                        (('Soja','7218'), ('Syngenta.Soja','NK PICASA 4.0')),
                        (('Soja','7599'), ('Sursem.Soja','NM55R')),
                        (('Soja','7598'), ('Sursem.Soja','NM70R')),
                        (('Girasol','19442'), ('Nidera.Girasol','NS 1109')),
                        (('Girasol','20444'), ('Nidera.Girasol','NS 1113')),
                        (('Soja','12707'), ('Nidera.Soja','NS 2018')),
                        (('Soja','12948'), ('Nidera.Soja','NS 2632')),
                        (('Soja','11272'), ('Nidera.Soja','NS 3215')),
                        (('Soja','15529'), ('Nidera.Soja','NS 3220 STS')),
                        (('Soja','14721'), ('Nidera.Soja','NS 3313')),
                        (('Soja','15464'), ('Nidera.Soja','NS 3809 IPRO')),
                        (('Soja','19675'), ('Nidera.Soja','NS 3821 STS')),
                        (('Soja','14722'), ('Nidera.Soja','NS 3909')),
                        (('Soja','11273'), ('Nidera.Soja','NS 4009')),
                        (('Soja','15586'), ('Nidera.Soja','NS 4309')),
                        (('Soja','13585'), ('Nidera.Soja','NS 4313')),
                        (('Soja','15158'), ('Nidera.Soja','NS 4319 IPRO')),
                        (('Soja','12952'), ('Nidera.Soja','NS 4611 STS')),
                        (('Soja','19679'), ('Nidera.Soja','NS 4621 IPRO STS')),
                        (('Soja','11275'), ('Nidera.Soja','NS 4903')),
                        (('Soja','13589'), ('Nidera.Soja','NS 4955')),
                        (('Soja','12198'), ('Nidera.Soja','NS 4997')),
                        (('Soja','19839'), ('Nidera.Soja','NS 5023 STS')),
                        (('Soja','16230'), ('Nidera.Soja','NS 5028 STS')),
                        (('Soja','18842'), ('Nidera.Soja','NS 5030 IPRO STS')),
                        (('Soja','12956'), ('Nidera.Soja','NS 5230')),
                        (('Soja','13591'), ('Nidera.Soja','NS 5258')),
                        (('Soja','15443'), ('Nidera.Soja','NS 5419 IPRO')),
                        (('Soja','19677'), ('Nidera.Soja','NS 5421 STS')),
                        (('Soja','12201'), ('Nidera.Soja','NS 6002')),
                        (('Soja','17543'), ('Nidera.Soja','NS 6120 IPRO')),
                        (('Soja','20030'), ('Nidera.Soja','NS 6212 IPRO')),
                        (('Soja','14349'), ('Nidera.Soja','NS 6248')),
                        (('Soja','12199'), ('Nidera.Soja','NS 6267')),
                        (('Soja','14356'), ('Nidera.Soja','NS 6419 IPRO')),
                        (('Soja','14348'), ('Nidera.Soja','NS 6483')),
                        (('Soja','16228'), ('Nidera.Soja','NS 6538 IPRO')),
                        (('Soja','18226'), ('Nidera.Soja','NS 6721 IPRO STS')),
                        (('Soja','16229'), ('Nidera.Soja','NS 6859 IPRO')),
                        (('Soja','14357'), ('Nidera.Soja','NS 6909 IPRO')),
                        (('Soja','14704'), ('Nidera.Soja','NS 7209 IPRO')),
                        (('Soja','12194'), ('Nidera.Soja','NS 7211')),
                        (('Soja','14705'), ('Nidera.Soja','NS 7300 IPRO')),
                        (('Maíz','20488'), ('Nidera.Maíz','NS 7621')),
                        (('Soja','15093'), ('Nidera.Soja','NS 7709 IPRO')),
                        (('Soja','14158'), ('Nidera.Soja','NS 7711 IPRO')),
                        (('Maíz','20483'), ('Nidera.Maíz','NS 7800')),
                        (('Soja','15589'), ('Nidera.Soja','NS 7809')),
                        (('Maíz','18119'), ('Nidera.Maíz','NS 7818')),
                        (('Maíz','19456'), ('Nidera.Maíz','NS 7921')),
                        (('Soja','10005'), ('Nidera.Soja','NS 8004')),
                        (('Soja','16877'), ('Nidera.Soja','NS 8018 IPRO STS')),
                        (('Soja','12196'), ('Nidera.Soja','NS 8262')),
                        (('Soja','12193'), ('Nidera.Soja','NS 8282')),
                        (('Soja','15584'), ('Nidera.Soja','NS 8288 STS')),
                        (('Soja','20028'), ('Nidera.Soja','NS 7922')),
                        (('Girasol','19628'), ('Brevant.Girasol','NTO2.5')),
                        (('Girasol','15882'), ('Brevant.Girasol','NTO3.6')),
                        (('Girasol','12742'), ('Brevant.Girasol','NTO4.0')),
                        (('Sorgo','18756'), ('NuSeed.Sorgo','NU 441 IG')),
                        (('Sorgo','12589'), ('NuSeed.Sorgo','NU Don Jacinto')),
                        (('Sorgo','17036'), ('NuSeed.Sorgo','NU Nugrain 440')),
                        (('Sorgo','17292'), ('NuSeed.Sorgo','NU Nugrass 900')),
                        (('Sorgo','17293'), ('NuSeed.Sorgo','NU Nusil500')),
                        (('Sorgo','17290'), ('NuSeed.Sorgo','NU Nusil600')),
                        (('Sorgo','12612'), ('NuSeed.Sorgo','NU Spring T60')),
                        (('Sorgo','16417'), ('NuSeed.Sorgo','NU SummerII')),
                        (('Colza','15220'), ('Nuseed.Colza','NUVETTE 2286')),
                        (('Maíz','19620'), ('MacroSeed.Maíz','NXM 1122')),
                        (('Maíz','19620'), ('NexSem.Maíz','NXM 1122')),
                        (('Poroto','1699'), ('EEAOC.Poroto','Negro - BAT 304')),
                        (('Poroto','1698'), ('EEAOC.Poroto','Negro - DOR 157')),
                        (('Poroto','1143'), ('EEAOC.Poroto','Negro - DOR 41')),
                        (('Poroto','11502'), ('EEAOC.Poroto','Negro - TUC 300')),
                        (('Poroto','3602'), ('EEAOC.Poroto','Negro - TUC 390')),
                        (('Poroto','3603'), ('EEAOC.Poroto','Negro - TUC 500')),
                        (('Poroto','7402'), ('EEAOC.Poroto','Negro - TUC 510')),
                        (('Poroto','9957'), ('EEAOC.Poroto','Negro - TUC 550')),
                        (('Sorgo','11141'), ('Biscayart.Sorgo','NiagaSil')),
                        (('Vicia','17721'), ('Gentos.Vicia','Nitro Max')),
                        (('Sorgo','11258'), ('Biscayart.Sorgo','Niágara BL')),
                        (('Sorgo','18185'), ('Biscayart.Sorgo','Niágara III')),
                        (('Trigo','9190'), ('Sursem.Trigo','Nogal')),
                        (('Maíz','16525'), ('Nord.Maíz','Nord ACRUX')),
                        (('Maíz','14537'), ('Nord.Maíz','Nord AVALON')),
                        (('Maíz','14980'), ('Nord.Maíz','Nord BALTOS')),
                        (('Maíz','16871'), ('Nord.Maíz','Nord BORAX')),
                        (('Maíz','15794'), ('Nord.Maíz','Nord BUYAN')),
                        (('Sorgo','10880'), ('Nord.Sorgo','Nord Palatable 10 BMR')),
                        (('Sorgo','14230'), ('Nord.Sorgo','Nord Palatable 10 MAX')),
                        (('Girasol','16448'), ('Nord.Girasol','Nord Sungro 70')),
                        (('Girasol','16448'), ('Nord.Girasol','Nord Sungro 80')),
                        (('Maíz','17936'), ('Nord.Maíz','Nord ZEFIR')),
                        (('Trigo','4924'), ('Buck.Trigo','Norteño')),
                        (('Garbanzo','4924'), ('Inta.Garbanzo','Norteño')),
                        (('Maíz','18054'), ('Nuseed.Maíz','Nucorn 2650')),
                        (('Maíz','17321'), ('Nuseed.Maíz','Nucorn 2881')),
                        (('Sorgo','18030'), ('NuSeed.Sorgo','Nugrain 300')),
                        (('Algodón','11949'), ('Gensus.Algodón','Nuopal')),
                        (('Girasol','14578'), ('NuSeed.Girasol','Nusol 2100')),
                        (('Girasol','14580'), ('NuSeed.Girasol','Nusol 4100')),
                        (('Girasol','17340'), ('NuSeed.Girasol','Nusol 4120')),
                        (('Girasol','20600'), ('NuSeed.Girasol','Nusol 4145')),
                        (('Girasol','17340'), ('NuSeed.Girasol','Nusol 4170')),
                        (('Girasol','16433'), ('NuSeed.Girasol','Nusol 4510')),
                        (('Trigo','11588'), ('Klein.Trigo','Nutria')),
                        (('Cebada','454'), ('INTA.Cebada','OLIVEROS LITORAL SAG ')),
                        (('Girasol','19885'), ('OrigoSemillas.Girasol','ORI 730 CL')),
                        (('Trigo','16048'), ('Buck.Trigo','Odisseo')),
                        (('Trigo','8488'), ('DonMario.Trigo','Onix')),
                        (('Cebada','15720'), ('Limagrain.Cebada','Overture')),
                        (('Maíz','20674'), ('Pioneer.Maíz','P 0622')),
                        (('Maíz','15312'), ('Pioneer.Maíz','P 1778')),
                        (('Maíz','15309'), ('Pioneer.Maíz','P 1780')),
                        (('Maíz','19132'), ('Pioneer.Maíz','P 1804')),
                        (('Maíz','18100'), ('Pioneer.Maíz','P 1815')),
                        (('Maíz','15306'), ('Pioneer.Maíz','P 1833')),
                        (('Maíz','12576'), ('Pioneer.Maíz','P 1845')),
                        (('Maíz','15307'), ('Pioneer.Maíz','P 1952')),
                        (('Maíz','13198'), ('Pioneer.Maíz','P 1979')),
                        (('Maíz','17228'), ('Pioneer.Maíz','P 2005')),
                        (('Maíz','21560'), ('Pioneer.Maíz','P 2021')),
                        (('Maíz','13200'), ('Pioneer.Maíz','P 2038')),
                        (('Maíz','13199'), ('Pioneer.Maíz','P 2049')),
                        (('Maíz','10877'), ('Pioneer.Maíz','P 2053')),
                        (('Maíz','12575'), ('Pioneer.Maíz','P 2058')),
                        (('Maíz','13865'), ('Pioneer.Maíz','P 2069')),
                        (('Maíz','17230'), ('Pioneer.Maíz','P 2089')),
                        (('Maíz','18788'), ('Pioneer.Maíz','P 2103')),
                        (('Maíz','17229'), ('Pioneer.Maíz','P 2109')),
                        (('Maíz','16467'), ('Pioneer.Maíz','P 2151')),
                        (('Maíz','19278'), ('Pioneer.Maíz','P 2167')),
                        (('Maíz','20422'), ('Pioneer.Maíz','P 2297')),
                        (('Maíz','19115'), ('Pioneer.Maíz','P 2353')),
                        (('Maíz','11920'), ('Pioneer.Maíz','P 30B39')),
                        (('Maíz','10978'), ('Pioneer.Maíz','P 30F35')),
                        (('Maíz','18278'), ('Pioneer.Maíz','P 30F53')),
                        (('Maíz','9816'), ('Pioneer.Maíz','P 30P70')),
                        (('Maíz','6882'), ('Pioneer.Maíz','P 30R76')),
                        (('Maíz','10159'), ('Pioneer.Maíz','P 31A08')),
                        (('Maíz','8560'), ('Pioneer.Maíz','P 31B18')),
                        (('Maíz','9155'), ('Pioneer.Maíz','P 31D06')),
                        (('Maíz','8562'), ('Pioneer.Maíz','P 31F25')),
                        (('Maíz','12598'), ('Pioneer.Maíz','P 31G71')),
                        (('Maíz','9626'), ('Pioneer.Maíz','P 31P77')),
                        (('Maíz','9628'), ('Pioneer.Maíz','P 31R31')),
                        (('Maíz','8140'), ('Pioneer.Maíz','P 31Y04')),
                        (('Maíz','11745'), ('Pioneer.Maíz','P 31Y05')),
                        (('Maíz','7364'), ('Pioneer.Maíz','P 32F07')),
                        (('Maíz','16587'), ('Pioneer.Maíz','P 32R48')),
                        (('Maíz','15860'), ('Pioneer.Maíz','P 38A57')),
                        (('Maíz','10949'), ('Pioneer.Maíz','P 39B77')),
                        (('Soja','19129'), ('Pioneer.Soja','P 43A04SE')),
                        (('Soja','19130'), ('Pioneer.Soja','P 46A03SE')),
                        (('Soja','17720'), ('Pioneer.Soja','P 50A02E')),
                        (('Soja','20210'), ('Pioneer.Soja','P 60A01 SCE')),
                        (('Girasol','14653'), ('Pioneer.Girasol','P 64ll95')),
                        (('Girasol','9287'), ('Pioneer.Girasol','P 65A25')),
                        (('Soja','20170'), ('Pioneer.Soja','P 75A06 SCE')),
                        (('Soja','19432'), ('Pioneer.Soja','P 80A02 SCE')),
                        (('Sorgo','14026'), ('Pioneer.Sorgo','P 80T25')),
                        (('Sorgo','10012'), ('Pioneer.Sorgo','P 81G67')),
                        (('Sorgo','7833'), ('Pioneer.Sorgo','P 84G62')),
                        (('Soja','6034'), ('Pioneer.Soja','P 93B34 RR')),
                        (('Soja','8295'), ('Pioneer.Soja','P 93B36 RR')),
                        (('Soja','7872'), ('Pioneer.Soja','P 93B85 RR')),
                        (('Soja','12174'), ('Pioneer.Soja','P 93M70 RR')),
                        (('Soja','10509'), ('Pioneer.Soja','P 93M92 RR')),
                        (('Soja','10510'), ('Pioneer.Soja','P 93M96 RR')),
                        (('Soja','7873'), ('Pioneer.Soja','P 94B54 RR')),
                        (('Soja','7052'), ('Pioneer.Soja','P 94B73 RR')),
                        (('Soja','9969'), ('Pioneer.Soja','P 94M30 RR')),
                        (('Soja','9469'), ('Pioneer.Soja','P 94M40 RR')),
                        (('Soja','10511'), ('Pioneer.Soja','P 94M80 RR')),
                        (('Maíz','20678'), ('Pioneer.Maíz','P 9946')),
                        (('Girasol','9619'), ('Pioneer.Girasol','P PAN7047')),
                        (('Alfalfa','1690'), ('Barenbrug.Alfalfa','P-105')),
                        (('Trigo','6745'), ('Prointa.Trigo','P. Don Umberto')),
                        (('Trigo','2057'), ('Prointa.Trigo','P. Federal')),
                        (('Trigo','6649'), ('Prointa.Trigo','P. Gaucho')),
                        (('Trigo','2905'), ('Prointa.Trigo','P. Imperial')),
                        (('Trigo','6746'), ('Prointa.Trigo','P. Molinero')),
                        (('Trigo','2118'), ('Prointa.Trigo','P. Oasis')),
                        (('Trigo','3700'), ('Prointa.Trigo','P. Puntal')),
                        (('Trigo','21707'), ('Bioseminis.Trigo','PARANÁ HB4')),
                        (('Alfalfa','19166'), ('PGG.Alfalfa','PGW 931')),
                        (('Colza','18068'), ('DSV.Colza','PHOENIX')),
                        (('Maíz','16965'), ('Pop.Arg.Maíz','POP 4512')),
                        (('Trigo','21627'), ('Buck.Trigo','PRETAL')),
                        (('Colza','13529'), ('HighTech.Colza','PRIMUS')),
                        (('Alfalfa','7259'), ('Produsem.Alfalfa','PRO INTA LUJÁN')),
                        (('Alfalfa','10049'), ('Produsem.Alfalfa','PRO INTA MORA')),
                        (('Alfalfa','7258'), ('Produsem.Alfalfa','PRO INTA PATRICIA')),
                        (('Sorgo','9045'), ('AdSur.Sorgo','PRODUCTOR 401')),
                        (('Sorgo','16648'), ('AdSur.Sorgo','PRODUCTOR-161BL')),
                        (('Sorgo','16647'), ('AdSur.Sorgo','PRODUCTOR-162FS')),
                        (('Sorgo','14796'), ('Pemán.Sorgo','PS 500 BMR')),
                        (('Sorgo','13496'), ('Pemán.Sorgo','PS 55')),
                        (('Colza','9629'), ('HighTech.Colza','PULSAR')),
                        (('Arveja','13122'), ('ARVES.Arveja','Pampa')),
                        (('Trigo','10788'), ('Klein.Trigo','Pantera')),
                        (('Soja','12258'), ('INTA.Soja','Paraná 5500')),
                        (('Soja','12257'), ('INTA.Soja','Paraná 6200')),
                        (('Soja','8048'), ('INTA.Soja','Paraná 629')),
                        (('Soja','8049'), ('INTA.Soja','Paraná 661')),
                        (('Alfalfa','10273'), ('Peman.Alfalfa','Patriarca')),
                        (('Alfalfa','12836'), ('Barenbrug.Alfalfa','Pegasis')),
                        (('Trigo','4997'), ('Klein.Trigo','Pegaso')),
                        (('Trigo','18146'), ('DonMario.Trigo','Pehuen')),
                        (('Trigo','19466'), ('Buck.Trigo','Peregrino')),
                        (('Trigo','18644'), ('Buck.Trigo','Perla')),
                        (('Sorgo','11139'), ('Biscayart.Sorgo','Pilcomayo 2')),
                        (('Sorgo','18576'), ('Biscayart.Sorgo','Pilcomayo III')),
                        (('Sorgo','15876'), ('Pemán.Sorgo','Pitavá')),
                        (('Maíz','17363'), ('BMHSemillas.Maíz','Piumassimo')),
                        (('Maíz','19889'), ('BMHSemillas.Maíz','Piusoldi')),
                        (('Trigo','13682'), ('Buck.Trigo','Pleno')),
                        (('Algodón','18240'), ('Gensus.Algodón','Pora 3 INTA')),
                        (('Trigo','17994'), ('Klein.Trigo','Potro')),
                        (('Cebada','12906'), ('Cargill.Cebada','Prestige')),
                        (('Trigo','16310'), ('Klein.Trigo','Prometeo')),
                        (('Maní','11190'), ('CriaderoElCarmen.Maní','Pronto')),
                        (('Maní','11190'), ('CriaderoElCarmen.Maní','Pronto AO')),
                        (('Trigo','9621'), ('Proseme.Trigo','Proseme Cannizzo')),
                        (('Trigo','9622'), ('Proseme.Trigo','Proseme Ciccio')),
                        (('Trigo','9623'), ('Proseme.Trigo','Proseme Coloseo')),
                        (('Trigo','8093'), ('Klein.Trigo','Proteo')),
                        (('Trigo','10017'), ('Buck.Trigo','Puelche')),
                        (('Maíz','20809'), ('Qseeds.Maíz','QS 72-01')),
                        (('Maíz','19776'), ('Qseeds.Maíz','QS 73-01')),
                        (('Maíz','18162'), ('Qseeds.Maíz','QS 75-01')),
                        (('Maíz','18163'), ('Qseeds.Maíz','QS 86-01')),
                        (('Cebada','1006'), ('CyMQuilmes.Cebada','QUILMES 271 ')),
                        (('Cebada','1007'), ('CyMQuilmes.Cebada','QUILMES ALFA ')),
                        (('Cebada','5872'), ('CyMQuilmes.Cebada','QUILMES AYELEN ')),
                        (('Cebada','11304'), ('CyMQuilmes.Cebada','QUILMES CARISMA ')),
                        (('Cebada','1141'), ('CyMQuilmes.Cebada','QUILMES CENTAURO ')),
                        (('Cebada','8669'), ('CyMQuilmes.Cebada','QUILMES KUYEN ')),
                        (('Cebada','5069'), ('CyMQuilmes.Cebada','QUILMES PAINE ')),
                        (('Cebada','3428'), ('CyMQuilmes.Cebada','QUILMES PALOMAR ')),
                        (('Cebada','1005'), ('CyMQuilmes.Cebada','QUILMES PAMPA ')),
                        (('Centeno','5399'), ('INTA.Centeno','Quehue')),
                        (('Trigo','19549'), ('RAGT.Trigo','Quiriko')),
                        (('Soja','11113'), ('CriaSantaRosa.Soja','RA 334')),
                        (('Soja','11111'), ('CriaSantaRosa.Soja','RA 338')),
                        (('Soja','12835'), ('CriaSantaRosa.Soja','RA 349')),
                        (('Soja','9895'), ('CriaSantaRosa.Soja','RA 426')),
                        (('Soja','12147'), ('CriaSantaRosa.Soja','RA 437')),
                        (('Soja','13440'), ('CriaSantaRosa.Soja','RA 444')),
                        (('Soja','13522'), ('CriaSantaRosa.Soja','RA 449')),
                        (('Soja','10285'), ('CriaSantaRosa.Soja','RA 524')),
                        (('Soja','11241'), ('CriaSantaRosa.Soja','RA 532')),
                        (('Soja','11242'), ('CriaSantaRosa.Soja','RA 536')),
                        (('Soja','10446'), ('CriaSantaRosa.Soja','RA 538')),
                        (('Soja','13562'), ('CriaSantaRosa.Soja','RA 541')),
                        (('Soja','13518'), ('CriaSantaRosa.Soja','RA 545')),
                        (('Soja','12151'), ('CriaSantaRosa.Soja','RA 549')),
                        (('Soja','14260'), ('CriaSantaRosa.Soja','RA 550')),
                        (('Soja','13520'), ('CriaSantaRosa.Soja','RA 556')),
                        (('Soja','15393'), ('CriaSantaRosa.Soja','RA 5715 IPRO')),
                        (('Soja','12150'), ('CriaSantaRosa.Soja','RA 644')),
                        (('Soja','11243'), ('CriaSantaRosa.Soja','RA 732')),
                        (('Soja','9890'), ('CriaSantaRosa.Soja','RA 733')),
                        (('Soja','13436'), ('CriaSantaRosa.Soja','RA 744')),
                        (('Soja','13561'), ('CriaSantaRosa.Soja','RA 844')),
                        (('Soja','7028'), ('Sursem.Soja','RAFAELA58')),
                        (('Cebada','11725'), ('INTA.Cebada','RAYEN INTA ')),
                        (('Maíz','19802'), ('Brevant.Maíz','RFG 22')),
                        (('Maíz','19802'), ('Brevant.Maíz','RFG22RRE')),
                        (('Girasol','21515'), ('RAGT.Girasol','RGT CHARLLOTTE')),
                        (('Girasol','13932'), ('RAGT.Girasol','RGT KAPLLAN')),
                        (('Girasol','13936'), ('RAGT.Girasol','RGT MOOGLLI')),
                        (('Girasol','700'), ('RAGT.Girasol','RGT OBELLISCO')),
                        (('Girasol','13931'), ('RAGT.Girasol','RGT SIKLLOS')),
                        (('Girasol','13933'), ('RAGT.Girasol','RGT VELLOX')),
                        (('Colza','9688'), ('Nuseed.Colza','RIVETTE')),
                        (('Soja','12161'), ('Sursem.Soja','RM 6900')),
                        (('Soja','9904'), ('Sursem.Soja','RMO4637')),
                        (('Soja','9908'), ('Sursem.Soja','RMO58')),
                        (('Soja','9901'), ('Sursem.Soja','RMO75')),
                        (('Soja','12658'), ('Sursem.Soja','RMO77')),
                        (('Soja','12588'), ('Sursem.Soja','RMO805')),
                        (('Alfalfa','13630'), ('LosPrados.Alfalfa','ROBERTA')),
                        (('Trigo','18645'), ('Buck.Trigo','Resplandor')),
                        (('Arveja','14839'), ('Bioseminis.Arveja','Reussite')),
                        (('Girasol','12473'), ('NuSeed.Girasol','Rhino')),
                        (('Centeno','15277'), ('INTA.Centeno','Ricardo')),
                        (('Trigo','13089'), ('Klein.Trigo','Roble')),
                        (('Poroto','1700'), ('EEAOC.Poroto','Rojo - PVAD 1101')),
                        (('Poroto','1701'), ('EEAOC.Poroto','Rojo - PVAD 1111')),
                        (('Poroto','7403'), ('EEAOC.Poroto','Rojo - TUC 310')),
                        (('Arveja','20651'), ('ARVES.Arveja','Rosita')),
                        (('Cebada','15873'), ('INTA.Cebada','SARA INTA ')),
                        (('Cebada','15952'), ('INTA.Cebada','SILERA INTA ')),
                        (('Trigo','9537'), ('Sursem.Trigo','SIRIRI')),
                        (('Colza','10478'), ('HighTech.Colza','SITRO')),
                        (('Soja','12866'), ('Klein.Soja','SK 3.5')),
                        (('Soja','11770'), ('Klein.Soja','SK 3.8')),
                        (('Soja','11146'), ('Klein.Soja','SK 4.7')),
                        (('Maíz','17512'), ('Nidera.Maíz','SMF 8007')),
                        (('Maíz','19486'), ('Nidera.Maíz','SMF 8080')),
                        (('Trigo','16445'), ('Sursem.Trigo','SN 90')),
                        (('Soja','8909'), ('Syngenta.Soja','SP 3900')),
                        (('Soja','9160'), ('Syngenta.Soja','SP 4500')),
                        (('Soja','9845'), ('Syngenta.Soja','SP 4X0')),
                        (('Soja','14879'), ('Syngenta.Soja','SP 4X3 IPRO')),
                        (('Soja','10560'), ('Syngenta.Soja','SP 4X4')),
                        (('Soja','15379'), ('Syngenta.Soja','SP 4X6 IPRO')),
                        (('Soja','11637'), ('Syngenta.Soja','SP 4X99')),
                        (('Soja','11185'), ('Syngenta.Soja','SP 5x2')),
                        (('Soja','10093'), ('Syngenta.Soja','SP 5x5')),
                        (('Soja','11178'), ('Syngenta.Soja','SP 5x9')),
                        (('Soja','13501'), ('Syngenta.Soja','SP 6X1')),
                        (('Soja','9866'), ('Syngenta.Soja','SP 6X2')),
                        (('Soja','9869'), ('Syngenta.Soja','SP 7x0')),
                        (('Soja','9856'), ('Syngenta.Soja','SP 8x0')),
                        (('Soja','12789'), ('Syngenta.Soja','SP 8x8')),
                        (('Maíz','7334'), ('SPS.Maíz','SPS 2727')),
                        (('Soja','14324'), ('Syngenta.Soja','SPS 6x6 IPRO')),
                        (('Soja','14334'), ('Syngenta.Soja','SPS 7x8 IPRO')),
                        (('Colza','9932'), ('Sursem.Colza','SRM 2836')),
                        (('Soja','12153'), ('Sursem.Soja','SRM 3300')),
                        (('Soja','12042'), ('Sursem.Soja','SRM 3410')),
                        (('Soja','12709'), ('Sursem.Soja','SRM 3801')),
                        (('Soja','14781'), ('Sursem.Soja','SRM 3988')),
                        (('Soja','14177'), ('Sursem.Soja','SRM 4222')),
                        (('Soja','8398'), ('Sursem.Soja','SRM 4500')),
                        (('Soja','12910'), ('Sursem.Soja','SRM 4602 STS')),
                        (('Soja','9837'), ('Sursem.Soja','SRM 4754')),
                        (('Soja','11290'), ('Sursem.Soja','SRM 4839')),
                        (('Soja','10485'), ('Sursem.Soja','SRM 5001')),
                        (('Soja','12908'), ('Sursem.Soja','SRM 5200')),
                        (('Soja','9232'), ('Sursem.Soja','SRM 5301')),
                        (('Soja','12158'), ('Sursem.Soja','SRM 5500')),
                        (('Maíz','12602'), ('Sursem.Maíz','SRM 553')),
                        (('Maíz','10299'), ('Sursem.Maíz','SRM 56-20')),
                        (('Maíz','14417'), ('Sursem.Maíz','SRM 56-22')),
                        (('Maíz','14541'), ('Sursem.Maíz','SRM 56-24')),
                        (('Soja','11294'), ('Sursem.Soja','SRM 5601')),
                        (('Maíz','12601'), ('Sursem.Maíz','SRM 567')),
                        (('Soja','12159'), ('Sursem.Soja','SRM 5700')),
                        (('Soja','15072'), ('Sursem.Soja','SRM 5835IPRO')),
                        (('Soja','13524'), ('Sursem.Soja','SRM 6256')),
                        (('Soja','9933'), ('Sursem.Soja','SRM 6403')),
                        (('Maíz','18074'), ('Sursem.Maíz','SRM 6620')),
                        (('Maíz','18076'), ('Sursem.Maíz','SRM 6670')),
                        (('Soja','12161'), ('Sursem.Soja','SRM 6900')),
                        (('Soja','12162'), ('Sursem.Soja','SRM 7200')),
                        (('Soja','12710'), ('Sursem.Soja','SRM 7800')),
                        (('Soja','11295'), ('Sursem.Soja','SRM 8201')),
                        (('Maíz','7758'), ('Sursem.Maíz','SRM POPER 42')),
                        (('Maíz','19606'), ('Stine.Maíz','ST 120-09 RG')),
                        (('Maíz','18860'), ('Stine.Maíz','ST 120-29 BTRG')),
                        (('Maíz','18230'), ('Stine.Maíz','ST 9734-20')),
                        (('Maíz','18229'), ('Stine.Maíz','ST 9734-G')),
                        (('Maíz','17565'), ('Stine.Maíz','ST 9739E-20')),
                        (('Maíz','21654'), ('Stine.Maíz','ST 9741-20')),
                        (('Maíz','18859'), ('Stine.Maíz','ST 9808 E-20')),
                        (('Maíz','19605'), ('Stine.Maíz','ST 9820-20')),
                        (('Maíz','20685'), ('Stine.Maíz','ST 9910-20')),
                        (('Sorgo','13699'), ('Advanta.Sorgo','SUGARGRAZE AR')),
                        (('Trigo','15179'), ('Buck.Trigo','SY 015')),
                        (('Trigo','15180'), ('Buck.Trigo','SY 041')),
                        (('Trigo','11657'), ('Buck.Trigo','SY 100')),
                        (('Trigo','19469'), ('Buck.Trigo','SY 109')),
                        (('Trigo','12337'), ('Buck.Trigo','SY 110')),
                        (('Trigo','17064'), ('Buck.Trigo','SY 120')),
                        (('Trigo','11658'), ('Buck.Trigo','SY 200')),
                        (('Trigo','15753'), ('Buck.Trigo','SY 211')),
                        (('Trigo','11656'), ('Buck.Trigo','SY 300')),
                        (('Trigo','15786'), ('Buck.Trigo','SY 330')),
                        (('Soja','11200'), ('Syngenta.Soja','SY 3X5')),
                        (('Soja','14833'), ('Syngenta.Soja','SY 5X8 IPRO')),
                        (('Soja','14817'), ('Syngenta.Soja','SY 5x1 RR')),
                        (('Soja','14818'), ('Syngenta.Soja','SY 6x8 IPRO')),
                        (('Soja','15525'), ('Syngenta.Soja','SY 7X1IPRO')),
                        (('Trigo','16707'), ('Buck.Trigo','SY OBELIX')),
                        (('Soja','16070'), ('Syngenta.Soja','SYN 1561IPRO')),
                        (('Girasol','17893'), ('Syngenta.Girasol','SYN 3990')),
                        (('Girasol','17892'), ('Syngenta.Girasol','SYN 4066')),
                        (('Maíz','19684'), ('Syngenta.Maíz','SYN 505')),
                        (('Maíz','18617'), ('Syngenta.Maíz','SYN 840')),
                        (('Maíz','18618'), ('Syngenta.Maíz','SYN 897')),
                        (('Maíz','17998'), ('Syngenta.Maíz','SYN 979')),
                        (('Trigo','15754'), ('Buck.Trigo','Saeta')),
                        (('Trigo','6818'), ('Klein.Trigo','Sagitario')),
                        (('Maíz','9476'), ('Albert.Maíz','Santa Fe 2')),
                        (('Trigo','18188'), ('DonMario.Trigo','Sauce')),
                        (('Cebada','6359'), ('Cargill.Cebada','Scarlett')),
                        (('Cebada','14183'), ('MalteriaPampa.Cebada','Scrabble')),
                        (('Trigo','6817'), ('Proseme.Trigo','Scudo')),
                        (('Trigo','20522'), ('Klein.Trigo','Selenio CL')),
                        (('Trigo','15216'), ('Klein.Trigo','Serpiente')),
                        (('Cebada','10134'), ('Limagrain.Cebada','Shakira')),
                        (('Cebada','10134'), ('MalteriaPampa.Cebada','Shakira')),
                        (('Arveja','15991'), ('Limagrain.Arveja','Shamrock')),
                        (('Sorgo','10019'), ('Pemán.Sorgo','Silero Inta Pemán')),
                        (('Maíz','16598'), ('Albert.Maíz','Silomax Full')),
                        (('Cebada','17891'), ('ACA.Cebada','Sinfonia')),
                        (('Alfalfa','8733'), ('Biscayart.Alfalfa','Super Aurora')),
                        (('Alfalfa','9923'), ('Biscayart.Alfalfa','Super Sonic')),
                        (('Alfalfa','13109'), ('Biscayart.Alfalfa','Super Star')),
                        (('Trigo','6686'), ('Buck.Trigo','Sureño')),
                        (('Sorgo','6488'), ('Biscayart.Sorgo','Sweetgreen')),
                        (('Cebada','12628'), ('Cargill.Cebada','Sylphide')),
                        (('Soja','11150'), ('LaTijereta.Soja','T 2137')),
                        (('Soja','11998'), ('LaTijereta.Soja','T 2246')),
                        (('Soja','12035'), ('LaTijereta.Soja','T 2249')),
                        (('Soja','12795'), ('LaTijereta.Soja','T 2259')),
                        (('Soja','13468'), ('LaTijereta.Soja','T 2266')),
                        (('Girasol','12733'), ('Agseed.Girasol','TB 11')),
                        (('Girasol','12724'), ('Agseed.Girasol','TB 14 IMI')),
                        (('Soja','11255'), ('LaTijereta.Soja','TJ 2138 R')),
                        (('Soja','7221'), ('LaTijereta.Soja','TJS 2044')),
                        (('Soja','7222'), ('LaTijereta.Soja','TJS 2049 RR')),
                        (('Soja','7696'), ('LaTijereta.Soja','TJS 2055 RR')),
                        (('Soja','4150'), ('LaTijereta.Soja','TJS 2070')),
                        (('Soja','11256'), ('LaTijereta.Soja','TJS 2136 RR')),
                        (('Soja','8440'), ('LaTijereta.Soja','TJS 2139')),
                        (('Soja','11254'), ('LaTijereta.Soja','TJS 2145 RR')),
                        (('Soja','11255'), ('LaTijereta.Soja','TJS 2148')),
                        (('Soja','8439'), ('LaTijereta.Soja','TJS 2156')),
                        (('Soja','11214'), ('LaTijereta.Soja','TJS 2158 R')),
                        (('Soja','8438'), ('LaTijereta.Soja','TJS 2164')),
                        (('Soja','11640'), ('LaTijereta.Soja','TJS 2165 R')),
                        (('Soja','8437'), ('LaTijereta.Soja','TJS 2170')),
                        (('Soja','10505'), ('LaTijereta.Soja','TJS 2171 R')),
                        (('Soja','8436'), ('LaTijereta.Soja','TJS 2178')),
                        (('Sorgo','13354'), ('Tobin.Sorgo','TOB 62 T')),
                        (('Sorgo','17355'), ('Tobin.Sorgo','TOB FACA BMR')),
                        (('Sorgo','10972'), ('Tobin.Sorgo','TOB Matrero')),
                        (('Sorgo','15573'), ('Tobin.Sorgo','TOB Padrillo Plus')),
                        (('Trigo','21632'), ('Bioseminis.Trigo','TRAFUL HB4')),
                        (('Cebada','17701'), ('INTA.Cebada','TRINIDAD INTA ')),
                        (('Sorgo','14384'), ('LaTijereta.Sorgo','TS 267')),
                        (('Sorgo','9767'), ('LaTijereta.Sorgo','TS 281')),
                        (('Sorgo','13012'), ('LaTijereta.Sorgo','TS 283')),
                        (('Trigo','10659'), ('Buck.Trigo','Taita')),
                        (('Sorgo','17101'), ('Pemán.Sorgo','Takurí')),
                        (('Trigo','9062'), ('Klein.Trigo','Tauro')),
                        (('Trigo','17458'), ('DonMario.Trigo','Tbio Audaz')),
                        (('Triticale','1648'), ('INTA.Triticale','Tehuelche')),
                        (('Trigo','10011'), ('DonMario.Trigo','Themix')),
                        (('Trigo','10787'), ('Klein.Trigo','Tigre')),
                        (('Alfalfa','10500'), ('PGG.Alfalfa','Tigresa')),
                        (('Trigo','14388'), ('Buck.Trigo','Tilcara')),
                        (('Trigo','14487'), ('Bioceres.Trigo','Timbo')),
                        (('Trigo','15215'), ('Klein.Trigo','Titanio')),
                        (('Trigo','5422'), ('Buck.Trigo','Topacio')),
                        (('Trigo','20491'), ('Illinois.Trigo','Tordo')),
                        (('Trigo','21632'), ('Bioceres.Trigo','Traful HB4')),
                        (('Cebada','15240'), ('ACA.Cebada','Traveler')),
                        (('Cebada','13141'), ('Cargill.Cebada','Umbrella')),
                        (('Cebada','1048'), ('INTA.Cebada','UÑAICHE INTA ')),
                        (('Sorgo','8596'), ('Advanta.Sorgo','VDH 314')),
                        (('Sorgo','8193'), ('Advanta.Sorgo','VDH 422')),
                        (('Sorgo','5855'), ('Advanta.Sorgo','VDH 701')),
                        (('Alfalfa','2114'), ('BayaCasal.Alfalfa','VICTORIA')),
                        (('Alfalfa','2114'), ('Produsem.Alfalfa','VICTORIA SP INTA')),
                        (('Soja','14851'), ('VTSeeds.Soja','VT 5335')),
                        (('Girasol','17063'), ('Argensun.Girasol','Valia 41')),
                        (('Girasol','17672'), ('Argensun.Girasol','Valia 73')),
                        (('Girasol','19573'), ('Argensun.Girasol','Valia 92')),
                        (('Girasol','17063'), ('Argensun.Girasol','Valia NTC 418')),
                        (('Girasol','21736'), ('Argensun.Girasol','Valia V22')),
                        (('Trigo','17993'), ('Klein.Trigo','Valor')),
                        (('Alfalfa','10952'), ('Barenbrug.Alfalfa','Verdor')),
                        (('Alfalfa','11940'), ('Barenbrug.Alfalfa','Verzy')),
                        (('Alfalfa','2114'), ('Biscayart.Alfalfa','Victoria')),
                        (('Arveja','5804'), ('AFA.Arveja','Viper')),
                        (('Alfalfa','10225'), ('WL.Alfalfa','WL1058')),
                        (('Alfalfa','9281'), ('WL.Alfalfa','WL611')),
                        (('Alfalfa','17145'), ('WL.Alfalfa','WL825 HVX.RR')),
                        (('Alfalfa','19713'), ('WL.Alfalfa','WL828')),
                        (('Alfalfa','17144'), ('WL.Alfalfa','WL835 HVX.RR')),
                        (('Alfalfa','8559'), ('WL.Alfalfa','WL903')),
                        (('Alfalfa','15124'), ('WL.Alfalfa','WL919')),
                        (('Girasol','19670'), ('NuSeed.Girasol','X4334 CL')),
                        (('Triticale','2776'), ('INTA.Triticale','Yagan')),
                        (('Arveja','14726'), ('Bioseminis.Arveja','Yams')),
                        (('Trigo','11595'), ('Klein.Trigo','Yarara')),
                        (('Trigo','5421'), ('Buck.Trigo','Yatasto')),
                        (('Girasol','19616'), ('ZetaSemillas.Girasol','ZT 74C15')),
                        (('Girasol','19614'), ('ZetaSemillas.Girasol','ZT 74H55 ')),
                        (('Girasol','19663'), ('ZetaSemillas.Girasol','ZT 74H70')),
                        (('Girasol','19615'), ('ZetaSemillas.Girasol','ZT 74L60')),
                        (('Girasol','20755'), ('ZetaSemillas.Girasol','ZT 74L62')),
                        (('Girasol','19662'), ('ZetaSemillas.Girasol','ZT 75L50')),
                        (('Trigo','15712'), ('Buck.Trigo','Zafiro')),
                        (('Trigo','20076'), ('LG.Trigo','Zaino')),
                        (('Trigo','9559'), ('Klein.Trigo','Zorro')),
                        (('Trigo','21945'), ('Neogen.Trigo','neo30t23')),
                        (('Trigo','21946'), ('Neogen.Trigo','neo50t23')),
                        (('Trigo','17457'), ('DonMario.Trigo','Ñandubay'))
                    ]
                    
                    for (cultivo, codigo_genetica), (semillero, genetica) in asignaciones:
                        df.loc[(df['Cultivo'] == cultivo) & (df['Codigo_Genetica'] == codigo_genetica), ['Semillero', 'Genetica']] = [semillero, genetica]
                
                    return df
                    
                df = validar_genetica_semillero(df, conn)

                # Convertir el Código de Fertilizante en integer # FUNCIONA
                def fertilizante_int(df, conn):
                    columnas = [
                        '1_Codigo_Registro_1', '1_Codigo_Registro_2',
                        '2_Codigo_Registro_1', '2_Codigo_Registro_2',
                        '3_Codigo_Registro_1', '3_Codigo_Registro_2',
                        '4_Codigo_Registro_1', '4_Codigo_Registro_2'
                    ]
                    
                    for col in columnas:
                        df[col] = df[col].fillna(0).astype('Int64')
                    
                    return df
                
                df = fertilizante_int(df, conn)

                # 3.7) Validación de Fertilizantes # FUNCIONA
                
                # Registro fertilizantes 
                def validar_id_senasa_fertilizantes(df, conn):
                    reg_fert_query = """SELECT id_senasa FROM datcrea_tablas.registro_fertilizantes WHERE id_senasa IS NOT NULL"""
                    reg_fert_df = pd.read_sql(reg_fert_query,con=conn)
                    id_senasa = reg_fert_df['id_senasa'].astype(int).tolist()
                    
                    df['1_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['1_Codigo_Registro_1']]
                    df['1_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['1_Codigo_Registro_2']]
                    df['2_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['2_Codigo_Registro_1']]
                    df['2_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['2_Codigo_Registro_2']]
                    df['3_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['3_Codigo_Registro_1']]
                    df['3_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['3_Codigo_Registro_2']]
                    df['4_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['4_Codigo_Registro_1']]
                    df['4_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['4_Codigo_Registro_2']]
                    return df
                df = validar_id_senasa_fertilizantes(df, conn)
                
                def validar_nombre_fertilizante(df, conn):
                
                    # Consulta para obtener los fertilizantes registrados con un id_senasa no nulo
                    reg_fert_query = "SELECT * FROM datcrea_tablas.registro_fertilizantes WHERE id_senasa IS NOT NULL"
                    registros_fertilizantes = pd.read_sql(reg_fert_query, con=conn)
                    
                    # Asegurarse de que id_senasa es de tipo entero
                    registros_fertilizantes['id_senasa'] = registros_fertilizantes['id_senasa'].astype(int)
                    
                    # Obtener la lista de fertilizantes únicos
                    fertilizantes_lista = registros_fertilizantes['fertilizante'].unique().tolist()
                    
                    # Lista de columnas a validar
                    columnas_a_validar = [
                        '1_Producto_1', '1_Producto_2', 
                        '2_Producto_1', '2_Producto_2', 
                        '3_Producto_1', '3_Producto_2', 
                        '4_Producto_1', '4_Producto_2'
                    ]
                    
                    # Validar cada columna en la lista
                    for col in columnas_a_validar:
                        df[col] = df[col].apply(lambda x: x if x in fertilizantes_lista else '')
                    
                    return df
                df = validar_nombre_fertilizante(df, conn)


                def validar_fertilizante(df,conn):
                    reg_fert_query = """SELECT * FROM datcrea_tablas.registro_fertilizantes WHERE id_senasa IS NOT NULL"""
                    registros_fertilizantes = pd.read_sql(reg_fert_query, con=conn)
                    registros_fertilizantes['id_senasa'] = registros_fertilizantes['id_senasa'].astype(int)
                    
                    # Crear un diccionario para almacenar los resultados
                    id_senasa_dict = {}
                    
                    
                    for index, row in registros_fertilizantes.iterrows():
                        fertilizante = row['fertilizante']
                        id_senasa = row['id_senasa']
                        if fertilizante in id_senasa_dict:
                            id_senasa_dict[fertilizante].append(id_senasa)
                        else:
                            id_senasa_dict[fertilizante] = [id_senasa]
                    
                    # Validar fertilizante en función del código
                    def validar_fertilizante_codigo(codigo):
                        
                        if pd.isna(codigo):
                            return ''
                        for producto, codigos in id_senasa_dict.items():
                            if codigo in codigos:
                                return producto
                        return ''
                    
                    # Aplicar validación en cada columna
                    for i in range(1, 5):  
                        df[f'{i}_Producto_1'] = df[f'{i}_Codigo_Registro_1'].apply(validar_fertilizante)
                        df[f'{i}_Producto_2'] = df[f'{i}_Codigo_Registro_2'].apply(validar_fertilizante)
                    
                    return df

                # VALIDAR DOSIS DE PRODUCTOS (si es que primero se validó el producto) # FUNCIONA
                def validar_dosis_productos(df):
                
                    columnas_productos_dosis = [
                        ('1_Producto_1', '1_Dosis_1'),
                        ('1_Producto_2', '1_Dosis_2'),
                        ('2_Producto_1', '2_Dosis_1'),
                        ('2_Producto_2', '2_Dosis_2'),
                        ('3_Producto_1', '3_Dosis_1'),
                        ('3_Producto_2', '3_Dosis_2'),
                        ('4_Producto_1', '4_Dosis_1'),
                        ('4_Producto_2', '4_Dosis_2')
                    ]
                
                    for producto_col, dosis_col in columnas_productos_dosis:
                        df.loc[df[producto_col] == '', dosis_col] = pd.NA
                
                    return df
                df = validar_dosis_productos(df)



                def validar_momentos(df, conn):
                
                    # Lista de momentos válidos
                    momentos_validos = [
                        'Presiembra',
                        'Refertilización',
                        'Siembra',
                        'Post siembra',
                        'Otro'
                    ]
                
                    # Columnas de momento a ajustar
                    columnas_momento = ['1_Momento', '2_Momento', '3_Momento', '4_Momento']
                
                    # Iterar sobre las columnas para ajustar los momentos
                    for columna in columnas_momento:
                        df[columna] = df[columna].apply(lambda x: x if x in momentos_validos else '')
                
                    return df
                
                df = validar_momentos(df,conn)

                def calcular_densidad(df, conn):
                        
                    df['DensidadPlantasHa'] = df['DensidadPlantasHa']/10000
                    df.rename(columns={'DensidadPlantasHa': 'DensidadPlantasM2'}, inplace=True)
                    return df
                
                df = calcular_densidad(df, conn)


                def reindex_columns(df):
                    cols = ['CUIC','Campo', 'Campo_Otro','Columna_I','Lote','Lote_Otro','vacio','vacio','vacio',
                            'Provincia','Departamento','vacio','Localidad','CalidadAmbiente','SuperficieSembrada',
                            'Tenencia','DestinoProduccion','vacio','Cultivo','vacio','SubgrupoCultivo','vacio',
                            'CultivoAntecesor','vacio','vacio','FechaSiembra','vacio','vacio','vacio','espaciamiento',
                            'DensidadSemillasM2','DensidadSemillasHA','DensidadPlantasM2','vacio',
                            'Semillero','vacio','Genetica','Version Biotecnológica','vacio','Fertilizacion',
                            '1_Producto_1', '1_Dosis_1', '1_Momento', '1_Forma',
                            '1_Producto_2','1_Dosis_2','1_Momento','1_Forma',
                            '2_Producto_1','2_Dosis_1','2_Momento','2_Forma',
                            '2_Producto_2','2_Dosis_2','2_Momento','2_Forma',
                            '3_Producto_1','3_Dosis_1','3_Momento','3_Forma',
                            '3_Producto_2','3_Dosis_2','3_Momento','3_Forma',
                            'vacio','Riego','Lamina_Riego','Napa','vacio','vacio','vacio','vacio','vacio','vacio',
                            'vacio','Superficie','vacio','Rendimiento']
                
                    return df.reindex(columns=cols)
                
                df = reindex_columns(df)

                
                # 5) Creación de columnas Campo_Otro y Lote_Otro (para caso SIMA) #FUNCIONA
                def otro_cuic_campo_lote(df, conn):
                    #df['CUIC'] = df['CUIC'].astype(str)
                    df['CUIC'] = df['CUIC'].fillna('Otro')
                    #df['Campo_Otro'] = df['Campo_Otro'].fillna('Otro')
                    #df['Lote_Otro'] = df['Lote_Otro'].fillna('Otro')
                    
                    # Llenar la columna 'Campo_Otro' con 'Otro' donde la columna 'Campo' tiene NaN
                    df.loc[df['Campo'].isna(), 'Campo_Otro'] = 'Otro'
                    # Dejar vacía la columna 'Campo_Otro' donde la columna 'Campo' tiene datos
                    df.loc[df['Campo'].notna(), 'Campo_Otro'] = ''
                    
                    # Llenar la columna 'Lote_Otro' con 'otro' donde la columna 'Lote' tiene NaN
                    df.loc[df['Lote'].isna(), 'Lote_Otro'] = 'Otro'
                    # Dejar vacía la columna 'Lote_Otro' donde la columna 'Lote' tiene datos
                    df.loc[df['Lote'].notna(), 'Lote_Otro'] = ''
                
                    return df
                
                df = otro_cuic_campo_lote(df, conn)


                # Localidades
                def validar_localidad(df):
                    
                    localidades_query = """SELECT localidad FROM datcrea_tablas.localidades"""
                    
                    localidades = pd.read_sql(localidades_query,con=conn)
                    localidades = localidades['localidad'].str.title().tolist()
                    ## 6.1) Validación Localidad
                    # Unificar criterio de escritura
                    df['Localidad'] = df['Localidad'].astype(str)
                    df['Localidad'] = df['Localidad'].str.title()
                    df['Localidad'] = ['' if x not in localidades else x for x in df['Localidad']]
                    return df
            

                df = validar_localidad(df)
                
                # Departamentos
                def validar_departamento(df,conn):
                
                    reemplazos = {
                        '25 DE MAYO': 'Veinticinco.de.Mayo',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'ADOLFO ALSINA': 'Adolfo.Alsina',
                        'ADOLFO GONZALES CHAVES': 'Adolfo.Gonzales.Chaves',
                        'ALBERTI': 'Alberti',
                        'ALMIRANTE BROWN': '',
                        'ARRECIFES': 'Arrecifes',
                        'AVELLANEDA': '',
                        'AYACUCHO': 'Ayacucho',
                        'AZUL': 'Azul',
                        'BAHIA BLANCA': 'Bahía.Blanca',
                        'BALCARCE': 'Balcarce',
                        'BARADERO': 'Baradero',
                        'BENITO JUAREZ': 'Benito.Juarez',
                        'BERAZATEGUI': 'Berazategui',
                        'BERISSO': 'Berisso',
                        'BOLIVAR': 'Bolivar',
                        'BRAGADO': 'Partido.Bragado',
                        'BRANDSEN': 'Brandsen',
                        'CAMPANA': 'Campana',
                        'CANUELAS': 'Cañuelas',
                        'CAPITAN SARMIENTO': 'Capitan.Sarmiento',
                        'CARLOS CASARES': 'Carlos.Casares',
                        'CARLOS TEJEDOR': 'Carlos.Tejedor',
                        'CARMEN DE ARECO': 'Carmen.de.Areco',
                        'CASTELLI': 'Castelli',
                        'CHACABUCO': 'Chacabuco',
                        'CHASCOMUS': 'Chascomus',
                        'CHIVILCOY': 'Chivilcoy',
                        'COLON': 'Colon',
                        'CORONEL DE MARINA L. ROSALES': 'Coronel.de.Marina.L.Rosales',
                        'CORONEL DORREGO': 'Coronel.Dorrego',
                        'CORONEL PRINGLES': 'Coronel.Pringles',
                        'CORONEL SUAREZ': 'Coronel.Suarez',
                        'DAIREAUX': 'Daireaux',
                        'DOLORES': 'Dolores',
                        'ENSENADA': 'Ensenada',
                        'ESCOBAR': 'Escobar',
                        'ESTEBAN ECHEVERRIA': 'Esteban.Echeverria',
                        'EXALTACION DE LA CRUZ': 'Exaltación.de.La.Cruz',
                        'EZEIZA': 'Ezeiza',
                        'FLORENCIO VARELA': 'Florencio.Varela',
                        'FLORENTINO AMEGHINO': 'Florentino.Ameghino',
                        'GENERAL ALVARADO': 'General.Alvarado',
                        'GENERAL ALVEAR': 'General.Alvear',
                        'GENERAL ARENALES': 'General.Arenales',
                        'GENERAL BELGRANO': 'General.Belgrano',
                        'GENERAL GUIDO': 'General.Guido',
                        'GENERAL JUAN MADARIAGA': 'General.Juan.Madariaga',
                        'GENERAL LA MADRID': 'General.La.Madrid',
                        'GENERAL LAS HERAS': 'General.Las.Heras',
                        'GENERAL LAVALLE': 'General.Lavalle',
                        'GENERAL PAZ': 'General.Paz',
                        'GENERAL PINTO': 'General.Pinto',
                        'GENERAL PUEYRREDON': 'General.Pueyrredon',
                        'GENERAL RODRIGUEZ': 'General.Rodriguez',
                        'GENERAL SAN MARTIN': 'General.San.Martin',
                        'GENERAL VIAMONTE': 'General.Viamonte',
                        'GENERAL VILLEGAS': 'General.Villegas',
                        'GUAMINI': 'Guamini',
                        'HIPOLITO YRIGOYEN': 'Hipolito.Yrigoyen',
                        'HURLINGHAM': 'Hurlingham',
                        'ITUZAINGO': 'Ituzaingo',
                        'JOSE C. PAZ': 'Jose.C.Paz',
                        'JUNIN': 'Junin',
                        'LA COSTA': 'La.Costa',
                        'LA MATANZA': 'La.Matanza',
                        'LA PLATA': 'La.Plata',
                        'LANUS': 'Lanús',
                        'LAPRIDA': 'Laprida',
                        'LAS FLORES': 'Las.Flores',
                        'LEANDRO N. ALEM': 'Leandro.N..Alem',
                        'LINCOLN': 'Lincoln',
                        'LOBERIA': 'Loberia',
                        'LOBOS': 'Lobos',
                        'LOMAS DE ZAMORA': 'Lomas.de.Zamora',
                        'LUJAN': 'Lujan',
                        'MAGDALENA': 'Magdalena',
                        'MAIPU': 'Maipu',
                        'MALVINAS ARGENTINAS': 'Malvinas.Argentinas',
                        'MAR CHIQUITA': 'Mar.Chiquita',
                        'MARCOS PAZ': 'Marcos.Paz',
                        'MERCEDES': 'Mercedes',
                        'MERLO': 'Merlo',
                        'MONTE': 'Monte',
                        'MONTE HERMOSO': 'Monte.Hermoso',
                        'MORENO': 'Moreno',
                        'MORON': 'Moron',
                        'NAVARRO': 'Navarro',
                        'NECOCHEA': 'Necochea',
                        'OLAVARRIA': 'Olavarria',
                        'PATAGONES': 'Patagones',
                        'PEHUAJO': 'Pehuajo',
                        'PELLEGRINI': 'Pellegrini',
                        'PERGAMINO': 'Partido.Pergamino',
                        'PILA': 'Pila',
                        'PILAR': 'Pilar',
                        'PINAMAR': 'Pinamar',
                        'PRESIDENTE PERON': 'Presidente.Peron',
                        'PUAN': 'Puan',
                        'PUNTA INDIO': 'Punta.Indio',
                        'QUILMES': 'Quilmes',
                        'RAMALLO': 'Ramallo',
                        'RAUCH': 'Rauch',
                        'RIVADAVIA': 'Rivadavia',
                        'ROJAS': 'Rojas',
                        'ROQUE PEREZ': 'Roque.Perez',
                        'SAAVEDRA': 'Saavedra',
                        'SALADILLO': 'Saladillo',
                        'SALLIQUELO': 'Salliqueló',
                        'SALTO': 'Partido.Salto',
                        'SAN ANDRES DE GILES': 'San.Andres.de.Giles',
                        'SAN ANTONIO DE ARECO': 'San.A.de.Areco',
                        'SAN CAYETANO': 'San.Cayetano',
                        'SAN FERNANDO': 'San.Fernando',
                        'SAN ISIDRO': 'San.Isidro',
                        'SAN MIGUEL': 'San.Miguel',
                        'SAN NICOLAS': 'San.Nicolas',
                        'SAN PEDRO': 'San.Pedro',
                        'SAN VICENTE': 'San.Vicente',
                        'SUIPACHA': 'Suipacha',
                        'TANDIL': 'Tandil',
                        'TAPALQUE': 'Tapalque',
                        'TIGRE': 'Tigre',
                        'TORDILLO': 'Tordillo',
                        'TORNQUIST': 'Tornquist',
                        'TRENQUE LAUQUEN': 'Trenque.Lauquen',
                        'TRES ARROYOS': 'Tres.Arroyos',
                        'TRES DE FEBRERO': 'Tres.de.Febrero',
                        'TRES LOMAS': 'Tres.Lomas',
                        'VICENTE LOPEZ': 'Vicente.Lopez',
                        'VILLA GESELL': 'Villa.Gesell',
                        'VILLARINO': 'Villarino',
                        'ZARATE': 'Zarate',
                        'CAPITAL FEDERAL': '',
                        'AMBATO': 'Ambato',
                        'ANCASTI': 'Ancasti',
                        'ANDALGALA': '',
                        'ANTOFAGASTA DE LA SIERRA': 'Antofagasta.de.La.Sierra',
                        'BELEN': 'Belen',
                        'CAPAYAN': 'Capayan',
                        'CAPITAL': 'Capital',
                        'EL ALTO': 'El.Alto',
                        'FRAY MAMERTO ESQUIU': 'Fray.Mamerto.Esquiu',
                        'LA PAZ': 'La.Paz',
                        'PACLIN': 'Paclin',
                        'POMAN': 'Poman',
                        'SANTA MARIA': 'Santa.Maria',
                        'SANTA ROSA': 'Santa.Rosa',
                        'TINOGASTA': 'Tinogasta',
                        'VALLE VIEJO': 'Valle.Viejo',
                        '1 DE MAYO': '',
                        '12 DE OCTUBRE': 'Doce.de.Octubre',
                        '2 DE ABRIL': 'Dos.de.Abril',
                        'ALMIRANTE BROWN': 'Almirante.Brown',
                        'BERMEJO': 'Bermejo',
                        'CHACABUCO': 'Chacabuco',
                        'COMPANIA DE MAYOR ING FRANKLIN': '',
                        'FRAY JUSTO SANTA MARIA DE ORO': 'Fray.Justo.Santa.Maria.de.Oro',
                        'GENERAL BELGRANO': 'General.Belgrano',
                        'GENERAL DONOVAN': 'General.Donovan',
                        'GENERAL GUEMES': 'General.Guemes',
                        'INDEPENDENCIA': 'Independencia',
                        'LIBERTAD': 'Libertad',
                        'LIBERTADOR GRL SAN MARTIN': '',
                        'MAIPU': 'Maipu',
                        'MAYOR LUIS J FONTANA': 'Mayor.Luis.J.Fontana',
                        'NUEVE DE JULIO': 'Nueve.de.Julio',
                        'O HIGGINS': 'O.Higgins',
                        'PRIMERA ARGENTINA': '',
                        'QUITILIPI': 'Quitilipi',
                        'SAN FERNANDO': 'San.Fernando',
                        'SAN LORENZO': 'San.Lorenzo',
                        'SAN MARTIN': 'San.Martin',
                        'SARMIENTO': 'Sarmiento',
                        'TAPENAGA': 'Tapenaga',
                        '25 DE MAYO': 'Veinticinco.de.Mayo',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'ANGACO': 'Angaco',
                        'CALINGASTA': 'Calingasta',
                        'CAPITAL': 'Capital',
                        'CAUCETE': 'Caucete',
                        'CHIMBAS': 'Chimbas',
                        'IGLESIA': 'Iglesia',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'POCITO': 'Pocito',
                        'RAWSON': 'Rawson',
                        'RIVADAVIA': 'Rivadavia',
                        'SAN MARTIN': 'San.Martin',
                        'SANTA LUCIA': 'Santa.Lucia',
                        'SARMIENTO': 'Sarmiento',
                        'ULLUM': 'Ullum',
                        'VALLE FERTIL': 'Valle.Fertil',
                        'ZONDA': 'Zonda',
                    }
                    
                    # Convertimos a mayúsculas para asegurar que las claves coincidan
                    df['Departamento'] = df['Departamento'].str.upper()
                    
                    # Reemplazamos los valores en la columna 'DEPARTAMENTO' basados en el diccionario
                    df['Departamento'] = df['Departamento'].replace(reemplazos)
                    
                    departamentos_query = """SELECT departamento FROM datcrea_tablas.localidades"""
                    departamento_df = pd.read_sql(departamentos_query,con=conn)
                    departamento_df = departamento_df.drop_duplicates()
                    departamentos = departamento_df['departamento'].tolist()
                    df['Departamento'] = ['' if x not in departamentos else x for x in df['Departamento']]
                    return df
                
                df = validar_departamento(df, conn)

                ## 6.3) Validación Provincias
                # PROVINCIAS
                def validar_provincia(df, conn):
                    reemplazos = {
                        'BUENOS AIRES': 'Buenos.Aires',
                        'CAPITAL FEDERAL': '',
                        'CATAMARCA': 'Catamarca',
                        'CHACO': 'Chaco',
                        'CHUBUT': '',
                        'CORDOBA': 'Córdoba',
                        'CORRIENTES': 'Corrientes',
                        'ENTRE RIOS': 'Entre.Ríos',
                        'FORMOSA': '',
                        'JUJUY': 'Jujuy',
                        'LA PAMPA': 'La.Pampa',
                        'LA RIOJA': '',
                        'MENDOZA': 'Mendoza',
                        'MISIONES': '',
                        'NEUQUEN': '',
                        'RIO NEGRO': '',
                        'SALTA': 'Salta',
                        'SAN JUAN': '',
                        'SAN LUIS': 'San.Luis',
                        'SANTA CRUZ': '',
                        'SANTA FE': 'Santa.Fe',
                        'SANTIAGO DEL ESTERO': 'Santiago.del.Estero',
                        'TIERRA DEL FUEGO': '',
                        'TUCUMAN': 'Tucumán',
                        'DESCONOCIDO': ''
                    }
                    
                    # Convertimos a mayúsculas para asegurar que las claves coincidan
                    df['Provincia'] = df['Provincia'].str.upper()
                    
                    # Reemplazamos los valores en la columna 'Provincia' basados en el diccionario
                    df['Provincia'] = df['Provincia'].replace(reemplazos)
                    #df['Provincia'] = df['Provincia'].astype(str)
                    df['Provincia'] = df['Provincia'].replace('nan', '')
                    
                    return df
                
                
                df = validar_provincia(df, conn)

                # TENENCIA #FUNCIONA
                def validar_tenencia(df,conn):
                    
                    tenencia = ['Propio', 'Aparceria', 'Arrendado', 'Cedido en alquiler']
                
                    df['Tenencia'] = df['Tenencia'].apply(lambda x: x if x in tenencia else '')
                
                    return df
                
                df = validar_tenencia(df,conn)

                
                # DESTINO #FUNCIONA
                def validar_destino(df,conn):
                    
                    destino = ['Grano', 
                            'Semilla (Convenio Comercial)', 
                            'Silo', 
                            'Heno',
                            'Pastoreo', 
                            'Pastoreo de rastrojos', 
                            'Cultivo de servicio', 
                            'Cultivo de servicio pastoreado',
                            'Conservación / Reserva']
                    
                    df['DestinoProduccion'] = df['DestinoProduccion'].apply(lambda x: x if x in destino else '')
                    return df
                
                df = validar_destino(df,conn)

                # CULTIVO ANTECESOR
                def validar_cultivo_antecesor(df,conn):
                    
                    # Lista de valores a reemplazar y sus correspondientes reemplazos
                    reemplazos = {
                        'acai - tardío': '',
                        'acai - de primera': '',
                        'acai - de segunda': '',
                        'achicoria - de primera': 'Achicoria',
                        'agave - tardío': '',
                        'agave - de segunda': '',
                        'agave - de primera': '',
                        'aguacate - de primera': '',
                        'ajo - de primera': 'Ajo',
                        'alfalfa - de segunda': 'Alfalfa',
                        'alfalfa - de primera': 'Alfalfa',
                        'alfalfa - tardío': 'Alfalfa',
                        'algodón - de segunda': 'Algodón',
                        'algodón - tardío': 'Algodón',
                        'algodón - de primera': 'Algodón',
                        'alpiste - de segunda': 'Alpiste',
                        'alpiste - de primera': 'Alpiste',
                        'anana - de primera': '',
                        'arroz - de primera': 'Arroz',
                        'arroz - tardío': 'Arroz',
                        'arroz - de segunda': 'Arroz',
                        'arveja - tardío': 'Arveja',
                        'arveja - de primera': 'Arveja',
                        'arveja - de segunda': 'Arveja',
                        'avena - de primera': 'Avena',
                        'avena - de segunda': 'Avena',
                        'avena - tardío': 'Avena',
                        'avena blanca - de segunda': 'Avena',
                        'avena blanca - de primera': 'Avena',
                        'avena strigosa - tardío': 'Avena',
                        'avena strigosa - de segunda': 'Avena',
                        'avena strigosa - de primera': 'Avena',
                        'banano - de primera': '',
                        'brócoli - de primera': '',
                        'café - safra': '',
                        'café - de segunda': '',
                        'café - de primera': '',
                        'camelina - de primera': 'Camelina',
                        'caña de azucar - de primera': 'Caña de Azucar',
                        'caña de azucar - tardío': 'Caña de Azucar',
                        'caña de azucar - de segunda': 'Caña de Azucar',
                        'canola - de primera': '',
                        'carinata - de primera': 'Carinata',
                        'carinata - de segunda': 'Carinata',
                        'cartamo - de primera': 'Cártamo',
                        'cebada - de primera': 'Cebada',
                        'cebada - tardío': 'Cebada',
                        'cebada - de segunda': 'Cebada',
                        'cebadilla - de primera': 'Cebadilla',
                        'centeno - de segunda': 'Centeno',
                        'centeno - de primera': 'Centeno',
                        'centeno - tardío': 'Centeno',
                        'chia - tardío': '',
                        'chia - de segunda': '',
                        'chia - de primera': '',
                        'cítricos - de segunda': '',
                        'cítricos - tardío': '',
                        'cítricos - de primera': '',
                        'colza - de primera': 'Colza',
                        'colza - de segunda': 'Colza',
                        'coriandro - de primera': 'Coriandro',
                        'crotalaria - de primera': '',
                        'cultivo de servicio - tardío': 'Cultivo de Servicio',
                        'cultivo de servicio - de primera': 'Cultivo de Servicio',
                        'cultivo de servicio - de segunda': 'Cultivo de Servicio',
                        'festuca - de primera': 'Pastura perenne',
                        'garbanzo - de segunda': 'Garbanzo',
                        'garbanzo - de primera': 'Garbanzo',
                        'girasol - tardío': 'Girasol',
                        'girasol - de primera': 'Girasol',
                        'girasol - de segunda': 'Girasol',
                        'kiwi - de primera': '',
                        'lechuga - de segunda': '',
                        'lenteja - de primera': 'Lenteja',
                        'lenteja - de segunda': 'Lenteja',
                        'lenteja - tardío': 'Lenteja',
                        'limon - de primera': '',
                        'lotus - de primera': 'Lotus',
                        'lupino - de primera': '',
                        'maiz' : 'Maíz',
                        'maíz - de primera': 'Maíz de 1° Temprano',
                        'maíz - tardío': 'Maíz de 1° Tardio',
                        'maíz - safrinha': '',
                        'maíz - de segunda': 'Maíz de 2°',
                        'mani - tardío': 'Maní',
                        'mani - de primera': 'Maní',
                        'mani - de segunda': 'Maní',
                        'melon - de primera': '',
                        'mijo - de segunda': 'Mijo',
                        'mijo - de primera': 'Mijo',
                        'mijo - tardío': 'Mijo',
                        'moha - de primera': 'Moha',
                        'moha - de segunda': 'Moha',
                        'moha - tardío': 'Moha',
                        'nabo - de segunda': 'Nabo',
                        'nabo - de primera': 'Nabo',
                        'olivo - tardío': '',
                        'papa - de segunda': 'Papa',
                        'papa - de primera': 'Papa',
                        'pasturas - tardío': 'Pastura perenne',
                        'pasturas - de primera': 'Pastura perenne',
                        'pasturas - de segunda': 'Pastura perenne',
                        'poroto - tardío': 'Poroto',
                        'poroto - de segunda': 'Poroto',
                        'poroto - de primera': 'Poroto',
                        'poroto mung - tardío': 'Poroto Mung',
                        'poroto mung - de segunda': 'Poroto Mung',
                        'poroto mung - de primera': 'Poroto Mung',
                        'raigrass - de segunda': 'Ryegrass',
                        'raigrass - de primera': 'Ryegrass',
                        'raigrass - tardío': 'Ryegrass',
                        'sandia - de primera': '',
                        'sesamo - de primera': '',
                        'soja' : 'Soja',
                        'soja - de segunda': 'Soja de 2°',
                        'soja - tardío': 'Soja de 1°',
                        'soja - primavera / verano': '',
                        'soja - safra': '',
                        'soja - semestre a': '',
                        'soja - de primera': 'Soja de 1°',
                        'sorgo - de primera': 'Sorgo',
                        'sorgo - tardío': 'Sorgo',
                        'sorgo - de segunda': 'Sorgo',
                        'tabaco - de primera': '',
                        'té - de primera': '',
                        'tomate - de segunda': '',
                        'tomate - de primera': '',
                        'trebol blanco - tardío': 'Trebol',
                        'trebol blanco - de primera': 'Trebol',
                        'trebol rojo - de primera': 'Trebol',
                        'trigo - de segunda': 'Trigo',
                        'trigo - de primera': 'Trigo',
                        'trigo - tardío': 'Trigo',
                        'triticale - de primera': 'Triticale',
                        'vicia - de segunda': 'Vicia',
                        'vicia - de primera': 'Vicia',
                        'vicia - tardío': 'Vicia',
                        'vid - de primera': 'Vid',
                        'vid - tardío': 'Vid',
                        'vid - de segunda': 'Vid',
                        'yerba mate - de primera': ''
                    }
                
                    # Convertir a minúsculas y reemplazar valores según el diccionario
                    df['CultivoAntecesor'] = df['CultivoAntecesor'].str.lower()
                    df['CultivoAntecesor'] = df['CultivoAntecesor'].replace(reemplazos)
                    
                    
                    return df
                    
                df = validar_cultivo_antecesor(df, conn)
                

                ## 6.7) Validación Calidad ambiente #FUNCIONA
                def validar_calidad_ambiente(df,conn):
                        
                    calidad_ambiente = ['Potencial Alto', 'Potencial Promedio', 'Potencial Bajo']
                    
                    df['CalidadAmbiente'] = df['CalidadAmbiente'].apply(lambda x: x if x in calidad_ambiente else '')
                
                    return df
                    
                df = validar_calidad_ambiente(df, conn)
                

                def validar_subgrupo(df, conn): #REVISAR #FUNCIONA
                    
                    subgrupo_query = """SELECT subgrupo FROM datcrea_tablas.subgrupos_cultivos"""
                    subgrupo_df = pd.read_sql(subgrupo_query,con=conn)
                    subgrupo = subgrupo_df['subgrupo'].tolist()
                    df['SubgrupoCultivo'] = ['' if x not in subgrupo else x for x in df['SubgrupoCultivo']]
                
                
                    #Carga los subgrupos de cultivosvy los organiza en un diccionario
                    subgrupos_cultivos_query = """SELECT cultivo, subgrupo FROM datcrea_tablas.subgrupos_cultivos WHERE cultivo IS NOT NULL"""
                    subgrupos_cultivos = pd.read_sql(subgrupos_cultivos_query, con=conn)
                
                    # Crear un diccionario para almacenar los resultados
                    subgrupos_cultivos_dict = {}
                    for index, row in subgrupos_cultivos.iterrows():
                        cultivo = row['cultivo']
                        subgrupo = row['subgrupo']
                        if cultivo in subgrupos_cultivos_dict:
                            subgrupos_cultivos_dict[cultivo].append(subgrupo)
                        else:
                            subgrupos_cultivos_dict[cultivo] = [subgrupo]
                
                    def validar_subgrupo(row):
                        
                        valor_cultivo = row['Cultivo']
                        valor_subgrupo = row['SubgrupoCultivo']
                
                        if valor_cultivo in subgrupos_cultivos_dict:
                            if valor_subgrupo in subgrupos_cultivos_dict[valor_cultivo]:
                                return valor_subgrupo
                        return ''
                
                    return df
                
                df = validar_subgrupo(df, conn)
                

                def validar_biotecnologia(df, conn):
                    
                    # Ejecutar la consulta SQL y cargar los resultados en un DataFrame
                    version_biotecno_query = """SELECT cultivo, nombre FROM datcrea_tablas.version_biotecnologica WHERE cultivo IS NOT NULL"""
                    version_biotecno = pd.read_sql(version_biotecno_query, con=conn)
                    
                    # Crear un diccionario para almacenar los resultados
                    version_biotecno_dict = {}
                    
                    
                    for index, row in version_biotecno.iterrows():
                        cultivo = row['cultivo']
                        nombre = row['nombre']
                        if cultivo in version_biotecno_dict:
                            version_biotecno_dict[cultivo].append(nombre)
                        else:
                            version_biotecno_dict[cultivo] = [nombre]
                    
                    def validar_biotecnologia(row):
                        valor_cultivo = row['Cultivo']
                        valor_subgrupo = row['Version Biotecnológica']
                        
                        if valor_cultivo in version_biotecno_dict:
                            if valor_subgrupo in version_biotecno_dict[valor_cultivo]:
                                return valor_subgrupo
                        return ''
                    return df
                
                df = validar_biotecnologia(df, conn)

                # FERTILIZACIÓN
                def validar_fertilizacion(df,conn):
                    
                    fertilizacion_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'fertilizacion'"""
                    fertilizacion = pd.read_sql(fertilizacion_query, con=conn)
                    fertilizacion = fertilizacion['valor'].tolist()
                    df['Fertilizacion'] = ['' if x not in fertilizacion else x for x in df['Fertilizacion']]
                    return df
                
                df = validar_fertilizacion(df, conn)
                


                ## 6.12) Reemplazo de 0 por Na
                # REEMPLAZAR CEROS
                def replace_ceros(df):
                    df['SuperficieSembrada'] = df['SuperficieSembrada'].replace(0, pd.NA)
                    df['espaciamiento'] = df['espaciamiento'].replace(0, pd.NA)
                    df['Rendimiento'] = df['Rendimiento'].replace(0, pd.NA)
                    df['Superficie'] = df['Superficie'].replace(0, pd.NA)
                    return df
                
                df = replace_ceros(df)
                
                # 6.13) Validación de Forma
                
                # VALIDAR FORMA DE FERTILIZACIÓN
                def validar_forma_fertilizacion(df,conn):
                    forma_fertilizacion_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'forma_fertilizacion'"""
                    forma_fertilizacion = pd.read_sql(forma_fertilizacion_query, con=conn)
                    forma_fertilizacion = forma_fertilizacion['valor'].tolist()
                    df['1_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['1_Forma']]
                    df['2_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['2_Forma']]
                    df['3_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['3_Forma']]
                    return df
                
                df = validar_forma_fertilizacion(df, conn)

                
                # Validación de Sistema de riego
                # SISTEMA RIEGO
                def validar_sistema_riego(df,conn):
                    sistema_riego_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'sistema_riego'"""
                    sistema_riego = pd.read_sql(sistema_riego_query, con=conn)
                    sistema_riego = sistema_riego['valor'].tolist()
                    df['Riego'] = ['' if x not in sistema_riego else x for x in df['Riego']]
                    return df
                
                df = validar_sistema_riego(df,conn)


                # INFLUENCIA NAPA
                def influencia_napa(df,conn):
                    
                    influencia_napa_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'influencia_napa'"""
                    influencia_napa = pd.read_sql(influencia_napa_query,con=conn)
                    influencia_napa = influencia_napa['valor'].tolist()
                    df['Napa'] = ['' if x not in influencia_napa else x for x in df['Napa']]
                    return df
                
                df = influencia_napa(df, conn)
                

                
                
                # Conexión a la base
                import pandas as pd
                from sqlalchemy import create_engine
                
                host="dw.crea.org.ar"
                port="54322"
                user="postgres"
                password="Tr0pema44cr34#"
                database="warehouse"
                
                conn2 = create_engine("postgresql+psycopg2://{user}:{password}@{host}:{port}/{database}".format(host=host,user=user,password=password,database=database,port=port))
                
                # Escribir en celdas excel
                region_grupo_query = """SELECT r.region_sigla, r.region_nombre,c.crea_numero,c.crea_nombre  
                FROM crm_crea.regiones AS r
                JOIN crm_crea.crea AS c 
                ON r.region_id = c.crea_region_id
                WHERE c.crea_baja IS NULL
                GROUP BY r.region_sigla, r.region_nombre, c.crea_numero,c.crea_nombre
                ORDER BY r.region_sigla, c.crea_nombre;"""
                region_grupo = pd.read_sql(region_grupo_query, con=conn2)
                
                #cuic = 'SSF145002M'
                
                region = region_grupo.loc[region_grupo['region_sigla'].str.startswith(cuic[:3]), 'region_nombre'].iloc[0]
                #region = replace region por el formato dat que sale de una lista.
                grupo = region_grupo.loc[region_grupo['crea_numero'].str.contains(cuic[3:6]), 'crea_nombre'].iloc[0]
                
                regiones_dat_dict = {
                    
                    'CENTRO':'CENTRO',
                    'CHACO SANTIAGUEÑO':'CHACO.SANTIAGUEÑO',
                    'CORDOBA NORTE':'CORDOBA.NORTE',
                    'ESTE':'ESTE',
                    'LITORAL NORTE':'LITORAL.NORTE',
                    'LITORAL SUR':'LITORAL.SUR',
                    'MAR Y SIERRAS':'MAR.Y.SIERRAS',
                    'NORTE DE BUENOS AIRES':'NORTE.BUENOS.AIRES',
                    'NOA':'NOA',
                    'NORTE DE SANTA FE':'NORTE.SANTA.FE',
                    'OESTE ARENOSO':'OESTE.ARENOSO',
                    'OESTE':'OESTE',
                    'PATAGONIA':'PATAGONIA',
                    'SEMIARIDA':'SEMIARIDA',
                    'SUDESTE':'SUDESTE',
                    'SANTA FE CENTRO':'SANTA.FE.CENTRO',
                    'SUR DE SANTA FE':'SUR.SANTA.FE',
                    'SUDOESTE':'SUDOESTE',
                    'VALLES CORDILLERANOS':'VALLES.CORDILLERANOS'
                    }
                # Reemplazo directo sin función
                for clave, valor in regiones_dat_dict.items():
                    region = region.replace(clave, valor)
                
                #print('REGIÓN: ', region)
                
                grupos_dat_dict = {'ALEJANDRO CHAJAN':'ALEJANDRO.CHAJAN',
                'BUENA ESPERANZA':'BUENA.ESPERANZA',
                'CANALS':'CANALS',
                'CARNERILLO':'CARNERILLO',
                'CAÑADA SECA':'CAÑADA.SECA',
                'CTALAMOCHITA':'CTALAMOCHITA',
                'HUINCA RENANCO':'HUINCA.RENANCO',
                'LA CESIRA TAMBERO':'LA.CESIRA.TAMBERO',
                'LA PORTADA':'LA.PORTADA',
                'LABOULAYE - BOUCHARDO':'LABOULAYE.BUCHARDO',
                'MELO SERRANO':'MELO.SERRANO',
                'RANQUELES':'RANQUELES',
                'RIO CUARTO':'RIO.CUARTO',
                'RIO QUINTO':'RÍO.QUINTO',
                'TAMBERO LABOULAYE':'TAMBERO.LABOULAYE',
                'TAMBERO VILLA MARIA':'VILLA.MARIA',
                'TEGUA':'TEGUA',
                'VALLE DEL CONLARA':'VALLE.DEL.CONLARA',
                'WASHINGTON MACKENNA':'WASHINGTON.MACKENNA',
                'CAMPO GALLO MONTE QUEMADO':'CAMPO.GALLO.MONTE.QUEMADO',
                'GUAYACAN':'GUAYACAN',
                'IBARRETA':'IBARRETA',
                'LOMITAS':'LOMITAS',
                'PALMARES':'PALMARES',
                'PAMPA DEL INFIERNO':'PAMPA.DEL.INFIERNO',
                'QUIMILI':'QUIMILI',
                'RENOVALES':'RENOVALES',
                'SACHAYOJ':'SACHAYOJ',
                'SANAVIRONES':'SANAVIRONES',
                'SEMIARIDO NORTE':'SEMIARIDO.NORTE',
                'SUDESTE SANTIAGUEÑO':'SES',
                'TINTINA':'TINTINA',
                'ARROYITO':'ARROYITO',
                'BARRANCA YACO':'BARRANCA.YACO',
                'CAROYA':'CAROYA',
                'CAÑADA DE LUQUE SITON':'CAÑADA.DE.LUQUE.SITON',
                'DEL ESTE':'DEL.ESTE',
                'GANADERO DEL NOROESTE':'GANADERO.DEL.NOROESTE',
                'JESUS MARIA':'JESUS.MARIA',
                'LAGUNA LARGA':'LAGUNA.LARGA',
                'LEOPOLDO LUGONES':'LEOPOLDO.LUGONES',
                'MONTE CRISTO':'MONTE.CRISTO',
                'RIO PRIMERO':'RIO.PRIMERO',
                'SIERRAS CHICAS':'SIERRAS.CHICAS',
                'TOTORAL':'TOTORAL',
                'ABASTO':'Otro',
                'CAÑUELAS':'CAÑUELAS',
                'GELAS':'GELAS',
                'LUJAN':'LUJAN',
                'NAVARRO II':'NAVARROII',
                'PIONEROS ABASTO':'Otro',
                'AVATI - I - ARROCERO':'AVATI.I.ARROCERO',
                'CURUZU CUATIA':'CURUZU.CUATIA',
                'ESQUINA':'ESQUINA',
                'MALEZALES':'Otro',
                'MERCEDES':'MERCEDES',
                'TIERRA COLORADA':'TIERRA.COLORADA',
                'URUNDAY':'URUNDAY',
                'ÑANDUBAY':'ÑANDUBAY',
                'BOVRIL EL SOLAR':'BOVRIL.EL.SOLAR',
                'CONCEPCION URUGUAY':'CONCEPCION.URUGUAY',
                'CONCORDIA CHAJARI':'CONCORDIA.CHAJARI',
                'GALARZA':'GALARZA',
                'GUALEGUAYCHU':'GUALEGUAYCHU',
                'ISLAS DEL IBICUY':'ISLAS.DEL.IBICUY',
                'LA PAZ':'LA.PAZ',
                'LARROQUE GUALEGUAY':'LARROQUE.GUALEGUAY',
                'MANDISOVI CONCORDIA':'MANDISOVI.CONCORDIA',
                'MONTOYA':'MONTOYA',
                'SAN JAIME':'SAN.JAIME',
                'VICTORIA':'VICTORIA',
                'VILLAGUAY':'VILLAGUAY',
                'ARROYO DE LOS HUESOS':'ARROYO.DE.LOS.HUESOS',
                'AZUL CHILLAR':'AZUL.CHILLAR',
                'CTE.N.OTAMENDI':'CTE.N.OTAMENDI',
                'DEFFERRARI':'DEFFERRARI',
                'FRONTERA':'FRONTERA',
                'FULTON':'FULTON',
                'LOBERIAS GRANDES':'LOBERIAS.GRANDES',
                'NECOCHEA QUEQUEN':'NECOCHEA.QUEQUEN',
                'QUEQUEN SALADO':'QUEQUEN.SALADO',
                'SAN CAYETANO - TRES ARROYOS':'SAN.CAYETANO.TRES.ARROYOS',
                'SAN FRANCISCO DE BELLOCQ':'SAN.FRANCISCO.DE.BELLOCQ',
                'SAN MANUEL':'SAN.MANUEL',
                'TAMBERO MAR Y SIERRA':'TAMBERO.MAR.Y.SIERRA',
                'TANDIL':'TANDIL',
                'TRES ARROYOS':'TRES.ARROYOS',
                'ZONA 4 LECHERA':'ZONA.4.LECHERA',
                'ALBERDI':'ALBERDI',
                'ALBERTI- PLA':'ALBERTI.PLA',
                'ARROYO DEL MEDIO':'ARROYO.DEL.MEDIO',
                'BRAGADO':'BRAGADO',
                'GIDAG':'GIDAG',
                'PERGAMINO':'PERGAMINO',
                'RAWSON TRES SARGENTOS':'RAWSON.TRES.SARGENTOS',
                'SAN ANTONIO DE ARECO':'SAN.ANTONIO.DE.ARECO',
                'SAN PEDRO VILLA LIA':'SAN.PEDRO.VILLA.LIA',
                'SEGUI LA ORIENTAL':'SEGUI.LA.ORIENTAL',
                'CAÑAVERALES DE TUCUMAN':'CAÑAVERALES.DE.TUCUMAN',
                'EL PALOMAR':'EL.PALOMAR',
                'EL RODEO':'EL.RODEO',
                'LA COCHA':'LA.COCHA',
                'LOS ALGARROBOS':'LOS.ALGARROBOS',
                'METAN':'METAN',
                'PALO SANTO':'Otro',
                'SAN PATRICIO':'SAN.PATRICIO',
                'SANTA ROSA CATAMARCA':'SANTA.ROSA.CATAMARCA',
                'SURCOS':'Otro',
                'SUYAY':'SUYAY',
                'VALLES TEMPLADOS':'Otro',
                'YUNGAS':'YUNGAS',
                'CUÑA BOSCOSA':'CUÑA.BOSCOSA',
                'MARGARITA CAMPO ALEMAN':'MARGARITA.CAMPO.ALEMAN',
                'RAMAYON':'RAMAYON',
                'SAN CRISTOBAL-LA LUCILA':'SAN.CRISTOBAL.LALUCILA',
                'VILLA ANA-ARANDU':'VILLA.ANA.ARANDÚ',
                'VILLA OCAMPO':'VILLA.OCAMPO',
                'AMERICA':'AMERICA',
                'AMERICA II':'AMERICA.II',
                'AMERICA LECHERO':'AMERICA.LECHERO',
                'ATREUCO':'ATREUCO',
                'CORRALERO':'CORRALERO',
                'PELLEGRINI - TRES LOMAS':'PELLEGRINI.TRES.LOMAS',
                'PICO BARON':'PICO.BARON',
                'PICO QUEMU':'PICO.QUEMU',
                'QUEMU CATRILO':'QUEMU.CATRILO',
                'TRENQUE LAUQUEN II':'TRENQUE.LAUQUEN.II',
                'AGROGANADERO 9 DE JULIO':'AGROGANADERO.9.DE.JULIO',
                'AMEGHINO':'AMEGHINO',
                'BOLIVAR':'BOLIVAR',
                'CASARES - 9 DE JULIO':'CASARES.9.DE.JULIO',
                'GENERAL PINTO':'GENERAL.PINTO',
                'GENERAL VILLEGAS':'GENERAL.VILLEGAS',
                'GUANACO LAS TOSCAS':'GUANACO.LAS.TOSCAS',
                'HENDERSON-DAIREAUX':'HENDERSON.DAIREAUX',
                'HERRERA VEGAS - PEHUAJO':'HERRERA.VEGAS.PEHUAJO',
                'INFOSURA':'INFOSURA',
                'LA VIA':'LA.VIA',
                'LINCOLN':'LINCOLN',
                'MONES CAZON PEHUAJO':'MONES.CAZON.PEHUAJO',
                'NUEVE DE JULIO':'NUEVE.DE.JULIO',
                'PEHUAJO CASARES':'PEHUAJO.CASARES',
                'PIROVANO LA LARGA':'PIROVANO.LA.LARGA',
                'SALAZAR MONES CAZON':'SALAZAR.MONES.CAZON',
                'TAMBERO AMEGHINO VILLEGAS':'TAMBERO.AMEGHINO.VILLEGAS',
                'TEJEDOR':'TEJEDOR',
                'TREINTA AGOSTO- MARI LAUQUEN':'TREINTA.AGOSTO.MARI.LAUQUEN',
                'ALTO VALLE - VALLE MEDIO':'ALTO.VALLE.VALLE.MEDIO',
                'CUENCA DEL AGRIO':'CUENCA.DEL.AGRIO',
                'ESQUEL':'ESQUEL',
                'LANIN':'LANIN',
                'SANTA CRUZ':'SANTACRUZ',
                'TIERRA DEL FUEGO':'TIERRADELFUEGO',
                'VIEDMA':'VIEDMA',
                'AOKEN AL':'AOKEN.AL',
                'CALEUCHE':'CALEUCHE',
                'GUATRACHE':'GUATRACHE',
                'HOLISTICO':'HOLISTICO',
                'PEHUENCHE':'PEHUENCHE',
                'SOVEN':'SOVEN',
                'TAMBERO GUATRACHE':'TAMBERO.GUATRACHE',
                'UTRACAN':'UTRACAN',
                '25 DE MAYO':'CREA.25.DE.MAYO',
                'ARROYO DE LAS FLORES':'ARROYO.DE.LAS.FLORES',
                'ARROYO LANGUEYU':'ARROYO.LANGUEYU',
                'AYACUCHO':'AYACUCHO',
                'CASTELLI - BELGRANO':'CASTELLI.BELGRANO',
                'DEL TUYU':'DEL.TUYU',
                'FORTIN MULITAS':'FORTIN.MULITAS',
                'LEZAMA':'LEZAMA',
                'MAIPU':'MAIPU',
                'MAR CHIQUITA':'MAR.CHIQUITA',
                'MONTE':'MONTE',
                'PILA':'PILA',
                'RAUCH - UDAQUIOLA':'RAUCH.UDAQUIOLA',
                'RIO SALADO':'RIO.SALADO',
                'ROQUE PEREZ SALADILLO':'ROQUE.PEREZ.SALADILLO',
                'TAPALQUE II':'TAPALQUE.II',
                'VALLIMANCA':'VALLIMANCA',
                'CASTELAR':'CASTELAR',
                'CENTRO OESTE SANTAFESINO':'C.O.S',
                'CUENCA':'CUENCA',
                'EL CEIBO':'EL.CEIBO',
                'ESPERANZA':'ESPERANZA',
                'GALVEZ':'GALVEZ',
                'RAFAELA':'RAFAELA',
                'SAN FRANCISCO':'SAN.FRANCISCO',
                'SAN GUILLERMO':'Otro',
                'SAN MARTIN DE LAS ESCOBAS - COLONIA BELGRANO':'S.M.E.C.B',
                'SUNCHALES':'Otro',
                'APLICADORES':'Otro',
                'ARMSTRONG MONTES DE OCA':'ARMSTRONG.MONTES.DE.OCA',
                'ASCENSION':'ASCENSION',
                'COLONIA MEDICI':'COLONIA.MEDICI',
                'COSTAS DEL CARCARAÑA':'COSTAS.DEL.CARCARAÑA',
                'EL ABROJO':'EL.ABROJO',
                'GENERAL ARENALES':'GENERAL.ARENALES',
                'GENERAL BALDISSERA':'GENERAL.BALDISSERA',
                'LA CALANDRIA':'LA.CALANDRIA',
                'LA MAROMA':'LA.MAROMA',
                'LAS PETACAS':'LAS.PETACAS',
                'MARIA TERESA':'MARIA.TERESA',
                'MONTE BUEY INRIVILLE':'MONTE.BUEY.INRIVILLE',
                'MONTE MAIZ':'MONTE.MAIZ',
                'POSTA ESPINILLOS':'POSTA.ESPINILLOS',
                'ROSARIO':'ROSARIO',
                'SAN JORGE LAS ROSAS':'SAN.JORGE.LAS.ROSAS',
                'SANTA ISABEL':'SANTA.ISABEL',
                'SANTA MARIA':'SANTA.MARIA',
                'TEODELINA':'TEODELINA',
                'BENITO JUAREZ':'BENITO.JUAREZ',
                'CARHUE HUANGUELEN':'CARHUE.HUANGUELEN',
                'CORONEL PRINGLES II':'CORONEL.PRINGLES.II',
                'CORONEL SUAREZ':'CORONEL.SUAREZ',
                'GENERAL LAMADRID':'GENERAL.LAMADRID',
                'LAPRIDA':'LAPRIDA',
                'NUESTRA SEÑORA DE LAS PAMPAS':'NUESTRA.SEÑORA.DE.LAS.PAMPAS',
                'OLAVARRIA':'OLAVARRIA',
                'PEDRO LURO':'PEDRO.LURO',
                'SAN ELOY - PIÑEYRO':'SAN.ELOY.PIÑEYRO',
                'VENTANIA':'VENTANIA',
                'ACONCAGUA':'Otro',
                'ARAUCO':'Otro',
                'CALCHAQUI':'Otro',
                'FRUTICOLA CUYO':'Otro',
                'HUARPE':'Otro',
                'LAS ACEQUIAS':'Otro',
                'LOS ANDES':'Otro',
                'NOGALERO DEL NORTE':'Otro',
                'OLIVICOLA SAN JUAN':'Otro',
                'VIGNERONS':'Otro'
                }
                
                # Reemplazo directo sin función
                for clave, valor in grupos_dat_dict.items():
                    grupo = grupo.replace(clave, valor)
                
                #print('GRUPO: ', grupo)

                # Filtro el DataFrame
                # CULTIVO VERANO
                def filtrar_cultivos_verano(df, conn):
                    query = """SELECT cultivo FROM datcrea_tablas.cultivos WHERE ciclo = 'v'"""
                    verano_df = pd.read_sql(query, conn)
                    cultivos_verano = verano_df['cultivo'].tolist()
                    df = df[df['Cultivo'].isin(cultivos_verano)]
                    return df
                
                verano = filtrar_cultivos_verano(df, conn)
                
                def exportar_sima_verano(verano):
                    
                    seccion = 200 # Máx registros por planilla xls
                    registros_campana = len(verano.index) # cuenta el total de registros
                    registros_wb = math.ceil(registros_campana / seccion) #cociente para saber cantidad de worksheets
                    
                    # Armo listado de subset de 200 (máx) registros cada uno
                    dataframes = []
                    
                    for i in range(registros_wb):
                        inicio = i * seccion
                        fin = (i + 1) * seccion
                        seccion_df = verano.iloc[inicio:fin]
                        dataframes.append(seccion_df)
                    
                    # Crear una workbook para cada dataframe
                    
                    # Ruta del archivo original
                    ruta_original = 'C:/Users/EPolacco/Documents/9 - DAT/SIMA/DAT-Cultivos-de-verano-2023-24-1.xlsx'
                    
                    # Cargar el archivo original
                    wb_original = load_workbook(filename=ruta_original)
                    
                    # Obtener la hoja original
                    planilla_original = wb_original['Planilla ']
                    
                    # Columnas excluídas porque tienen fórmulas y validaciones.
                    columns_to_exclude = {'I', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'U', 'W', 'Y', 'Z', 'AC',
                                        'AF', 'AG', 'AH', 'AK', 'AO', 'CC', 'CS', 'CU', 'CV', 'CW', 'CX',
                                        'CY', 'CZ', 'DI', 'DJ', 'DK','DL', 'DM', 'DN','DO','DP','DQ','DR', 
                                        'DS','DT','DU', 'DV', 'DW', 'DY', 'DZ', 'EA',
                                        'IQ','IR','IS','IT','IU','IV','IW','IX','IY','IZ', 'JA', 'JB', 'JC',
                                        'JD', 'JE'}
                    
                    
                    # Iterar sobre los df y guardar en archivos xls separados
                    for i, seccion_df in enumerate(dataframes, 1):
                        # Crear un workbook para cada sección
                        if i == 1:
                            wb = wb_original  # Utilizar el archivo original para el primer df
                            sheet = planilla_original
                        else:
                            # Crear una copia del archivo original para los df subsiguientes
                            ruta_copia = f'C:/Users/EPolacco/Documents/9 - DAT/SIMA/{cuic}_DAT-Cultivos-de-verano-2023-24-1_SECCION_{i}.xlsx'
                            shutil.copy(ruta_original, ruta_copia)
                            
                            # Cargar la copia
                            wb = load_workbook(filename=ruta_copia)
                            sheet = wb['Planilla ']
                            sheet.title = f'Planilla '
                    
                        # Borrar datos existentes en la hoja, después de la fila 14 porque sino se pisan los datos
                        
                        for row in sheet.iter_rows(min_row=14, max_row=sheet.max_row, min_col=6, max_col=sheet.max_column):
                            for cell in row:
                                if cell.column_letter not in columns_to_exclude:
                                    cell.value = None
                    
                    
                        # Obtener las filas de datos del df
                        rows = dataframe_to_rows(seccion_df,index=False, header=False)
                        
                        # Copiar en worksheet a partir de la fila 14 y columna 6
                        for r_idx, row in enumerate(rows, 14):
                            for c_idx, value in enumerate(row, 6):
                                if not pd.isna(value):
                                    sheet.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Región
                        sheet['F3'] = region
                        # Grupo
                        sheet['F6'] = grupo
                        # Guardar el workbook, si no es el primer df
                        if i != 1:
                            wb.save(ruta_copia)
                    
                    
                    # Guardar workbook con el primer df
                    ruta_alternativa = f'C:/Users/EPolacco/Documents/9 - DAT/SIMA/{cuic}_DAT-Cultivos-de-verano-2023-24-1_SECCION_1.xlsx'
                    wb_original.save(ruta_alternativa)
                
                exportar_sima_verano(verano)
            
            else: #Invierno
                # Importo librerías
                from openpyxl import load_workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                import pandas as pd
                import numpy as np
                from datetime import datetime
                import math
                import shutil

                # Conexión a la base
                from sqlalchemy import create_engine
                import mysql.connector
                
                host='66.97.33.201'
                password='Dat##2023Tabla$'
                port=33306
                user='datcrea_tablas'
                database='datcrea_tablas'
                
                conn = create_engine("mysql+mysqlconnector://{user}:{password}@{host}:{port}/{database}".format(host=host,user=user,password=password,database=database,port=port),connect_args={'auth_plugin': 'mysql_native_password'})
                
            
                #ruta_completa = directorio + archivo
                #df = pd.read_excel(ruta_completa)
                cuic = df['CUIC'][0]
                
                def filtrar_campana(df, campana_filtro):
                    
                    filtro = df['Ciclo'] == campana_filtro
                    df = df[filtro]
                    return df

                df = filtrar_campana(df, campana_filtro)

                
                def formato_fecha(df):
                    df['FechaSiembra'] = df['FechaSiembra'].str[:10] #Elimino horas en fechas
                    df['FechaSiembra'] = pd.to_datetime(df['FechaSiembra'])
                    df['FechaSiembra'] = df['FechaSiembra'].dt.strftime('%d/%m/%Y')    
                    return df
                    
                df = formato_fecha(df)


                # 3.2) Validar código genética #FUNCIONA
                # Genética
                # Consultar los nro_registro desde la base de datos
                def validar_codigo_genetica(df, conn):
                    query = "SELECT nro_registro FROM datcrea_tablas.materiales WHERE nro_registro IS NOT NULL"
                    df_materiales = pd.read_sql(query, conn)
                    materiales = df_materiales['nro_registro'].astype(int).tolist()
                    df['Codigo_Genetica'] = df['Codigo_Genetica'].fillna(0)
                    df['Codigo_Genetica'] = df['Codigo_Genetica'].astype('Int64')
                    df['Codigo_Genetica'] = ['' if x not in materiales else x for x in df['Codigo_Genetica']]
                    return df

                df = validar_codigo_genetica(df, conn)



                ## 3.3) Validación Cultivo #FUNCIONA
                # Cultivos
                def validar_cultivos(df, conn):
                    cultivos_query = """SELECT cultivo FROM datcrea_tablas.cultivos"""
                    cultivos_df = pd.read_sql(cultivos_query,con=conn)
                    cultivos = cultivos_df['cultivo'].tolist()
                    df['Cultivo'] = df['Cultivo'].replace(['Soja - De segunda', 'Mani'], ['Soja', 'Maní'])
                    df['Cultivo'] = ['' if x not in cultivos else x for x in df['Cultivo']]
                    return df
                
                df = validar_cultivos(df,conn)

                ## 3.4) Validación Genética
                
                # Cultivos
                def validar_genetica(df,conn): #FUNCIONA
                    
                    genetica_query = """SELECT material FROM datcrea_tablas.materiales"""
                    genetica_df = pd.read_sql(genetica_query,con=conn)
                    
                    genetica = genetica_df['material'].tolist()
                    df['Genetica'] = ['' if x not in genetica else x for x in df['Genetica']]
                    return df
                    
                df = validar_genetica(df,conn)

                ## 3.9) Validación Semillero #FUNCIONA
                def validar_semillero(df, conn): 
                    
                    semillero_query = """SELECT semillero FROM datcrea_tablas.semilleros"""
                    semillero = pd.read_sql(semillero_query,con=conn)
                    semillero = semillero['semillero'].tolist()
                    df['Semillero'] = ['' if x not in semillero else x for x in df['Semillero']]
                    return df
                
                df = validar_semillero(df, conn)

                def validar_genetica_semillero(df, conn): #FUNCIONA
                    # VALIDA GENÉTICA Y SEMILLERO EN BASE A CÓDIGO GENÉTICO Y CULTIVO
                    asignaciones = [
                        (('Girasol', '18615'), ('Nidera.Girasol', '106 CL HO')),
                        (('Trigo', '9993'), ('Aca.Trigo', '1801F')),
                        (('Trigo', '11270'), ('Aca.Trigo', '1901F')),
                        (('Soja', '19320'), ('Stine.Soja', '25EB32')),
                        (('Soja', '19321'), ('Stine.Soja', '29EB02')),
                        (('Maíz', '16570'), ('Nuseed.Maíz', '3790 RR2-CL')),
                        (('Soja','19324'), ('Stine.Soja','38EB03')),
                        (('Soja','19348'), ('Nidera.Soja','4031 E NS')),
                        (('Soja','19326'), ('Stine.Soja','40EB20')),
                        (('Soja','17688'), ('MacroSeed.Soja','40MS01 E STS')),
                        (('Soja','20271'), ('Stine.Soja','45EB52')),
                        (('Soja','20277'), ('Stine.Soja','47EA32')),
                        (('Soja','15994'), ('MacroSeed.Soja','47MS01 STS')),
                        (('Soja','19331'), ('Stine.Soja','48EB20')),
                        (('Soja','18457'), ('MacroSeed.Soja','48MS01 E')),
                        (('Soja','15267'), ('LaTijereta.Soja','4914 IPRO')),
                        (('Soja','20958'), ('Nidera.Soja','4941 E STS NS')),
                        (('Soja','16661'), ('MacroSeed.Soja','50MS01 STS')),
                        (('Soja','20069'), ('MacroSeed.Soja','51MS01 E')),
                        (('Soja','17113'), ('MacroSeed.Soja','53MS01 IPRO')),
                        (('Soja','17640'), ('MacroSeed.Soja','53MS02 STS')),
                        (('Soja','16545'), ('MacroSeed.Soja','59MS01 IPRO STS')),
                        (('Soja','15479'), ('MacroSeed.Soja','60MS01 STS')),
                        (('Soja','19100'), ('MacroSeed.Soja','61MS01 STS')),
                        (('Maíz','20857'), ('Nidera.Maíz','6223 CE')),
                        (('Soja','20857'), ('Nidera.Soja','6223 CE')),
                        (('Soja','16717'), ('MacroSeed.Soja','62MS01 IPRO STS')),
                        (('Soja','20099'), ('MacroSeed.Soja','62MS02 E')),
                        (('Soja','20175'), ('MacroSeed.Soja','63MS01 CE')),
                        (('Soja','20194'), ('Credenz.Soja','6522')),
                        (('Soja','17639'), ('MacroSeed.Soja','66MS01')),
                        (('Soja','20171'), ('MacroSeed.Soja','68MS01 CE')),
                        (('Soja','16669'), ('MacroSeed.Soja','78MS01 IPRO')),
                        (('Soja','6915'), ('Nidera.Soja','A 3302 RG')),
                        (('Soja','9986'), ('Nidera.Soja','A 3731 RG')),
                        (('Soja','6165'), ('Nidera.Soja','A 3901 RG')),
                        (('Soja','9217'), ('Nidera.Soja','A 4209 RG')),
                        (('Soja','7246'), ('Nidera.Soja','A 4303 RG')),
                        (('Soja','8272'), ('Nidera.Soja','A 4505 RG')),
                        (('Soja','9214'), ('Nidera.Soja','A 4553 RG')),
                        (('Soja','8756'), ('Nidera.Soja','A 4613 RG')),
                        (('Soja','10082'), ('Nidera.Soja','A 5009 RG')),
                        (('Soja','5829'), ('Nidera.Soja','A 5409 RG')),
                        (('Soja','6742'), ('Nidera.Soja','A 5503')),
                        (('Soja','11072'), ('Nidera.Soja','A 5509 RG')),
                        (('Soja','11073'), ('Nidera.Soja','A 5909 RG')),
                        (('Soja','8759'), ('Nidera.Soja','A 6126 RG')),
                        (('Soja','8273'), ('Nidera.Soja','A 6411 RG')),
                        (('Soja','10083'), ('Nidera.Soja','A 6509 RG')),
                        (('Soja','6739'), ('Nidera.Soja','A 7636 RG')),
                        (('Soja','5828'), ('Nidera.Soja','A 8000 RG')),
                        (('Colza','11426'), ('HighTech.Colza','ABILITY')),
                        (('Trigo','10021'), ('Aca.Trigo','ACA 201')),
                        (('Trigo','11267'), ('Aca.Trigo','ACA 202')),
                        (('Girasol','10219'), ('ACA.Girasol','ACA 203')),
                        (('Trigo','7632'), ('Aca.Trigo','ACA 302')),
                        (('Trigo','7633'), ('Aca.Trigo','ACA 303')),
                        (('Trigo','15747'), ('Aca.Trigo','ACA 303 plus')),
                        (('Trigo','8470'), ('Aca.Trigo','ACA 304')),
                        (('Trigo','20497'), ('Aca.Trigo','ACA 308')),
                        (('Trigo','9530'), ('Aca.Trigo','ACA 315')),
                        (('Trigo','11271'), ('Aca.Trigo','ACA 320')),
                        (('Girasol','15388'), ('ACA.Girasol','ACA 350')),
                        (('Trigo','14471'), ('Aca.Trigo','ACA 360')),
                        (('Soja','7203'), ('ACA.Soja','ACA 360 GR')),
                        (('Trigo','21453'), ('Aca.Trigo','ACA 363')),
                        (('Trigo','21451'), ('Aca.Trigo','ACA 364')),
                        (('Soja','7674'), ('ACA.Soja','ACA 420 GR')),
                        (('Soja','8395'), ('ACA.Soja','ACA 460 GR')),
                        (('Maíz','13902'), ('ACA.Maíz','ACA 470')),
                        (('Soja','7202'), ('ACA.Soja','ACA 470 GR')),
                        (('Maíz','14836'), ('ACA.Maíz','ACA 474')),
                        (('Maíz','14087'), ('ACA.Maíz','ACA 480')),
                        (('Soja','7697'), ('ACA.Soja','ACA 480 GR')),
                        (('Soja','3937'), ('ACA.Soja','ACA 490')),
                        (('Maíz','16320'), ('ACA.Maíz','ACA 493')),
                        (('Soja','16543'), ('ACA.Soja','ACA 4949 IPRO')),
                        (('Soja','15547'), ('ACA.Soja','ACA 5020 IPRO')),
                        (('Maíz','14089'), ('ACA.Maíz','ACA 514 Flint')),
                        (('Maíz','14088'), ('ACA.Maíz','ACA 530 Flint')),
                        (('Soja','7204'), ('ACA.Soja','ACA 530 GR')),
                        (('Sorgo','14103'), ('ACA.Sorgo','ACA 548')),
                        (('Sorgo','1846'), ('ACA.Sorgo','ACA 550')),
                        (('Sorgo','5975'), ('ACA.Sorgo','ACA 558')),
                        (('Soja','3940'), ('ACA.Soja','ACA 560')),
                        (('Sorgo','10448'), ('ACA.Sorgo','ACA 561')),
                        (('Sorgo','17570'), ('ACA.Sorgo','ACA 563')),
                        (('Soja','7698'), ('ACA.Soja','ACA 570 GR')),
                        (('Maíz','16019'), ('ACA.Maíz','ACA 580 Flint')),
                        (('Soja','15265'), ('ACA.Soja','ACA 5814 IPRO')),
                        (('Soja','15394'), ('ACA.Soja','ACA 5825 IPRO')),
                        (('Trigo','8067'), ('Aca.Trigo','ACA 601')),
                        (('Trigo','14472'), ('Aca.Trigo','ACA 602')),
                        (('Trigo','19477'), ('Aca.Trigo','ACA 604')),
                        (('Alfalfa','9455'), ('ACA.Alfalfa','ACA 605')),
                        (('Maíz','16020'), ('ACA.Maíz','ACA 622 Flint')),
                        (('Soja','14613'), ('ACA.Soja','ACA 6513 IPRO')),
                        (('Soja','20178'), ('ACA.Soja','ACA 70A70 CE')),
                        (('Soja','16023'), ('ACA.Soja','ACA 7890 IPRO')),
                        (('Trigo','7979'), ('Aca.Trigo','ACA 801')),
                        (('Soja','14949'), ('ACA.Soja','ACA 8080 IPRO')),
                        (('Girasol','9307'), ('ACA.Girasol','ACA 861')),
                        (('Girasol','15884'), ('ACA.Girasol','ACA 869')),
                        (('Girasol','17526'), ('ACA.Girasol','ACA 870')),
                        (('Girasol','2330'), ('ACA.Girasol','ACA 884')),
                        (('Girasol','11822'), ('ACA.Girasol','ACA 887')),
                        (('Girasol','15885'), ('ACA.Girasol','ACA 889')),
                        (('Trigo','9520'), ('Aca.Trigo','ACA 901')),
                        (('Alfalfa','10740'), ('ACA.Alfalfa','ACA 903')),
                        (('Trigo','10740'), ('Aca.Trigo','ACA 903')),
                        (('Trigo','10739'), ('Aca.Trigo','ACA 905PA')),
                        (('Trigo','12340'), ('Aca.Trigo','ACA 906')),
                        (('Trigo','14473'), ('Aca.Trigo','ACA 908')),
                        (('Trigo','15733'), ('Aca.Trigo','ACA 909')),
                        (('Trigo','19479'), ('Aca.Trigo','ACA 917')),
                        (('Maíz','16139'), ('ACA.Maíz','ACA VG 48')),
                        (('Alfalfa','2974'), ('LosPrados.Alfalfa','ACONCAGUA')),
                        (('Sorgo','12653'), ('AdSur.Sorgo','AD-91 SUCROL')),
                        (('Sorgo','15370'), ('Advanta.Sorgo','ADV 1100')),
                        (('Sorgo','17874'), ('Advanta.Sorgo','ADV 1114')),
                        (('Sorgo','12726'), ('Advanta.Sorgo','ADV 114')),
                        (('Sorgo','16870'), ('Advanta.Sorgo','ADV 1250 IG')),
                        (('Sorgo','16415'), ('Advanta.Sorgo','ADV 1302')),
                        (('Sorgo','16414'), ('Advanta.Sorgo','ADV 1305')),
                        (('Sorgo','16972'), ('Advanta.Sorgo','ADV 1350 IG')),
                        (('Sorgo','14804'), ('Advanta.Sorgo','ADV 2010')),
                        (('Maíz','19860'), ('Advanta.Maíz','ADV 23.0')),
                        (('Sorgo','18582'), ('Advanta.Sorgo','ADV 2450 IG')),
                        (('Sorgo','14138'), ('Advanta.Sorgo','ADV 2499')),
                        (('Sorgo','17702'), ('Advanta.Sorgo','ADV 2701')),
                        (('Sorgo','14137'), ('Advanta.Sorgo','ADV 2800')),
                        (('Girasol','14857'), ('Advanta.Girasol','ADV 5200')),
                        (('Girasol','17932'), ('Advanta.Girasol','ADV 5304')),
                        (('Girasol','21663'), ('Advanta.Girasol','ADV 5310')),
                        (('Girasol','20601'), ('Advanta.Girasol','ADV 5407')),
                        (('Girasol','16183'), ('Advanta.Girasol','ADV 5500')),
                        (('Girasol','19063'), ('Advanta.Girasol','ADV 5566')),
                        (('Maíz','14793'), ('Advanta.Maíz','ADV 8101')),
                        (('Maíz','18158'), ('Advanta.Maíz','ADV 8560 T')),
                        (('Maíz','19859'), ('Advanta.Maíz','ADV 8570 T')),
                        (('Sorgo','14396'), ('Agseed.Sorgo','AG 1816')),
                        (('Sorgo','14393'), ('Agseed.Sorgo','AG 1817')),
                        (('Maíz','16719'), ('Agseed.Maíz','AG 7004')),
                        (('Maíz','11799'), ('Agseed.Maíz','AG 8000')),
                        (('Maíz','9577'), ('Agseed.Maíz','AG 9005')),
                        (('Maíz','17727'), ('Agseed.Maíz','AG 9300')),
                        (('Maíz','18407'), ('Agseed.Maíz','AG 9400')),
                        (('Sorgo','6874'), ('Agseed.Sorgo','AG SILO 200')),
                        (('Trigo','13123'), ('Buck.Trigo','AGP 127')),
                        (('Trigo','11565'), ('Buck.Trigo','AGP FAST')),
                        (('Girasol','11685'), ('Advanta.Girasol','AGUARA 6')),
                        (('Trigo','22406'), ('Bioseminis.Trigo','AGUARIBAY')),
                        (('Trigo','21629'), ('Buck.Trigo','AIMARÁ')),
                        (('Cebada','4675'), ('INTA.Cebada','ALICIA INTA ')),
                        (('Soja','8276'), ('ALM.Soja','ALM 3530')),
                        (('Soja','10565'), ('ALM.Soja','ALM 3830')),
                        (('Soja','8078'), ('ALM.Soja','ALM 4650')),
                        (('Soja','9850'), ('ALM.Soja','ALM 4930')),
                        (('Soja','8787'), ('Sursem.Soja','ANDREA 60')),
                        (('Soja','8785'), ('Sursem.Soja','ANDREA 63')),
                        (('Soja','8342'), ('Sursem.Soja','ANTA81')),
                        (('Maíz','16964'), ('AlumniSeed.Maíz','AP2505')),
                        (('Maíz','19373'), ('AlumniSeed.Maíz','AP2509')),
                        (('Maíz','16965'), ('AlumniSeed.Maíz','AP4512')),
                        (('Maíz','12343'), ('AlumniSeed.Maíz','AP6002')),
                        (('Maíz','16963'), ('AlumniSeed.Maíz','AP6005')),
                        (('Maíz','18612'), ('AlumniSeed.Maíz','AP8204')),
                        (('Maíz','18613'), ('AlumniSeed.Maíz','AP8205')),
                        (('Trigo','22343'), ('Bioseminis.Trigo','ARAZÁ')),
                        (('Trigo','21523'), ('Bioseminis.Trigo','ARCE')),
                        (('Maíz','17593'), ('Argenetics.Maíz','ARG 7712 BT RR')),
                        (('Maíz','19480'), ('Argenetics.Maíz','ARG 7716 BT RR')),
                        (('Maíz','15925'), ('Argenetics.Maíz','ARG 7730 BT')),
                        (('Maíz','15924'), ('Argenetics.Maíz','ARG 7732 BT CL')),
                        (('Maíz','17449'), ('Argenetics.Maíz','ARG 7742 FCL')),
                        (('Maíz','17535'), ('Argenetics.Maíz','ARG 8410 BT')),
                        (('Maíz','17450'), ('Argenetics.Maíz','ARG 8500 T')),
                        (('Maíz','17452'), ('Argenetics.Maíz','ARG 8800 T')),
                        (('Maíz','17454'), ('Argenetics.Maíz','ARG 8900 W')),
                        (('Maíz','15416'), ('Argenetics.Maíz','ARGENPOP 141')),
                        (('Maíz','15420'), ('Argenetics.Maíz','ARGENPOP 142')),
                        (('Sorgo','13412'), ('Argenetics.Sorgo','ARGENSOR 110 T')),
                        (('Sorgo','11806'), ('Argenetics.Sorgo','ARGENSOR 121')),
                        (('Sorgo','14106'), ('Argenetics.Sorgo','ARGENSOR 125 B')),
                        (('Sorgo','13614'), ('Argenetics.Sorgo','ARGENSOR 130 T')),
                        (('Sorgo','15428'), ('Argenetics.Sorgo','ARGENSOR 134 T')),
                        (('Sorgo','11805'), ('Argenetics.Sorgo','ARGENSOR 151 DP')),
                        (('Sorgo','14783'), ('Argenetics.Sorgo','ARGENSOR 155 DP')),
                        (('Maíz','13859'), ('Arvales.Maíz','ARV 2155')),
                        (('Maíz','12510'), ('Arvales.Maíz','ARV 2180')),
                        (('Maíz','12498'), ('Arvales.Maíz','ARV 2194')),
                        (('Maíz','12499'), ('Arvales.Maíz','ARV 2310')),
                        (('Sorgo','14782'), ('Arvales.Sorgo','ARV 300')),
                        (('Sorgo','13300'), ('Arvales.Sorgo','ARV 390')),
                        (('Soja','12013'), ('ASP.Soja','AS 3601')),
                        (('Soja','12400'), ('ASP.Soja','AS 3911')),
                        (('Soja','9857'), ('ASP.Soja','AS 4201')),
                        (('Soja','12037'), ('ASP.Soja','AS 4402')),
                        (('Soja','16421'), ('ASP.Soja','AS 4731')),
                        (('Soja','9764'), ('ASP.Soja','AS 4801')),
                        (('Soja','13543'), ('ASP.Soja','AS 4931')),
                        (('Soja','10573'), ('ASP.Soja','AS 5308i')),
                        (('Maní','14799'), ('INTA.Maní','ASEM 400 INTA')),
                        (('Maní','13703'), ('INTA.Maní','ASEM VICTOR INTA')),
                        (('Alfalfa','2394'), ('BayaCasal.Alfalfa','AURORA')),
                        (('Soja','16521'), ('ASGROW.Soja','AW 3806 IPRO')),
                        (('Soja','18422'), ('ASGROW.Soja','AW 3920 IPRO')),
                        (('Soja','18427'), ('ASGROW.Soja','AW 4320 IPRO')),
                        (('Soja','16408'), ('ASGROW.Soja','AW 4326 IPRO')),
                        (('Soja','18435'), ('ASGROW.Soja','AW 4610 IPRO')),
                        (('Soja','16420'), ('ASGROW.Soja','AW 4736 IPRO')),
                        (('Soja','16885'), ('ASGROW.Soja','AW 4927 IPRO')),
                        (('Soja','18441'), ('ASGROW.Soja','AW 5021 IPRO')),
                        (('Soja','15320'), ('ASGROW.Soja','AW 5714 IPRO')),
                        (('Soja','15318'), ('ASGROW.Soja','AW 5815 IPRO')),
                        (('Soja','18437'), ('ASGROW.Soja','AW 5920 IPRO')),
                        (('Soja','14645'), ('ASGROW.Soja','AW 6211 IPRO')),
                        (('Soja','18425'), ('ASGROW.Soja','AW 6320 IPRO')),
                        (('Soja','17787'), ('ASGROW.Soja','AW 7720 IPRO')),
                        (('Maíz','15325'), ('Nidera.Maíz','AX 7761')),
                        (('Maíz','16558'), ('Nidera.Maíz','AX 7784')),
                        (('Maíz','15326'), ('Nidera.Maíz','AX 7822')),
                        (('Maíz','15862'), ('Nidera.Maíz','AX 7918')),
                        (('Maíz','15862'), ('Nidera.Maíz','AX 7919')),
                        (('Maíz','16557'), ('Nidera.Maíz','AX 8010')),
                        (('Maíz','10269'), ('Nidera.Maíz','AX 852')),
                        (('Maíz','12677'), ('Nidera.Maíz','AX 887')),
                        (('Soja','8588'), ('Sursem.Soja','AYELEN22')),
                        (('Soja','7592'), ('Sursem.Soja','AZUL35')),
                        (('Colza','15253'), ('Limagrain.Colza','Albatros')),
                        (('Trigo','18190'), ('DonMario.Trigo','Alerce')),
                        (('Alfalfa','19212'), ('Gentos.Alfalfa','Alfalfa Latte 920')),
                        (('Alfalfa','13853'), ('Gentos.Alfalfa','Alfalfa Nobel 620')),
                        (('Alfalfa','15109'), ('Gentos.Alfalfa','Alfalfa Nobel 720')),
                        (('Trigo','15699'), ('DonMario.Trigo','Algarrobo')),
                        (('Trigo','14576'), ('LG.Trigo','Alhambra')),
                        (('Cebada','18704'), ('CyMQuilmes.Cebada','Alhue')),
                        (('Soja','12135'), ('INTA.Soja','Alim 5.09')),
                        (('Alfalfa','19944'), ('Forratec.Alfalfa','AlphaTec 621')),
                        (('Alfalfa','19945'), ('Forratec.Alfalfa','AlphaTec 821')),
                        (('Alfalfa','19947'), ('Forratec.Alfalfa','AlphaTec 921')),
                        (('Cebada','13129'), ('MalteriaPampa.Cebada','Andreia')),
                        (('Trigo','12406'), ('DonMario.Trigo','Arex')),
                        (('Girasol','10945'), ('Argenetics.Girasol','Argensol 20')),
                        (('Girasol','17453'), ('Argenetics.Girasol','Argensol 20 Max')),
                        (('Girasol','13316'), ('Argenetics.Girasol','Argensol 40')),
                        (('Girasol','18034'), ('Argenetics.Girasol','Argensol 54 AO')),
                        (('Girasol','17448'), ('Argenetics.Girasol','Argensol 72 CL')),
                        (('Girasol','10945'), ('Argenetics.Girasol','Argensol 76')),
                        (('Trigo','17150'), ('LG.Trigo','Arlask')),
                        (('Trigo','19732'), ('DonMario.Trigo','Aromo')),
                        (('Trigo','5378'), ('Buck.Trigo','Arriero')),
                        (('Alfalfa','18394'), ('Barenbrug.Alfalfa','Artemis')),
                        (('Trigo','19379'), ('LG.Trigo','Aryal')),
                        (('Trigo','10823'), ('DonMario.Trigo','Atlax')),
                        (('Trigo','17458'), ('DonMario.Trigo','Audaz')),
                        (('Poroto','6992'), ('INTA.Leales.Poroto','Azabache INTA')),
                        (('Maíz','19835'), ('Basso.Maíz','BAS 2206')),
                        (('Maíz','12736'), ('Basso.Maíz','BAS 5802')),
                        (('Maíz','16271'), ('Basso.Maíz','BAS 5803')),
                        (('Maíz','12645'), ('Basso.Maíz','BAS 6101')),
                        (('Maíz','12737'), ('Basso.Maíz','BAS 6102')),
                        (('Maíz','16270'), ('Basso.Maíz','BAS 6104')),
                        (('Trigo','15764'), ('Bioceres.Trigo','BASILIO')),
                        (('Trigo','21706'), ('Bioseminis.Trigo','BERMEJO HB4')),
                        (('Trigo','16330'), ('Bioceres.Trigo','BIO 1008')),
                        (('Soja','12860'), ('Bioceres.Soja','BIO 3.5')),
                        (('Soja','11769'), ('Bioceres.Soja','BIO 3.8')),
                        (('Soja','13547'), ('Bioceres.Soja','BIO 3.9')),
                        (('Soja','10574'), ('Bioceres.Soja','BIO 4.2')),
                        (('Soja','12465'), ('Bioceres.Soja','BIO 4.6')),
                        (('Soja','11302'), ('Bioceres.Soja','BIO 4.8')),
                        (('Soja','13541'), ('Bioceres.Soja','BIO 4.9')),
                        (('Soja','12862'), ('Bioceres.Soja','BIO 5.4')),
                        (('Colza','10066'), ('Nuseed.Colza','BIOAUREO 2386')),
                        (('Colza','10067'), ('Nuseed.Colza','BIOAUREO 2486')),
                        (('Soja','15568'), ('Bioceres.Soja','BIOCERES 3.41')),
                        (('Soja','15967'), ('Bioceres.Soja','BIOCERES 4.11')),
                        (('Soja','15968'), ('Bioceres.Soja','BIOCERES 4.51')),
                        (('Soja','18412'), ('Bioceres.Soja','BIOCERES 4.52')),
                        (('Soja','15966'), ('Bioceres.Soja','BIOCERES 4.91')),
                        (('Soja','15969'), ('Bioceres.Soja','BIOCERES 5.11')),
                        (('Soja','14894'), ('Bioceres.Soja','BIOCERES 5.21')),
                        (('Soja','16769'), ('Bioceres.Soja','BIOCERES 5.92')),
                        (('Colza','9054'), ('Nuseed.Colza','BIOLZA 440')),
                        (('Soja','9920'), ('Buck.Soja','BK42')),
                        (('Soja','18588'), ('BullMark.Soja','BK44P41')),
                        (('Soja','21001'), ('BullMark.Soja','BK44P41 STS')),
                        (('Maíz','13253'), ('Forratec.Maíz','BMR 126')),
                        (('Girasol','20671'), ('Brevant.Girasol','BRV 3304')),
                        (('Soja','21502'), ('Brevant.Soja','BRV 53722 SE')),
                        (('Soja','17857'), ('Brevant.Soja','BRV 54321E')),
                        (('Soja','17855'), ('Brevant.Soja','BRV 54621SE')),
                        (('Soja','19140'), ('Brevant.Soja','BRV 55021SE')),
                        (('Soja','17855'), ('Brevant.Soja','BRV 55621SE')),
                        (('Soja','20211'), ('Brevant.Soja','BRV 56123SCE')),
                        (('Soja','20088'), ('Brevant.Soja','BRV 56222 E')),
                        (('Maíz','19942'), ('Brevant.Maíz','BRV 8380')),
                        (('Maíz','20927'), ('Brevant.Maíz','BRV 8421')),
                        (('Maíz','20112'), ('Brevant.Maíz','BRV 8472')),
                        (('Cebada','21359'), ('Buck.Cebada','BUCK 316')),
                        (('Trigo','9107'), ('Buck.Trigo','BUCK 75 Aniversario')),
                        (('Trigo','20475'), ('Buck.Trigo','BUCK BRAVÍO CL2')),
                        (('Trigo','6246'), ('Nidera.Trigo','Baguette 10')),
                        (('Trigo','6246'), ('Nidera.Trigo','Baguette 11')),
                        (('Trigo','6246'), ('Nidera.Trigo','Baguette 13')),
                        (('Trigo','10602'), ('Nidera.Trigo','Baguette 17')),
                        (('Trigo','10603'), ('Nidera.Trigo','Baguette 18')),
                        (('Trigo','8996'), ('Nidera.Trigo','Baguette 19')),
                        (('Trigo','7815'), ('Nidera.Trigo','Baguette 21')),
                        (('Trigo','10152'), ('Nidera.Trigo','Baguette 30')),
                        (('Trigo','10451'), ('Nidera.Trigo','Baguette 31')),
                        (('Trigo','17153'), ('Nidera.Trigo','Baguette 450')),
                        (('Trigo','13642'), ('Nidera.Trigo','Baguette 501')),
                        (('Trigo','21533'), ('Nidera.Trigo','Baguette 525')),
                        (('Trigo','17979'), ('Nidera.Trigo','Baguette 550')),
                        (('Trigo','13028'), ('Nidera.Trigo','Baguette 560 CL')),
                        (('Trigo','13130'), ('Nidera.Trigo','Baguette 601')),
                        (('Trigo','6246'), ('Nidera.Trigo','Baguette 610')),
                        (('Trigo','17147'), ('Nidera.Trigo','Baguette 620')),
                        (('Trigo','17146'), ('Nidera.Trigo','Baguette 680')),
                        (('Trigo','13642'), ('Nidera.Trigo','Baguette 701')),
                        (('Trigo','16313'), ('Nidera.Trigo','Baguette 750')),
                        (('Trigo','13511'), ('Nidera.Trigo','Baguette 802')),
                        (('Trigo','20447'), ('Nidera.Trigo','Baguette 820')),
                        (('Trigo','9808'), ('Nidera.Trigo','Baguette 9')),
                        (('Trigo','22466'), ('Klein.Trigo','Ballesta')),
                        (('Arveja','15956'), ('Bioseminis.Arveja','Balltrap')),
                        (('Trigo','9942'), ('Buck.Trigo','Baqueano')),
                        (('Alfalfa','5384'), ('Barenbrug.Alfalfa','Baralfa 85')),
                        (('Sorgo','15923'), ('Barenbrug.Sorgo','Bardoble')),
                        (('Cebada','6360'), ('Cargill.Cebada','Barke')),
                        (('Sorgo','15412'), ('Barenbrug.Sorgo','Barkilos')),
                        (('Sorgo','15413'), ('Barenbrug.Sorgo','Barluz')),
                        (('Sorgo','15411'), ('Barenbrug.Sorgo','Barplus')),
                        (('Trigo','15202'), ('Buck.Trigo','Bellaco')),
                        (('Poroto','1761'), ('EEAOC.Poroto','Blanco - TUC 122')),
                        (('Poroto','1760'), ('EEAOC.Poroto','Blanco - TUC 27')),
                        (('Poroto','1762'), ('EEAOC.Poroto','Blanco - TUC 56')),
                        (('Arveja','13345'), ('Inv.Suipachense.Arveja','Bluestar')),
                        (('Trigo','21897'), ('RAGT.Trigo','Borsalino')),
                        (('Trigo','6981'), ('Buck.Trigo','Brasil')),
                        (('Girasol','15435'), ('Buck.Girasol','Buck 355')),
                        (('Girasol','17864'), ('Buck.Girasol','Buck 363')),
                        (('Cebada','563'), ('INTA.Cebada','CALCU INTA ')),
                        (('Cebada','3733'), ('INTA.Cebada','CARLA INTA - MP ')),
                        (('Cebada','755'), ('INTA.Cebada','CAÑUMIL INTA ')),
                        (('Trigo','14470'), ('Aca.Trigo','CEDRO')),
                        (('Sorgo','1622'), ('Agseed.Sorgo','CENTELLA')),
                        (('Sorgo','14574'), ('Agseed.Sorgo','CENTELLA PLUS')),
                        (('Girasol','11698'), ('Advanta.Girasol','CF 101')),
                        (('Girasol','12529'), ('Advanta.Girasol','CF 202')),
                        (('Girasol','8423'), ('Advanta.Girasol','CF 27')),
                        (('Girasol','9312'), ('Advanta.Girasol','CF 31')),
                        (('Trigo','10749'), ('Inia.Trigo','CH 12507')),
                        (('Soja','8235'), ('AGD.Soja','CHAMPAQUI 5.40')),
                        (('Soja','8604'), ('AGD.Soja','CHAMPAQUI 5.80')),
                        (('Soja','8603'), ('AGD.Soja','CHAMPAQUI 5.90')),
                        (('Colza','19722'), ('DSV.Colza','CHIP CL')),
                        (('Trigo','13062'), ('Aca.Trigo','CIPRES')),
                        (('Colza','18066'), ('DSV.Colza','CLARUS')),
                        (('Trigo','18778'), ('Buck.Trigo','COLIHUE')),
                        (('Alfalfa','13329'), ('LosPrados.Alfalfa','CONSTANZA')),
                        (('Soja','9874'), ('AGD.Soja','CQ 4.55')),
                        (('Soja','9865'), ('AGD.Soja','CQ 4.90')),
                        (('Cebada','8689'), ('INTA.Cebada','CRESPA FCA ')),
                        (('Trigo','19048'), ('Buck.Trigo','CUARZO')),
                        (('Alfalfa','10185'), ('CalWestSeeds.Alfalfa','CW 194')),
                        (('Alfalfa','16036'), ('CalWestSeeds.Alfalfa','CW 197')),
                        (('Alfalfa','13339'), ('CalWestSeeds.Alfalfa','CW 660')),
                        (('Alfalfa','14806'), ('CalWestSeeds.Alfalfa','CW 809')),
                        (('Alfalfa','14805'), ('CalWestSeeds.Alfalfa','CW Premium')),
                        (('Soja','19675'), ('Credenz.Soja','CZ 3621 STS')),
                        (('Soja','15041'), ('Credenz.Soja','CZ 4306 B')),
                        (('Soja','18303'), ('Credenz.Soja','CZ 4721 STS')),
                        (('Soja','17084'), ('Credenz.Soja','CZ 4908 IPRO')),
                        (('Soja','17242'), ('Credenz.Soja','CZ 5407 IPRO')),
                        (('Soja','17108'), ('Credenz.Soja','CZ 5907 IPRO')),
                        (('Soja','16604'), ('Credenz.Soja','CZ 6806 IPRO')),
                        (('Soja','14963'), ('Credenz.Soja','CZ 7905 IPRO')),
                        (('Girasol','9615'), ('CriaderoElCencerro.Girasol','Cacique')),
                        (('Trigo','9615'), ('Klein.Trigo','Cacique')),
                        (('Girasol','13842'), ('CriaderoElCencerro.Girasol','Cacique 312')),
                        (('Girasol','19172'), ('CriaderoElCencerro.Girasol','Cacique 320')),
                        (('Cebada','14474'), ('DonMario.Cebada','Cambium')),
                        (('Trigo','14474'), ('DonMario.Trigo','Cambium')),
                        (('Centeno','8695'), ('INTA.Centeno','Camilo')),
                        (('Alfalfa','11003'), ('CriaderoElCencerro.Alfalfa','Candela')),
                        (('Trigo','8501'), ('Klein.Trigo','Capricornio')),
                        (('Alfalfa','9753'), ('Biscayart.Alfalfa','Carabela')),
                        (('Trigo','10050'), ('Klein.Trigo','Carpincho')),
                        (('Trigo','9061'), ('Klein.Trigo','Castor')),
                        (('Trigo','20929'), ('DonMario.Trigo','Catalpa')),
                        (('Trigo','5423'), ('Buck.Trigo','Caudillo')),
                        (('Alfalfa','10886'), ('Biscayart.Alfalfa','Cautiva II')),
                        (('Alfalfa','14294'), ('Biscayart.Alfalfa','Cautiva III')),
                        (('Trigo','15365'), ('DonMario.Trigo','Ceibo')),
                        (('Trigo','9042'), ('Buck.Trigo','Chacarero')),
                        (('Trigo','7746'), ('Klein.Trigo','Chaja')),
                        (('Cebada','17521'), ('Cargill.Cebada','Charles')),
                        (('Trigo','2038'), ('Buck.Trigo','Charrua')),
                        (('Cebada','17572'), ('Cargill.Cebada','Cheers')),
                        (('Trigo','18693'), ('Klein.Trigo','Cien años')),
                        (('Trigo','16333'), ('Buck.Trigo','Claraz')),
                        (('Arveja','1392'), ('INTA.Arveja','Cobri')),
                        (('Poroto','7404'), ('EEAOC.Poroto','Cranberry - TUC 241')),
                        (('Trigo','8601'), ('DonMario.Trigo','Cronox')),
                        (('Trigo','18373'), ('Buck.Trigo','Cumelen')),
                        (('Arveja','15963'), ('Bioseminis.Arveja','Curling')),
                        (('Soja','7269'), ('AgriSeed.Soja','DALIA 390')),
                        (('Soja','7663'), ('AgriSeed.Soja','DALIA 455')),
                        (('Soja','12976'), ('AgriSeed.Soja','DALIA 490')),
                        (('Soja','7664'), ('AgriSeed.Soja','DALIA 500')),
                        (('Soja','8811'), ('AgriSeed.Soja','DALIA 550')),
                        (('Soja','14838'), ('AgriSeed.Soja','DALIA 610')),
                        (('Soja','10581'), ('AgriSeed.Soja','DALIA 620')),
                        (('Soja','10583'), ('AgriSeed.Soja','DALIA 680')),
                        (('Soja','9431'), ('AgriSeed.Soja','DALIA 700')),
                        (('Soja','9432'), ('AgriSeed.Soja','DALIA 740')),
                        (('Soja','9444'), ('AgriSeed.Soja','DALIA 750')),
                        (('Soja','10584'), ('AgriSeed.Soja','DALIA 780')),
                        (('Colza','17133'), ('Nuseed.Colza','DIAMOND')),
                        (('Colza','13528'), ('HighTech.Colza','DIMENSION')),
                        (('Girasol','11926'), ('Syngenta.Girasol','DK 4065')),
                        (('Sorgo','14383'), ('Dekalb.Sorgo','DK 53')),
                        (('Sorgo','7624'), ('Dekalb.Sorgo','DK 61T')),
                        (('Sorgo','13640'), ('Dekalb.Sorgo','DK 64T')),
                        (('Maíz','14516'), ('Dekalb.Maíz','DK 66-10')),
                        (('Maíz','16385'), ('Dekalb.Maíz','DK 69-10')),
                        (('Maíz','13231'), ('Dekalb.Maíz','DK 692')),
                        (('Maíz','9457'), ('Dekalb.Maíz','DK 70-10')),
                        (('Maíz','18769'), ('Dekalb.Maíz','DK 70-20')),
                        (('Maíz','21630'), ('Dekalb.Maíz','DK 72-08')),
                        (('Maíz','14755'), ('Dekalb.Maíz','DK 72-10')),
                        (('Maíz','14802'), ('Dekalb.Maíz','DK 72-50')),
                        (('Maíz','21007'), ('Dekalb.Maíz','DK 72-70')),
                        (('Maíz','22202'), ('Dekalb.Maíz','DK 72-72')),
                        (('Maíz','22135'), ('Dekalb.Maíz','DK 73-03')),
                        (('Maíz','15854'), ('Dekalb.Maíz','DK 73-10')),
                        (('Maíz','19457'), ('Dekalb.Maíz','DK 73-20')),
                        (('Maíz','21606'), ('Dekalb.Maíz','DK 73-30')),
                        (('Maíz','22661'), ('Dekalb.Maíz','DK 74-47')),
                        (('Maíz','21639'), ('Dekalb.Maíz','DK 77-02')),
                        (('Maíz','17233'), ('Dekalb.Maíz','DK 77-10')),
                        (('Maíz','17302'), ('Dekalb.Maíz','DK 78-20')),
                        (('Maíz','15837'), ('Dekalb.Maíz','DK 79-10')),
                        (('Maíz','10976'), ('DonMario.Maíz','DM 2747')),
                        (('Maíz','10187'), ('DonMario.Maíz','DM 2753')),
                        (('Maíz','10176'), ('DonMario.Maíz','DM 2765')),
                        (('Soja','21557'), ('DonMario.Soja','DM 33E22 SE')),
                        (('Soja','20002'), ('DonMario.Soja','DM 33R22')),
                        (('Soja','15267'), ('DonMario.Soja','DM 4014 IPRO')),
                        (('Soja','22560'), ('DonMario.Soja','DM 40E23 SE')),
                        (('Soja','16637'), ('DonMario.Soja','DM 46R18 STS')),
                        (('Soja','17048'), ('DonMario.Soja','DM 46i17 IPRO')),
                        (('Soja','19679'), ('DonMario.Soja','DM 46i20 IPRO STS')),
                        (('Soja','22571'), ('DonMario.Soja','DM 47E23')),
                        (('Soja','6079'), ('DonMario.Soja','DM 4800')),
                        (('Soja','14309'), ('DonMario.Soja','DM 4915 IPRO')),
                        (('Soja','20068'), ('DonMario.Soja','DM 50E22 SE')),
                        (('Soja','17632'), ('DonMario.Soja','DM 52R19')),
                        (('Soja','15937'), ('DonMario.Soja','DM 53i53 IPRO')),
                        (('Soja','18322'), ('DonMario.Soja','DM 55R20 STS')),
                        (('Soja','20105'), ('DonMario.Soja','DM 60K60')),
                        (('Soja','20105'), ('DonMario.Soja','DM 60K60 SCE')),
                        (('Soja','17107'), ('DonMario.Soja','DM 60i62 IPRO')),
                        (('Soja','14228'), ('DonMario.Soja','DM 6262 RSF IPRO')),
                        (('Soja','16480'), ('DonMario.Soja','DM 62R63 STS')),
                        (('Soja','16210'), ('DonMario.Soja','DM 63i64 Garra IPRO STS')),
                        (('Soja','20090'), ('DonMario.Soja','DM 64E64 SE')),
                        (('Soja','21081'), ('DonMario.Soja','DM 64K64 SCE')),
                        (('Soja','17658'), ('DonMario.Soja','DM 66R69 STS')),
                        (('Soja','18349'), ('DonMario.Soja','DM 68K68 SCE')),
                        (('Soja','18349'), ('DonMario.Soja','DM 68K68 STS')),
                        (('Soja','20172'), ('DonMario.Soja','DM 75K75 CE')),
                        (('Soja','16670'), ('DonMario.Soja','DM 75i75 IPRO')),
                        (('Soja','16023'), ('DonMario.Soja','DM 7870 IPRO')),
                        (('Soja','19401'), ('DonMario.Soja','DM 80K80 SCE')),
                        (('Alfalfa','9745'), ('BayaCasal.Alfalfa','DON ENRIQUE')),
                        (('Algodón','15182'), ('Gensus.Algodón','DP 1238')),
                        (('Algodón','8402'), ('Gensus.Algodón','DP 402')),
                        (('Maíz','15676'), ('Forratec.Maíz','DUO 24')),
                        (('Maíz','15793'), ('Forratec.Maíz','DUO 28')),
                        (('Maíz','17271'), ('Forratec.Maíz','DUO 30')),
                        (('Maíz','15230'), ('Forratec.Maíz','DUO 575')),
                        (('Cebada','15653'), ('MalteriaOriental.Cebada','Danielle')),
                        (('Trigo','6816'), ('Klein.Trigo','Delfin')),
                        (('Trigo','17282'), ('Buck.Trigo','Destello')),
                        (('Centeno','8209'), ('INTA.Centeno','Don Alberto')),
                        (('Alfalfa','15565'), ('AdSur.Alfalfa','Don Carlos')),
                        (('Centeno','921'), ('INTA.Centeno','Don Enrique')),
                        (('Trigo','921'), ('Klein.Trigo','Don Enrique')),
                        (('Centeno','12213'), ('INTA.Centeno','Don Ewald')),
                        (('Centeno','3400'), ('INTA.Centeno','Don Guillermo')),
                        (('Centeno','3233'), ('INTA.Centeno','Don Lisandro')),
                        (('Centeno','8209'), ('INTA.Centeno','Don Norberto')),
                        (('Alfalfa','15557'), ('AdSur.Alfalfa','Don Ramon')),
                        (('Tricepiro','3278'), ('INTA.Tricepiro','Don Rene')),
                        (('Soja','11512'), ('Nidera.Soja','EBC 4900 RG')),
                        (('Alfalfa','9752'), ('BayaCasal.Alfalfa','EBC 90')),
                        (('Alfalfa','18018'), ('BayaCasal.Alfalfa','EBC 909 MAX')),
                        (('Maní','18383'), ('CriaderoElCarmen.Maní','EC 214')),
                        (('Maíz','14846'), ('Produsem.Maíz','EG 808 ')),
                        (('Maíz','18356'), ('Produsem.Maíz','EG 809')),
                        (('Girasol','14074'), ('LGseeds.Girasol','ES SHERPA')),
                        (('Arveja','1628'), ('Inia.Arveja','ElRemate')),
                        (('Poroto','12786'), ('INTA.Leales.Poroto','Escarlata INTA')),
                        (('Trigo','6199'), ('Klein.Trigo','Escorpion')),
                        (('Trigo','6817'), ('Klein.Trigo','Escudo')),
                        (('Trigo','6753'), ('Buck.Trigo','Esmeralda')),
                        (('Trigo','4318'), ('Klein.Trigo','Estrella')),
                        (('Cebada','13324'), ('Nidera.Cebada','Explorer')),
                        (('Sorgo','13284'), ('Biscayart.Sorgo','Expreso 131 GR')),
                        (('Sorgo','13285'), ('Biscayart.Sorgo','Expreso 636')),
                        (('Trigo','22465'), ('Klein.Trigo','Extremo')),
                        (('Sorgo','4428'), ('GAPP.Sorgo','F - 700')),
                        (('Sorgo','11010'), ('Forratec.Sorgo','F 1200')),
                        (('Sorgo','13174'), ('Forratec.Sorgo','F 1300 FS')),
                        (('Sorgo','14789'), ('Forratec.Sorgo','F 1307')),
                        (('Sorgo','11186'), ('Forratec.Sorgo','F 1400')),
                        (('Sorgo','14790'), ('Forratec.Sorgo','F 1405')),
                        (('Sorgo','14791'), ('Forratec.Sorgo','F 1497')),
                        (('Sorgo','14757'), ('Forratec.Sorgo','F 2490')),
                        (('Sorgo','14758'), ('Forratec.Sorgo','F 3590')),
                        (('Sorgo','15341'), ('Forratec.Sorgo','F 750 Plus')),
                        (('Trigo','19535'), ('Klein.Trigo','FAVORITO II')),
                        (('Colza','8335'), ('Nuseed.Colza','FILIAL UOFA')),
                        (('Trigo','12979'), ('Agseed.Trigo','FLORIPAN 100')),
                        (('Trigo','13026'), ('Agseed.Trigo','FLORIPAN 200')),
                        (('Trigo','13027'), ('Agseed.Trigo','FLORIPAN 300')),
                        (('Trigo','13027'), ('Agseed.Trigo','FLORIPAN 301')),
                        (('Soja','9416'), ('Credenz.Soja','FN 3.90')),
                        (('Soja','12176'), ('Credenz.Soja','FN 365 AP')),
                        (('Soja','8260'), ('Credenz.Soja','FN 4.10')),
                        (('Soja','11246'), ('Credenz.Soja','FN 4.50')),
                        (('Soja','13573'), ('Credenz.Soja','FN 4.95')),
                        (('Soja','11247'), ('Credenz.Soja','FN 5.20')),
                        (('Soja','12182'), ('Credenz.Soja','FN 5.25')),
                        (('Soja','13569'), ('Credenz.Soja','FN 5.55')),
                        (('Soja','11248'), ('Credenz.Soja','FN 5.75')),
                        (('Soja','12179'), ('Credenz.Soja','FN 6.25')),
                        (('Soja','13563'), ('Credenz.Soja','FN 6.55')),
                        (('Alfalfa','13833'), ('LosPrados.Alfalfa','FRANCESCA')),
                        (('Trigo','20358'), ('Aca.Trigo','FRESNO')),
                        (('Maíz','18696'), ('Forratec.Maíz','FT 2122')),
                        (('Maíz','17556'), ('Forratec.Maíz','FT 4180')),
                        (('Maíz','19608'), ('Forratec.Maíz','FT 4212')),
                        (('Arveja','6873'), ('Bioseminis.Arveja','Facon')),
                        (('Trigo','6083'), ('Buck.Trigo','Farol')),
                        (('Cebada','17586'), ('ACA.Cebada','Fatima')),
                        (('Centeno','8881'), ('INTA.Centeno','Fausto')),
                        (('Garbanzo','13061'), ('Inta.Garbanzo','Felipe')),
                        (('Trigo','13775'), ('Klein.Trigo','Flamenco')),
                        (('Trigo','8092'), ('Klein.Trigo','Flecha')),
                        (('Trigo','20345'), ('Buck.Trigo','Fulgor')),
                        (('Trigo','13998'), ('DonMario.Trigo','Fuste')),
                        (('Sorgo','19862'), ('Gentos.Sorgo','G2.85 BMR')),
                        (('Alfalfa','12939'), ('PGG.Alfalfa','G686')),
                        (('Alfalfa','12938'), ('PGG.Alfalfa','G969')),
                        (('Sorgo','8306'), ('GAPP.Sorgo','GAPP 202')),
                        (('Trigo','15271'), ('Aca.Trigo','GARDELL')),
                        (('Sorgo','14633'), ('GAPP.Sorgo','GG 409')),
                        (('Sorgo','13851'), ('GAPP.Sorgo','GP 307')),
                        (('Sorgo','12485'), ('GAPP.Sorgo','GP 702 BMR')),
                        (('Sorgo','14519'), ('GAPP.Sorgo','GPF 203 BMR')),
                        (('Poroto','6967'), ('INTA.Leales.Poroto','Gateado INTA')),
                        (('Trigo','8502'), ('Klein.Trigo','Gavilan')),
                        (('Sorgo','11019'), ('Gentos.Sorgo','Gentos 125')),
                        (('Sorgo','15470'), ('Gentos.Sorgo','Gentos 130 AA')),
                        (('Sorgo','14776'), ('Gentos.Sorgo','Gentos 135 BMR')),
                        (('Sorgo','11020'), ('Gentos.Sorgo','Gentos 75 BMR')),
                        (('Sorgo','11021'), ('Gentos.Sorgo','Gentos 95 BMR')),
                        (('Trigo','17944'), ('Bioceres.Trigo','Gingko')),
                        (('Trigo','12413'), ('Klein.Trigo','Gladiador')),
                        (('Trigo','10015'), ('Buck.Trigo','Glutino')),
                        (('Maní','7907'), ('CriaderoElCarmen.Maní','Granoleico')),
                        (('Trigo','6685'), ('Buck.Trigo','Guapo')),
                        (('Algodón','18241'), ('Gensus.Algodón','Guaraní INTA')),
                        (('Maní','14642'), ('CriaderoElCarmen.Maní','Guasu (Virginia)')),
                        (('Trigo','7355'), ('Buck.Trigo','Guatimozin')),
                        (('Trigo','17945'), ('Bioceres.Trigo','Guayabo')),
                        (('Algodón','7537'), ('Gensus.Algodón','Guazuncho 2000')),
                        (('Algodón','2126'), ('Gensus.Algodón','Guazuncho 4 INTA')),
                        (('Trigo','10051'), ('Klein.Trigo','Guerrero')),
                        (('Trigo','19553'), ('Klein.Trigo','Géminis')),
                        (('Maíz','9072'), ('DonMario.Maíz','H 2740')),
                        (('Alfalfa','5670'), ('BayaCasal.Alfalfa','HAYGRAZAER')),
                        (('Soja','12761'), ('Horus.Soja','HO 3890')),
                        (('Soja','14242'), ('Horus.Soja','HO 3998')),
                        (('Soja','16418'), ('Horus.Soja','HO 4119 IPRO')),
                        (('Soja','11987'), ('Horus.Soja','HO 4880')),
                        (('Soja','16419'), ('Horus.Soja','HO 4919 IPRO')),
                        (('Soja','13470'), ('Horus.Soja','HO 5010')),
                        (('Soja','14969'), ('Horus.Soja','HO 5310 IPRO')),
                        (('Soja','14258'), ('Horus.Soja','HO 5910')),
                        (('Soja','15910'), ('Horus.Soja','HO 6110 IPRO')),
                        (('Soja','17274'), ('Horus.Soja','HO 6620 IPRO STS')),
                        (('Soja','14326'), ('Horus.Soja','HO 6997 IPRO')),
                        (('Soja','15917'), ('Horus.Soja','HO 7510 IPRO')),
                        (('Soja','16666'), ('Horus.Soja','HO59136 IPRO')),
                        (('Colza','10479'), ('HighTech.Colza','HORNET')),
                        (('Soja','17648'), ('Horus.Soja','HO 74134 IPRO STS')),
                        (('Cebada','14590'), ('INTA.Cebada','HUILEN INTA ')),
                        (('Cebada','777'), ('INTA.Cebada','HUITRU INTA ')),
                        (('Colza','11534'), ('ADVANTA.Colza','HYOLA 433')),
                        (('Colza','14646'), ('ADVANTA.Colza','HYOLA 575 CL')),
                        (('Colza','9102'), ('ADVANTA.Colza','HYOLA 61')),
                        (('Colza','11533'), ('ADVANTA.Colza','HYOLA 76')),
                        (('Colza','15761'), ('ADVANTA.Colza','HYOLA 830 CC')),
                        (('Colza','12567'), ('ADVANTA.Colza','HYOLA 971 CL')),
                        (('Cebada','13713'), ('SABMiller.Cebada','Henrike')),
                        (('Trigo','17283'), ('DonMario.Trigo','Ho Atuel')),
                        (('Trigo','20492'), ('Illinois.Trigo','Hornero')),
                        (('Trigo','3853'), ('Prointa.Trigo','Huenpan')),
                        (('Trigo','17219'), ('Klein.Trigo','Huracan')),
                        (('Maíz','9648'), ('Illinois.Maíz','I 550')),
                        (('Colza','14919'), ('HighTech.Colza','INSPIRATION')),
                        (('Cebada','14591'), ('INTA.Cebada','INTA 7302 ')),
                        (('Trigo','21709'), ('Bioseminis.Trigo','IRUYA HB4')),
                        (('Soja','21503'), ('Illinois.Soja','IS 38.2 SE')),
                        (('Soja','10518'), ('Illinois.Soja','IS 3808')),
                        (('Soja','12766'), ('Illinois.Soja','IS 3909')),
                        (('Soja','12463'), ('Illinois.Soja','IS 4510')),
                        (('Soja','19118'), ('Illinois.Soja','IS 46.1 SE')),
                        (('Soja','21545'), ('Illinois.Soja','IS 46.2 RR/STS')),
                        (('Soja','11741'), ('Illinois.Soja','IS 4777')),
                        (('Soja','21504'), ('Illinois.Soja','IS 48.2 E')),
                        (('Soja','18329'), ('Illinois.Soja','IS 52.0 RR/STS')),
                        (('Soja','10515'), ('Illinois.Soja','IS 5250i')),
                        (('Soja','19147'), ('Illinois.Soja','IS 60.1 SE')),
                        (('Soja','18318'), ('Illinois.Soja','IS 62.1 IPRO STS')),
                        (('Soja','19399'), ('Illinois.Soja','IS-69.2 CE')),
                        (('Cebada','15238'), ('INTA.Cebada','IVANKA INTA ')),
                        (('Girasol','21711'), ('Basf.Girasol','InSun 211B22')),
                        (('Trigo','18026'), ('Bioseminis.Trigo','JACARANDÁ')),
                        (('Trigo','7766'), ('Klein.Trigo','Jabalí')),
                        (('Cebada','15146'), ('Cargill.Cebada','Jennifer')),
                        (('Cebada','15146'), ('SABMiller.Cebada','Jennifer')),
                        (('Trigo','21953'), ('LosGrobo.Trigo','Juramento')),
                        (('Soja','16869'), ('Kumagro.Soja','K 3717 STS')),
                        (('Soja','15175'), ('Kumagro.Soja','K 4001 STS')),
                        (('Soja','16868'), ('Kumagro.Soja','K 4017 STS')),
                        (('Soja','16114'), ('Kumagro.Soja','K 4616 STS')),
                        (('Soja','15657'), ('Kumagro.Soja','K 5102 STS')),
                        (('Soja','16839'), ('Kumagro.Soja','K 6000')),
                        (('Soja','15654'), ('Kumagro.Soja','K 6501 STS')),
                        (('Soja','15656'), ('Kumagro.Soja','K 7102')),
                        (('Maíz','21749'), ('KWS.Maíz','K19-120')),
                        (('Maíz','19891'), ('KWS.Maíz','K9606')),
                        (('Maíz','12876'), ('KWS.Maíz','KM 1301')),
                        (('Maíz','9792'), ('KWS.Maíz','KM 3601')),
                        (('Maíz','15951'), ('KWS.Maíz','KM 3720')),
                        (('Maíz','15501'), ('KWS.Maíz','KM 3800')),
                        (('Maíz','17447'), ('KWS.Maíz','KM 3916 GLS')),
                        (('Maíz','18217'), ('KWS.Maíz','KM 3927')),
                        (('Maíz','14431'), ('KWS.Maíz','KM 4020')),
                        (('Maíz','15949'), ('KWS.Maíz','KM 4200')),
                        (('Maíz','22843'), ('KWS.Maíz','KM 4216')),
                        (('Maíz','12826'), ('KWS.Maíz','KM 4321')),
                        (('Maíz','13447'), ('KWS.Maíz','KM 4360 AS')),
                        (('Maíz','17696'), ('KWS.Maíz','KM 4480')),
                        (('Maíz','14890'), ('KWS.Maíz','KM 4500')),
                        (('Maíz','18969'), ('KWS.Maíz','KM 4580')),
                        (('Sorgo','14707'), ('KWS.Sorgo','KSGR 28')),
                        (('Sorgo','15116'), ('KWS.Sorgo','KSGR 42')),
                        (('Maíz','21750'), ('KWS.Maíz','KWS 13-160')),
                        (('Maíz','21745'), ('KWS.Maíz','KWS 605')),
                        (('Girasol','10950'), ('KWS.Girasol','KWSol 362')),
                        (('Girasol','14709'), ('KWS.Girasol','KWSol 480')),
                        (('Alfalfa','7212'), ('Biscayart.Alfalfa','Key II')),
                        (('Garbanzo','13060'), ('Inta.Garbanzo','Kiara')),
                        (('Arveja','18631'), ('Limagrain.Arveja','Kingfisher')),
                        (('Sorgo','8715'), ('Pemán.Sorgo','Kuntur INTA-Pemán')),
                        (('Cebada','697'), ('INTA.Cebada','LA PLATA BORDEBA FA ')),
                        (('Cebada','631'), ('INTA.Cebada','LA PLATA CAFPTA ')),
                        (('Trigo','22344'), ('Bioseminis.Trigo','LAUREL')),
                        (('Soja','12759'), ('MacroSeed.Soja','LDC 3.7')),
                        (('Soja','11151'), ('MacroSeed.Soja','LDC 3.8  STS')),
                        (('Soja','13546'), ('MacroSeed.Soja','LDC 5.3')),
                        (('Soja','11149'), ('MacroSeed.Soja','LDC 5.9 STS')),
                        (('Soja','14271'), ('MacroSeed.Soja','LDC 8.5')),
                        (('Trigo','7454'), ('Inia.Trigo','LE 2210 (INIA TIJERETA) (TCL)')),
                        (('Trigo','1322'), ('Inia.Trigo','LE 2249 (INIA CHURRINCHE) (TCI)')),
                        (('Trigo','10751'), ('Sursem.Trigo','LE 2330')),
                        (('Trigo','1922'), ('Inia.Trigo','LE 2331 (INIA DON ALBERTO)')),
                        (('Trigo','22588'), ('Inia.Trigo','LE 2333 (INIA CARPINTERO)')),
                        (('Soja','10316'), ('Sursem.Soja','LEO7800')),
                        (('Maíz','22307'), ('LGseeds.Maíz','LG 30.600')),
                        (('Maíz','19602'), ('LGseeds.Maíz','LG 30.680')),
                        (('Maíz','15866'), ('LGseeds.Maíz','LG 30.860')),
                        (('Girasol','18102'), ('LGseeds.Girasol','LG 50.750')),
                        (('Girasol','19006'), ('LGseeds.Girasol','LG 50.760')),
                        (('Girasol','17566'), ('LGseeds.Girasol','LG 5626 HO')),
                        (('Girasol','15345'), ('LGseeds.Girasol','LG 5710')),
                        (('Cebada','21419'), ('Limagrain.Cebada','LG ANDANTE')),
                        (('Trigo','21689'), ('LG.Trigo','LG BAYO')),
                        (('Cebada','22246'), ('Limagrain.Cebada','LG BELCANTO')),
                        (('Trigo','21373'), ('LG.Trigo','LG MORO')),
                        (('Trigo','21802'), ('LG.Trigo','LG PICAZO')),
                        (('Cebada','21089'), ('Limagrain.Cebada','LG Zodiac')),
                        (('Maíz','13897'), ('LGseeds.Maíz','LGSA 30.850')),
                        (('Maíz','18110'), ('LGseeds.Maíz','LGSA 30775')),
                        (('Colza','10480'), ('HighTech.Colza','LILIAN')),
                        (('Maíz','13193'), ('LaTijereta.Maíz','LT 611')),
                        (('Maíz','12479'), ('LaTijereta.Maíz','LT 617')),
                        (('Maíz','10936'), ('LaTijereta.Maíz','LT 618')),
                        (('Maíz','9093'), ('LaTijereta.Maíz','LT 620')),
                        (('Maíz','13996'), ('LaTijereta.Maíz','LT 621')),
                        (('Maíz','11692'), ('LaTijereta.Maíz','LT 622')),
                        (('Maíz','13835'), ('LaTijereta.Maíz','LT 623')),
                        (('Maíz','11766'), ('LaTijereta.Maíz','LT 624')),
                        (('Maíz','9709'), ('LaTijereta.Maíz','LT 625')),
                        (('Maíz','12853'), ('LaTijereta.Maíz','LT 626')),
                        (('Maíz','10935'), ('LaTijereta.Maíz','LT 632')),
                        (('Maíz','19372'), ('LaTijereta.Maíz','LT 718')),
                        (('Maíz','15688'), ('LaTijereta.Maíz','LT 719')),
                        (('Maíz','21344'), ('LaTijereta.Maíz','LT 720')),
                        (('Maíz','19489'), ('LaTijereta.Maíz','LT 721')),
                        (('Maíz','15858'), ('LaTijereta.Maíz','LT 722')),
                        (('Maíz','22306'), ('LaTijereta.Maíz','LT 723')),
                        (('Maíz','21771'), ('LaTijereta.Maíz','LT 725')),
                        (('Maíz','15540'), ('LaTijereta.Maíz','LT 780')),
                        (('Maíz','13331'), ('LaTijereta.Maíz','LT 790')),
                        (('Maíz','17232'), ('LaTijereta.Maíz','LT 795')),
                        (('Maíz','16051'), ('LaTijereta.Maíz','LT 800')),
                        (('Trigo','16413'), ('Klein.Trigo','Lanza')),
                        (('Trigo','14467'), ('LG.Trigo','Lapacho')),
                        (('Trigo','14467'), ('Sursem.Trigo','Lapacho (Nogal 111)')),
                        (('Trigo','22344'), ('Bioceres.Trigo','Laurel')),
                        (('Poroto','10035'), ('INTA.Leales.Poroto','Leales 10 INTA')),
                        (('Poroto','10035'), ('INTA.Leales.Poroto','Leales 14 INTA')),
                        (('Poroto','11108'), ('INTA.Leales.Poroto','Leales 15 INTA')),
                        (('Poroto','10036'), ('INTA.Leales.Poroto','Leales 17 INTA')),
                        (('Poroto','10037'), ('INTA.Leales.Poroto','Leales 22 INTA')),
                        (('Poroto','11109'), ('INTA.Leales.Poroto','Leales 24 INTA')),
                        (('Poroto','10037'), ('INTA.Leales.Poroto','Leales 26 INTA')),
                        (('Poroto','14401'), ('INTA.Leales.Poroto','Leales B30 INTA')),
                        (('Poroto','14402'), ('INTA.Leales.Poroto','Leales B40 INTA')),
                        (('Poroto','14403'), ('INTA.Leales.Poroto','Leales C1 INTA')),
                        (('Poroto','14404'), ('INTA.Leales.Poroto','Leales CR5 INTA')),
                        (('Poroto','14405'), ('INTA.Leales.Poroto','Leales R4 INTA')),
                        (('Trigo','14486'), ('DonMario.Trigo','Lenga')),
                        (('Trigo','13055'), ('DonMario.Trigo','Lenox')),
                        (('Trigo','22463'), ('Klein.Trigo','Leyenda')),
                        (('Trigo','13778'), ('Klein.Trigo','Liebre')),
                        (('Trigo','20676'), ('LosGrobo.Trigo','Limay')),
                        (('Sorgo','14711'), ('LaTijereta.Sorgo','Litio')),
                        (('Trigo','13380'), ('DonMario.Trigo','Lyon')),
                        (('Soja','14994'), ('LaTijereta.Soja','M 5410 IPRO')),
                        (('Maíz','15381'), ('RedSurcos.Maíz','M5890 BT')),
                        (('Soja','14713'), ('ASGROW.Soja','M6210 IPRO')),
                        (('Soja','14714'), ('ASGROW.Soja','M6410 IPRO')),
                        (('Maní','17238'), ('Maniagro.Maní','MA-02')),
                        (('Maní','17237'), ('Maniagro.Maní','MA-121')),
                        (('Maní','17240'), ('Maniagro.Maní','MA-757')),
                        (('Maní','17234'), ('Maniagro.Maní','MA-767')),
                        (('Maní','17235'), ('Maniagro.Maní','MA-88')),
                        (('Maní','17236'), ('Maniagro.Maní','MA-90')),
                        (('Colza','16397'), ('INTA.Colza','MACACHA INTA')),
                        (('Cebada','606'), ('INTA.Cebada','MAGNIF 102 INTA ')),
                        (('Cebada','607'), ('INTA.Cebada','MAGNIF 105 INTA ')),
                        (('Sorgo','8781'), ('Argenetics.Sorgo','MALON')),
                        (('Colza','18065'), ('DSV.Colza','MARATHON')),
                        (('Soja','8788'), ('Sursem.Soja','MARIA 50')),
                        (('Vicia','22134'), ('Zinma.Vicia','MASSA')),
                        (('Alfalfa','5732'), ('LosPrados.Alfalfa','MAYACO')),
                        (('Cebada','7970'), ('INTA.Cebada','MELIPAL INTA ')),
                        (('Soja','8339'), ('Sursem.Soja','MERCEDES 76')),
                        (('Girasol','14594'), ('Brevant.Girasol','MG 303GP')),
                        (('Girasol','13281'), ('Brevant.Girasol','MG 305')),
                        (('Girasol','13920'), ('Brevant.Girasol','MG 360')),
                        (('Sorgo','14701'), ('Nord.Sorgo','MGS 76')),
                        (('Sorgo','14702'), ('Nord.Sorgo','MGS 85')),
                        (('Soja','15050'), ('INTA.Soja','MJ42 STS')),
                        (('Cebada','8688'), ('INTA.Cebada','MOROCHA FCA ')),
                        (('Cebada','11440'), ('MalteriaPampa.Cebada','MP 1012')),
                        (('Cebada','9609'), ('MalteriaPampa.Cebada','MP 1109')),
                        (('Cebada','13014'), ('MalteriaPampa.Cebada','MP 2122')),
                        (('Cebada','7880'), ('MalteriaPampa.Cebada','MP 546')),
                        (('Cebada','7880'), ('MalteriaPampa.Cebada','MP 546 ')),
                        (('Sorgo','15371'), ('Brevant.Sorgo','MS 100')),
                        (('Sorgo','9689'), ('Brevant.Sorgo','MS 102')),
                        (('Sorgo','12772'), ('Brevant.Sorgo','MS 105')),
                        (('Sorgo','15276'), ('Brevant.Sorgo','MS 106')),
                        (('Sorgo','7890'), ('Brevant.Sorgo','MS 108')),
                        (('Sorgo','9244'), ('Brevant.Sorgo','MS 109')),
                        (('Soja','14868'), ('MacroSeed.Soja','MS 4.0 IPRO')),
                        (('Soja','14881'), ('MacroSeed.Soja','MS 4.9 IPRO')),
                        (('Soja','14961'), ('MacroSeed.Soja','MS 6.3 IPRO')),
                        (('Soja','15914'), ('MacroSeed.Soja','MS 6.9 IPRO')),
                        (('Soja','14327'), ('MacroSeed.Soja','MS 7.4 IPRO')),
                        (('Trigo','16336'), ('MacroSeed.Trigo','MS INTA 116')),
                        (('Trigo','18706'), ('MacroSeed.Trigo','MS INTA 119')),
                        (('Trigo','16328'), ('MacroSeed.Trigo','MS INTA 415')),
                        (('Trigo','16335'), ('MacroSeed.Trigo','MS INTA 416')),
                        (('Trigo','21350'), ('MacroSeed.Trigo','MS INTA 521')),
                        (('Trigo','13715'), ('MacroSeed.Trigo','MS INTA 615')),
                        (('Trigo','15746'), ('MacroSeed.Trigo','MS INTA 815')),
                        (('Maíz','17518'), ('Stine.Maíz','MST 120-19')),
                        (('Trigo','9523'), ('Buck.Trigo','Malevo')),
                        (('Trigo','11566'), ('Buck.Trigo','Mangrullo')),
                        (('Trigo','7279'), ('Klein.Trigo','Martillo')),
                        (('Trigo','7744'), ('Buck.Trigo','Mataco')),
                        (('Sorgo','10972'), ('Tobin.Sorgo','Matrero')),
                        (('Arveja','14615'), ('AFA.Arveja','Meadow')),
                        (('Trigo','17140'), ('Klein.Trigo','Mercurio')),
                        (('Trigo','10658'), ('Buck.Trigo','Meteoro')),
                        (('Cebada','18905'), ('Produsem.Cebada','Militza INTA')),
                        (('Trigo','17218'), ('Klein.Trigo','Minerva')),
                        (('Cebada','17571'), ('CyMQuilmes.Cebada','Montoya')),
                        (('Girasol','5861'), ('Nidera.Girasol','N Aromo105')),
                        (('Girasol','11259'), ('Nidera.Girasol','N Aromo11')),
                        (('Girasol','8707'), ('Nidera.Girasol','N Paraiso102')),
                        (('Soja','8337'), ('Sursem.Soja','N49R')),
                        (('Cebada','15936'), ('INTA.Cebada','NELIDA INTA ')),
                        (('Soja','22613'), ('NeoGen.Soja','NEO 35S23 SE')),
                        (('Soja','21501'), ('NeoGen.Soja','NEO 40S22 SE')),
                        (('Soja','20003'), ('NeoGen.Soja','NEO 45S22 RR STS')),
                        (('Soja','20043'), ('NeoGen.Soja','NEO 46S22 SE')),
                        (('Alfalfa','6668'), ('LosPrados.Alfalfa','NEVADA')),
                        (('Maíz','16352'), ('Brevant.Maíz','NEXT 20.6')),
                        (('Maíz','16353'), ('Brevant.Maíz','NEXT 22.6')),
                        (('Maíz','17583'), ('Brevant.Maíz','NEXT 25.8')),
                        (('Maíz','10494'), ('Syngenta.Maíz','NK 135')),
                        (('Soja','9426'), ('Syngenta.Soja','NK 32-00')),
                        (('Soja','8885'), ('Syngenta.Soja','NK 34-00')),
                        (('Soja','10455'), ('Syngenta.Soja','NK 35-00')),
                        (('Soja','9333'), ('Syngenta.Soja','NK 37-00')),
                        (('Soja','11075'), ('Syngenta.Soja','NK 38-00')),
                        (('Soja','11196'), ('Syngenta.Soja','NK 39-00')),
                        (('Girasol','20445'), ('Syngenta.Girasol','NK 3969')),
                        (('Soja','9429'), ('Syngenta.Soja','NK 43-00')),
                        (('Soja','8888'), ('Syngenta.Soja','NK 47-00')),
                        (('Soja','10454'), ('Syngenta.Soja','NK 48-00')),
                        (('Soja','18227'), ('Syngenta.Soja','NK 51X22 IPRO STS')),
                        (('Soja','19674'), ('Syngenta.Soja','NK 52x21 STS')),
                        (('Soja','18218'), ('Syngenta.Soja','NK 60x21 IPRO STS')),
                        (('Maíz','20484'), ('Syngenta.Maíz','NK 800')),
                        (('Maíz','20486'), ('Syngenta.Maíz','NK 842')),
                        (('Maíz','8590'), ('Syngenta.Maíz','NK 870')),
                        (('Maíz','10137'), ('Syngenta.Maíz','NK 880')),
                        (('Maíz','19483'), ('Syngenta.Maíz','NK 885')),
                        (('Maíz','19482'), ('Syngenta.Maíz','NK 890')),
                        (('Maíz','12332'), ('Syngenta.Maíz','NK 900')),
                        (('Maíz','12331'), ('Syngenta.Maíz','NK 907')),
                        (('Maíz','10195'), ('Syngenta.Maíz','NK 910')),
                        (('Maíz','8129'), ('Syngenta.Maíz','NK 940')),
                        (('Soja','8889'), ('Syngenta.Soja','NK COKER 6.6')),
                        (('Soja','7654'), ('Syngenta.Soja','NK COKER 6.8 RR')),
                        (('Soja','7659'), ('Syngenta.Soja','NK COKER 7.5 R')),
                        (('Soja','9883'), ('Syngenta.Soja','NK COKER 8.0')),
                        (('Soja','7219'), ('Syngenta.Soja','NK MIREYA 4.2')),
                        (('Soja','7218'), ('Syngenta.Soja','NK PICASA 4.0')),
                        (('Soja','7599'), ('Sursem.Soja','NM55R')),
                        (('Soja','7598'), ('Sursem.Soja','NM70R')),
                        (('Girasol','19442'), ('Nidera.Girasol','NS 1109')),
                        (('Girasol','20444'), ('Nidera.Girasol','NS 1113')),
                        (('Soja','12707'), ('Nidera.Soja','NS 2018')),
                        (('Soja','12948'), ('Nidera.Soja','NS 2632')),
                        (('Soja','11272'), ('Nidera.Soja','NS 3215')),
                        (('Soja','15529'), ('Nidera.Soja','NS 3220 STS')),
                        (('Soja','14721'), ('Nidera.Soja','NS 3313')),
                        (('Soja','15464'), ('Nidera.Soja','NS 3809 IPRO')),
                        (('Soja','19675'), ('Nidera.Soja','NS 3821 STS')),
                        (('Soja','14722'), ('Nidera.Soja','NS 3909')),
                        (('Soja','11273'), ('Nidera.Soja','NS 4009')),
                        (('Soja','15586'), ('Nidera.Soja','NS 4309')),
                        (('Soja','13585'), ('Nidera.Soja','NS 4313')),
                        (('Soja','15158'), ('Nidera.Soja','NS 4319 IPRO')),
                        (('Soja','12952'), ('Nidera.Soja','NS 4611 STS')),
                        (('Soja','19679'), ('Nidera.Soja','NS 4621 IPRO STS')),
                        (('Soja','11275'), ('Nidera.Soja','NS 4903')),
                        (('Soja','13589'), ('Nidera.Soja','NS 4955')),
                        (('Soja','12198'), ('Nidera.Soja','NS 4997')),
                        (('Soja','19839'), ('Nidera.Soja','NS 5023 STS')),
                        (('Soja','16230'), ('Nidera.Soja','NS 5028 STS')),
                        (('Soja','18842'), ('Nidera.Soja','NS 5030 IPRO STS')),
                        (('Soja','12956'), ('Nidera.Soja','NS 5230')),
                        (('Soja','13591'), ('Nidera.Soja','NS 5258')),
                        (('Soja','15443'), ('Nidera.Soja','NS 5419 IPRO')),
                        (('Soja','19677'), ('Nidera.Soja','NS 5421 STS')),
                        (('Soja','12201'), ('Nidera.Soja','NS 6002')),
                        (('Soja','17543'), ('Nidera.Soja','NS 6120 IPRO')),
                        (('Soja','20030'), ('Nidera.Soja','NS 6212 IPRO')),
                        (('Soja','14349'), ('Nidera.Soja','NS 6248')),
                        (('Soja','12199'), ('Nidera.Soja','NS 6267')),
                        (('Soja','14356'), ('Nidera.Soja','NS 6419 IPRO')),
                        (('Soja','14348'), ('Nidera.Soja','NS 6483')),
                        (('Soja','16228'), ('Nidera.Soja','NS 6538 IPRO')),
                        (('Soja','18226'), ('Nidera.Soja','NS 6721 IPRO STS')),
                        (('Soja','16229'), ('Nidera.Soja','NS 6859 IPRO')),
                        (('Soja','14357'), ('Nidera.Soja','NS 6909 IPRO')),
                        (('Soja','14704'), ('Nidera.Soja','NS 7209 IPRO')),
                        (('Soja','12194'), ('Nidera.Soja','NS 7211')),
                        (('Soja','14705'), ('Nidera.Soja','NS 7300 IPRO')),
                        (('Maíz','20488'), ('Nidera.Maíz','NS 7621')),
                        (('Soja','15093'), ('Nidera.Soja','NS 7709 IPRO')),
                        (('Soja','14158'), ('Nidera.Soja','NS 7711 IPRO')),
                        (('Maíz','20483'), ('Nidera.Maíz','NS 7800')),
                        (('Soja','15589'), ('Nidera.Soja','NS 7809')),
                        (('Maíz','18119'), ('Nidera.Maíz','NS 7818')),
                        (('Maíz','19456'), ('Nidera.Maíz','NS 7921')),
                        (('Soja','10005'), ('Nidera.Soja','NS 8004')),
                        (('Soja','16877'), ('Nidera.Soja','NS 8018 IPRO STS')),
                        (('Soja','12196'), ('Nidera.Soja','NS 8262')),
                        (('Soja','12193'), ('Nidera.Soja','NS 8282')),
                        (('Soja','15584'), ('Nidera.Soja','NS 8288 STS')),
                        (('Soja','20028'), ('Nidera.Soja','NS 7922')),
                        (('Girasol','19628'), ('Brevant.Girasol','NTO2.5')),
                        (('Girasol','15882'), ('Brevant.Girasol','NTO3.6')),
                        (('Girasol','12742'), ('Brevant.Girasol','NTO4.0')),
                        (('Sorgo','18756'), ('NuSeed.Sorgo','NU 441 IG')),
                        (('Sorgo','12589'), ('NuSeed.Sorgo','NU Don Jacinto')),
                        (('Sorgo','17036'), ('NuSeed.Sorgo','NU Nugrain 440')),
                        (('Sorgo','17292'), ('NuSeed.Sorgo','NU Nugrass 900')),
                        (('Sorgo','17293'), ('NuSeed.Sorgo','NU Nusil500')),
                        (('Sorgo','17290'), ('NuSeed.Sorgo','NU Nusil600')),
                        (('Sorgo','12612'), ('NuSeed.Sorgo','NU Spring T60')),
                        (('Sorgo','16417'), ('NuSeed.Sorgo','NU SummerII')),
                        (('Colza','15220'), ('Nuseed.Colza','NUVETTE 2286')),
                        (('Maíz','19620'), ('MacroSeed.Maíz','NXM 1122')),
                        (('Maíz','19620'), ('NexSem.Maíz','NXM 1122')),
                        (('Poroto','1699'), ('EEAOC.Poroto','Negro - BAT 304')),
                        (('Poroto','1698'), ('EEAOC.Poroto','Negro - DOR 157')),
                        (('Poroto','1143'), ('EEAOC.Poroto','Negro - DOR 41')),
                        (('Poroto','11502'), ('EEAOC.Poroto','Negro - TUC 300')),
                        (('Poroto','3602'), ('EEAOC.Poroto','Negro - TUC 390')),
                        (('Poroto','3603'), ('EEAOC.Poroto','Negro - TUC 500')),
                        (('Poroto','7402'), ('EEAOC.Poroto','Negro - TUC 510')),
                        (('Poroto','9957'), ('EEAOC.Poroto','Negro - TUC 550')),
                        (('Sorgo','11141'), ('Biscayart.Sorgo','NiagaSil')),
                        (('Vicia','17721'), ('Gentos.Vicia','Nitro Max')),
                        (('Sorgo','11258'), ('Biscayart.Sorgo','Niágara BL')),
                        (('Sorgo','18185'), ('Biscayart.Sorgo','Niágara III')),
                        (('Trigo','9190'), ('Sursem.Trigo','Nogal')),
                        (('Maíz','16525'), ('Nord.Maíz','Nord ACRUX')),
                        (('Maíz','14537'), ('Nord.Maíz','Nord AVALON')),
                        (('Maíz','14980'), ('Nord.Maíz','Nord BALTOS')),
                        (('Maíz','16871'), ('Nord.Maíz','Nord BORAX')),
                        (('Maíz','15794'), ('Nord.Maíz','Nord BUYAN')),
                        (('Sorgo','10880'), ('Nord.Sorgo','Nord Palatable 10 BMR')),
                        (('Sorgo','14230'), ('Nord.Sorgo','Nord Palatable 10 MAX')),
                        (('Girasol','16448'), ('Nord.Girasol','Nord Sungro 70')),
                        (('Girasol','16448'), ('Nord.Girasol','Nord Sungro 80')),
                        (('Maíz','17936'), ('Nord.Maíz','Nord ZEFIR')),
                        (('Trigo','4924'), ('Buck.Trigo','Norteño')),
                        (('Garbanzo','4924'), ('Inta.Garbanzo','Norteño')),
                        (('Maíz','18054'), ('Nuseed.Maíz','Nucorn 2650')),
                        (('Maíz','17321'), ('Nuseed.Maíz','Nucorn 2881')),
                        (('Sorgo','18030'), ('NuSeed.Sorgo','Nugrain 300')),
                        (('Algodón','11949'), ('Gensus.Algodón','Nuopal')),
                        (('Girasol','14578'), ('NuSeed.Girasol','Nusol 2100')),
                        (('Girasol','14580'), ('NuSeed.Girasol','Nusol 4100')),
                        (('Girasol','17340'), ('NuSeed.Girasol','Nusol 4120')),
                        (('Girasol','20600'), ('NuSeed.Girasol','Nusol 4145')),
                        (('Girasol','17340'), ('NuSeed.Girasol','Nusol 4170')),
                        (('Girasol','16433'), ('NuSeed.Girasol','Nusol 4510')),
                        (('Trigo','11588'), ('Klein.Trigo','Nutria')),
                        (('Cebada','454'), ('INTA.Cebada','OLIVEROS LITORAL SAG ')),
                        (('Girasol','19885'), ('OrigoSemillas.Girasol','ORI 730 CL')),
                        (('Trigo','16048'), ('Buck.Trigo','Odisseo')),
                        (('Trigo','8488'), ('DonMario.Trigo','Onix')),
                        (('Cebada','15720'), ('Limagrain.Cebada','Overture')),
                        (('Maíz','20674'), ('Pioneer.Maíz','P 0622')),
                        (('Maíz','15312'), ('Pioneer.Maíz','P 1778')),
                        (('Maíz','15309'), ('Pioneer.Maíz','P 1780')),
                        (('Maíz','19132'), ('Pioneer.Maíz','P 1804')),
                        (('Maíz','18100'), ('Pioneer.Maíz','P 1815')),
                        (('Maíz','15306'), ('Pioneer.Maíz','P 1833')),
                        (('Maíz','12576'), ('Pioneer.Maíz','P 1845')),
                        (('Maíz','15307'), ('Pioneer.Maíz','P 1952')),
                        (('Maíz','13198'), ('Pioneer.Maíz','P 1979')),
                        (('Maíz','17228'), ('Pioneer.Maíz','P 2005')),
                        (('Maíz','21560'), ('Pioneer.Maíz','P 2021')),
                        (('Maíz','13200'), ('Pioneer.Maíz','P 2038')),
                        (('Maíz','13199'), ('Pioneer.Maíz','P 2049')),
                        (('Maíz','10877'), ('Pioneer.Maíz','P 2053')),
                        (('Maíz','12575'), ('Pioneer.Maíz','P 2058')),
                        (('Maíz','13865'), ('Pioneer.Maíz','P 2069')),
                        (('Maíz','17230'), ('Pioneer.Maíz','P 2089')),
                        (('Maíz','18788'), ('Pioneer.Maíz','P 2103')),
                        (('Maíz','17229'), ('Pioneer.Maíz','P 2109')),
                        (('Maíz','16467'), ('Pioneer.Maíz','P 2151')),
                        (('Maíz','19278'), ('Pioneer.Maíz','P 2167')),
                        (('Maíz','20422'), ('Pioneer.Maíz','P 2297')),
                        (('Maíz','19115'), ('Pioneer.Maíz','P 2353')),
                        (('Maíz','11920'), ('Pioneer.Maíz','P 30B39')),
                        (('Maíz','10978'), ('Pioneer.Maíz','P 30F35')),
                        (('Maíz','18278'), ('Pioneer.Maíz','P 30F53')),
                        (('Maíz','9816'), ('Pioneer.Maíz','P 30P70')),
                        (('Maíz','6882'), ('Pioneer.Maíz','P 30R76')),
                        (('Maíz','10159'), ('Pioneer.Maíz','P 31A08')),
                        (('Maíz','8560'), ('Pioneer.Maíz','P 31B18')),
                        (('Maíz','9155'), ('Pioneer.Maíz','P 31D06')),
                        (('Maíz','8562'), ('Pioneer.Maíz','P 31F25')),
                        (('Maíz','12598'), ('Pioneer.Maíz','P 31G71')),
                        (('Maíz','9626'), ('Pioneer.Maíz','P 31P77')),
                        (('Maíz','9628'), ('Pioneer.Maíz','P 31R31')),
                        (('Maíz','8140'), ('Pioneer.Maíz','P 31Y04')),
                        (('Maíz','11745'), ('Pioneer.Maíz','P 31Y05')),
                        (('Maíz','7364'), ('Pioneer.Maíz','P 32F07')),
                        (('Maíz','16587'), ('Pioneer.Maíz','P 32R48')),
                        (('Maíz','15860'), ('Pioneer.Maíz','P 38A57')),
                        (('Maíz','10949'), ('Pioneer.Maíz','P 39B77')),
                        (('Soja','19129'), ('Pioneer.Soja','P 43A04SE')),
                        (('Soja','19130'), ('Pioneer.Soja','P 46A03SE')),
                        (('Soja','17720'), ('Pioneer.Soja','P 50A02E')),
                        (('Soja','20210'), ('Pioneer.Soja','P 60A01 SCE')),
                        (('Girasol','14653'), ('Pioneer.Girasol','P 64ll95')),
                        (('Girasol','9287'), ('Pioneer.Girasol','P 65A25')),
                        (('Soja','20170'), ('Pioneer.Soja','P 75A06 SCE')),
                        (('Soja','19432'), ('Pioneer.Soja','P 80A02 SCE')),
                        (('Sorgo','14026'), ('Pioneer.Sorgo','P 80T25')),
                        (('Sorgo','10012'), ('Pioneer.Sorgo','P 81G67')),
                        (('Sorgo','7833'), ('Pioneer.Sorgo','P 84G62')),
                        (('Soja','6034'), ('Pioneer.Soja','P 93B34 RR')),
                        (('Soja','8295'), ('Pioneer.Soja','P 93B36 RR')),
                        (('Soja','7872'), ('Pioneer.Soja','P 93B85 RR')),
                        (('Soja','12174'), ('Pioneer.Soja','P 93M70 RR')),
                        (('Soja','10509'), ('Pioneer.Soja','P 93M92 RR')),
                        (('Soja','10510'), ('Pioneer.Soja','P 93M96 RR')),
                        (('Soja','7873'), ('Pioneer.Soja','P 94B54 RR')),
                        (('Soja','7052'), ('Pioneer.Soja','P 94B73 RR')),
                        (('Soja','9969'), ('Pioneer.Soja','P 94M30 RR')),
                        (('Soja','9469'), ('Pioneer.Soja','P 94M40 RR')),
                        (('Soja','10511'), ('Pioneer.Soja','P 94M80 RR')),
                        (('Maíz','20678'), ('Pioneer.Maíz','P 9946')),
                        (('Girasol','9619'), ('Pioneer.Girasol','P PAN7047')),
                        (('Alfalfa','1690'), ('Barenbrug.Alfalfa','P-105')),
                        (('Trigo','6745'), ('Prointa.Trigo','P. Don Umberto')),
                        (('Trigo','2057'), ('Prointa.Trigo','P. Federal')),
                        (('Trigo','6649'), ('Prointa.Trigo','P. Gaucho')),
                        (('Trigo','2905'), ('Prointa.Trigo','P. Imperial')),
                        (('Trigo','6746'), ('Prointa.Trigo','P. Molinero')),
                        (('Trigo','2118'), ('Prointa.Trigo','P. Oasis')),
                        (('Trigo','3700'), ('Prointa.Trigo','P. Puntal')),
                        (('Trigo','21707'), ('Bioseminis.Trigo','PARANÁ HB4')),
                        (('Alfalfa','19166'), ('PGG.Alfalfa','PGW 931')),
                        (('Colza','18068'), ('DSV.Colza','PHOENIX')),
                        (('Maíz','16965'), ('Pop.Arg.Maíz','POP 4512')),
                        (('Trigo','21627'), ('Buck.Trigo','PRETAL')),
                        (('Colza','13529'), ('HighTech.Colza','PRIMUS')),
                        (('Alfalfa','7259'), ('Produsem.Alfalfa','PRO INTA LUJÁN')),
                        (('Alfalfa','10049'), ('Produsem.Alfalfa','PRO INTA MORA')),
                        (('Alfalfa','7258'), ('Produsem.Alfalfa','PRO INTA PATRICIA')),
                        (('Sorgo','9045'), ('AdSur.Sorgo','PRODUCTOR 401')),
                        (('Sorgo','16648'), ('AdSur.Sorgo','PRODUCTOR-161BL')),
                        (('Sorgo','16647'), ('AdSur.Sorgo','PRODUCTOR-162FS')),
                        (('Sorgo','14796'), ('Pemán.Sorgo','PS 500 BMR')),
                        (('Sorgo','13496'), ('Pemán.Sorgo','PS 55')),
                        (('Colza','9629'), ('HighTech.Colza','PULSAR')),
                        (('Arveja','13122'), ('ARVES.Arveja','Pampa')),
                        (('Trigo','10788'), ('Klein.Trigo','Pantera')),
                        (('Soja','12258'), ('INTA.Soja','Paraná 5500')),
                        (('Soja','12257'), ('INTA.Soja','Paraná 6200')),
                        (('Soja','8048'), ('INTA.Soja','Paraná 629')),
                        (('Soja','8049'), ('INTA.Soja','Paraná 661')),
                        (('Alfalfa','10273'), ('Peman.Alfalfa','Patriarca')),
                        (('Alfalfa','12836'), ('Barenbrug.Alfalfa','Pegasis')),
                        (('Trigo','4997'), ('Klein.Trigo','Pegaso')),
                        (('Trigo','18146'), ('DonMario.Trigo','Pehuen')),
                        (('Trigo','19466'), ('Buck.Trigo','Peregrino')),
                        (('Trigo','18644'), ('Buck.Trigo','Perla')),
                        (('Sorgo','11139'), ('Biscayart.Sorgo','Pilcomayo 2')),
                        (('Sorgo','18576'), ('Biscayart.Sorgo','Pilcomayo III')),
                        (('Sorgo','15876'), ('Pemán.Sorgo','Pitavá')),
                        (('Maíz','17363'), ('BMHSemillas.Maíz','Piumassimo')),
                        (('Maíz','19889'), ('BMHSemillas.Maíz','Piusoldi')),
                        (('Trigo','13682'), ('Buck.Trigo','Pleno')),
                        (('Algodón','18240'), ('Gensus.Algodón','Pora 3 INTA')),
                        (('Trigo','17994'), ('Klein.Trigo','Potro')),
                        (('Cebada','12906'), ('Cargill.Cebada','Prestige')),
                        (('Trigo','16310'), ('Klein.Trigo','Prometeo')),
                        (('Maní','11190'), ('CriaderoElCarmen.Maní','Pronto')),
                        (('Maní','11190'), ('CriaderoElCarmen.Maní','Pronto AO')),
                        (('Trigo','9621'), ('Proseme.Trigo','Proseme Cannizzo')),
                        (('Trigo','9622'), ('Proseme.Trigo','Proseme Ciccio')),
                        (('Trigo','9623'), ('Proseme.Trigo','Proseme Coloseo')),
                        (('Trigo','8093'), ('Klein.Trigo','Proteo')),
                        (('Trigo','10017'), ('Buck.Trigo','Puelche')),
                        (('Maíz','20809'), ('Qseeds.Maíz','QS 72-01')),
                        (('Maíz','19776'), ('Qseeds.Maíz','QS 73-01')),
                        (('Maíz','18162'), ('Qseeds.Maíz','QS 75-01')),
                        (('Maíz','18163'), ('Qseeds.Maíz','QS 86-01')),
                        (('Cebada','1006'), ('CyMQuilmes.Cebada','QUILMES 271 ')),
                        (('Cebada','1007'), ('CyMQuilmes.Cebada','QUILMES ALFA ')),
                        (('Cebada','5872'), ('CyMQuilmes.Cebada','QUILMES AYELEN ')),
                        (('Cebada','11304'), ('CyMQuilmes.Cebada','QUILMES CARISMA ')),
                        (('Cebada','1141'), ('CyMQuilmes.Cebada','QUILMES CENTAURO ')),
                        (('Cebada','8669'), ('CyMQuilmes.Cebada','QUILMES KUYEN ')),
                        (('Cebada','5069'), ('CyMQuilmes.Cebada','QUILMES PAINE ')),
                        (('Cebada','3428'), ('CyMQuilmes.Cebada','QUILMES PALOMAR ')),
                        (('Cebada','1005'), ('CyMQuilmes.Cebada','QUILMES PAMPA ')),
                        (('Centeno','5399'), ('INTA.Centeno','Quehue')),
                        (('Trigo','19549'), ('RAGT.Trigo','Quiriko')),
                        (('Soja','11113'), ('CriaSantaRosa.Soja','RA 334')),
                        (('Soja','11111'), ('CriaSantaRosa.Soja','RA 338')),
                        (('Soja','12835'), ('CriaSantaRosa.Soja','RA 349')),
                        (('Soja','9895'), ('CriaSantaRosa.Soja','RA 426')),
                        (('Soja','12147'), ('CriaSantaRosa.Soja','RA 437')),
                        (('Soja','13440'), ('CriaSantaRosa.Soja','RA 444')),
                        (('Soja','13522'), ('CriaSantaRosa.Soja','RA 449')),
                        (('Soja','10285'), ('CriaSantaRosa.Soja','RA 524')),
                        (('Soja','11241'), ('CriaSantaRosa.Soja','RA 532')),
                        (('Soja','11242'), ('CriaSantaRosa.Soja','RA 536')),
                        (('Soja','10446'), ('CriaSantaRosa.Soja','RA 538')),
                        (('Soja','13562'), ('CriaSantaRosa.Soja','RA 541')),
                        (('Soja','13518'), ('CriaSantaRosa.Soja','RA 545')),
                        (('Soja','12151'), ('CriaSantaRosa.Soja','RA 549')),
                        (('Soja','14260'), ('CriaSantaRosa.Soja','RA 550')),
                        (('Soja','13520'), ('CriaSantaRosa.Soja','RA 556')),
                        (('Soja','15393'), ('CriaSantaRosa.Soja','RA 5715 IPRO')),
                        (('Soja','12150'), ('CriaSantaRosa.Soja','RA 644')),
                        (('Soja','11243'), ('CriaSantaRosa.Soja','RA 732')),
                        (('Soja','9890'), ('CriaSantaRosa.Soja','RA 733')),
                        (('Soja','13436'), ('CriaSantaRosa.Soja','RA 744')),
                        (('Soja','13561'), ('CriaSantaRosa.Soja','RA 844')),
                        (('Soja','7028'), ('Sursem.Soja','RAFAELA58')),
                        (('Cebada','11725'), ('INTA.Cebada','RAYEN INTA ')),
                        (('Maíz','19802'), ('Brevant.Maíz','RFG 22')),
                        (('Maíz','19802'), ('Brevant.Maíz','RFG22RRE')),
                        (('Girasol','21515'), ('RAGT.Girasol','RGT CHARLLOTTE')),
                        (('Girasol','13932'), ('RAGT.Girasol','RGT KAPLLAN')),
                        (('Girasol','13936'), ('RAGT.Girasol','RGT MOOGLLI')),
                        (('Girasol','700'), ('RAGT.Girasol','RGT OBELLISCO')),
                        (('Girasol','13931'), ('RAGT.Girasol','RGT SIKLLOS')),
                        (('Girasol','13933'), ('RAGT.Girasol','RGT VELLOX')),
                        (('Colza','9688'), ('Nuseed.Colza','RIVETTE')),
                        (('Soja','12161'), ('Sursem.Soja','RM 6900')),
                        (('Soja','9904'), ('Sursem.Soja','RMO4637')),
                        (('Soja','9908'), ('Sursem.Soja','RMO58')),
                        (('Soja','9901'), ('Sursem.Soja','RMO75')),
                        (('Soja','12658'), ('Sursem.Soja','RMO77')),
                        (('Soja','12588'), ('Sursem.Soja','RMO805')),
                        (('Alfalfa','13630'), ('LosPrados.Alfalfa','ROBERTA')),
                        (('Trigo','18645'), ('Buck.Trigo','Resplandor')),
                        (('Arveja','14839'), ('Bioseminis.Arveja','Reussite')),
                        (('Girasol','12473'), ('NuSeed.Girasol','Rhino')),
                        (('Centeno','15277'), ('INTA.Centeno','Ricardo')),
                        (('Trigo','13089'), ('Klein.Trigo','Roble')),
                        (('Poroto','1700'), ('EEAOC.Poroto','Rojo - PVAD 1101')),
                        (('Poroto','1701'), ('EEAOC.Poroto','Rojo - PVAD 1111')),
                        (('Poroto','7403'), ('EEAOC.Poroto','Rojo - TUC 310')),
                        (('Arveja','20651'), ('ARVES.Arveja','Rosita')),
                        (('Cebada','15873'), ('INTA.Cebada','SARA INTA ')),
                        (('Cebada','15952'), ('INTA.Cebada','SILERA INTA ')),
                        (('Trigo','9537'), ('Sursem.Trigo','SIRIRI')),
                        (('Colza','10478'), ('HighTech.Colza','SITRO')),
                        (('Soja','12866'), ('Klein.Soja','SK 3.5')),
                        (('Soja','11770'), ('Klein.Soja','SK 3.8')),
                        (('Soja','11146'), ('Klein.Soja','SK 4.7')),
                        (('Maíz','17512'), ('Nidera.Maíz','SMF 8007')),
                        (('Maíz','19486'), ('Nidera.Maíz','SMF 8080')),
                        (('Trigo','16445'), ('Sursem.Trigo','SN 90')),
                        (('Soja','8909'), ('Syngenta.Soja','SP 3900')),
                        (('Soja','9160'), ('Syngenta.Soja','SP 4500')),
                        (('Soja','9845'), ('Syngenta.Soja','SP 4X0')),
                        (('Soja','14879'), ('Syngenta.Soja','SP 4X3 IPRO')),
                        (('Soja','10560'), ('Syngenta.Soja','SP 4X4')),
                        (('Soja','15379'), ('Syngenta.Soja','SP 4X6 IPRO')),
                        (('Soja','11637'), ('Syngenta.Soja','SP 4X99')),
                        (('Soja','11185'), ('Syngenta.Soja','SP 5x2')),
                        (('Soja','10093'), ('Syngenta.Soja','SP 5x5')),
                        (('Soja','11178'), ('Syngenta.Soja','SP 5x9')),
                        (('Soja','13501'), ('Syngenta.Soja','SP 6X1')),
                        (('Soja','9866'), ('Syngenta.Soja','SP 6X2')),
                        (('Soja','9869'), ('Syngenta.Soja','SP 7x0')),
                        (('Soja','9856'), ('Syngenta.Soja','SP 8x0')),
                        (('Soja','12789'), ('Syngenta.Soja','SP 8x8')),
                        (('Maíz','7334'), ('SPS.Maíz','SPS 2727')),
                        (('Soja','14324'), ('Syngenta.Soja','SPS 6x6 IPRO')),
                        (('Soja','14334'), ('Syngenta.Soja','SPS 7x8 IPRO')),
                        (('Colza','9932'), ('Sursem.Colza','SRM 2836')),
                        (('Soja','12153'), ('Sursem.Soja','SRM 3300')),
                        (('Soja','12042'), ('Sursem.Soja','SRM 3410')),
                        (('Soja','12709'), ('Sursem.Soja','SRM 3801')),
                        (('Soja','14781'), ('Sursem.Soja','SRM 3988')),
                        (('Soja','14177'), ('Sursem.Soja','SRM 4222')),
                        (('Soja','8398'), ('Sursem.Soja','SRM 4500')),
                        (('Soja','12910'), ('Sursem.Soja','SRM 4602 STS')),
                        (('Soja','9837'), ('Sursem.Soja','SRM 4754')),
                        (('Soja','11290'), ('Sursem.Soja','SRM 4839')),
                        (('Soja','10485'), ('Sursem.Soja','SRM 5001')),
                        (('Soja','12908'), ('Sursem.Soja','SRM 5200')),
                        (('Soja','9232'), ('Sursem.Soja','SRM 5301')),
                        (('Soja','12158'), ('Sursem.Soja','SRM 5500')),
                        (('Maíz','12602'), ('Sursem.Maíz','SRM 553')),
                        (('Maíz','10299'), ('Sursem.Maíz','SRM 56-20')),
                        (('Maíz','14417'), ('Sursem.Maíz','SRM 56-22')),
                        (('Maíz','14541'), ('Sursem.Maíz','SRM 56-24')),
                        (('Soja','11294'), ('Sursem.Soja','SRM 5601')),
                        (('Maíz','12601'), ('Sursem.Maíz','SRM 567')),
                        (('Soja','12159'), ('Sursem.Soja','SRM 5700')),
                        (('Soja','15072'), ('Sursem.Soja','SRM 5835IPRO')),
                        (('Soja','13524'), ('Sursem.Soja','SRM 6256')),
                        (('Soja','9933'), ('Sursem.Soja','SRM 6403')),
                        (('Maíz','18074'), ('Sursem.Maíz','SRM 6620')),
                        (('Maíz','18076'), ('Sursem.Maíz','SRM 6670')),
                        (('Soja','12161'), ('Sursem.Soja','SRM 6900')),
                        (('Soja','12162'), ('Sursem.Soja','SRM 7200')),
                        (('Soja','12710'), ('Sursem.Soja','SRM 7800')),
                        (('Soja','11295'), ('Sursem.Soja','SRM 8201')),
                        (('Maíz','7758'), ('Sursem.Maíz','SRM POPER 42')),
                        (('Maíz','19606'), ('Stine.Maíz','ST 120-09 RG')),
                        (('Maíz','18860'), ('Stine.Maíz','ST 120-29 BTRG')),
                        (('Maíz','18230'), ('Stine.Maíz','ST 9734-20')),
                        (('Maíz','18229'), ('Stine.Maíz','ST 9734-G')),
                        (('Maíz','17565'), ('Stine.Maíz','ST 9739E-20')),
                        (('Maíz','21654'), ('Stine.Maíz','ST 9741-20')),
                        (('Maíz','18859'), ('Stine.Maíz','ST 9808 E-20')),
                        (('Maíz','19605'), ('Stine.Maíz','ST 9820-20')),
                        (('Maíz','20685'), ('Stine.Maíz','ST 9910-20')),
                        (('Sorgo','13699'), ('Advanta.Sorgo','SUGARGRAZE AR')),
                        (('Trigo','15179'), ('Buck.Trigo','SY 015')),
                        (('Trigo','15180'), ('Buck.Trigo','SY 041')),
                        (('Trigo','11657'), ('Buck.Trigo','SY 100')),
                        (('Trigo','19469'), ('Buck.Trigo','SY 109')),
                        (('Trigo','12337'), ('Buck.Trigo','SY 110')),
                        (('Trigo','17064'), ('Buck.Trigo','SY 120')),
                        (('Trigo','11658'), ('Buck.Trigo','SY 200')),
                        (('Trigo','15753'), ('Buck.Trigo','SY 211')),
                        (('Trigo','11656'), ('Buck.Trigo','SY 300')),
                        (('Trigo','15786'), ('Buck.Trigo','SY 330')),
                        (('Soja','11200'), ('Syngenta.Soja','SY 3X5')),
                        (('Soja','14833'), ('Syngenta.Soja','SY 5X8 IPRO')),
                        (('Soja','14817'), ('Syngenta.Soja','SY 5x1 RR')),
                        (('Soja','14818'), ('Syngenta.Soja','SY 6x8 IPRO')),
                        (('Soja','15525'), ('Syngenta.Soja','SY 7X1IPRO')),
                        (('Trigo','16707'), ('Buck.Trigo','SY OBELIX')),
                        (('Soja','16070'), ('Syngenta.Soja','SYN 1561IPRO')),
                        (('Girasol','17893'), ('Syngenta.Girasol','SYN 3990')),
                        (('Girasol','17892'), ('Syngenta.Girasol','SYN 4066')),
                        (('Maíz','19684'), ('Syngenta.Maíz','SYN 505')),
                        (('Maíz','18617'), ('Syngenta.Maíz','SYN 840')),
                        (('Maíz','18618'), ('Syngenta.Maíz','SYN 897')),
                        (('Maíz','17998'), ('Syngenta.Maíz','SYN 979')),
                        (('Trigo','15754'), ('Buck.Trigo','Saeta')),
                        (('Trigo','6818'), ('Klein.Trigo','Sagitario')),
                        (('Maíz','9476'), ('Albert.Maíz','Santa Fe 2')),
                        (('Trigo','18188'), ('DonMario.Trigo','Sauce')),
                        (('Cebada','6359'), ('Cargill.Cebada','Scarlett')),
                        (('Cebada','14183'), ('MalteriaPampa.Cebada','Scrabble')),
                        (('Trigo','6817'), ('Proseme.Trigo','Scudo')),
                        (('Trigo','20522'), ('Klein.Trigo','Selenio CL')),
                        (('Trigo','15216'), ('Klein.Trigo','Serpiente')),
                        (('Cebada','10134'), ('Limagrain.Cebada','Shakira')),
                        (('Cebada','10134'), ('MalteriaPampa.Cebada','Shakira')),
                        (('Arveja','15991'), ('Limagrain.Arveja','Shamrock')),
                        (('Sorgo','10019'), ('Pemán.Sorgo','Silero Inta Pemán')),
                        (('Maíz','16598'), ('Albert.Maíz','Silomax Full')),
                        (('Cebada','17891'), ('ACA.Cebada','Sinfonia')),
                        (('Alfalfa','8733'), ('Biscayart.Alfalfa','Super Aurora')),
                        (('Alfalfa','9923'), ('Biscayart.Alfalfa','Super Sonic')),
                        (('Alfalfa','13109'), ('Biscayart.Alfalfa','Super Star')),
                        (('Trigo','6686'), ('Buck.Trigo','Sureño')),
                        (('Sorgo','6488'), ('Biscayart.Sorgo','Sweetgreen')),
                        (('Cebada','12628'), ('Cargill.Cebada','Sylphide')),
                        (('Soja','11150'), ('LaTijereta.Soja','T 2137')),
                        (('Soja','11998'), ('LaTijereta.Soja','T 2246')),
                        (('Soja','12035'), ('LaTijereta.Soja','T 2249')),
                        (('Soja','12795'), ('LaTijereta.Soja','T 2259')),
                        (('Soja','13468'), ('LaTijereta.Soja','T 2266')),
                        (('Girasol','12733'), ('Agseed.Girasol','TB 11')),
                        (('Girasol','12724'), ('Agseed.Girasol','TB 14 IMI')),
                        (('Soja','11255'), ('LaTijereta.Soja','TJ 2138 R')),
                        (('Soja','7221'), ('LaTijereta.Soja','TJS 2044')),
                        (('Soja','7222'), ('LaTijereta.Soja','TJS 2049 RR')),
                        (('Soja','7696'), ('LaTijereta.Soja','TJS 2055 RR')),
                        (('Soja','4150'), ('LaTijereta.Soja','TJS 2070')),
                        (('Soja','11256'), ('LaTijereta.Soja','TJS 2136 RR')),
                        (('Soja','8440'), ('LaTijereta.Soja','TJS 2139')),
                        (('Soja','11254'), ('LaTijereta.Soja','TJS 2145 RR')),
                        (('Soja','11255'), ('LaTijereta.Soja','TJS 2148')),
                        (('Soja','8439'), ('LaTijereta.Soja','TJS 2156')),
                        (('Soja','11214'), ('LaTijereta.Soja','TJS 2158 R')),
                        (('Soja','8438'), ('LaTijereta.Soja','TJS 2164')),
                        (('Soja','11640'), ('LaTijereta.Soja','TJS 2165 R')),
                        (('Soja','8437'), ('LaTijereta.Soja','TJS 2170')),
                        (('Soja','10505'), ('LaTijereta.Soja','TJS 2171 R')),
                        (('Soja','8436'), ('LaTijereta.Soja','TJS 2178')),
                        (('Sorgo','13354'), ('Tobin.Sorgo','TOB 62 T')),
                        (('Sorgo','17355'), ('Tobin.Sorgo','TOB FACA BMR')),
                        (('Sorgo','10972'), ('Tobin.Sorgo','TOB Matrero')),
                        (('Sorgo','15573'), ('Tobin.Sorgo','TOB Padrillo Plus')),
                        (('Trigo','21632'), ('Bioseminis.Trigo','TRAFUL HB4')),
                        (('Cebada','17701'), ('INTA.Cebada','TRINIDAD INTA ')),
                        (('Sorgo','14384'), ('LaTijereta.Sorgo','TS 267')),
                        (('Sorgo','9767'), ('LaTijereta.Sorgo','TS 281')),
                        (('Sorgo','13012'), ('LaTijereta.Sorgo','TS 283')),
                        (('Trigo','10659'), ('Buck.Trigo','Taita')),
                        (('Sorgo','17101'), ('Pemán.Sorgo','Takurí')),
                        (('Trigo','9062'), ('Klein.Trigo','Tauro')),
                        (('Trigo','17458'), ('DonMario.Trigo','Tbio Audaz')),
                        (('Triticale','1648'), ('INTA.Triticale','Tehuelche')),
                        (('Trigo','10011'), ('DonMario.Trigo','Themix')),
                        (('Trigo','10787'), ('Klein.Trigo','Tigre')),
                        (('Alfalfa','10500'), ('PGG.Alfalfa','Tigresa')),
                        (('Trigo','14388'), ('Buck.Trigo','Tilcara')),
                        (('Trigo','14487'), ('Bioceres.Trigo','Timbo')),
                        (('Trigo','15215'), ('Klein.Trigo','Titanio')),
                        (('Trigo','5422'), ('Buck.Trigo','Topacio')),
                        (('Trigo','20491'), ('Illinois.Trigo','Tordo')),
                        (('Trigo','21632'), ('Bioceres.Trigo','Traful HB4')),
                        (('Cebada','15240'), ('ACA.Cebada','Traveler')),
                        (('Cebada','13141'), ('Cargill.Cebada','Umbrella')),
                        (('Cebada','1048'), ('INTA.Cebada','UÑAICHE INTA ')),
                        (('Sorgo','8596'), ('Advanta.Sorgo','VDH 314')),
                        (('Sorgo','8193'), ('Advanta.Sorgo','VDH 422')),
                        (('Sorgo','5855'), ('Advanta.Sorgo','VDH 701')),
                        (('Alfalfa','2114'), ('BayaCasal.Alfalfa','VICTORIA')),
                        (('Alfalfa','2114'), ('Produsem.Alfalfa','VICTORIA SP INTA')),
                        (('Soja','14851'), ('VTSeeds.Soja','VT 5335')),
                        (('Girasol','17063'), ('Argensun.Girasol','Valia 41')),
                        (('Girasol','17672'), ('Argensun.Girasol','Valia 73')),
                        (('Girasol','19573'), ('Argensun.Girasol','Valia 92')),
                        (('Girasol','17063'), ('Argensun.Girasol','Valia NTC 418')),
                        (('Girasol','21736'), ('Argensun.Girasol','Valia V22')),
                        (('Trigo','17993'), ('Klein.Trigo','Valor')),
                        (('Alfalfa','10952'), ('Barenbrug.Alfalfa','Verdor')),
                        (('Alfalfa','11940'), ('Barenbrug.Alfalfa','Verzy')),
                        (('Alfalfa','2114'), ('Biscayart.Alfalfa','Victoria')),
                        (('Arveja','5804'), ('AFA.Arveja','Viper')),
                        (('Alfalfa','10225'), ('WL.Alfalfa','WL1058')),
                        (('Alfalfa','9281'), ('WL.Alfalfa','WL611')),
                        (('Alfalfa','17145'), ('WL.Alfalfa','WL825 HVX.RR')),
                        (('Alfalfa','19713'), ('WL.Alfalfa','WL828')),
                        (('Alfalfa','17144'), ('WL.Alfalfa','WL835 HVX.RR')),
                        (('Alfalfa','8559'), ('WL.Alfalfa','WL903')),
                        (('Alfalfa','15124'), ('WL.Alfalfa','WL919')),
                        (('Girasol','19670'), ('NuSeed.Girasol','X4334 CL')),
                        (('Triticale','2776'), ('INTA.Triticale','Yagan')),
                        (('Arveja','14726'), ('Bioseminis.Arveja','Yams')),
                        (('Trigo','11595'), ('Klein.Trigo','Yarara')),
                        (('Trigo','5421'), ('Buck.Trigo','Yatasto')),
                        (('Girasol','19616'), ('ZetaSemillas.Girasol','ZT 74C15')),
                        (('Girasol','19614'), ('ZetaSemillas.Girasol','ZT 74H55 ')),
                        (('Girasol','19663'), ('ZetaSemillas.Girasol','ZT 74H70')),
                        (('Girasol','19615'), ('ZetaSemillas.Girasol','ZT 74L60')),
                        (('Girasol','20755'), ('ZetaSemillas.Girasol','ZT 74L62')),
                        (('Girasol','19662'), ('ZetaSemillas.Girasol','ZT 75L50')),
                        (('Trigo','15712'), ('Buck.Trigo','Zafiro')),
                        (('Trigo','20076'), ('LG.Trigo','Zaino')),
                        (('Trigo','9559'), ('Klein.Trigo','Zorro')),
                        (('Trigo','21945'), ('Neogen.Trigo','neo30t23')),
                        (('Trigo','21946'), ('Neogen.Trigo','neo50t23')),
                        (('Trigo','17457'), ('DonMario.Trigo','Ñandubay'))
                    ]
                    
                    for (cultivo, codigo_genetica), (semillero, genetica) in asignaciones:
                        df.loc[(df['Cultivo'] == cultivo) & (df['Codigo_Genetica'] == codigo_genetica), ['Semillero', 'Genetica']] = [semillero, genetica]
                
                    return df
                    
                df = validar_genetica_semillero(df, conn)

                # Convertir el Código de Fertilizante en integer # FUNCIONA
                def fertilizante_int(df, conn):
                    columnas = [
                        '1_Codigo_Registro_1', '1_Codigo_Registro_2',
                        '2_Codigo_Registro_1', '2_Codigo_Registro_2',
                        '3_Codigo_Registro_1', '3_Codigo_Registro_2',
                        '4_Codigo_Registro_1', '4_Codigo_Registro_2'
                    ]
                    
                    for col in columnas:
                        df[col] = df[col].fillna(0).astype('Int64')
                    
                    return df
                
                df = fertilizante_int(df, conn)

                # 3.7) Validación de Fertilizantes # FUNCIONA
                
                # Registro fertilizantes 
                def validar_id_senasa_fertilizantes(df, conn):
                    reg_fert_query = """SELECT id_senasa FROM datcrea_tablas.registro_fertilizantes WHERE id_senasa IS NOT NULL"""
                    reg_fert_df = pd.read_sql(reg_fert_query,con=conn)
                    id_senasa = reg_fert_df['id_senasa'].astype(int).tolist()
                    
                    df['1_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['1_Codigo_Registro_1']]
                    df['1_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['1_Codigo_Registro_2']]
                    df['2_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['2_Codigo_Registro_1']]
                    df['2_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['2_Codigo_Registro_2']]
                    df['3_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['3_Codigo_Registro_1']]
                    df['3_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['3_Codigo_Registro_2']]
                    df['4_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['4_Codigo_Registro_1']]
                    df['4_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['4_Codigo_Registro_2']]
                    return df
                df = validar_id_senasa_fertilizantes(df, conn)
                
                def validar_nombre_fertilizante(df, conn):
                
                    # Consulta para obtener los fertilizantes registrados con un id_senasa no nulo
                    reg_fert_query = "SELECT * FROM datcrea_tablas.registro_fertilizantes WHERE id_senasa IS NOT NULL"
                    registros_fertilizantes = pd.read_sql(reg_fert_query, con=conn)
                    
                    # Asegurarse de que id_senasa es de tipo entero
                    registros_fertilizantes['id_senasa'] = registros_fertilizantes['id_senasa'].astype(int)
                    
                    # Obtener la lista de fertilizantes únicos
                    fertilizantes_lista = registros_fertilizantes['fertilizante'].unique().tolist()
                    
                    # Lista de columnas a validar
                    columnas_a_validar = [
                        '1_Producto_1', '1_Producto_2', 
                        '2_Producto_1', '2_Producto_2', 
                        '3_Producto_1', '3_Producto_2', 
                        '4_Producto_1', '4_Producto_2'
                    ]
                    
                    # Validar cada columna en la lista
                    for col in columnas_a_validar:
                        df[col] = df[col].apply(lambda x: x if x in fertilizantes_lista else '')
                    
                    return df
                df = validar_nombre_fertilizante(df, conn)


                def validar_fertilizante(df,conn):
                    reg_fert_query = """SELECT * FROM datcrea_tablas.registro_fertilizantes WHERE id_senasa IS NOT NULL"""
                    registros_fertilizantes = pd.read_sql(reg_fert_query, con=conn)
                    registros_fertilizantes['id_senasa'] = registros_fertilizantes['id_senasa'].astype(int)
                    
                    # Crear un diccionario para almacenar los resultados
                    id_senasa_dict = {}
                    
                    
                    for index, row in registros_fertilizantes.iterrows():
                        fertilizante = row['fertilizante']
                        id_senasa = row['id_senasa']
                        if fertilizante in id_senasa_dict:
                            id_senasa_dict[fertilizante].append(id_senasa)
                        else:
                            id_senasa_dict[fertilizante] = [id_senasa]
                    
                    # Validar fertilizante en función del código
                    def validar_fertilizante_codigo(codigo):
                        
                        if pd.isna(codigo):
                            return ''
                        for producto, codigos in id_senasa_dict.items():
                            if codigo in codigos:
                                return producto
                        return ''
                    
                    # Aplicar validación en cada columna
                    for i in range(1, 5):  
                        df[f'{i}_Producto_1'] = df[f'{i}_Codigo_Registro_1'].apply(validar_fertilizante)
                        df[f'{i}_Producto_2'] = df[f'{i}_Codigo_Registro_2'].apply(validar_fertilizante)
                    
                    return df

                # VALIDAR DOSIS DE PRODUCTOS (si es que primero se validó el producto) # FUNCIONA
                def validar_dosis_productos(df):
                
                    columnas_productos_dosis = [
                        ('1_Producto_1', '1_Dosis_1'),
                        ('1_Producto_2', '1_Dosis_2'),
                        ('2_Producto_1', '2_Dosis_1'),
                        ('2_Producto_2', '2_Dosis_2'),
                        ('3_Producto_1', '3_Dosis_1'),
                        ('3_Producto_2', '3_Dosis_2'),
                        ('4_Producto_1', '4_Dosis_1'),
                        ('4_Producto_2', '4_Dosis_2')
                    ]
                
                    for producto_col, dosis_col in columnas_productos_dosis:
                        df.loc[df[producto_col] == '', dosis_col] = pd.NA
                
                    return df
                df = validar_dosis_productos(df)



                def validar_momentos(df, conn):
                
                    # Lista de momentos válidos
                    momentos_validos = [
                        'Presiembra',
                        'Refertilización',
                        'Siembra',
                        'Post siembra',
                        'Otro'
                    ]
                
                    # Columnas de momento a ajustar
                    columnas_momento = ['1_Momento', '2_Momento', '3_Momento', '4_Momento']
                
                    # Iterar sobre las columnas para ajustar los momentos
                    for columna in columnas_momento:
                        df[columna] = df[columna].apply(lambda x: x if x in momentos_validos else '')
                
                    return df
                
                df = validar_momentos(df,conn)

                def calcular_densidad(df, conn):
                        
                    df['DensidadPlantasHa'] = df['DensidadPlantasHa']/10000
                    df.rename(columns={'DensidadPlantasHa': 'DensidadPlantasM2'}, inplace=True)
                    return df
                
                df = calcular_densidad(df, conn)


                def reindex_columns(df):
                    # Columnas del procesamiento anterior
                    cols = ['CUIC','Campo','Campo_Otro','Columna_I','Lote','Lote_Otro','vacio','vacio','vacio',
                        'Provincia','Departamento','vacio','Localidad','CalidadAmbiente','SuperficieSembrada',
                        'Tenencia','DestinoProduccion','vacio','Cultivo','vacio','SubgrupoCultivo','vacio',
                        'CultivoAntecesor','vacio','FechaSiembra','vacio','vacio','vacio','espaciamiento',
                        'DensidadKgHa','DensidadSemillasM2','DensidadSemillasHA','DensidadPlantasM2','vacio','vacio',
                        'Semillero','Genetica','vacio','Fertilizacion','1_Producto_1','1_Dosis_1','1_Momento',
                        '1_Forma','1_Producto_2','1_Dosis_2','2_Momento','1_Forma','2_Producto_1','2_Dosis_1',
                        '3_Momento','2_Forma','2_Producto_2','2_Dosis_2','4_Momento','2_Forma','3_Producto_1',
                        '3_Dosis_1','5_Momento','3_Forma','3_Producto_2','3_Dosis_2','6_Momento','3_Forma',
                        'vacio','Riego','Lamina_Riego','Napa','vacio','vacio','vacio','vacio','vacio','vacio',
                        'vacio','Superficie','vacio','Rendimiento']
                    return df.reindex(columns = cols)


                
                df = reindex_columns(df)

                
                # 5) Creación de columnas Campo_Otro y Lote_Otro (para caso SIMA) #FUNCIONA
                def otro_cuic_campo_lote(df, conn):
                    #df['CUIC'] = df['CUIC'].astype(str)
                    df['CUIC'] = df['CUIC'].fillna('Otro')
                    #df['Campo_Otro'] = df['Campo_Otro'].fillna('Otro')
                    #df['Lote_Otro'] = df['Lote_Otro'].fillna('Otro')
                    
                    # Llenar la columna 'Campo_Otro' con 'Otro' donde la columna 'Campo' tiene NaN
                    df.loc[df['Campo'].isna(), 'Campo_Otro'] = 'Otro'
                    # Dejar vacía la columna 'Campo_Otro' donde la columna 'Campo' tiene datos
                    df.loc[df['Campo'].notna(), 'Campo_Otro'] = ''
                    
                    # Llenar la columna 'Lote_Otro' con 'otro' donde la columna 'Lote' tiene NaN
                    df.loc[df['Lote'].isna(), 'Lote_Otro'] = 'Otro'
                    # Dejar vacía la columna 'Lote_Otro' donde la columna 'Lote' tiene datos
                    df.loc[df['Lote'].notna(), 'Lote_Otro'] = ''
                
                    return df
                
                df = otro_cuic_campo_lote(df, conn)


                # Localidades
                def validar_localidad(df):
                    
                    localidades_query = """SELECT localidad FROM datcrea_tablas.localidades"""
                    
                    localidades = pd.read_sql(localidades_query,con=conn)
                    localidades = localidades['localidad'].str.title().tolist()
                    ## 6.1) Validación Localidad
                    # Unificar criterio de escritura
                    df['Localidad'] = df['Localidad'].astype(str)
                    df['Localidad'] = df['Localidad'].str.title()
                    df['Localidad'] = ['' if x not in localidades else x for x in df['Localidad']]
                    return df
            

                df = validar_localidad(df)
                
                # Departamentos
                def validar_departamento(df,conn):
                
                    reemplazos = {
                        '25 DE MAYO': 'Veinticinco.de.Mayo',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'ADOLFO ALSINA': 'Adolfo.Alsina',
                        'ADOLFO GONZALES CHAVES': 'Adolfo.Gonzales.Chaves',
                        'ALBERTI': 'Alberti',
                        'ALMIRANTE BROWN': '',
                        'ARRECIFES': 'Arrecifes',
                        'AVELLANEDA': '',
                        'AYACUCHO': 'Ayacucho',
                        'AZUL': 'Azul',
                        'BAHIA BLANCA': 'Bahía.Blanca',
                        'BALCARCE': 'Balcarce',
                        'BARADERO': 'Baradero',
                        'BENITO JUAREZ': 'Benito.Juarez',
                        'BERAZATEGUI': 'Berazategui',
                        'BERISSO': 'Berisso',
                        'BOLIVAR': 'Bolivar',
                        'BRAGADO': 'Partido.Bragado',
                        'BRANDSEN': 'Brandsen',
                        'CAMPANA': 'Campana',
                        'CANUELAS': 'Cañuelas',
                        'CAPITAN SARMIENTO': 'Capitan.Sarmiento',
                        'CARLOS CASARES': 'Carlos.Casares',
                        'CARLOS TEJEDOR': 'Carlos.Tejedor',
                        'CARMEN DE ARECO': 'Carmen.de.Areco',
                        'CASTELLI': 'Castelli',
                        'CHACABUCO': 'Chacabuco',
                        'CHASCOMUS': 'Chascomus',
                        'CHIVILCOY': 'Chivilcoy',
                        'COLON': 'Colon',
                        'CORONEL DE MARINA L. ROSALES': 'Coronel.de.Marina.L.Rosales',
                        'CORONEL DORREGO': 'Coronel.Dorrego',
                        'CORONEL PRINGLES': 'Coronel.Pringles',
                        'CORONEL SUAREZ': 'Coronel.Suarez',
                        'DAIREAUX': 'Daireaux',
                        'DOLORES': 'Dolores',
                        'ENSENADA': 'Ensenada',
                        'ESCOBAR': 'Escobar',
                        'ESTEBAN ECHEVERRIA': 'Esteban.Echeverria',
                        'EXALTACION DE LA CRUZ': 'Exaltación.de.La.Cruz',
                        'EZEIZA': 'Ezeiza',
                        'FLORENCIO VARELA': 'Florencio.Varela',
                        'FLORENTINO AMEGHINO': 'Florentino.Ameghino',
                        'GENERAL ALVARADO': 'General.Alvarado',
                        'GENERAL ALVEAR': 'General.Alvear',
                        'GENERAL ARENALES': 'General.Arenales',
                        'GENERAL BELGRANO': 'General.Belgrano',
                        'GENERAL GUIDO': 'General.Guido',
                        'GENERAL JUAN MADARIAGA': 'General.Juan.Madariaga',
                        'GENERAL LA MADRID': 'General.La.Madrid',
                        'GENERAL LAS HERAS': 'General.Las.Heras',
                        'GENERAL LAVALLE': 'General.Lavalle',
                        'GENERAL PAZ': 'General.Paz',
                        'GENERAL PINTO': 'General.Pinto',
                        'GENERAL PUEYRREDON': 'General.Pueyrredon',
                        'GENERAL RODRIGUEZ': 'General.Rodriguez',
                        'GENERAL SAN MARTIN': 'General.San.Martin',
                        'GENERAL VIAMONTE': 'General.Viamonte',
                        'GENERAL VILLEGAS': 'General.Villegas',
                        'GUAMINI': 'Guamini',
                        'HIPOLITO YRIGOYEN': 'Hipolito.Yrigoyen',
                        'HURLINGHAM': 'Hurlingham',
                        'ITUZAINGO': 'Ituzaingo',
                        'JOSE C. PAZ': 'Jose.C.Paz',
                        'JUNIN': 'Junin',
                        'LA COSTA': 'La.Costa',
                        'LA MATANZA': 'La.Matanza',
                        'LA PLATA': 'La.Plata',
                        'LANUS': 'Lanús',
                        'LAPRIDA': 'Laprida',
                        'LAS FLORES': 'Las.Flores',
                        'LEANDRO N. ALEM': 'Leandro.N..Alem',
                        'LINCOLN': 'Lincoln',
                        'LOBERIA': 'Loberia',
                        'LOBOS': 'Lobos',
                        'LOMAS DE ZAMORA': 'Lomas.de.Zamora',
                        'LUJAN': 'Lujan',
                        'MAGDALENA': 'Magdalena',
                        'MAIPU': 'Maipu',
                        'MALVINAS ARGENTINAS': 'Malvinas.Argentinas',
                        'MAR CHIQUITA': 'Mar.Chiquita',
                        'MARCOS PAZ': 'Marcos.Paz',
                        'MERCEDES': 'Mercedes',
                        'MERLO': 'Merlo',
                        'MONTE': 'Monte',
                        'MONTE HERMOSO': 'Monte.Hermoso',
                        'MORENO': 'Moreno',
                        'MORON': 'Moron',
                        'NAVARRO': 'Navarro',
                        'NECOCHEA': 'Necochea',
                        'OLAVARRIA': 'Olavarria',
                        'PATAGONES': 'Patagones',
                        'PEHUAJO': 'Pehuajo',
                        'PELLEGRINI': 'Pellegrini',
                        'PERGAMINO': 'Partido.Pergamino',
                        'PILA': 'Pila',
                        'PILAR': 'Pilar',
                        'PINAMAR': 'Pinamar',
                        'PRESIDENTE PERON': 'Presidente.Peron',
                        'PUAN': 'Puan',
                        'PUNTA INDIO': 'Punta.Indio',
                        'QUILMES': 'Quilmes',
                        'RAMALLO': 'Ramallo',
                        'RAUCH': 'Rauch',
                        'RIVADAVIA': 'Rivadavia',
                        'ROJAS': 'Rojas',
                        'ROQUE PEREZ': 'Roque.Perez',
                        'SAAVEDRA': 'Saavedra',
                        'SALADILLO': 'Saladillo',
                        'SALLIQUELO': 'Salliqueló',
                        'SALTO': 'Partido.Salto',
                        'SAN ANDRES DE GILES': 'San.Andres.de.Giles',
                        'SAN ANTONIO DE ARECO': 'San.A.de.Areco',
                        'SAN CAYETANO': 'San.Cayetano',
                        'SAN FERNANDO': 'San.Fernando',
                        'SAN ISIDRO': 'San.Isidro',
                        'SAN MIGUEL': 'San.Miguel',
                        'SAN NICOLAS': 'San.Nicolas',
                        'SAN PEDRO': 'San.Pedro',
                        'SAN VICENTE': 'San.Vicente',
                        'SUIPACHA': 'Suipacha',
                        'TANDIL': 'Tandil',
                        'TAPALQUE': 'Tapalque',
                        'TIGRE': 'Tigre',
                        'TORDILLO': 'Tordillo',
                        'TORNQUIST': 'Tornquist',
                        'TRENQUE LAUQUEN': 'Trenque.Lauquen',
                        'TRES ARROYOS': 'Tres.Arroyos',
                        'TRES DE FEBRERO': 'Tres.de.Febrero',
                        'TRES LOMAS': 'Tres.Lomas',
                        'VICENTE LOPEZ': 'Vicente.Lopez',
                        'VILLA GESELL': 'Villa.Gesell',
                        'VILLARINO': 'Villarino',
                        'ZARATE': 'Zarate',
                        'CAPITAL FEDERAL': '',
                        'AMBATO': 'Ambato',
                        'ANCASTI': 'Ancasti',
                        'ANDALGALA': '',
                        'ANTOFAGASTA DE LA SIERRA': 'Antofagasta.de.La.Sierra',
                        'BELEN': 'Belen',
                        'CAPAYAN': 'Capayan',
                        'CAPITAL': 'Capital',
                        'EL ALTO': 'El.Alto',
                        'FRAY MAMERTO ESQUIU': 'Fray.Mamerto.Esquiu',
                        'LA PAZ': 'La.Paz',
                        'PACLIN': 'Paclin',
                        'POMAN': 'Poman',
                        'SANTA MARIA': 'Santa.Maria',
                        'SANTA ROSA': 'Santa.Rosa',
                        'TINOGASTA': 'Tinogasta',
                        'VALLE VIEJO': 'Valle.Viejo',
                        '1 DE MAYO': '',
                        '12 DE OCTUBRE': 'Doce.de.Octubre',
                        '2 DE ABRIL': 'Dos.de.Abril',
                        'ALMIRANTE BROWN': 'Almirante.Brown',
                        'BERMEJO': 'Bermejo',
                        'CHACABUCO': 'Chacabuco',
                        'COMPANIA DE MAYOR ING FRANKLIN': '',
                        'FRAY JUSTO SANTA MARIA DE ORO': 'Fray.Justo.Santa.Maria.de.Oro',
                        'GENERAL BELGRANO': 'General.Belgrano',
                        'GENERAL DONOVAN': 'General.Donovan',
                        'GENERAL GUEMES': 'General.Guemes',
                        'INDEPENDENCIA': 'Independencia',
                        'LIBERTAD': 'Libertad',
                        'LIBERTADOR GRL SAN MARTIN': '',
                        'MAIPU': 'Maipu',
                        'MAYOR LUIS J FONTANA': 'Mayor.Luis.J.Fontana',
                        'NUEVE DE JULIO': 'Nueve.de.Julio',
                        'O HIGGINS': 'O.Higgins',
                        'PRIMERA ARGENTINA': '',
                        'QUITILIPI': 'Quitilipi',
                        'SAN FERNANDO': 'San.Fernando',
                        'SAN LORENZO': 'San.Lorenzo',
                        'SAN MARTIN': 'San.Martin',
                        'SARMIENTO': 'Sarmiento',
                        'TAPENAGA': 'Tapenaga',
                        '25 DE MAYO': 'Veinticinco.de.Mayo',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'ANGACO': 'Angaco',
                        'CALINGASTA': 'Calingasta',
                        'CAPITAL': 'Capital',
                        'CAUCETE': 'Caucete',
                        'CHIMBAS': 'Chimbas',
                        'IGLESIA': 'Iglesia',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'POCITO': 'Pocito',
                        'RAWSON': 'Rawson',
                        'RIVADAVIA': 'Rivadavia',
                        'SAN MARTIN': 'San.Martin',
                        'SANTA LUCIA': 'Santa.Lucia',
                        'SARMIENTO': 'Sarmiento',
                        'ULLUM': 'Ullum',
                        'VALLE FERTIL': 'Valle.Fertil',
                        'ZONDA': 'Zonda',
                    }
                    
                    # Convertimos a mayúsculas para asegurar que las claves coincidan
                    df['Departamento'] = df['Departamento'].str.upper()
                    
                    # Reemplazamos los valores en la columna 'DEPARTAMENTO' basados en el diccionario
                    df['Departamento'] = df['Departamento'].replace(reemplazos)
                    
                    departamentos_query = """SELECT departamento FROM datcrea_tablas.localidades"""
                    departamento_df = pd.read_sql(departamentos_query,con=conn)
                    departamento_df = departamento_df.drop_duplicates()
                    departamentos = departamento_df['departamento'].tolist()
                    df['Departamento'] = ['' if x not in departamentos else x for x in df['Departamento']]
                    return df
                
                df = validar_departamento(df, conn)

                ## 6.3) Validación Provincias
                # PROVINCIAS
                def validar_provincia(df, conn):
                    reemplazos = {
                        'BUENOS AIRES': 'Buenos.Aires',
                        'CAPITAL FEDERAL': '',
                        'CATAMARCA': 'Catamarca',
                        'CHACO': 'Chaco',
                        'CHUBUT': '',
                        'CORDOBA': 'Córdoba',
                        'CORRIENTES': 'Corrientes',
                        'ENTRE RIOS': 'Entre.Ríos',
                        'FORMOSA': '',
                        'JUJUY': 'Jujuy',
                        'LA PAMPA': 'La.Pampa',
                        'LA RIOJA': '',
                        'MENDOZA': 'Mendoza',
                        'MISIONES': '',
                        'NEUQUEN': '',
                        'RIO NEGRO': '',
                        'SALTA': 'Salta',
                        'SAN JUAN': '',
                        'SAN LUIS': 'San.Luis',
                        'SANTA CRUZ': '',
                        'SANTA FE': 'Santa.Fe',
                        'SANTIAGO DEL ESTERO': 'Santiago.del.Estero',
                        'TIERRA DEL FUEGO': '',
                        'TUCUMAN': 'Tucumán',
                        'DESCONOCIDO': ''
                    }
                    
                    # Convertimos a mayúsculas para asegurar que las claves coincidan
                    df['Provincia'] = df['Provincia'].str.upper()
                    
                    # Reemplazamos los valores en la columna 'Provincia' basados en el diccionario
                    df['Provincia'] = df['Provincia'].replace(reemplazos)
                    #df['Provincia'] = df['Provincia'].astype(str)
                    df['Provincia'] = df['Provincia'].replace('nan', '')
                    
                    return df
                
                
                df = validar_provincia(df, conn)

                # TENENCIA #FUNCIONA
                def validar_tenencia(df,conn):
                    
                    tenencia = ['Propio', 'Aparceria', 'Arrendado', 'Cedido en alquiler']
                
                    df['Tenencia'] = df['Tenencia'].apply(lambda x: x if x in tenencia else '')
                
                    return df
                
                df = validar_tenencia(df,conn)

                
                # DESTINO #FUNCIONA
                def validar_destino(df,conn):
                    
                    destino = ['Grano', 
                            'Semilla (Convenio Comercial)', 
                            'Silo', 
                            'Heno',
                            'Pastoreo', 
                            'Pastoreo de rastrojos', 
                            'Cultivo de servicio', 
                            'Cultivo de servicio pastoreado',
                            'Conservación / Reserva']
                    
                    df['DestinoProduccion'] = df['DestinoProduccion'].apply(lambda x: x if x in destino else '')
                    return df
                
                df = validar_destino(df,conn)

                # CULTIVO ANTECESOR
                def validar_cultivo_antecesor(df,conn):
                    
                    # Lista de valores a reemplazar y sus correspondientes reemplazos
                    reemplazos = {
                        'acai - tardío': '',
                        'acai - de primera': '',
                        'acai - de segunda': '',
                        'achicoria - de primera': 'Achicoria',
                        'agave - tardío': '',
                        'agave - de segunda': '',
                        'agave - de primera': '',
                        'aguacate - de primera': '',
                        'ajo - de primera': 'Ajo',
                        'alfalfa - de segunda': 'Alfalfa',
                        'alfalfa - de primera': 'Alfalfa',
                        'alfalfa - tardío': 'Alfalfa',
                        'algodón - de segunda': 'Algodón',
                        'algodón - tardío': 'Algodón',
                        'algodón - de primera': 'Algodón',
                        'alpiste - de segunda': 'Alpiste',
                        'alpiste - de primera': 'Alpiste',
                        'anana - de primera': '',
                        'arroz - de primera': 'Arroz',
                        'arroz - tardío': 'Arroz',
                        'arroz - de segunda': 'Arroz',
                        'arveja - tardío': 'Arveja',
                        'arveja - de primera': 'Arveja',
                        'arveja - de segunda': 'Arveja',
                        'avena - de primera': 'Avena',
                        'avena - de segunda': 'Avena',
                        'avena - tardío': 'Avena',
                        'avena blanca - de segunda': 'Avena',
                        'avena blanca - de primera': 'Avena',
                        'avena strigosa - tardío': 'Avena',
                        'avena strigosa - de segunda': 'Avena',
                        'avena strigosa - de primera': 'Avena',
                        'banano - de primera': '',
                        'brócoli - de primera': '',
                        'café - safra': '',
                        'café - de segunda': '',
                        'café - de primera': '',
                        'camelina - de primera': 'Camelina',
                        'caña de azucar - de primera': 'Caña de Azucar',
                        'caña de azucar - tardío': 'Caña de Azucar',
                        'caña de azucar - de segunda': 'Caña de Azucar',
                        'canola - de primera': '',
                        'carinata - de primera': 'Carinata',
                        'carinata - de segunda': 'Carinata',
                        'cartamo - de primera': 'Cártamo',
                        'cebada - de primera': 'Cebada',
                        'cebada - tardío': 'Cebada',
                        'cebada - de segunda': 'Cebada',
                        'cebadilla - de primera': 'Cebadilla',
                        'centeno - de segunda': 'Centeno',
                        'centeno - de primera': 'Centeno',
                        'centeno - tardío': 'Centeno',
                        'chia - tardío': '',
                        'chia - de segunda': '',
                        'chia - de primera': '',
                        'cítricos - de segunda': '',
                        'cítricos - tardío': '',
                        'cítricos - de primera': '',
                        'colza - de primera': 'Colza',
                        'colza - de segunda': 'Colza',
                        'coriandro - de primera': 'Coriandro',
                        'crotalaria - de primera': '',
                        'cultivo de servicio - tardío': 'Cultivo de Servicio',
                        'cultivo de servicio - de primera': 'Cultivo de Servicio',
                        'cultivo de servicio - de segunda': 'Cultivo de Servicio',
                        'festuca - de primera': 'Pastura perenne',
                        'garbanzo - de segunda': 'Garbanzo',
                        'garbanzo - de primera': 'Garbanzo',
                        'girasol - tardío': 'Girasol',
                        'girasol - de primera': 'Girasol',
                        'girasol - de segunda': 'Girasol',
                        'kiwi - de primera': '',
                        'lechuga - de segunda': '',
                        'lenteja - de primera': 'Lenteja',
                        'lenteja - de segunda': 'Lenteja',
                        'lenteja - tardío': 'Lenteja',
                        'limon - de primera': '',
                        'lotus - de primera': 'Lotus',
                        'lupino - de primera': '',
                        'maiz' : 'Maíz',
                        'maíz - de primera': 'Maíz de 1° Temprano',
                        'maíz - tardío': 'Maíz de 1° Tardio',
                        'maíz - safrinha': '',
                        'maíz - de segunda': 'Maíz de 2°',
                        'mani - tardío': 'Maní',
                        'mani - de primera': 'Maní',
                        'mani - de segunda': 'Maní',
                        'melon - de primera': '',
                        'mijo - de segunda': 'Mijo',
                        'mijo - de primera': 'Mijo',
                        'mijo - tardío': 'Mijo',
                        'moha - de primera': 'Moha',
                        'moha - de segunda': 'Moha',
                        'moha - tardío': 'Moha',
                        'nabo - de segunda': 'Nabo',
                        'nabo - de primera': 'Nabo',
                        'olivo - tardío': '',
                        'papa - de segunda': 'Papa',
                        'papa - de primera': 'Papa',
                        'pasturas - tardío': 'Pastura perenne',
                        'pasturas - de primera': 'Pastura perenne',
                        'pasturas - de segunda': 'Pastura perenne',
                        'poroto - tardío': 'Poroto',
                        'poroto - de segunda': 'Poroto',
                        'poroto - de primera': 'Poroto',
                        'poroto mung - tardío': 'Poroto Mung',
                        'poroto mung - de segunda': 'Poroto Mung',
                        'poroto mung - de primera': 'Poroto Mung',
                        'raigrass - de segunda': 'Ryegrass',
                        'raigrass - de primera': 'Ryegrass',
                        'raigrass - tardío': 'Ryegrass',
                        'sandia - de primera': '',
                        'sesamo - de primera': '',
                        'soja' : 'Soja',
                        'soja - de segunda': 'Soja de 2°',
                        'soja - tardío': 'Soja de 1°',
                        'soja - primavera / verano': '',
                        'soja - safra': '',
                        'soja - semestre a': '',
                        'soja - de primera': 'Soja de 1°',
                        'sorgo - de primera': 'Sorgo',
                        'sorgo - tardío': 'Sorgo',
                        'sorgo - de segunda': 'Sorgo',
                        'tabaco - de primera': '',
                        'té - de primera': '',
                        'tomate - de segunda': '',
                        'tomate - de primera': '',
                        'trebol blanco - tardío': 'Trebol',
                        'trebol blanco - de primera': 'Trebol',
                        'trebol rojo - de primera': 'Trebol',
                        'trigo - de segunda': 'Trigo',
                        'trigo - de primera': 'Trigo',
                        'trigo - tardío': 'Trigo',
                        'triticale - de primera': 'Triticale',
                        'vicia - de segunda': 'Vicia',
                        'vicia - de primera': 'Vicia',
                        'vicia - tardío': 'Vicia',
                        'vid - de primera': 'Vid',
                        'vid - tardío': 'Vid',
                        'vid - de segunda': 'Vid',
                        'yerba mate - de primera': ''
                    }
                
                    # Convertir a minúsculas y reemplazar valores según el diccionario
                    df['CultivoAntecesor'] = df['CultivoAntecesor'].str.lower()
                    df['CultivoAntecesor'] = df['CultivoAntecesor'].replace(reemplazos)
                    
                    
                    return df
                    
                df = validar_cultivo_antecesor(df, conn)
                

                ## 6.7) Validación Calidad ambiente #FUNCIONA
                def validar_calidad_ambiente(df,conn):
                        
                    calidad_ambiente = ['Potencial Alto', 'Potencial Promedio', 'Potencial Bajo']
                    
                    df['CalidadAmbiente'] = df['CalidadAmbiente'].apply(lambda x: x if x in calidad_ambiente else '')
                
                    return df
                    
                df = validar_calidad_ambiente(df, conn)
                

                def validar_subgrupo(df, conn): #REVISAR #FUNCIONA
                    
                    subgrupo_query = """SELECT subgrupo FROM datcrea_tablas.subgrupos_cultivos"""
                    subgrupo_df = pd.read_sql(subgrupo_query,con=conn)
                    subgrupo = subgrupo_df['subgrupo'].tolist()
                    df['SubgrupoCultivo'] = ['' if x not in subgrupo else x for x in df['SubgrupoCultivo']]
                
                
                    #Carga los subgrupos de cultivosvy los organiza en un diccionario
                    subgrupos_cultivos_query = """SELECT cultivo, subgrupo FROM datcrea_tablas.subgrupos_cultivos WHERE cultivo IS NOT NULL"""
                    subgrupos_cultivos = pd.read_sql(subgrupos_cultivos_query, con=conn)
                
                    # Crear un diccionario para almacenar los resultados
                    subgrupos_cultivos_dict = {}
                    for index, row in subgrupos_cultivos.iterrows():
                        cultivo = row['cultivo']
                        subgrupo = row['subgrupo']
                        if cultivo in subgrupos_cultivos_dict:
                            subgrupos_cultivos_dict[cultivo].append(subgrupo)
                        else:
                            subgrupos_cultivos_dict[cultivo] = [subgrupo]
                
                    def validar_subgrupo(row):
                        
                        valor_cultivo = row['Cultivo']
                        valor_subgrupo = row['SubgrupoCultivo']
                
                        if valor_cultivo in subgrupos_dict:
                            if valor_subgrupo in subgrupos_dict[valor_cultivo]:
                                return valor_subgrupo
                        return ''
                
                    return df
                
                df = validar_subgrupo(df, conn)
                

                # FERTILIZACIÓN
                def validar_fertilizacion(df,conn):
                    
                    fertilizacion_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'fertilizacion'"""
                    fertilizacion = pd.read_sql(fertilizacion_query, con=conn)
                    fertilizacion = fertilizacion['valor'].tolist()
                    df['Fertilizacion'] = ['' if x not in fertilizacion else x for x in df['Fertilizacion']]
                    return df
                
                df = validar_fertilizacion(df, conn)
                


                ## 6.12) Reemplazo de 0 por Na
                # REEMPLAZAR CEROS
                def replace_ceros(df):
                    df['SuperficieSembrada'] = df['SuperficieSembrada'].replace(0, pd.NA)
                    df['espaciamiento'] = df['espaciamiento'].replace(0, pd.NA)
                    df['Rendimiento'] = df['Rendimiento'].replace(0, pd.NA)
                    df['Superficie'] = df['Superficie'].replace(0, pd.NA)
                    return df
                
                df = replace_ceros(df)
                
                # 6.13) Validación de Forma
                
                # VALIDAR FORMA DE FERTILIZACIÓN
                def validar_forma_fertilizacion(df,conn):
                    forma_fertilizacion_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'forma_fertilizacion'"""
                    forma_fertilizacion = pd.read_sql(forma_fertilizacion_query, con=conn)
                    forma_fertilizacion = forma_fertilizacion['valor'].tolist()
                    df['1_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['1_Forma']]
                    df['2_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['2_Forma']]
                    df['3_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['3_Forma']]
                    return df
                
                df = validar_forma_fertilizacion(df, conn)

                
                # Validación de Sistema de riego
                # SISTEMA RIEGO
                def validar_sistema_riego(df,conn):
                    sistema_riego_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'sistema_riego'"""
                    sistema_riego = pd.read_sql(sistema_riego_query, con=conn)
                    sistema_riego = sistema_riego['valor'].tolist()
                    df['Riego'] = ['' if x not in sistema_riego else x for x in df['Riego']]
                    return df
                
                df = validar_sistema_riego(df,conn)


                # INFLUENCIA NAPA
                def influencia_napa(df,conn):
                    
                    influencia_napa_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'influencia_napa'"""
                    influencia_napa = pd.read_sql(influencia_napa_query,con=conn)
                    influencia_napa = influencia_napa['valor'].tolist()
                    df['Napa'] = ['' if x not in influencia_napa else x for x in df['Napa']]
                    return df
                
                df = influencia_napa(df, conn)
                

                
                
                # Conexión a la base
                import pandas as pd
                from sqlalchemy import create_engine
                
                host="dw.crea.org.ar"
                port="54322"
                user="postgres"
                password="Tr0pema44cr34#"
                database="warehouse"
                
                conn2 = create_engine("postgresql+psycopg2://{user}:{password}@{host}:{port}/{database}".format(host=host,user=user,password=password,database=database,port=port))
                
                # Escribir en celdas excel
                region_grupo_query = """SELECT r.region_sigla, r.region_nombre,c.crea_numero,c.crea_nombre  
                FROM crm_crea.regiones AS r
                JOIN crm_crea.crea AS c 
                ON r.region_id = c.crea_region_id
                WHERE c.crea_baja IS NULL
                GROUP BY r.region_sigla, r.region_nombre, c.crea_numero,c.crea_nombre
                ORDER BY r.region_sigla, c.crea_nombre;"""
                region_grupo = pd.read_sql(region_grupo_query, con=conn2)
                
                #cuic = 'SSF145002M'
                
                region = region_grupo.loc[region_grupo['region_sigla'].str.startswith(cuic[:3]), 'region_nombre'].iloc[0]
                #region = replace region por el formato dat que sale de una lista.
                grupo = region_grupo.loc[region_grupo['crea_numero'].str.contains(cuic[3:6]), 'crea_nombre'].iloc[0]
                
                regiones_dat_dict = {
                    
                    'CENTRO':'CENTRO',
                    'CHACO SANTIAGUEÑO':'CHACO.SANTIAGUEÑO',
                    'CORDOBA NORTE':'CORDOBA.NORTE',
                    'ESTE':'ESTE',
                    'LITORAL NORTE':'LITORAL.NORTE',
                    'LITORAL SUR':'LITORAL.SUR',
                    'MAR Y SIERRAS':'MAR.Y.SIERRAS',
                    'NORTE DE BUENOS AIRES':'NORTE.BUENOS.AIRES',
                    'NOA':'NOA',
                    'NORTE DE SANTA FE':'NORTE.SANTA.FE',
                    'OESTE ARENOSO':'OESTE.ARENOSO',
                    'OESTE':'OESTE',
                    'PATAGONIA':'PATAGONIA',
                    'SEMIARIDA':'SEMIARIDA',
                    'SUDESTE':'SUDESTE',
                    'SANTA FE CENTRO':'SANTA.FE.CENTRO',
                    'SUR DE SANTA FE':'SUR.SANTA.FE',
                    'SUDOESTE':'SUDOESTE',
                    'VALLES CORDILLERANOS':'VALLES.CORDILLERANOS'
                    }
                # Reemplazo directo sin función
                for clave, valor in regiones_dat_dict.items():
                    region = region.replace(clave, valor)
                
                #print('REGIÓN: ', region)
                
                grupos_dat_dict = {'ALEJANDRO CHAJAN':'ALEJANDRO.CHAJAN',
                'BUENA ESPERANZA':'BUENA.ESPERANZA',
                'CANALS':'CANALS',
                'CARNERILLO':'CARNERILLO',
                'CAÑADA SECA':'CAÑADA.SECA',
                'CTALAMOCHITA':'CTALAMOCHITA',
                'HUINCA RENANCO':'HUINCA.RENANCO',
                'LA CESIRA TAMBERO':'LA.CESIRA.TAMBERO',
                'LA PORTADA':'LA.PORTADA',
                'LABOULAYE - BOUCHARDO':'LABOULAYE.BUCHARDO',
                'MELO SERRANO':'MELO.SERRANO',
                'RANQUELES':'RANQUELES',
                'RIO CUARTO':'RIO.CUARTO',
                'RIO QUINTO':'RÍO.QUINTO',
                'TAMBERO LABOULAYE':'TAMBERO.LABOULAYE',
                'TAMBERO VILLA MARIA':'VILLA.MARIA',
                'TEGUA':'TEGUA',
                'VALLE DEL CONLARA':'VALLE.DEL.CONLARA',
                'WASHINGTON MACKENNA':'WASHINGTON.MACKENNA',
                'CAMPO GALLO MONTE QUEMADO':'CAMPO.GALLO.MONTE.QUEMADO',
                'GUAYACAN':'GUAYACAN',
                'IBARRETA':'IBARRETA',
                'LOMITAS':'LOMITAS',
                'PALMARES':'PALMARES',
                'PAMPA DEL INFIERNO':'PAMPA.DEL.INFIERNO',
                'QUIMILI':'QUIMILI',
                'RENOVALES':'RENOVALES',
                'SACHAYOJ':'SACHAYOJ',
                'SANAVIRONES':'SANAVIRONES',
                'SEMIARIDO NORTE':'SEMIARIDO.NORTE',
                'SUDESTE SANTIAGUEÑO':'SES',
                'TINTINA':'TINTINA',
                'ARROYITO':'ARROYITO',
                'BARRANCA YACO':'BARRANCA.YACO',
                'CAROYA':'CAROYA',
                'CAÑADA DE LUQUE SITON':'CAÑADA.DE.LUQUE.SITON',
                'DEL ESTE':'DEL.ESTE',
                'GANADERO DEL NOROESTE':'GANADERO.DEL.NOROESTE',
                'JESUS MARIA':'JESUS.MARIA',
                'LAGUNA LARGA':'LAGUNA.LARGA',
                'LEOPOLDO LUGONES':'LEOPOLDO.LUGONES',
                'MONTE CRISTO':'MONTE.CRISTO',
                'RIO PRIMERO':'RIO.PRIMERO',
                'SIERRAS CHICAS':'SIERRAS.CHICAS',
                'TOTORAL':'TOTORAL',
                'ABASTO':'Otro',
                'CAÑUELAS':'CAÑUELAS',
                'GELAS':'GELAS',
                'LUJAN':'LUJAN',
                'NAVARRO II':'NAVARROII',
                'PIONEROS ABASTO':'Otro',
                'AVATI - I - ARROCERO':'AVATI.I.ARROCERO',
                'CURUZU CUATIA':'CURUZU.CUATIA',
                'ESQUINA':'ESQUINA',
                'MALEZALES':'Otro',
                'MERCEDES':'MERCEDES',
                'TIERRA COLORADA':'TIERRA.COLORADA',
                'URUNDAY':'URUNDAY',
                'ÑANDUBAY':'ÑANDUBAY',
                'BOVRIL EL SOLAR':'BOVRIL.EL.SOLAR',
                'CONCEPCION URUGUAY':'CONCEPCION.URUGUAY',
                'CONCORDIA CHAJARI':'CONCORDIA.CHAJARI',
                'GALARZA':'GALARZA',
                'GUALEGUAYCHU':'GUALEGUAYCHU',
                'ISLAS DEL IBICUY':'ISLAS.DEL.IBICUY',
                'LA PAZ':'LA.PAZ',
                'LARROQUE GUALEGUAY':'LARROQUE.GUALEGUAY',
                'MANDISOVI CONCORDIA':'MANDISOVI.CONCORDIA',
                'MONTOYA':'MONTOYA',
                'SAN JAIME':'SAN.JAIME',
                'VICTORIA':'VICTORIA',
                'VILLAGUAY':'VILLAGUAY',
                'ARROYO DE LOS HUESOS':'ARROYO.DE.LOS.HUESOS',
                'AZUL CHILLAR':'AZUL.CHILLAR',
                'CTE.N.OTAMENDI':'CTE.N.OTAMENDI',
                'DEFFERRARI':'DEFFERRARI',
                'FRONTERA':'FRONTERA',
                'FULTON':'FULTON',
                'LOBERIAS GRANDES':'LOBERIAS.GRANDES',
                'NECOCHEA QUEQUEN':'NECOCHEA.QUEQUEN',
                'QUEQUEN SALADO':'QUEQUEN.SALADO',
                'SAN CAYETANO - TRES ARROYOS':'SAN.CAYETANO.TRES.ARROYOS',
                'SAN FRANCISCO DE BELLOCQ':'SAN.FRANCISCO.DE.BELLOCQ',
                'SAN MANUEL':'SAN.MANUEL',
                'TAMBERO MAR Y SIERRA':'TAMBERO.MAR.Y.SIERRA',
                'TANDIL':'TANDIL',
                'TRES ARROYOS':'TRES.ARROYOS',
                'ZONA 4 LECHERA':'ZONA.4.LECHERA',
                'ALBERDI':'ALBERDI',
                'ALBERTI- PLA':'ALBERTI.PLA',
                'ARROYO DEL MEDIO':'ARROYO.DEL.MEDIO',
                'BRAGADO':'BRAGADO',
                'GIDAG':'GIDAG',
                'PERGAMINO':'PERGAMINO',
                'RAWSON TRES SARGENTOS':'RAWSON.TRES.SARGENTOS',
                'SAN ANTONIO DE ARECO':'SAN.ANTONIO.DE.ARECO',
                'SAN PEDRO VILLA LIA':'SAN.PEDRO.VILLA.LIA',
                'SEGUI LA ORIENTAL':'SEGUI.LA.ORIENTAL',
                'CAÑAVERALES DE TUCUMAN':'CAÑAVERALES.DE.TUCUMAN',
                'EL PALOMAR':'EL.PALOMAR',
                'EL RODEO':'EL.RODEO',
                'LA COCHA':'LA.COCHA',
                'LOS ALGARROBOS':'LOS.ALGARROBOS',
                'METAN':'METAN',
                'PALO SANTO':'Otro',
                'SAN PATRICIO':'SAN.PATRICIO',
                'SANTA ROSA CATAMARCA':'SANTA.ROSA.CATAMARCA',
                'SURCOS':'Otro',
                'SUYAY':'SUYAY',
                'VALLES TEMPLADOS':'Otro',
                'YUNGAS':'YUNGAS',
                'CUÑA BOSCOSA':'CUÑA.BOSCOSA',
                'MARGARITA CAMPO ALEMAN':'MARGARITA.CAMPO.ALEMAN',
                'RAMAYON':'RAMAYON',
                'SAN CRISTOBAL-LA LUCILA':'SAN.CRISTOBAL.LALUCILA',
                'VILLA ANA-ARANDU':'VILLA.ANA.ARANDÚ',
                'VILLA OCAMPO':'VILLA.OCAMPO',
                'AMERICA':'AMERICA',
                'AMERICA II':'AMERICA.II',
                'AMERICA LECHERO':'AMERICA.LECHERO',
                'ATREUCO':'ATREUCO',
                'CORRALERO':'CORRALERO',
                'PELLEGRINI - TRES LOMAS':'PELLEGRINI.TRES.LOMAS',
                'PICO BARON':'PICO.BARON',
                'PICO QUEMU':'PICO.QUEMU',
                'QUEMU CATRILO':'QUEMU.CATRILO',
                'TRENQUE LAUQUEN II':'TRENQUE.LAUQUEN.II',
                'AGROGANADERO 9 DE JULIO':'AGROGANADERO.9.DE.JULIO',
                'AMEGHINO':'AMEGHINO',
                'BOLIVAR':'BOLIVAR',
                'CASARES - 9 DE JULIO':'CASARES.9.DE.JULIO',
                'GENERAL PINTO':'GENERAL.PINTO',
                'GENERAL VILLEGAS':'GENERAL.VILLEGAS',
                'GUANACO LAS TOSCAS':'GUANACO.LAS.TOSCAS',
                'HENDERSON-DAIREAUX':'HENDERSON.DAIREAUX',
                'HERRERA VEGAS - PEHUAJO':'HERRERA.VEGAS.PEHUAJO',
                'INFOSURA':'INFOSURA',
                'LA VIA':'LA.VIA',
                'LINCOLN':'LINCOLN',
                'MONES CAZON PEHUAJO':'MONES.CAZON.PEHUAJO',
                'NUEVE DE JULIO':'NUEVE.DE.JULIO',
                'PEHUAJO CASARES':'PEHUAJO.CASARES',
                'PIROVANO LA LARGA':'PIROVANO.LA.LARGA',
                'SALAZAR MONES CAZON':'SALAZAR.MONES.CAZON',
                'TAMBERO AMEGHINO VILLEGAS':'TAMBERO.AMEGHINO.VILLEGAS',
                'TEJEDOR':'TEJEDOR',
                'TREINTA AGOSTO- MARI LAUQUEN':'TREINTA.AGOSTO.MARI.LAUQUEN',
                'ALTO VALLE - VALLE MEDIO':'ALTO.VALLE.VALLE.MEDIO',
                'CUENCA DEL AGRIO':'CUENCA.DEL.AGRIO',
                'ESQUEL':'ESQUEL',
                'LANIN':'LANIN',
                'SANTA CRUZ':'SANTACRUZ',
                'TIERRA DEL FUEGO':'TIERRADELFUEGO',
                'VIEDMA':'VIEDMA',
                'AOKEN AL':'AOKEN.AL',
                'CALEUCHE':'CALEUCHE',
                'GUATRACHE':'GUATRACHE',
                'HOLISTICO':'HOLISTICO',
                'PEHUENCHE':'PEHUENCHE',
                'SOVEN':'SOVEN',
                'TAMBERO GUATRACHE':'TAMBERO.GUATRACHE',
                'UTRACAN':'UTRACAN',
                '25 DE MAYO':'CREA.25.DE.MAYO',
                'ARROYO DE LAS FLORES':'ARROYO.DE.LAS.FLORES',
                'ARROYO LANGUEYU':'ARROYO.LANGUEYU',
                'AYACUCHO':'AYACUCHO',
                'CASTELLI - BELGRANO':'CASTELLI.BELGRANO',
                'DEL TUYU':'DEL.TUYU',
                'FORTIN MULITAS':'FORTIN.MULITAS',
                'LEZAMA':'LEZAMA',
                'MAIPU':'MAIPU',
                'MAR CHIQUITA':'MAR.CHIQUITA',
                'MONTE':'MONTE',
                'PILA':'PILA',
                'RAUCH - UDAQUIOLA':'RAUCH.UDAQUIOLA',
                'RIO SALADO':'RIO.SALADO',
                'ROQUE PEREZ SALADILLO':'ROQUE.PEREZ.SALADILLO',
                'TAPALQUE II':'TAPALQUE.II',
                'VALLIMANCA':'VALLIMANCA',
                'CASTELAR':'CASTELAR',
                'CENTRO OESTE SANTAFESINO':'C.O.S',
                'CUENCA':'CUENCA',
                'EL CEIBO':'EL.CEIBO',
                'ESPERANZA':'ESPERANZA',
                'GALVEZ':'GALVEZ',
                'RAFAELA':'RAFAELA',
                'SAN FRANCISCO':'SAN.FRANCISCO',
                'SAN GUILLERMO':'Otro',
                'SAN MARTIN DE LAS ESCOBAS - COLONIA BELGRANO':'S.M.E.C.B',
                'SUNCHALES':'Otro',
                'APLICADORES':'Otro',
                'ARMSTRONG MONTES DE OCA':'ARMSTRONG.MONTES.DE.OCA',
                'ASCENSION':'ASCENSION',
                'COLONIA MEDICI':'COLONIA.MEDICI',
                'COSTAS DEL CARCARAÑA':'COSTAS.DEL.CARCARAÑA',
                'EL ABROJO':'EL.ABROJO',
                'GENERAL ARENALES':'GENERAL.ARENALES',
                'GENERAL BALDISSERA':'GENERAL.BALDISSERA',
                'LA CALANDRIA':'LA.CALANDRIA',
                'LA MAROMA':'LA.MAROMA',
                'LAS PETACAS':'LAS.PETACAS',
                'MARIA TERESA':'MARIA.TERESA',
                'MONTE BUEY INRIVILLE':'MONTE.BUEY.INRIVILLE',
                'MONTE MAIZ':'MONTE.MAIZ',
                'POSTA ESPINILLOS':'POSTA.ESPINILLOS',
                'ROSARIO':'ROSARIO',
                'SAN JORGE LAS ROSAS':'SAN.JORGE.LAS.ROSAS',
                'SANTA ISABEL':'SANTA.ISABEL',
                'SANTA MARIA':'SANTA.MARIA',
                'TEODELINA':'TEODELINA',
                'BENITO JUAREZ':'BENITO.JUAREZ',
                'CARHUE HUANGUELEN':'CARHUE.HUANGUELEN',
                'CORONEL PRINGLES II':'CORONEL.PRINGLES.II',
                'CORONEL SUAREZ':'CORONEL.SUAREZ',
                'GENERAL LAMADRID':'GENERAL.LAMADRID',
                'LAPRIDA':'LAPRIDA',
                'NUESTRA SEÑORA DE LAS PAMPAS':'NUESTRA.SEÑORA.DE.LAS.PAMPAS',
                'OLAVARRIA':'OLAVARRIA',
                'PEDRO LURO':'PEDRO.LURO',
                'SAN ELOY - PIÑEYRO':'SAN.ELOY.PIÑEYRO',
                'VENTANIA':'VENTANIA',
                'ACONCAGUA':'Otro',
                'ARAUCO':'Otro',
                'CALCHAQUI':'Otro',
                'FRUTICOLA CUYO':'Otro',
                'HUARPE':'Otro',
                'LAS ACEQUIAS':'Otro',
                'LOS ANDES':'Otro',
                'NOGALERO DEL NORTE':'Otro',
                'OLIVICOLA SAN JUAN':'Otro',
                'VIGNERONS':'Otro'
                }
                
                # Reemplazo directo sin función
                for clave, valor in grupos_dat_dict.items():
                    grupo = grupo.replace(clave, valor)
                
                #print('GRUPO: ', grupo)

                # Filtro el DataFrame
                # CULTIVO VERANO
                def filtrar_cultivos_invierno(df, conn):
                    query = """SELECT cultivo FROM datcrea_tablas.cultivos WHERE ciclo = 'i'"""
                    invierno_df = pd.read_sql(query, conn)
                    cultivos_invierno = invierno_df['cultivo'].tolist()
                    df = df[df['Cultivo'].isin(cultivos_invierno)]
                    return df
                
                invierno = filtrar_cultivos_invierno(df, conn)
                
                def exportar_sima_invierno(invierno):
                    seccion = 200 # Máx registros por planilla xls
                    registros_campana = len(invierno.index) # cuenta el total de registros
                    registros_wb = math.ceil(registros_campana / seccion) #cociente para saber cantidad de worksheets
                    
                    # Armo listado de subset de 200 (máx) registros cada uno
                    dataframes = []
                    
                    for i in range(registros_wb):
                        inicio = i * seccion
                        fin = (i + 1) * seccion
                        seccion_df = invierno.iloc[inicio:fin]
                        dataframes.append(seccion_df)
                    
                    # Crear una workbook para cada dataframe
                    
                    # Ruta del archivo original
                    ruta_original = 'C:/Users/EPolacco/Documents/9 - DAT/SIMA/DAT-Cultivos-de-invierno-2023-24.xlsx'
                    
                    # Cargar el archivo original
                    wb_original = load_workbook(filename=ruta_original)
                    
                    # Obtener la hoja original
                    planilla_original = wb_original['Planilla ']
                    
                    columns_to_exclude = {'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'U', 'W', 'Y', 'Z',
                                        'AF', 'AG', 'AK', 'AN', 'CB', 'CK', 'CN', 'CM', 'CO', 'CP', 'CQ', 
                                        'CR', 'DU', 'DV', 'DW', 'IA','IB','IC','ID','IE','IF','IG','IH',
                                        'II','IJ','IK','IL','IM','IN','IO' }
                    
                    
                    # Iterar sobre los df y guardar en archivos xls separados
                    for i, seccion_df in enumerate(dataframes, 1):
                        # Crear un workbook para cada sección
                        if i == 1:
                            wb = wb_original  # Utilizar el archivo original para el primer df
                            sheet = planilla_original
                        else:
                            # Crear una copia del archivo original para los df subsiguientes
                            ruta_copia = f'C:/Users/EPolacco/Documents/9 - DAT/SIMA/{cuic}_DAT-Cultivos-de-invierno-2023-24_SECCION_{i}.xlsx'
                            shutil.copy(ruta_original, ruta_copia)
                            
                            # Cargar la copia
                            wb = load_workbook(filename=ruta_copia)
                            sheet = wb['Planilla ']
                            sheet.title = f'Planilla '
                    
                        # Borrar datos existentes en la hoja, después de la fila 14 porque sino se pisan los datos
                        
                        for row in sheet.iter_rows(min_row=15, max_row=sheet.max_row, min_col=6, max_col=sheet.max_column):
                            for cell in row:
                                if cell.column_letter not in columns_to_exclude:
                                    cell.value = None
                    
                    
                        # Obtener las filas de datos del df
                        rows = dataframe_to_rows(seccion_df,index=False, header=False)
                        
                        # Copiar en worksheet a partir de la fila 15 y columna 7
                        for r_idx, row in enumerate(rows, 15):
                            for c_idx, value in enumerate(row, 6):
                                if not pd.isna(value):
                                    sheet.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Región
                        sheet['F3'] = region
                        # Grupo
                        sheet['F6'] = grupo
                        
                        # Guardar el workbook, si no es el primer df
                        if i != 1:
                            wb.save(ruta_copia)
                    
                    
                    # Guardar workbook con el primer df
                    ruta_alternativa = f'C:/Users/EPolacco/Documents/9 - DAT/SIMA/{cuic}_DAT-Cultivos-de-invierno-2023-24_SECCION_1.xlsx'
                    wb_original.save(ruta_alternativa)
                    
                
                exportar_sima_invierno(invierno)


        convertir_a_sima(archivo_excel, estacion)
        st.success('Proceso terminado')