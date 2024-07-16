import streamlit as st
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import math
import shutil
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import mysql.connector


st.markdown("# SIMA")
#st.sidebar.markdown("# SIMA")

estaciones = ['Invierno', 'Verano']
estacion = st.radio("Seleccioná la estación de cultivo", estaciones)
st.write("La opción elegida es: ", estacion)
campana_filtro = st.selectbox("Seleccioná campaña a procesar",['2023-2024'])
# Bloque 1: Arrastrar archivo y leerlo
archivo_excel = st.file_uploader("Arrastra el archivo acá", type=["xlsx", "xls"])

if archivo_excel is not None:
    progress_bar = st.progress(0)
    
    for porcentaje_completado in range(100):
        time.sleep(0.1)  # Ajusta el tiempo de espera según el tamaño del archivo
        progress_bar.progress(porcentaje_completado + 1)
    st.success("Archivo subido correctamente")

    is_clicked = st.button("PROCESAR")
    if is_clicked:
        
        @st.cache_data(experimental_allow_widgets=True)
        def convertir_a_sima(archivo_excel,estacion, campana_filtro):
            if estacion == "Verano":
                           
                # Importo librerías

                from openpyxl import load_workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                import pandas as pd
                import numpy as np
                from datetime import datetime
                import math
                import shutil

                # Conexión a la base
                from sqlalchemy import create_engine
                import mysql.connector
                
                host='66.97.33.201'
                password='Dat##2023Tabla$'
                port=33306
                user='datcrea_tablas'
                database='datcrea_tablas'
                
                conn = create_engine("mysql+mysqlconnector://{user}:{password}@{host}:{port}/{database}".format(host=host,user=user,password=password,database=database,port=port),connect_args={'auth_plugin': 'mysql_native_password'})
                
            
                #fullpath = directorio + archivo
                df = pd.read_excel(archivo_excel) #fullpath

                if not pd.isnull(df['CUIC'][0]) and df['CUIC'][0] != '':
                    cuic = df['CUIC'][0]
                else:
                    cuic = 'Otro'

                
                def filtrar_campana(df, campana_filtro):
                    
                    filtro = df['Ciclo'] == campana_filtro
                    df = df[filtro]
                    return df

                df = filtrar_campana(df, campana_filtro)

                
                def formato_fecha(df):
                    df['FechaSiembra'] = df['FechaSiembra'].str[:10] #Elimino horas en fechas
                    df['FechaSiembra'] = pd.to_datetime(df['FechaSiembra'])
                    df['FechaSiembra'] = df['FechaSiembra'].dt.strftime('%d/%m/%Y')    
                    return df
                    
                df = formato_fecha(df)


                # 3.2) Validar código genética #FUNCIONA
                # Genética
                # Consultar los nro_registro desde la base de datos
                def validar_codigo_genetica(df, conn):
                    query = "SELECT nro_registro FROM datcrea_tablas.materiales WHERE nro_registro IS NOT NULL"
                    df_materiales = pd.read_sql(query, conn)
                    materiales = df_materiales['nro_registro'].astype(int).tolist()
                    df['Codigo_Genetica'] = df['Codigo_Genetica'].fillna(0)
                    df['Codigo_Genetica'] = df['Codigo_Genetica'].astype('Int64')
                    df['Codigo_Genetica'] = ['' if x not in materiales else x for x in df['Codigo_Genetica']]
                    return df

                df = validar_codigo_genetica(df, conn)


                ## 3.3) Validación Cultivo #FUNCIONA
                # Cultivos
                def validar_cultivos(df, conn):
                    cultivos_query = """SELECT DISTINCT(cultivo) FROM datcrea_tablas.cultivos"""
                    cultivos_df = pd.read_sql(cultivos_query,con=conn)
                    cultivos = cultivos_df['cultivo'].tolist()
                    df['Cultivo'] = df['Cultivo'].replace(['Soja - De segunda', 'Mani'], ['Soja', 'Maní'])
                    df['Cultivo'] = ['' if x not in cultivos else x for x in df['Cultivo']]
                    return df
                
                df = validar_cultivos(df,conn)

                ## 3.4) Validación Genética
                
                def validar_genetica(df,conn): 
                    
                    genetica_query = """SELECT material_inase FROM datcrea_tablas.materiales 
                                        WHERE nro_registro IS NOT NULL"""
                    genetica_df = pd.read_sql(genetica_query,con=conn)
                    
                    genetica = genetica_df['material_inase'].tolist()
                    df['Genetica'] = ['' if x not in genetica else x for x in df['Genetica']]
                    return df
                df = validar_genetica(df,conn)

                ## 3.9) Validación Semillero #FUNCIONA
                def validar_semillero(df, conn): 
                    
                    semillero_query = """SELECT semillero FROM datcrea_tablas.semilleros"""
                    semillero = pd.read_sql(semillero_query,con=conn)
                    semillero = semillero['semillero'].tolist()
                    df['Semillero'] = ['' if x not in semillero else x for x in df['Semillero']]
                    return df
                
                df = validar_semillero(df, conn)

                def validar_genetica_semillero(df, conn): #FUNCIONA
                    
                    # VALIDA GENÉTICA Y SEMILLERO EN BASE A CULTIVO Y CÓDIGO GENÉTICO
                    semillero_query = """SELECT cultivo, nro_registro, semillero, material 
                    FROM datcrea_tablas.materiales where nro_registro IS NOT NULL;"""
                    semillero_genetica = pd.read_sql(semillero_query,con=conn)
                    semillero_genetica['nro_registro'] = semillero_genetica['nro_registro'].astype(int)
                    
                    for index, row in semillero_genetica.iterrows():
                        cultivo = row['cultivo']
                        codigo_genetica = row['nro_registro']
                        semillero = row['semillero']
                        genetica = row['material']
                        
                        # Actualizar 'df' donde coincidan cultivo y codigo_genetica
                        df.loc[(df['Cultivo'] == cultivo) & (df['Codigo_Genetica'] == codigo_genetica), ['Semillero', 'Genetica']] = [semillero, genetica]
                    
                    return df
                    
                df = validar_genetica_semillero(df, conn)

                def validar_biotecnologia(df, conn):
                    
                
                    version_bio_query = """SELECT semillero, nro_registro AS codigo_genetica,
                        CASE
                            WHEN material_inase LIKE '%HCL-MGRR2%' THEN 'HCLMGRR2'
                            WHEN material_inase LIKE '%HCL-MG%' THEN 'HCLMG'
                            WHEN material_inase LIKE '%HCLRR2%' THEN 'HCLRR'
                            WHEN material_inase LIKE '%VIPTERA3 CL%' THEN 'CLVIP3'
                            WHEN material_inase LIKE '%CLVT3P%' THEN 'CLVT3P'
                            WHEN material_inase LIKE '%GLSTACK%' THEN 'GL STACK'
                            WHEN material_inase LIKE '%MGRR2%' THEN 'MGRR2'
                            WHEN material_inase LIKE '%MG%' THEN 'MG'
                            WHEN material_inase LIKE '%PWUE%' THEN 'PW Ultra Enlist'
                            WHEN material_inase LIKE '%PWU%' THEN 'PW ULTRA'
                            WHEN material_inase LIKE '%PWE%' THEN 'PW Enlist'
                            WHEN material_inase LIKE '%PW%' THEN 'PW'
                            WHEN material_inase LIKE '%VIPTERA3%' THEN 'VIPTERA 3'
                            WHEN material_inase LIKE '%VIPTERA2%' THEN 'VIPTERA 2'
                            WHEN material_inase LIKE '%VIPTERA4%' THEN 'VIPTERA 4'
                            WHEN material_inase LIKE '%VIPTERA%' THEN 'VIPTERA'
                            WHEN material_inase LIKE '%VT4P%' THEN 'VT4P'
                            WHEN material_inase LIKE '%VT3P%' THEN 'VT3P'
                            WHEN material_inase LIKE '%TDMAX%' THEN 'TDmax'
                            WHEN material_inase LIKE '%TD/TG%' THEN 'TD/TG'
                            WHEN material_inase LIKE '%TRE%' THEN 'TRECEPTA'
                            WHEN material_inase LIKE '%TG PLUS%' THEN 'TGplus'
                            WHEN material_inase LIKE '%YHR%' THEN 'YHR'
                            WHEN material_inase LIKE '%VYHR%' THEN 'VYHR'
                            WHEN material_inase LIKE '%RR2%' THEN 'RR'
                            WHEN material_inase LIKE '%PRO4%' THEN 'VT4P'
                            WHEN material_inase LIKE '%' THEN 'VT3P / RIB' -- Para cadena vacía
                            WHEN material_inase LIKE '%RE%' THEN 'Refugio (sin proteína insecticida)'
                            ELSE NULL -- Opcional: manejar otro caso si ninguno de los patrones coincide
                        END AS version_biotecnologica
                    FROM materiales
                    WHERE nro_registro IS NOT NULL"""
                    version_bio = pd.read_sql(version_bio_query,con=conn)
                    version_bio['codigo_genetica'] = version_bio['codigo_genetica'].astype(int)
                    df['Version Biotecnológica'] = None
                    for index, row in version_bio.iterrows():
                        codigo_genetica = row['codigo_genetica']
                        version_biotecnologica = row['version_biotecnologica']
                        
                        df.loc[(df['Codigo_Genetica'] == codigo_genetica), 'Version Biotecnológica'] = version_biotecnologica
                    return df
                df = validar_biotecnologia(df, conn)

                
                # Convertir el Código de Fertilizante en integer # FUNCIONA
                def fertilizante_int(df, conn):
                    columnas = [
                        '1_Codigo_Registro_1', '1_Codigo_Registro_2',
                        '2_Codigo_Registro_1', '2_Codigo_Registro_2',
                        '3_Codigo_Registro_1', '3_Codigo_Registro_2',
                        '4_Codigo_Registro_1', '4_Codigo_Registro_2'
                    ]
                    
                    for col in columnas:
                        df[col] = df[col].fillna(0).astype('Int64')
                    
                    return df
                
                df = fertilizante_int(df, conn)

                # 3.7) Validación de Fertilizantes # FUNCIONA
                
                # Registro fertilizantes 
                def validar_id_senasa_fertilizantes(df, conn):
                    reg_fert_query = """SELECT id_senasa 
                                        FROM datcrea_tablas.registro_fertilizantes 
                                        WHERE id_senasa IS NOT NULL"""
                    reg_fert_df = pd.read_sql(reg_fert_query,con=conn)
                    id_senasa = reg_fert_df['id_senasa'].astype(int).tolist()
                    
                    df['1_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['1_Codigo_Registro_1']]
                    df['1_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['1_Codigo_Registro_2']]
                    df['2_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['2_Codigo_Registro_1']]
                    df['2_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['2_Codigo_Registro_2']]
                    df['3_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['3_Codigo_Registro_1']]
                    df['3_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['3_Codigo_Registro_2']]
                    df['4_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['4_Codigo_Registro_1']]
                    df['4_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['4_Codigo_Registro_2']]
                    return df
                df = validar_id_senasa_fertilizantes(df, conn)
                
                def validar_nombre_fertilizante(df, conn):
                    # Consulta para obtener los fertilizantes registrados con un id_senasa no nulo
                    reg_fert_query = "SELECT id_senasa, fertilizante FROM datcrea_tablas.registro_fertilizantes WHERE id_senasa IS NOT NULL"
                    registros_fertilizantes = pd.read_sql(reg_fert_query, con=conn)
                    
                    # Asegurarse de que id_senasa es de tipo entero
                    registros_fertilizantes['id_senasa'] = registros_fertilizantes['id_senasa'].astype(int)
                
                    # Preparar un diccionario para mapear id_senasa a fertilizante
                    id_senasa_to_fertilizante = dict(zip(registros_fertilizantes['id_senasa'], registros_fertilizantes['fertilizante']))
                
                    # Actualizar las columnas en df usando el diccionario de mapeo
                    for i in range(1, 5):  # Iterar sobre las 4 series de columnas
                        for j in range(1, 3):  # Iterar sobre las 2 columnas por serie
                            codigo_columna = f"{i}_Codigo_Registro_{j}"
                            producto_columna = f"{i}_Producto_{j}"
                            df[producto_columna] = df[codigo_columna].map(id_senasa_to_fertilizante).fillna(df[producto_columna])
                    
                    return df
                df = validar_nombre_fertilizante(df, conn)


                def validar_fertilizante(df,conn):
                    reg_fert_query = """SELECT * 
                                        FROM datcrea_tablas.registro_fertilizantes 
                                        WHERE id_senasa IS NOT NULL"""
                    registros_fertilizantes = pd.read_sql(reg_fert_query, con=conn)
                    registros_fertilizantes['id_senasa'] = registros_fertilizantes['id_senasa'].astype(int)
                    
                    # Crear un diccionario para almacenar los resultados
                    id_senasa_dict = {}
                    
                    for index, row in registros_fertilizantes.iterrows():
                        fertilizante = row['fertilizante']
                        id_senasa = row['id_senasa']
                        if fertilizante in id_senasa_dict:
                            id_senasa_dict[fertilizante].append(id_senasa)
                        else:
                            id_senasa_dict[fertilizante] = [id_senasa]
                    
                    # Validar fertilizante en función del código
                    def validar_fertilizante_codigo(codigo):
                        
                        if pd.isna(codigo):
                            return ''
                        for producto, codigos in id_senasa_dict.items():
                            if codigo in codigos:
                                return producto
                        return ''
                    
                    # Aplicar validación en cada columna
                    for i in range(1, 5):  
                        df[f'{i}_Producto_1'] = df[f'{i}_Codigo_Registro_1'].apply(validar_fertilizante)
                        df[f'{i}_Producto_2'] = df[f'{i}_Codigo_Registro_2'].apply(validar_fertilizante)
                    
                    return df

                # VALIDAR DOSIS DE PRODUCTOS (si es que primero se validó el producto) # FUNCIONA
                def validar_dosis_productos(df):
                
                    columnas_productos_dosis = [
                        ('1_Producto_1', '1_Dosis_1'),
                        ('1_Producto_2', '1_Dosis_2'),
                        ('2_Producto_1', '2_Dosis_1'),
                        ('2_Producto_2', '2_Dosis_2'),
                        ('3_Producto_1', '3_Dosis_1'),
                        ('3_Producto_2', '3_Dosis_2'),
                        ('4_Producto_1', '4_Dosis_1'),
                        ('4_Producto_2', '4_Dosis_2')
                    ]
                
                    for producto_col, dosis_col in columnas_productos_dosis:
                        df.loc[df[producto_col] == '', dosis_col] = pd.NA
                
                    return df
                df = validar_dosis_productos(df)


                def validar_momentos(df, conn):
                        
                    momento_fertilizacion_query = """SELECT valor FROM datcrea_tablas.variables_generales 
                                                        WHERE variable = 'momento_fertilizacion'
                                                        AND valor NOT IN ('Postsiembra')"""
                    momento_fertilizacion_df = pd.read_sql(momento_fertilizacion_query,con=conn)
                    momento_fertilizacion = momento_fertilizacion_df['valor'].tolist()
                
                    # Columnas de momento a ajustar
                    columnas_momento = ['1_Momento', '2_Momento', '3_Momento', '4_Momento']
                
                    # Iterar sobre las columnas para ajustar los momentos
                    for columna in columnas_momento:
                        df[columna] = df[columna].apply(lambda x: x if x in momento_fertilizacion else '')
                
                    return df
                
                df = validar_momentos(df,conn)

                def calcular_densidad(df, conn):
                        
                    df['DensidadPlantasHa'] = df['DensidadPlantasHa']/10000
                    df.rename(columns={'DensidadPlantasHa': 'DensidadPlantasM2'}, inplace=True)
                    return df
                
                df = calcular_densidad(df, conn)


                def reindex_columns(df):
                    cols = ['CUIC','Campo', 'Campo_Otro','Columna_I','Lote','Lote_Otro','vacio','vacio','vacio',
                            'Provincia','Departamento','vacio','Localidad','CalidadAmbiente','SuperficieSembrada',
                            'Tenencia','DestinoProduccion','vacio','Cultivo','vacio','SubgrupoCultivo','vacio',
                            'CultivoAntecesor','vacio','vacio','FechaSiembra','vacio','vacio','vacio','espaciamiento',
                            'DensidadSemillasM2','DensidadSemillasHA','DensidadPlantasM2','vacio',
                            'Semillero','vacio','Genetica','Version Biotecnológica','vacio','Fertilizacion',
                            '1_Producto_1', '1_Dosis_1', '1_Momento', '1_Forma',
                            '1_Producto_2','1_Dosis_2','2_Momento','1_Forma',
                            '2_Producto_1','2_Dosis_1','3_Momento','2_Forma',
                            '2_Producto_2','2_Dosis_2','4_Momento','2_Forma',
                            '3_Producto_1','3_Dosis_1','5_Momento','3_Forma',
                            '3_Producto_2','3_Dosis_2','6_Momento','3_Forma',
                            'vacio','Riego','Lamina_Riego','Napa','vacio','vacio','vacio','vacio','vacio','vacio',
                            'vacio','Superficie','vacio','Rendimiento']
                
                    return df.reindex(columns=cols)
                
                df = reindex_columns(df)

                
                # 5) Creación de columnas Campo_Otro y Lote_Otro (para caso SIMA) #FUNCIONA
                def otro_cuic_campo_lote(df, conn):
                    #df['CUIC'] = df['CUIC'].astype(str)
                    df['CUIC'] = df['CUIC'].fillna('Otro')
                    #df['Campo_Otro'] = df['Campo_Otro'].fillna('Otro')
                    #df['Lote_Otro'] = df['Lote_Otro'].fillna('Otro')
                    
                    # Llenar la columna 'Campo_Otro' con 'Otro' donde la columna 'Campo' tiene NaN
                    df.loc[df['Campo'].isna(), 'Campo_Otro'] = 'Otro'
                    # Dejar vacía la columna 'Campo_Otro' donde la columna 'Campo' tiene datos
                    df.loc[df['Campo'].notna(), 'Campo_Otro'] = ''
                    
                    # Llenar la columna 'Lote_Otro' con 'otro' donde la columna 'Lote' tiene NaN
                    df.loc[df['Lote'].isna(), 'Lote_Otro'] = 'Otro'
                    # Dejar vacía la columna 'Lote_Otro' donde la columna 'Lote' tiene datos
                    df.loc[df['Lote'].notna(), 'Lote_Otro'] = ''
                
                    return df
                
                df = otro_cuic_campo_lote(df, conn)


                # Localidades
                def validar_localidad(df):
                    
                    localidades_query = """SELECT localidad FROM datcrea_tablas.localidades"""
                    
                    localidades = pd.read_sql(localidades_query,con=conn)
                    localidades = localidades['localidad'].str.title().tolist()
                    ## 6.1) Validación Localidad
                    # Unificar criterio de escritura
                    df['Localidad'] = df['Localidad'].astype(str)
                    df['Localidad'] = df['Localidad'].str.title()
                    df['Localidad'] = ['' if x not in localidades else x for x in df['Localidad']]
                    return df
            

                df = validar_localidad(df)
                
                # Departamentos
                def validar_departamento(df,conn):
                
                    reemplazos = {
                        '25 DE MAYO': 'Veinticinco.de.Mayo',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'ADOLFO ALSINA': 'Adolfo.Alsina',
                        'ADOLFO GONZALES CHAVES': 'Adolfo.Gonzales.Chaves',
                        'ALBERTI': 'Alberti',
                        'ALMIRANTE BROWN': '',
                        'ARRECIFES': 'Arrecifes',
                        'AVELLANEDA': '',
                        'AYACUCHO': 'Ayacucho',
                        'AZUL': 'Azul',
                        'BAHIA BLANCA': 'Bahía.Blanca',
                        'BALCARCE': 'Balcarce',
                        'BARADERO': 'Baradero',
                        'BENITO JUAREZ': 'Benito.Juarez',
                        'BERAZATEGUI': 'Berazategui',
                        'BERISSO': 'Berisso',
                        'BOLIVAR': 'Bolivar',
                        'BRAGADO': 'Partido.Bragado',
                        'BRANDSEN': 'Brandsen',
                        'CAMPANA': 'Campana',
                        'CANUELAS': 'Cañuelas',
                        'CAPITAN SARMIENTO': 'Capitan.Sarmiento',
                        'CARLOS CASARES': 'Carlos.Casares',
                        'CARLOS TEJEDOR': 'Carlos.Tejedor',
                        'CARMEN DE ARECO': 'Carmen.de.Areco',
                        'CASTELLI': 'Castelli',
                        'CHACABUCO': 'Chacabuco',
                        'CHASCOMUS': 'Chascomus',
                        'CHIVILCOY': 'Chivilcoy',
                        'COLON': 'Colon',
                        'CORONEL DE MARINA L. ROSALES': 'Coronel.de.Marina.L.Rosales',
                        'CORONEL DORREGO': 'Coronel.Dorrego',
                        'CORONEL PRINGLES': 'Coronel.Pringles',
                        'CORONEL SUAREZ': 'Coronel.Suarez',
                        'DAIREAUX': 'Daireaux',
                        'DOLORES': 'Dolores',
                        'ENSENADA': 'Ensenada',
                        'ESCOBAR': 'Escobar',
                        'ESTEBAN ECHEVERRIA': 'Esteban.Echeverria',
                        'EXALTACION DE LA CRUZ': 'Exaltación.de.La.Cruz',
                        'EZEIZA': 'Ezeiza',
                        'FLORENCIO VARELA': 'Florencio.Varela',
                        'FLORENTINO AMEGHINO': 'Florentino.Ameghino',
                        'GENERAL ALVARADO': 'General.Alvarado',
                        'GENERAL ALVEAR': 'General.Alvear',
                        'GENERAL ARENALES': 'General.Arenales',
                        'GENERAL BELGRANO': 'General.Belgrano',
                        'GENERAL GUIDO': 'General.Guido',
                        'GENERAL JUAN MADARIAGA': 'General.Juan.Madariaga',
                        'GENERAL LA MADRID': 'General.La.Madrid',
                        'GENERAL LAS HERAS': 'General.Las.Heras',
                        'GENERAL LAVALLE': 'General.Lavalle',
                        'GENERAL PAZ': 'General.Paz',
                        'GENERAL PINTO': 'General.Pinto',
                        'GENERAL PUEYRREDON': 'General.Pueyrredon',
                        'GENERAL RODRIGUEZ': 'General.Rodriguez',
                        'GENERAL SAN MARTIN': 'General.San.Martin',
                        'GENERAL VIAMONTE': 'General.Viamonte',
                        'GENERAL VILLEGAS': 'General.Villegas',
                        'GUAMINI': 'Guamini',
                        'HIPOLITO YRIGOYEN': 'Hipolito.Yrigoyen',
                        'HURLINGHAM': 'Hurlingham',
                        'ITUZAINGO': 'Ituzaingo',
                        'JOSE C. PAZ': 'Jose.C.Paz',
                        'JUNIN': 'Junin',
                        'LA COSTA': 'La.Costa',
                        'LA MATANZA': 'La.Matanza',
                        'LA PLATA': 'La.Plata',
                        'LANUS': 'Lanús',
                        'LAPRIDA': 'Laprida',
                        'LAS FLORES': 'Las.Flores',
                        'LEANDRO N. ALEM': 'Leandro.N..Alem',
                        'LINCOLN': 'Lincoln',
                        'LOBERIA': 'Loberia',
                        'LOBOS': 'Lobos',
                        'LOMAS DE ZAMORA': 'Lomas.de.Zamora',
                        'LUJAN': 'Lujan',
                        'MAGDALENA': 'Magdalena',
                        'MAIPU': 'Maipu',
                        'MALVINAS ARGENTINAS': 'Malvinas.Argentinas',
                        'MAR CHIQUITA': 'Mar.Chiquita',
                        'MARCOS PAZ': 'Marcos.Paz',
                        'MERCEDES': 'Mercedes',
                        'MERLO': 'Merlo',
                        'MONTE': 'Monte',
                        'MONTE HERMOSO': 'Monte.Hermoso',
                        'MORENO': 'Moreno',
                        'MORON': 'Moron',
                        'NAVARRO': 'Navarro',
                        'NECOCHEA': 'Necochea',
                        'OLAVARRIA': 'Olavarria',
                        'PATAGONES': 'Patagones',
                        'PEHUAJO': 'Pehuajo',
                        'PELLEGRINI': 'Pellegrini',
                        'PERGAMINO': 'Partido.Pergamino',
                        'PILA': 'Pila',
                        'PILAR': 'Pilar',
                        'PINAMAR': 'Pinamar',
                        'PRESIDENTE PERON': 'Presidente.Peron',
                        'PUAN': 'Puan',
                        'PUNTA INDIO': 'Punta.Indio',
                        'QUILMES': 'Quilmes',
                        'RAMALLO': 'Ramallo',
                        'RAUCH': 'Rauch',
                        'RIVADAVIA': 'Rivadavia',
                        'ROJAS': 'Rojas',
                        'ROQUE PEREZ': 'Roque.Perez',
                        'SAAVEDRA': 'Saavedra',
                        'SALADILLO': 'Saladillo',
                        'SALLIQUELO': 'Salliqueló',
                        'SALTO': 'Partido.Salto',
                        'SAN ANDRES DE GILES': 'San.Andres.de.Giles',
                        'SAN ANTONIO DE ARECO': 'San.A.de.Areco',
                        'SAN CAYETANO': 'San.Cayetano',
                        'SAN FERNANDO': 'San.Fernando',
                        'SAN ISIDRO': 'San.Isidro',
                        'SAN MIGUEL': 'San.Miguel',
                        'SAN NICOLAS': 'San.Nicolas',
                        'SAN PEDRO': 'San.Pedro',
                        'SAN VICENTE': 'San.Vicente',
                        'SUIPACHA': 'Suipacha',
                        'TANDIL': 'Tandil',
                        'TAPALQUE': 'Tapalque',
                        'TIGRE': 'Tigre',
                        'TORDILLO': 'Tordillo',
                        'TORNQUIST': 'Tornquist',
                        'TRENQUE LAUQUEN': 'Trenque.Lauquen',
                        'TRES ARROYOS': 'Tres.Arroyos',
                        'TRES DE FEBRERO': 'Tres.de.Febrero',
                        'TRES LOMAS': 'Tres.Lomas',
                        'VICENTE LOPEZ': 'Vicente.Lopez',
                        'VILLA GESELL': 'Villa.Gesell',
                        'VILLARINO': 'Villarino',
                        'ZARATE': 'Zarate',
                        'CAPITAL FEDERAL': '',
                        'AMBATO': 'Ambato',
                        'ANCASTI': 'Ancasti',
                        'ANDALGALA': '',
                        'ANTOFAGASTA DE LA SIERRA': 'Antofagasta.de.La.Sierra',
                        'BELEN': 'Belen',
                        'CAPAYAN': 'Capayan',
                        'CAPITAL': 'Capital',
                        'EL ALTO': 'El.Alto',
                        'FRAY MAMERTO ESQUIU': 'Fray.Mamerto.Esquiu',
                        'LA PAZ': 'La.Paz',
                        'PACLIN': 'Paclin',
                        'POMAN': 'Poman',
                        'SANTA MARIA': 'Santa.Maria',
                        'SANTA ROSA': 'Santa.Rosa',
                        'TINOGASTA': 'Tinogasta',
                        'VALLE VIEJO': 'Valle.Viejo',
                        '1 DE MAYO': '',
                        '12 DE OCTUBRE': 'Doce.de.Octubre',
                        '2 DE ABRIL': 'Dos.de.Abril',
                        'ALMIRANTE BROWN': 'Almirante.Brown',
                        'BERMEJO': 'Bermejo',
                        'CHACABUCO': 'Chacabuco',
                        'COMPANIA DE MAYOR ING FRANKLIN': '',
                        'FRAY JUSTO SANTA MARIA DE ORO': 'Fray.Justo.Santa.Maria.de.Oro',
                        'GENERAL BELGRANO': 'General.Belgrano',
                        'GENERAL DONOVAN': 'General.Donovan',
                        'GENERAL GUEMES': 'General.Guemes',
                        'INDEPENDENCIA': 'Independencia',
                        'LIBERTAD': 'Libertad',
                        'LIBERTADOR GRL SAN MARTIN': '',
                        'MAIPU': 'Maipu',
                        'MAYOR LUIS J FONTANA': 'Mayor.Luis.J.Fontana',
                        'NUEVE DE JULIO': 'Nueve.de.Julio',
                        'O HIGGINS': 'O.Higgins',
                        'PRIMERA ARGENTINA': '',
                        'QUITILIPI': 'Quitilipi',
                        'SAN FERNANDO': 'San.Fernando',
                        'SAN LORENZO': 'San.Lorenzo',
                        'SAN MARTIN': 'San.Martin',
                        'SARMIENTO': 'Sarmiento',
                        'TAPENAGA': 'Tapenaga',
                        '25 DE MAYO': 'Veinticinco.de.Mayo',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'ANGACO': 'Angaco',
                        'CALINGASTA': 'Calingasta',
                        'CAPITAL': 'Capital',
                        'CAUCETE': 'Caucete',
                        'CHIMBAS': 'Chimbas',
                        'IGLESIA': 'Iglesia',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'POCITO': 'Pocito',
                        'RAWSON': 'Rawson',
                        'RIVADAVIA': 'Rivadavia',
                        'SAN MARTIN': 'San.Martin',
                        'SANTA LUCIA': 'Santa.Lucia',
                        'SARMIENTO': 'Sarmiento',
                        'ULLUM': 'Ullum',
                        'VALLE FERTIL': 'Valle.Fertil',
                        'ZONDA': 'Zonda',
                    }
                    
                    # Convertimos a mayúsculas para asegurar que las claves coincidan
                    df['Departamento'] = df['Departamento'].str.upper()
                    
                    # Reemplazamos los valores en la columna 'DEPARTAMENTO' basados en el diccionario
                    df['Departamento'] = df['Departamento'].replace(reemplazos)
                    
                    departamentos_query = """SELECT departamento FROM datcrea_tablas.localidades"""
                    departamento_df = pd.read_sql(departamentos_query,con=conn)
                    departamento_df = departamento_df.drop_duplicates()
                    departamentos = departamento_df['departamento'].tolist()
                    df['Departamento'] = ['' if x not in departamentos else x for x in df['Departamento']]
                    return df
                
                df = validar_departamento(df, conn)

                ## 6.3) Validación Provincias
                # PROVINCIAS
                def validar_provincia(df, conn):
                    reemplazos = {
                        'BUENOS AIRES': 'Buenos.Aires',
                        'CAPITAL FEDERAL': '',
                        'CATAMARCA': 'Catamarca',
                        'CHACO': 'Chaco',
                        'CHUBUT': '',
                        'CORDOBA': 'Córdoba',
                        'CORRIENTES': 'Corrientes',
                        'ENTRE RIOS': 'Entre.Ríos',
                        'FORMOSA': '',
                        'JUJUY': 'Jujuy',
                        'LA PAMPA': 'La.Pampa',
                        'LA RIOJA': '',
                        'MENDOZA': 'Mendoza',
                        'MISIONES': '',
                        'NEUQUEN': '',
                        'RIO NEGRO': '',
                        'SALTA': 'Salta',
                        'SAN JUAN': '',
                        'SAN LUIS': 'San.Luis',
                        'SANTA CRUZ': '',
                        'SANTA FE': 'Santa.Fe',
                        'SANTIAGO DEL ESTERO': 'Santiago.del.Estero',
                        'TIERRA DEL FUEGO': '',
                        'TUCUMAN': 'Tucumán',
                        'DESCONOCIDO': ''
                    }
                    
                    # Convertimos a mayúsculas para asegurar que las claves coincidan
                    df['Provincia'] = df['Provincia'].str.upper()
                    
                    # Reemplazamos los valores en la columna 'Provincia' basados en el diccionario
                    df['Provincia'] = df['Provincia'].replace(reemplazos)
                    #df['Provincia'] = df['Provincia'].astype(str)
                    df['Provincia'] = df['Provincia'].replace('nan', '')
                    
                    return df
                
                
                df = validar_provincia(df, conn)

                # TENENCIA #FUNCIONA
                def validar_tenencia(df,conn):
                    
                    tenencia_query = """SELECT valor FROM datcrea_tablas.variables_generales 
                                    WHERE variable = 'tenencia'"""
                    tenencia_df = pd.read_sql(tenencia_query,con=conn)
                    tenencia = tenencia_df['valor'].tolist()
                
                    df['Tenencia'] = df['Tenencia'].apply(lambda x: x if x in tenencia else '')
                
                    return df
                
                df = validar_tenencia(df,conn)

                
                # DESTINO #FUNCIONA
                def validar_destino(df,conn):
                    
                    destino_query = """SELECT valor FROM datcrea_tablas.variables_generales 
                                        WHERE variable = 'destino'"""
                    destino_df = pd.read_sql(destino_query,con=conn)
                    destino = destino_df['valor'].tolist()
                    df['DestinoProduccion'] = df['DestinoProduccion'].apply(lambda x: x if x in destino else '')
                    
                    return df
                
                df = validar_destino(df,conn)

                # CULTIVO ANTECESOR
                def validar_cultivo_antecesor(df,conn):
                    
                    # Lista de valores a reemplazar y sus correspondientes reemplazos
                    reemplazos = {
                        'acai - tardío': '',
                        'acai - de primera': '',
                        'acai - de segunda': '',
                        'achicoria - de primera': 'Achicoria',
                        'agave - tardío': '',
                        'agave - de segunda': '',
                        'agave - de primera': '',
                        'aguacate - de primera': '',
                        'ajo - de primera': 'Ajo',
                        'alfalfa - de segunda': 'Alfalfa',
                        'alfalfa - de primera': 'Alfalfa',
                        'alfalfa - tardío': 'Alfalfa',
                        'algodón - de segunda': 'Algodón',
                        'algodón - tardío': 'Algodón',
                        'algodón - de primera': 'Algodón',
                        'alpiste - de segunda': 'Alpiste',
                        'alpiste - de primera': 'Alpiste',
                        'anana - de primera': '',
                        'arroz - de primera': 'Arroz',
                        'arroz - tardío': 'Arroz',
                        'arroz - de segunda': 'Arroz',
                        'arveja - tardío': 'Arveja',
                        'arveja - de primera': 'Arveja',
                        'arveja - de segunda': 'Arveja',
                        'avena - de primera': 'Avena',
                        'avena - de segunda': 'Avena',
                        'avena - tardío': 'Avena',
                        'avena blanca - de segunda': 'Avena',
                        'avena blanca - de primera': 'Avena',
                        'avena strigosa - tardío': 'Avena',
                        'avena strigosa - de segunda': 'Avena',
                        'avena strigosa - de primera': 'Avena',
                        'banano - de primera': '',
                        'brócoli - de primera': '',
                        'café - safra': '',
                        'café - de segunda': '',
                        'café - de primera': '',
                        'camelina - de primera': 'Camelina',
                        'caña de azucar - de primera': 'Caña de Azucar',
                        'caña de azucar - tardío': 'Caña de Azucar',
                        'caña de azucar - de segunda': 'Caña de Azucar',
                        'canola - de primera': '',
                        'carinata - de primera': 'Carinata',
                        'carinata - de segunda': 'Carinata',
                        'cartamo - de primera': 'Cártamo',
                        'cebada - de primera': 'Cebada',
                        'cebada - tardío': 'Cebada',
                        'cebada - de segunda': 'Cebada',
                        'cebadilla - de primera': 'Cebadilla',
                        'centeno - de segunda': 'Centeno',
                        'centeno - de primera': 'Centeno',
                        'centeno - tardío': 'Centeno',
                        'chia - tardío': '',
                        'chia - de segunda': '',
                        'chia - de primera': '',
                        'cítricos - de segunda': '',
                        'cítricos - tardío': '',
                        'cítricos - de primera': '',
                        'colza - de primera': 'Colza',
                        'colza - de segunda': 'Colza',
                        'coriandro - de primera': 'Coriandro',
                        'crotalaria - de primera': '',
                        'cultivo de servicio - tardío': 'Cultivo de Servicio',
                        'cultivo de servicio - de primera': 'Cultivo de Servicio',
                        'cultivo de servicio - de segunda': 'Cultivo de Servicio',
                        'festuca - de primera': 'Pastura perenne',
                        'garbanzo - de segunda': 'Garbanzo',
                        'garbanzo - de primera': 'Garbanzo',
                        'girasol - tardío': 'Girasol',
                        'girasol - de primera': 'Girasol',
                        'girasol - de segunda': 'Girasol',
                        'kiwi - de primera': '',
                        'lechuga - de segunda': '',
                        'lenteja - de primera': 'Lenteja',
                        'lenteja - de segunda': 'Lenteja',
                        'lenteja - tardío': 'Lenteja',
                        'limon - de primera': '',
                        'lotus - de primera': 'Lotus',
                        'lupino - de primera': '',
                        'maiz' : 'Maíz',
                        'maíz - de primera': 'Maíz de 1° Temprano',
                        'maíz - tardío': 'Maíz de 1° Tardio',
                        'maíz - safrinha': '',
                        'maíz - de segunda': 'Maíz de 2°',
                        'mani - tardío': 'Maní',
                        'mani - de primera': 'Maní',
                        'mani - de segunda': 'Maní',
                        'melon - de primera': '',
                        'mijo - de segunda': 'Mijo',
                        'mijo - de primera': 'Mijo',
                        'mijo - tardío': 'Mijo',
                        'moha - de primera': 'Moha',
                        'moha - de segunda': 'Moha',
                        'moha - tardío': 'Moha',
                        'nabo - de segunda': 'Nabo',
                        'nabo - de primera': 'Nabo',
                        'olivo - tardío': '',
                        'papa - de segunda': 'Papa',
                        'papa - de primera': 'Papa',
                        'pasturas - tardío': 'Pastura perenne',
                        'pasturas - de primera': 'Pastura perenne',
                        'pasturas - de segunda': 'Pastura perenne',
                        'poroto - tardío': 'Poroto',
                        'poroto - de segunda': 'Poroto',
                        'poroto - de primera': 'Poroto',
                        'poroto mung - tardío': 'Poroto Mung',
                        'poroto mung - de segunda': 'Poroto Mung',
                        'poroto mung - de primera': 'Poroto Mung',
                        'raigrass - de segunda': 'Ryegrass',
                        'raigrass - de primera': 'Ryegrass',
                        'raigrass - tardío': 'Ryegrass',
                        'sandia - de primera': '',
                        'sesamo - de primera': '',
                        'soja' : 'Soja',
                        'soja - de segunda': 'Soja de 2°',
                        'soja - tardío': 'Soja de 1°',
                        'soja - primavera / verano': '',
                        'soja - safra': '',
                        'soja - semestre a': '',
                        'soja - de primera': 'Soja de 1°',
                        'sorgo - de primera': 'Sorgo',
                        'sorgo - tardío': 'Sorgo',
                        'sorgo - de segunda': 'Sorgo',
                        'tabaco - de primera': '',
                        'té - de primera': '',
                        'tomate - de segunda': '',
                        'tomate - de primera': '',
                        'trebol blanco - tardío': 'Trebol',
                        'trebol blanco - de primera': 'Trebol',
                        'trebol rojo - de primera': 'Trebol',
                        'trigo - de segunda': 'Trigo',
                        'trigo - de primera': 'Trigo',
                        'trigo - tardío': 'Trigo',
                        'triticale - de primera': 'Triticale',
                        'vicia - de segunda': 'Vicia',
                        'vicia - de primera': 'Vicia',
                        'vicia - tardío': 'Vicia',
                        'vid - de primera': 'Vid',
                        'vid - tardío': 'Vid',
                        'vid - de segunda': 'Vid',
                        'yerba mate - de primera': ''
                    }
                
                    # Convertir a minúsculas y reemplazar valores según el diccionario
                    df['CultivoAntecesor'] = df['CultivoAntecesor'].str.lower()
                    df['CultivoAntecesor'] = df['CultivoAntecesor'].replace(reemplazos)
                    
                    
                    return df
                    
                df = validar_cultivo_antecesor(df, conn)
                

                ## 6.7) Validación Calidad ambiente #FUNCIONA
                def validar_calidad_ambiente(df,conn):
                                
                    calidad_ambiente_query = """SELECT valor 
                                                FROM datcrea_tablas.variables_generales 
                                                WHERE variable = 'calidad_ambiente'"""
                    calidad_ambiente_df = pd.read_sql(calidad_ambiente_query,con=conn)
                    calidad_ambiente = calidad_ambiente_df['valor'].tolist()
                    
                    df['CalidadAmbiente'] = df['CalidadAmbiente'].apply(lambda x: x if x in calidad_ambiente else '')
                
                    return df
                    
                df = validar_calidad_ambiente(df, conn)
                

                def validar_subgrupo(df, conn): #REVISAR #FUNCIONA
                    
                    subgrupo_query = """SELECT subgrupo FROM datcrea_tablas.subgrupos_cultivos"""
                    subgrupo_df = pd.read_sql(subgrupo_query,con=conn)
                    subgrupo = subgrupo_df['subgrupo'].tolist()
                    df['SubgrupoCultivo'] = ['' if x not in subgrupo else x for x in df['SubgrupoCultivo']]
                
                
                    #Carga los subgrupos de cultivosvy los organiza en un diccionario
                    subgrupos_cultivos_query = """SELECT cultivo, subgrupo FROM datcrea_tablas.subgrupos_cultivos WHERE cultivo IS NOT NULL"""
                    subgrupos_cultivos = pd.read_sql(subgrupos_cultivos_query, con=conn)
                
                    # Crear un diccionario para almacenar los resultados
                    subgrupos_cultivos_dict = {}
                    for index, row in subgrupos_cultivos.iterrows():
                        cultivo = row['cultivo']
                        subgrupo = row['subgrupo']
                        if cultivo in subgrupos_cultivos_dict:
                            subgrupos_cultivos_dict[cultivo].append(subgrupo)
                        else:
                            subgrupos_cultivos_dict[cultivo] = [subgrupo]
                
                    def validar_subgrupo(row):
                        
                        valor_cultivo = row['Cultivo']
                        valor_subgrupo = row['SubgrupoCultivo']
                
                        if valor_cultivo in subgrupos_dict:
                            if valor_subgrupo in subgrupos_dict[valor_cultivo]:
                                return valor_subgrupo
                        return ''
                
                    return df
                
                df = validar_subgrupo(df, conn)
                
                # FERTILIZACIÓN
                def validar_fertilizacion(df,conn):
                    
                    fertilizacion_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'fertilizacion'"""
                    fertilizacion = pd.read_sql(fertilizacion_query, con=conn)
                    fertilizacion = fertilizacion['valor'].tolist()
                    df['Fertilizacion'] = ['' if x not in fertilizacion else x for x in df['Fertilizacion']]
                    return df
                
                df = validar_fertilizacion(df, conn)
                


                ## 6.12) Reemplazo de 0 por Na
                # REEMPLAZAR CEROS
                def replace_ceros(df):
                    df['SuperficieSembrada'] = df['SuperficieSembrada'].replace(0, pd.NA)
                    df['espaciamiento'] = df['espaciamiento'].replace(0, pd.NA)
                    df['Rendimiento'] = df['Rendimiento'].replace(0, pd.NA)
                    df['Superficie'] = df['Superficie'].replace(0, pd.NA)
                    return df
                
                df = replace_ceros(df)
                
                # 6.13) Validación de Forma
                
                # VALIDAR FORMA DE FERTILIZACIÓN
                def validar_forma_fertilizacion(df,conn):
                    forma_fertilizacion_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'forma_fertilizacion'"""
                    forma_fertilizacion = pd.read_sql(forma_fertilizacion_query, con=conn)
                    forma_fertilizacion = forma_fertilizacion['valor'].tolist()
                    df['1_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['1_Forma']]
                    df['2_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['2_Forma']]
                    df['3_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['3_Forma']]
                    return df
                
                df = validar_forma_fertilizacion(df, conn)

                
                # Validación de Sistema de riego
                # SISTEMA RIEGO
                def validar_sistema_riego(df,conn):
                    sistema_riego_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'sistema_riego'"""
                    sistema_riego = pd.read_sql(sistema_riego_query, con=conn)
                    sistema_riego = sistema_riego['valor'].tolist()
                    df['Riego'] = ['' if x not in sistema_riego else x for x in df['Riego']]
                    return df
                
                df = validar_sistema_riego(df,conn)

                # INFLUENCIA NAPA
                def influencia_napa(df,conn):
                    
                    influencia_napa_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'influencia_napa'"""
                    influencia_napa = pd.read_sql(influencia_napa_query,con=conn)
                    influencia_napa = influencia_napa['valor'].tolist()
                    df['Napa'] = ['' if x not in influencia_napa else x for x in df['Napa']]
                    return df
                
                df = influencia_napa(df, conn)
                
                
                # Conexión a la base
                import pandas as pd
                from sqlalchemy import create_engine
                
                host="dw.crea.org.ar"
                port="54322"
                user="postgres"
                password="Tr0pema44cr34#"
                database="warehouse"
                
                conn2 = create_engine("postgresql+psycopg2://{user}:{password}@{host}:{port}/{database}".format(host=host,user=user,password=password,database=database,port=port))
                
                # Escribir en celdas excel
                region_grupo_query = """SELECT r.region_sigla, r.region_nombre,c.crea_numero,c.crea_nombre  
                FROM crm_crea.regiones AS r
                JOIN crm_crea.crea AS c 
                ON r.region_id = c.crea_region_id
                WHERE c.crea_baja IS NULL
                GROUP BY r.region_sigla, r.region_nombre, c.crea_numero,c.crea_nombre
                ORDER BY r.region_sigla, c.crea_nombre;"""
                region_grupo = pd.read_sql(region_grupo_query, con=conn2)
                
                if not pd.isnull(df['CUIC'][0]) and df['CUIC'][0] != 'Otro':
                    cuic = df['CUIC'][0]
                
                    region = region_grupo.loc[region_grupo['region_sigla'].str.startswith(cuic[:3]), 'region_nombre'].iloc[0]
                    #region = replace region por el formato dat que sale de una lista.
                    grupo = region_grupo.loc[region_grupo['crea_numero'].str.contains(cuic[3:6]), 'crea_nombre'].iloc[0]
                    
                    regiones_dat_dict = {
                        
                        'CENTRO':'CENTRO',
                        'CHACO SANTIAGUEÑO':'CHACO.SANTIAGUEÑO',
                        'CORDOBA NORTE':'CORDOBA.NORTE',
                        'ESTE':'ESTE',
                        'LITORAL NORTE':'LITORAL.NORTE',
                        'LITORAL SUR':'LITORAL.SUR',
                        'MAR Y SIERRAS':'MAR.Y.SIERRAS',
                        'NORTE DE BUENOS AIRES':'NORTE.BUENOS.AIRES',
                        'NOA':'NOA',
                        'NORTE DE SANTA FE':'NORTE.SANTA.FE',
                        'OESTE ARENOSO':'OESTE.ARENOSO',
                        'OESTE':'OESTE',
                        'PATAGONIA':'PATAGONIA',
                        'SEMIARIDA':'SEMIARIDA',
                        'SUDESTE':'SUDESTE',
                        'SANTA FE CENTRO':'SANTA.FE.CENTRO',
                        'SUR DE SANTA FE':'SUR.SANTA.FE',
                        'SUDOESTE':'SUDOESTE',
                        'VALLES CORDILLERANOS':'VALLES.CORDILLERANOS'
                        }
                    # Reemplazo directo sin función
                    for clave, valor in regiones_dat_dict.items():
                        region = region.replace(clave, valor)
                    
                    #print('REGIÓN: ', region)
                    
                    grupos_dat_dict = {'ALEJANDRO CHAJAN':'ALEJANDRO.CHAJAN',
                    'BUENA ESPERANZA':'BUENA.ESPERANZA',
                    'CANALS':'CANALS',
                    'CARNERILLO':'CARNERILLO',
                    'CAÑADA SECA':'CAÑADA.SECA',
                    'CTALAMOCHITA':'CTALAMOCHITA',
                    'HUINCA RENANCO':'HUINCA.RENANCO',
                    'LA CESIRA TAMBERO':'LA.CESIRA.TAMBERO',
                    'LA PORTADA':'LA.PORTADA',
                    'LABOULAYE - BOUCHARDO':'LABOULAYE.BUCHARDO',
                    'MELO SERRANO':'MELO.SERRANO',
                    'RANQUELES':'RANQUELES',
                    'RIO CUARTO':'RIO.CUARTO',
                    'RIO QUINTO':'RÍO.QUINTO',
                    'TAMBERO LABOULAYE':'TAMBERO.LABOULAYE',
                    'TAMBERO VILLA MARIA':'VILLA.MARIA',
                    'TEGUA':'TEGUA',
                    'VALLE DEL CONLARA':'VALLE.DEL.CONLARA',
                    'WASHINGTON MACKENNA':'WASHINGTON.MACKENNA',
                    'CAMPO GALLO MONTE QUEMADO':'CAMPO.GALLO.MONTE.QUEMADO',
                    'GUAYACAN':'GUAYACAN',
                    'IBARRETA':'IBARRETA',
                    'LOMITAS':'LOMITAS',
                    'PALMARES':'PALMARES',
                    'PAMPA DEL INFIERNO':'PAMPA.DEL.INFIERNO',
                    'QUIMILI':'QUIMILI',
                    'RENOVALES':'RENOVALES',
                    'SACHAYOJ':'SACHAYOJ',
                    'SANAVIRONES':'SANAVIRONES',
                    'SEMIARIDO NORTE':'SEMIARIDO.NORTE',
                    'SUDESTE SANTIAGUEÑO':'SES',
                    'TINTINA':'TINTINA',
                    'ARROYITO':'ARROYITO',
                    'BARRANCA YACO':'BARRANCA.YACO',
                    'CAROYA':'CAROYA',
                    'CAÑADA DE LUQUE SITON':'CAÑADA.DE.LUQUE.SITON',
                    'DEL ESTE':'DEL.ESTE',
                    'GANADERO DEL NOROESTE':'GANADERO.DEL.NOROESTE',
                    'JESUS MARIA':'JESUS.MARIA',
                    'LAGUNA LARGA':'LAGUNA.LARGA',
                    'LEOPOLDO LUGONES':'LEOPOLDO.LUGONES',
                    'MONTE CRISTO':'MONTE.CRISTO',
                    'RIO PRIMERO':'RIO.PRIMERO',
                    'SIERRAS CHICAS':'SIERRAS.CHICAS',
                    'TOTORAL':'TOTORAL',
                    'ABASTO':'Otro',
                    'CAÑUELAS':'CAÑUELAS',
                    'GELAS':'GELAS',
                    'LUJAN':'LUJAN',
                    'NAVARRO II':'NAVARROII',
                    'PIONEROS ABASTO':'Otro',
                    'AVATI - I - ARROCERO':'AVATI.I.ARROCERO',
                    'CURUZU CUATIA':'CURUZU.CUATIA',
                    'ESQUINA':'ESQUINA',
                    'MALEZALES':'Otro',
                    'MERCEDES':'MERCEDES',
                    'TIERRA COLORADA':'TIERRA.COLORADA',
                    'URUNDAY':'URUNDAY',
                    'ÑANDUBAY':'ÑANDUBAY',
                    'BOVRIL EL SOLAR':'BOVRIL.EL.SOLAR',
                    'CONCEPCION URUGUAY':'CONCEPCION.URUGUAY',
                    'CONCORDIA CHAJARI':'CONCORDIA.CHAJARI',
                    'GALARZA':'GALARZA',
                    'GUALEGUAYCHU':'GUALEGUAYCHU',
                    'ISLAS DEL IBICUY':'ISLAS.DEL.IBICUY',
                    'LA PAZ':'LA.PAZ',
                    'LARROQUE GUALEGUAY':'LARROQUE.GUALEGUAY',
                    'MANDISOVI CONCORDIA':'MANDISOVI.CONCORDIA',
                    'MONTOYA':'MONTOYA',
                    'SAN JAIME':'SAN.JAIME',
                    'VICTORIA':'VICTORIA',
                    'VILLAGUAY':'VILLAGUAY',
                    'ARROYO DE LOS HUESOS':'ARROYO.DE.LOS.HUESOS',
                    'AZUL CHILLAR':'AZUL.CHILLAR',
                    'CTE.N.OTAMENDI':'CTE.N.OTAMENDI',
                    'DEFFERRARI':'DEFFERRARI',
                    'FRONTERA':'FRONTERA',
                    'FULTON':'FULTON',
                    'LOBERIAS GRANDES':'LOBERIAS.GRANDES',
                    'NECOCHEA QUEQUEN':'NECOCHEA.QUEQUEN',
                    'QUEQUEN SALADO':'QUEQUEN.SALADO',
                    'SAN CAYETANO - TRES ARROYOS':'SAN.CAYETANO.TRES.ARROYOS',
                    'SAN FRANCISCO DE BELLOCQ':'SAN.FRANCISCO.DE.BELLOCQ',
                    'SAN MANUEL':'SAN.MANUEL',
                    'TAMBERO MAR Y SIERRA':'TAMBERO.MAR.Y.SIERRA',
                    'TANDIL':'TANDIL',
                    'TRES ARROYOS':'TRES.ARROYOS',
                    'ZONA 4 LECHERA':'ZONA.4.LECHERA',
                    'ALBERDI':'ALBERDI',
                    'ALBERTI- PLA':'ALBERTI.PLA',
                    'ARROYO DEL MEDIO':'ARROYO.DEL.MEDIO',
                    'BRAGADO':'BRAGADO',
                    'GIDAG':'GIDAG',
                    'PERGAMINO':'PERGAMINO',
                    'RAWSON TRES SARGENTOS':'RAWSON.TRES.SARGENTOS',
                    'SAN ANTONIO DE ARECO':'SAN.ANTONIO.DE.ARECO',
                    'SAN PEDRO VILLA LIA':'SAN.PEDRO.VILLA.LIA',
                    'SEGUI LA ORIENTAL':'SEGUI.LA.ORIENTAL',
                    'CAÑAVERALES DE TUCUMAN':'CAÑAVERALES.DE.TUCUMAN',
                    'EL PALOMAR':'EL.PALOMAR',
                    'EL RODEO':'EL.RODEO',
                    'LA COCHA':'LA.COCHA',
                    'LOS ALGARROBOS':'LOS.ALGARROBOS',
                    'METAN':'METAN',
                    'PALO SANTO':'Otro',
                    'SAN PATRICIO':'SAN.PATRICIO',
                    'SANTA ROSA CATAMARCA':'SANTA.ROSA.CATAMARCA',
                    'SURCOS':'Otro',
                    'SUYAY':'SUYAY',
                    'VALLES TEMPLADOS':'Otro',
                    'YUNGAS':'YUNGAS',
                    'CUÑA BOSCOSA':'CUÑA.BOSCOSA',
                    'MARGARITA CAMPO ALEMAN':'MARGARITA.CAMPO.ALEMAN',
                    'RAMAYON':'RAMAYON',
                    'SAN CRISTOBAL-LA LUCILA':'SAN.CRISTOBAL.LALUCILA',
                    'VILLA ANA-ARANDU':'VILLA.ANA.ARANDÚ',
                    'VILLA OCAMPO':'VILLA.OCAMPO',
                    'AMERICA':'AMERICA',
                    'AMERICA II':'AMERICA.II',
                    'AMERICA LECHERO':'AMERICA.LECHERO',
                    'ATREUCO':'ATREUCO',
                    'CORRALERO':'CORRALERO',
                    'PELLEGRINI - TRES LOMAS':'PELLEGRINI.TRES.LOMAS',
                    'PICO BARON':'PICO.BARON',
                    'PICO QUEMU':'PICO.QUEMU',
                    'QUEMU CATRILO':'QUEMU.CATRILO',
                    'TRENQUE LAUQUEN II':'TRENQUE.LAUQUEN.II',
                    'AGROGANADERO 9 DE JULIO':'AGROGANADERO.9.DE.JULIO',
                    'AMEGHINO':'AMEGHINO',
                    'BOLIVAR':'BOLIVAR',
                    'CASARES - 9 DE JULIO':'CASARES.9.DE.JULIO',
                    'GENERAL PINTO':'GENERAL.PINTO',
                    'GENERAL VILLEGAS':'GENERAL.VILLEGAS',
                    'GUANACO LAS TOSCAS':'GUANACO.LAS.TOSCAS',
                    'HENDERSON-DAIREAUX':'HENDERSON.DAIREAUX',
                    'HERRERA VEGAS - PEHUAJO':'HERRERA.VEGAS.PEHUAJO',
                    'INFOSURA':'INFOSURA',
                    'LA VIA':'LA.VIA',
                    'LINCOLN':'LINCOLN',
                    'MONES CAZON PEHUAJO':'MONES.CAZON.PEHUAJO',
                    'NUEVE DE JULIO':'NUEVE.DE.JULIO',
                    'PEHUAJO CASARES':'PEHUAJO.CASARES',
                    'PIROVANO LA LARGA':'PIROVANO.LA.LARGA',
                    'SALAZAR MONES CAZON':'SALAZAR.MONES.CAZON',
                    'TAMBERO AMEGHINO VILLEGAS':'TAMBERO.AMEGHINO.VILLEGAS',
                    'TEJEDOR':'TEJEDOR',
                    'TREINTA AGOSTO- MARI LAUQUEN':'TREINTA.AGOSTO.MARI.LAUQUEN',
                    'ALTO VALLE - VALLE MEDIO':'ALTO.VALLE.VALLE.MEDIO',
                    'CUENCA DEL AGRIO':'CUENCA.DEL.AGRIO',
                    'ESQUEL':'ESQUEL',
                    'LANIN':'LANIN',
                    'SANTA CRUZ':'SANTACRUZ',
                    'TIERRA DEL FUEGO':'TIERRADELFUEGO',
                    'VIEDMA':'VIEDMA',
                    'AOKEN AL':'AOKEN.AL',
                    'CALEUCHE':'CALEUCHE',
                    'GUATRACHE':'GUATRACHE',
                    'HOLISTICO':'HOLISTICO',
                    'PEHUENCHE':'PEHUENCHE',
                    'SOVEN':'SOVEN',
                    'TAMBERO GUATRACHE':'TAMBERO.GUATRACHE',
                    'UTRACAN':'UTRACAN',
                    '25 DE MAYO':'CREA.25.DE.MAYO',
                    'ARROYO DE LAS FLORES':'ARROYO.DE.LAS.FLORES',
                    'ARROYO LANGUEYU':'ARROYO.LANGUEYU',
                    'AYACUCHO':'AYACUCHO',
                    'CASTELLI - BELGRANO':'CASTELLI.BELGRANO',
                    'DEL TUYU':'DEL.TUYU',
                    'FORTIN MULITAS':'FORTIN.MULITAS',
                    'LEZAMA':'LEZAMA',
                    'MAIPU':'MAIPU',
                    'MAR CHIQUITA':'MAR.CHIQUITA',
                    'MONTE':'MONTE',
                    'PILA':'PILA',
                    'RAUCH - UDAQUIOLA':'RAUCH.UDAQUIOLA',
                    'RIO SALADO':'RIO.SALADO',
                    'ROQUE PEREZ SALADILLO':'ROQUE.PEREZ.SALADILLO',
                    'TAPALQUE II':'TAPALQUE.II',
                    'VALLIMANCA':'VALLIMANCA',
                    'CASTELAR':'CASTELAR',
                    'CENTRO OESTE SANTAFESINO':'C.O.S',
                    'CUENCA':'CUENCA',
                    'EL CEIBO':'EL.CEIBO',
                    'ESPERANZA':'ESPERANZA',
                    'GALVEZ':'GALVEZ',
                    'RAFAELA':'RAFAELA',
                    'SAN FRANCISCO':'SAN.FRANCISCO',
                    'SAN GUILLERMO':'Otro',
                    'SAN MARTIN DE LAS ESCOBAS - COLONIA BELGRANO':'S.M.E.C.B',
                    'SUNCHALES':'Otro',
                    'APLICADORES':'Otro',
                    'ARMSTRONG MONTES DE OCA':'ARMSTRONG.MONTES.DE.OCA',
                    'ASCENSION':'ASCENSION',
                    'COLONIA MEDICI':'COLONIA.MEDICI',
                    'COSTAS DEL CARCARAÑA':'COSTAS.DEL.CARCARAÑA',
                    'EL ABROJO':'EL.ABROJO',
                    'GENERAL ARENALES':'GENERAL.ARENALES',
                    'GENERAL BALDISSERA':'GENERAL.BALDISSERA',
                    'LA CALANDRIA':'LA.CALANDRIA',
                    'LA MAROMA':'LA.MAROMA',
                    'LAS PETACAS':'LAS.PETACAS',
                    'MARIA TERESA':'MARIA.TERESA',
                    'MONTE BUEY INRIVILLE':'MONTE.BUEY.INRIVILLE',
                    'MONTE MAIZ':'MONTE.MAIZ',
                    'POSTA ESPINILLOS':'POSTA.ESPINILLOS',
                    'ROSARIO':'ROSARIO',
                    'SAN JORGE LAS ROSAS':'SAN.JORGE.LAS.ROSAS',
                    'SANTA ISABEL':'SANTA.ISABEL',
                    'SANTA MARIA':'SANTA.MARIA',
                    'TEODELINA':'TEODELINA',
                    'BENITO JUAREZ':'BENITO.JUAREZ',
                    'CARHUE HUANGUELEN':'CARHUE.HUANGUELEN',
                    'CORONEL PRINGLES II':'CORONEL.PRINGLES.II',
                    'CORONEL SUAREZ':'CORONEL.SUAREZ',
                    'GENERAL LAMADRID':'GENERAL.LAMADRID',
                    'LAPRIDA':'LAPRIDA',
                    'NUESTRA SEÑORA DE LAS PAMPAS':'NUESTRA.SEÑORA.DE.LAS.PAMPAS',
                    'OLAVARRIA':'OLAVARRIA',
                    'PEDRO LURO':'PEDRO.LURO',
                    'SAN ELOY - PIÑEYRO':'SAN.ELOY.PIÑEYRO',
                    'VENTANIA':'VENTANIA',
                    'ACONCAGUA':'Otro',
                    'ARAUCO':'Otro',
                    'CALCHAQUI':'Otro',
                    'FRUTICOLA CUYO':'Otro',
                    'HUARPE':'Otro',
                    'LAS ACEQUIAS':'Otro',
                    'LOS ANDES':'Otro',
                    'NOGALERO DEL NORTE':'Otro',
                    'OLIVICOLA SAN JUAN':'Otro',
                    'VIGNERONS':'Otro'
                    }
                    
                    # Reemplazo directo sin función
                    for clave, valor in grupos_dat_dict.items():
                        grupo = grupo.replace(clave, valor)

                else:
                    cuic = 'Otro'
                    region = ''
                    grupo = ''
                
                #print('GRUPO: ', grupo)

                # Filtro el DataFrame
                # CULTIVO VERANO
                def filtrar_cultivos_verano(df, conn):
                    query = """SELECT cultivo FROM datcrea_tablas.cultivos WHERE ciclo = 'v'"""
                    verano_df = pd.read_sql(query, conn)
                    cultivos_verano = verano_df['cultivo'].tolist()
                    df = df[df['Cultivo'].isin(cultivos_verano)]
                    return df
                
                verano = filtrar_cultivos_verano(df, conn)
                
                def exportar_sima_verano(verano):
                    
                    seccion = 200 # Máx registros por planilla xls
                    registros_campana = len(verano.index) # cuenta el total de registros
                    registros_wb = math.ceil(registros_campana / seccion) #cociente para saber cantidad de worksheets
                    
                    # Armo listado de subset de 200 (máx) registros cada uno
                    dataframes = []
                    
                    for i in range(registros_wb):
                        inicio = i * seccion
                        fin = (i + 1) * seccion
                        seccion_df = verano.iloc[inicio:fin]
                        dataframes.append(seccion_df)
                    
                    # Crear una workbook para cada dataframe
                    
                    # Ruta del archivo original
                    ruta_original = 'C:/Users/EPolacco/Documents/9 - DAT/DAT-Cultivos-de-verano-2023-24.xlsx'
                    
                    # Cargar el archivo original
                    wb_original = load_workbook(filename=ruta_original)
                    
                    # Obtener la hoja original
                    planilla_original = wb_original['Planilla ']
                    
                    # Columnas excluídas porque tienen fórmulas y validaciones.
                    columns_to_exclude = {'I', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'U', 'W', 'Y', 'Z', 'AC',
                                        'AF', 'AG', 'AH', 'AK', 'AO', 'CC', 'CS', 'CU', 'CV', 'CW', 'CX',
                                        'CY', 'CZ', 'DI', 'DJ', 'DK','DL', 'DM', 'DN','DO','DP','DQ','DR', 
                                        'DS','DT','DU', 'DV', 'DW', 'DY', 'DZ', 'EA',
                                        'IQ','IR','IS','IT','IU','IV','IW','IX','IY','IZ', 'JA', 'JB', 'JC',
                                        'JD', 'JE'}
                    
                    
                    # Iterar sobre los df y guardar en archivos xls separados
                    for i, seccion_df in enumerate(dataframes, 1):
                        # Crear un workbook para cada sección
                        if i == 1:
                            wb = wb_original  # Utilizar el archivo original para el primer df
                            sheet = planilla_original
                        else:
                            # Crear una copia del archivo original para los df subsiguientes
                            ruta_copia = f'C:/Users/EPolacco/Documents/9 - DAT/SIMA/{cuic}_DAT-Cultivos-de-verano-2023-24_SECCION_{i}.xlsx'
                            shutil.copy(ruta_original, ruta_copia)
                            
                            # Cargar la copia
                            wb = load_workbook(filename=ruta_copia)
                            sheet = wb['Planilla ']
                            sheet.title = f'Planilla '
                    
                        # Borrar datos existentes en la hoja, después de la fila 14 porque sino se pisan los datos
                        
                        for row in sheet.iter_rows(min_row=14, max_row=sheet.max_row, min_col=6, max_col=sheet.max_column):
                            for cell in row:
                                if cell.column_letter not in columns_to_exclude:
                                    cell.value = None
                    
                    
                        # Obtener las filas de datos del df
                        rows = dataframe_to_rows(seccion_df,index=False, header=False)
                        
                        # Copiar en worksheet a partir de la fila 14 y columna 6
                        for r_idx, row in enumerate(rows, 14):
                            for c_idx, value in enumerate(row, 6):
                                if not pd.isna(value):
                                    sheet.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Región
                        sheet['F3'] = region
                        # Grupo
                        sheet['F6'] = grupo
                        # Guardar el workbook, si no es el primer df
                        if i != 1:
                            wb.save(ruta_copia)
                    
                    
                    # Guardar workbook con el primer df
                    ruta_alternativa = f'C:/Users/EPolacco/Documents/9 - DAT/SIMA/{cuic}_DAT-Cultivos-de-verano-2023-24_SECCION_1.xlsx'
                    wb_original.save(ruta_alternativa)
                
                exportar_sima_verano(verano)
                    
            
            else: #Invierno
                # Importo librerías
     
                from openpyxl import load_workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                import pandas as pd
                import numpy as np
                from datetime import datetime
                import math
                import shutil

                # Conexión a la base
                from sqlalchemy import create_engine
                import mysql.connector
                
                host='66.97.33.201'
                password='Dat##2023Tabla$'
                port=33306
                user='datcrea_tablas'
                database='datcrea_tablas'
                
                conn = create_engine("mysql+mysqlconnector://{user}:{password}@{host}:{port}/{database}".format(host=host,user=user,password=password,database=database,port=port),connect_args={'auth_plugin': 'mysql_native_password'})
                
            
                #ruta_completa = directorio + archivo
                #df = pd.read_excel(ruta_completa)
                if not pd.isnull(df['CUIC'][0]) and df['CUIC'][0] != '':
                    cuic = df['CUIC'][0]
                else:
                    cuic = 'Otro'
                
                def filtrar_campana(df, campana_filtro):
                    
                    filtro = df['Ciclo'] == campana_filtro
                    df = df[filtro]
                    return df

                df = filtrar_campana(df, campana_filtro)

                
                def formato_fecha(df):
                    df['FechaSiembra'] = df['FechaSiembra'].str[:10] #Elimino horas en fechas
                    df['FechaSiembra'] = pd.to_datetime(df['FechaSiembra'])
                    df['FechaSiembra'] = df['FechaSiembra'].dt.strftime('%d/%m/%Y')    
                    return df
                    
                df = formato_fecha(df)


                # 3.2) Validar código genética #FUNCIONA
                # Genética
                # Consultar los nro_registro desde la base de datos
                def validar_codigo_genetica(df, conn):
                    query = "SELECT nro_registro FROM datcrea_tablas.materiales WHERE nro_registro IS NOT NULL"
                    df_materiales = pd.read_sql(query, conn)
                    materiales = df_materiales['nro_registro'].astype(int).tolist()
                    df['Codigo_Genetica'] = df['Codigo_Genetica'].fillna(0)
                    df['Codigo_Genetica'] = df['Codigo_Genetica'].astype('Int64')
                    df['Codigo_Genetica'] = ['' if x not in materiales else x for x in df['Codigo_Genetica']]
                    return df

                df = validar_codigo_genetica(df, conn)



                ## 3.3) Validación Cultivo #FUNCIONA
                # Cultivos
                def validar_cultivos(df, conn):
                    cultivos_query = """SELECT DISTINCT(cultivo) FROM datcrea_tablas.cultivos"""
                    cultivos_df = pd.read_sql(cultivos_query,con=conn)
                    cultivos = cultivos_df['cultivo'].tolist()
                    df['Cultivo'] = df['Cultivo'].replace(['Soja - De segunda', 'Mani'], ['Soja', 'Maní'])
                    df['Cultivo'] = ['' if x not in cultivos else x for x in df['Cultivo']]
                    return df
                
                df = validar_cultivos(df,conn)

                ## 3.4) Validación Genética
                
                # Cultivos
                def validar_genetica(df,conn): 
                    
                    genetica_query = """SELECT material_inase FROM datcrea_tablas.materiales 
                                        WHERE nro_registro IS NOT NULL"""
                    genetica_df = pd.read_sql(genetica_query,con=conn)
                    
                    genetica = genetica_df['material_inase'].tolist()
                    df['Genetica'] = ['' if x not in genetica else x for x in df['Genetica']]
                    return df
                df = validar_genetica(df,conn)

                ## 3.9) Validación Semillero #FUNCIONA
                def validar_semillero(df, conn): 
                    
                    semillero_query = """SELECT semillero FROM datcrea_tablas.semilleros"""
                    semillero = pd.read_sql(semillero_query,con=conn)
                    semillero = semillero['semillero'].tolist()
                    df['Semillero'] = ['' if x not in semillero else x for x in df['Semillero']]
                    return df
                
                df = validar_semillero(df, conn)

                def validar_genetica_semillero(df, conn): #FUNCIONA
                    
                    # VALIDA GENÉTICA Y SEMILLERO EN BASE A CULTIVO Y CÓDIGO GENÉTICO
                    semillero_query = """SELECT cultivo, nro_registro, semillero, material 
                    FROM datcrea_tablas.materiales where nro_registro IS NOT NULL;"""
                    semillero_genetica = pd.read_sql(semillero_query,con=conn)
                    semillero_genetica['nro_registro'] = semillero_genetica['nro_registro'].astype(int)
                    
                    for index, row in semillero_genetica.iterrows():
                        cultivo = row['cultivo']
                        codigo_genetica = row['nro_registro']
                        semillero = row['semillero']
                        genetica = row['material']
                        
                        # Actualizar 'df' donde coincidan cultivo y codigo_genetica
                        df.loc[(df['Cultivo'] == cultivo) & (df['Codigo_Genetica'] == codigo_genetica), ['Semillero', 'Genetica']] = [semillero, genetica]
                    
                    return df
                    
                df = validar_genetica_semillero(df, conn)

                # Convertir el Código de Fertilizante en integer # FUNCIONA
                def fertilizante_int(df, conn):
                    columnas = [
                        '1_Codigo_Registro_1', '1_Codigo_Registro_2',
                        '2_Codigo_Registro_1', '2_Codigo_Registro_2',
                        '3_Codigo_Registro_1', '3_Codigo_Registro_2',
                        '4_Codigo_Registro_1', '4_Codigo_Registro_2'
                    ]
                    
                    for col in columnas:
                        df[col] = df[col].fillna(0).astype('Int64')
                    
                    return df
                
                df = fertilizante_int(df, conn)

                # 3.7) Validación de Fertilizantes # FUNCIONA
                
                # Registro fertilizantes 
                def validar_id_senasa_fertilizantes(df, conn):
                    reg_fert_query = """SELECT id_senasa 
                                        FROM datcrea_tablas.registro_fertilizantes 
                                        WHERE id_senasa IS NOT NULL"""
                    reg_fert_df = pd.read_sql(reg_fert_query,con=conn)
                    id_senasa = reg_fert_df['id_senasa'].astype(int).tolist()
                    
                    df['1_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['1_Codigo_Registro_1']]
                    df['1_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['1_Codigo_Registro_2']]
                    df['2_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['2_Codigo_Registro_1']]
                    df['2_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['2_Codigo_Registro_2']]
                    df['3_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['3_Codigo_Registro_1']]
                    df['3_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['3_Codigo_Registro_2']]
                    df['4_Codigo_Registro_1'] = ['' if x not in id_senasa else x for x in df['4_Codigo_Registro_1']]
                    df['4_Codigo_Registro_2'] = ['' if x not in id_senasa else x for x in df['4_Codigo_Registro_2']]
                    return df
                df = validar_id_senasa_fertilizantes(df, conn)
                
                def validar_nombre_fertilizante(df, conn):
                    # Consulta para obtener los fertilizantes registrados con un id_senasa no nulo
                    reg_fert_query = "SELECT id_senasa, fertilizante FROM datcrea_tablas.registro_fertilizantes WHERE id_senasa IS NOT NULL"
                    registros_fertilizantes = pd.read_sql(reg_fert_query, con=conn)
                    
                    # Asegurarse de que id_senasa es de tipo entero
                    registros_fertilizantes['id_senasa'] = registros_fertilizantes['id_senasa'].astype(int)
                
                    # Preparar un diccionario para mapear id_senasa a fertilizante
                    id_senasa_to_fertilizante = dict(zip(registros_fertilizantes['id_senasa'], registros_fertilizantes['fertilizante']))
                
                    # Actualizar las columnas en df usando el diccionario de mapeo
                    for i in range(1, 5):  # Iterar sobre las 4 series de columnas
                        for j in range(1, 3):  # Iterar sobre las 2 columnas por serie
                            codigo_columna = f"{i}_Codigo_Registro_{j}"
                            producto_columna = f"{i}_Producto_{j}"
                            df[producto_columna] = df[codigo_columna].map(id_senasa_to_fertilizante).fillna(df[producto_columna])
                    
                    return df
                df = validar_nombre_fertilizante(df, conn)


                def validar_fertilizante(df,conn):
                    reg_fert_query = """SELECT * 
                                        FROM datcrea_tablas.registro_fertilizantes 
                                        WHERE id_senasa IS NOT NULL"""
                    registros_fertilizantes = pd.read_sql(reg_fert_query, con=conn)
                    registros_fertilizantes['id_senasa'] = registros_fertilizantes['id_senasa'].astype(int)
                    
                    # Crear un diccionario para almacenar los resultados
                    id_senasa_dict = {}
                    
                    for index, row in registros_fertilizantes.iterrows():
                        fertilizante = row['fertilizante']
                        id_senasa = row['id_senasa']
                        if fertilizante in id_senasa_dict:
                            id_senasa_dict[fertilizante].append(id_senasa)
                        else:
                            id_senasa_dict[fertilizante] = [id_senasa]
                    
                    # Validar fertilizante en función del código
                    def validar_fertilizante_codigo(codigo):
                        
                        if pd.isna(codigo):
                            return ''
                        for producto, codigos in id_senasa_dict.items():
                            if codigo in codigos:
                                return producto
                        return ''
                    
                    # Aplicar validación en cada columna
                    for i in range(1, 5):  
                        df[f'{i}_Producto_1'] = df[f'{i}_Codigo_Registro_1'].apply(validar_fertilizante)
                        df[f'{i}_Producto_2'] = df[f'{i}_Codigo_Registro_2'].apply(validar_fertilizante)
                    
                    return df

                # VALIDAR DOSIS DE PRODUCTOS (si es que primero se validó el producto) # FUNCIONA
                def validar_dosis_productos(df):
                        
                    columnas_productos_dosis = [
                        ('1_Producto_1', '1_Dosis_1'),
                        ('1_Producto_2', '1_Dosis_2'),
                        ('2_Producto_1', '2_Dosis_1'),
                        ('2_Producto_2', '2_Dosis_2'),
                        ('3_Producto_1', '3_Dosis_1'),
                        ('3_Producto_2', '3_Dosis_2'),
                        ('4_Producto_1', '4_Dosis_1'),
                        ('4_Producto_2', '4_Dosis_2')
                    ]
                
                    for producto_col, dosis_col in columnas_productos_dosis:
                        df.loc[df[producto_col] == '', dosis_col] = pd.NA
                
                    return df
                df = validar_dosis_productos(df)

                def validar_momentos(df, conn):
                        
                    momento_fertilizacion_query = """SELECT valor FROM datcrea_tablas.variables_generales 
                                                        WHERE variable = 'momento_fertilizacion'
                                                        AND valor NOT IN ('Postsiembra')"""
                    momento_fertilizacion_df = pd.read_sql(momento_fertilizacion_query,con=conn)
                    momento_fertilizacion = momento_fertilizacion_df['valor'].tolist()
                
                    # Columnas de momento a ajustar
                    columnas_momento = ['1_Momento', '2_Momento', '3_Momento', '4_Momento']
                
                    # Iterar sobre las columnas para ajustar los momentos
                    for columna in columnas_momento:
                        df[columna] = df[columna].apply(lambda x: x if x in momento_fertilizacion else '')
                
                    return df
                
                df = validar_momentos(df,conn)

                def calcular_densidad(df, conn):
                        
                    df['DensidadPlantasHa'] = df['DensidadPlantasHa']/10000
                    df.rename(columns={'DensidadPlantasHa': 'DensidadPlantasM2'}, inplace=True)
                    return df
                
                df = calcular_densidad(df, conn)


                def reindex_columns(df):
                    # Columnas del procesamiento anterior
                    cols = ['CUIC','Campo','Campo_Otro','Columna_I','Lote','Lote_Otro','vacio','vacio','vacio',
                        'Provincia','Departamento','vacio','Localidad','CalidadAmbiente','SuperficieSembrada',
                        'Tenencia','DestinoProduccion','vacio','Cultivo','vacio','SubgrupoCultivo','vacio',
                        'CultivoAntecesor','vacio','FechaSiembra','vacio','vacio','vacio','espaciamiento',
                        'DensidadKgHa','DensidadSemillasM2','DensidadSemillasHA','DensidadPlantasM2','vacio','vacio',
                        'Semillero','Genetica','vacio','Fertilizacion','1_Producto_1','1_Dosis_1','1_Momento',
                        '1_Forma','1_Producto_2','1_Dosis_2','2_Momento','1_Forma','2_Producto_1','2_Dosis_1',
                        '3_Momento','2_Forma','2_Producto_2','2_Dosis_2','4_Momento','2_Forma','3_Producto_1',
                        '3_Dosis_1','5_Momento','3_Forma','3_Producto_2','3_Dosis_2','6_Momento','3_Forma',
                        'vacio','Riego','Lamina_Riego','Napa','vacio','vacio','vacio','vacio','vacio','vacio',
                        'vacio','Superficie','vacio','Rendimiento']
                    return df.reindex(columns = cols)


                
                df = reindex_columns(df)

                
                # 5) Creación de columnas Campo_Otro y Lote_Otro (para caso SIMA) #FUNCIONA
                def otro_cuic_campo_lote(df, conn):
                    #df['CUIC'] = df['CUIC'].astype(str)
                    df['CUIC'] = df['CUIC'].fillna('Otro')
                    #df['Campo_Otro'] = df['Campo_Otro'].fillna('Otro')
                    #df['Lote_Otro'] = df['Lote_Otro'].fillna('Otro')
                    
                    # Llenar la columna 'Campo_Otro' con 'Otro' donde la columna 'Campo' tiene NaN
                    df.loc[df['Campo'].isna(), 'Campo_Otro'] = 'Otro'
                    # Dejar vacía la columna 'Campo_Otro' donde la columna 'Campo' tiene datos
                    df.loc[df['Campo'].notna(), 'Campo_Otro'] = ''
                    
                    # Llenar la columna 'Lote_Otro' con 'otro' donde la columna 'Lote' tiene NaN
                    df.loc[df['Lote'].isna(), 'Lote_Otro'] = 'Otro'
                    # Dejar vacía la columna 'Lote_Otro' donde la columna 'Lote' tiene datos
                    df.loc[df['Lote'].notna(), 'Lote_Otro'] = ''
                
                    return df
                
                df = otro_cuic_campo_lote(df, conn)


                # Localidades
                def validar_localidad(df):
                    
                    localidades_query = """SELECT localidad FROM datcrea_tablas.localidades"""
                    
                    localidades = pd.read_sql(localidades_query,con=conn)
                    localidades = localidades['localidad'].str.title().tolist()
                    ## 6.1) Validación Localidad
                    # Unificar criterio de escritura
                    df['Localidad'] = df['Localidad'].astype(str)
                    df['Localidad'] = df['Localidad'].str.title()
                    df['Localidad'] = ['' if x not in localidades else x for x in df['Localidad']]
                    return df
            

                df = validar_localidad(df)
                
                # Departamentos
                def validar_departamento(df,conn):
                
                    reemplazos = {
                        '25 DE MAYO': 'Veinticinco.de.Mayo',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'ADOLFO ALSINA': 'Adolfo.Alsina',
                        'ADOLFO GONZALES CHAVES': 'Adolfo.Gonzales.Chaves',
                        'ALBERTI': 'Alberti',
                        'ALMIRANTE BROWN': '',
                        'ARRECIFES': 'Arrecifes',
                        'AVELLANEDA': '',
                        'AYACUCHO': 'Ayacucho',
                        'AZUL': 'Azul',
                        'BAHIA BLANCA': 'Bahía.Blanca',
                        'BALCARCE': 'Balcarce',
                        'BARADERO': 'Baradero',
                        'BENITO JUAREZ': 'Benito.Juarez',
                        'BERAZATEGUI': 'Berazategui',
                        'BERISSO': 'Berisso',
                        'BOLIVAR': 'Bolivar',
                        'BRAGADO': 'Partido.Bragado',
                        'BRANDSEN': 'Brandsen',
                        'CAMPANA': 'Campana',
                        'CANUELAS': 'Cañuelas',
                        'CAPITAN SARMIENTO': 'Capitan.Sarmiento',
                        'CARLOS CASARES': 'Carlos.Casares',
                        'CARLOS TEJEDOR': 'Carlos.Tejedor',
                        'CARMEN DE ARECO': 'Carmen.de.Areco',
                        'CASTELLI': 'Castelli',
                        'CHACABUCO': 'Chacabuco',
                        'CHASCOMUS': 'Chascomus',
                        'CHIVILCOY': 'Chivilcoy',
                        'COLON': 'Colon',
                        'CORONEL DE MARINA L. ROSALES': 'Coronel.de.Marina.L.Rosales',
                        'CORONEL DORREGO': 'Coronel.Dorrego',
                        'CORONEL PRINGLES': 'Coronel.Pringles',
                        'CORONEL SUAREZ': 'Coronel.Suarez',
                        'DAIREAUX': 'Daireaux',
                        'DOLORES': 'Dolores',
                        'ENSENADA': 'Ensenada',
                        'ESCOBAR': 'Escobar',
                        'ESTEBAN ECHEVERRIA': 'Esteban.Echeverria',
                        'EXALTACION DE LA CRUZ': 'Exaltación.de.La.Cruz',
                        'EZEIZA': 'Ezeiza',
                        'FLORENCIO VARELA': 'Florencio.Varela',
                        'FLORENTINO AMEGHINO': 'Florentino.Ameghino',
                        'GENERAL ALVARADO': 'General.Alvarado',
                        'GENERAL ALVEAR': 'General.Alvear',
                        'GENERAL ARENALES': 'General.Arenales',
                        'GENERAL BELGRANO': 'General.Belgrano',
                        'GENERAL GUIDO': 'General.Guido',
                        'GENERAL JUAN MADARIAGA': 'General.Juan.Madariaga',
                        'GENERAL LA MADRID': 'General.La.Madrid',
                        'GENERAL LAS HERAS': 'General.Las.Heras',
                        'GENERAL LAVALLE': 'General.Lavalle',
                        'GENERAL PAZ': 'General.Paz',
                        'GENERAL PINTO': 'General.Pinto',
                        'GENERAL PUEYRREDON': 'General.Pueyrredon',
                        'GENERAL RODRIGUEZ': 'General.Rodriguez',
                        'GENERAL SAN MARTIN': 'General.San.Martin',
                        'GENERAL VIAMONTE': 'General.Viamonte',
                        'GENERAL VILLEGAS': 'General.Villegas',
                        'GUAMINI': 'Guamini',
                        'HIPOLITO YRIGOYEN': 'Hipolito.Yrigoyen',
                        'HURLINGHAM': 'Hurlingham',
                        'ITUZAINGO': 'Ituzaingo',
                        'JOSE C. PAZ': 'Jose.C.Paz',
                        'JUNIN': 'Junin',
                        'LA COSTA': 'La.Costa',
                        'LA MATANZA': 'La.Matanza',
                        'LA PLATA': 'La.Plata',
                        'LANUS': 'Lanús',
                        'LAPRIDA': 'Laprida',
                        'LAS FLORES': 'Las.Flores',
                        'LEANDRO N. ALEM': 'Leandro.N..Alem',
                        'LINCOLN': 'Lincoln',
                        'LOBERIA': 'Loberia',
                        'LOBOS': 'Lobos',
                        'LOMAS DE ZAMORA': 'Lomas.de.Zamora',
                        'LUJAN': 'Lujan',
                        'MAGDALENA': 'Magdalena',
                        'MAIPU': 'Maipu',
                        'MALVINAS ARGENTINAS': 'Malvinas.Argentinas',
                        'MAR CHIQUITA': 'Mar.Chiquita',
                        'MARCOS PAZ': 'Marcos.Paz',
                        'MERCEDES': 'Mercedes',
                        'MERLO': 'Merlo',
                        'MONTE': 'Monte',
                        'MONTE HERMOSO': 'Monte.Hermoso',
                        'MORENO': 'Moreno',
                        'MORON': 'Moron',
                        'NAVARRO': 'Navarro',
                        'NECOCHEA': 'Necochea',
                        'OLAVARRIA': 'Olavarria',
                        'PATAGONES': 'Patagones',
                        'PEHUAJO': 'Pehuajo',
                        'PELLEGRINI': 'Pellegrini',
                        'PERGAMINO': 'Partido.Pergamino',
                        'PILA': 'Pila',
                        'PILAR': 'Pilar',
                        'PINAMAR': 'Pinamar',
                        'PRESIDENTE PERON': 'Presidente.Peron',
                        'PUAN': 'Puan',
                        'PUNTA INDIO': 'Punta.Indio',
                        'QUILMES': 'Quilmes',
                        'RAMALLO': 'Ramallo',
                        'RAUCH': 'Rauch',
                        'RIVADAVIA': 'Rivadavia',
                        'ROJAS': 'Rojas',
                        'ROQUE PEREZ': 'Roque.Perez',
                        'SAAVEDRA': 'Saavedra',
                        'SALADILLO': 'Saladillo',
                        'SALLIQUELO': 'Salliqueló',
                        'SALTO': 'Partido.Salto',
                        'SAN ANDRES DE GILES': 'San.Andres.de.Giles',
                        'SAN ANTONIO DE ARECO': 'San.A.de.Areco',
                        'SAN CAYETANO': 'San.Cayetano',
                        'SAN FERNANDO': 'San.Fernando',
                        'SAN ISIDRO': 'San.Isidro',
                        'SAN MIGUEL': 'San.Miguel',
                        'SAN NICOLAS': 'San.Nicolas',
                        'SAN PEDRO': 'San.Pedro',
                        'SAN VICENTE': 'San.Vicente',
                        'SUIPACHA': 'Suipacha',
                        'TANDIL': 'Tandil',
                        'TAPALQUE': 'Tapalque',
                        'TIGRE': 'Tigre',
                        'TORDILLO': 'Tordillo',
                        'TORNQUIST': 'Tornquist',
                        'TRENQUE LAUQUEN': 'Trenque.Lauquen',
                        'TRES ARROYOS': 'Tres.Arroyos',
                        'TRES DE FEBRERO': 'Tres.de.Febrero',
                        'TRES LOMAS': 'Tres.Lomas',
                        'VICENTE LOPEZ': 'Vicente.Lopez',
                        'VILLA GESELL': 'Villa.Gesell',
                        'VILLARINO': 'Villarino',
                        'ZARATE': 'Zarate',
                        'CAPITAL FEDERAL': '',
                        'AMBATO': 'Ambato',
                        'ANCASTI': 'Ancasti',
                        'ANDALGALA': '',
                        'ANTOFAGASTA DE LA SIERRA': 'Antofagasta.de.La.Sierra',
                        'BELEN': 'Belen',
                        'CAPAYAN': 'Capayan',
                        'CAPITAL': 'Capital',
                        'EL ALTO': 'El.Alto',
                        'FRAY MAMERTO ESQUIU': 'Fray.Mamerto.Esquiu',
                        'LA PAZ': 'La.Paz',
                        'PACLIN': 'Paclin',
                        'POMAN': 'Poman',
                        'SANTA MARIA': 'Santa.Maria',
                        'SANTA ROSA': 'Santa.Rosa',
                        'TINOGASTA': 'Tinogasta',
                        'VALLE VIEJO': 'Valle.Viejo',
                        '1 DE MAYO': '',
                        '12 DE OCTUBRE': 'Doce.de.Octubre',
                        '2 DE ABRIL': 'Dos.de.Abril',
                        'ALMIRANTE BROWN': 'Almirante.Brown',
                        'BERMEJO': 'Bermejo',
                        'CHACABUCO': 'Chacabuco',
                        'COMPANIA DE MAYOR ING FRANKLIN': '',
                        'FRAY JUSTO SANTA MARIA DE ORO': 'Fray.Justo.Santa.Maria.de.Oro',
                        'GENERAL BELGRANO': 'General.Belgrano',
                        'GENERAL DONOVAN': 'General.Donovan',
                        'GENERAL GUEMES': 'General.Guemes',
                        'INDEPENDENCIA': 'Independencia',
                        'LIBERTAD': 'Libertad',
                        'LIBERTADOR GRL SAN MARTIN': '',
                        'MAIPU': 'Maipu',
                        'MAYOR LUIS J FONTANA': 'Mayor.Luis.J.Fontana',
                        'NUEVE DE JULIO': 'Nueve.de.Julio',
                        'O HIGGINS': 'O.Higgins',
                        'PRIMERA ARGENTINA': '',
                        'QUITILIPI': 'Quitilipi',
                        'SAN FERNANDO': 'San.Fernando',
                        'SAN LORENZO': 'San.Lorenzo',
                        'SAN MARTIN': 'San.Martin',
                        'SARMIENTO': 'Sarmiento',
                        'TAPENAGA': 'Tapenaga',
                        '25 DE MAYO': 'Veinticinco.de.Mayo',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'ANGACO': 'Angaco',
                        'CALINGASTA': 'Calingasta',
                        'CAPITAL': 'Capital',
                        'CAUCETE': 'Caucete',
                        'CHIMBAS': 'Chimbas',
                        'IGLESIA': 'Iglesia',
                        '9 DE JULIO': 'Nueve.de.Julio',
                        'POCITO': 'Pocito',
                        'RAWSON': 'Rawson',
                        'RIVADAVIA': 'Rivadavia',
                        'SAN MARTIN': 'San.Martin',
                        'SANTA LUCIA': 'Santa.Lucia',
                        'SARMIENTO': 'Sarmiento',
                        'ULLUM': 'Ullum',
                        'VALLE FERTIL': 'Valle.Fertil',
                        'ZONDA': 'Zonda',
                    }
                    
                    # Convertimos a mayúsculas para asegurar que las claves coincidan
                    df['Departamento'] = df['Departamento'].str.upper()
                    
                    # Reemplazamos los valores en la columna 'DEPARTAMENTO' basados en el diccionario
                    df['Departamento'] = df['Departamento'].replace(reemplazos)
                    
                    departamentos_query = """SELECT departamento FROM datcrea_tablas.localidades"""
                    departamento_df = pd.read_sql(departamentos_query,con=conn)
                    departamento_df = departamento_df.drop_duplicates()
                    departamentos = departamento_df['departamento'].tolist()
                    df['Departamento'] = ['' if x not in departamentos else x for x in df['Departamento']]
                    return df
                
                df = validar_departamento(df, conn)

                ## 6.3) Validación Provincias
                # PROVINCIAS
                def validar_provincia(df, conn):
                    reemplazos = {
                        'BUENOS AIRES': 'Buenos.Aires',
                        'CAPITAL FEDERAL': '',
                        'CATAMARCA': 'Catamarca',
                        'CHACO': 'Chaco',
                        'CHUBUT': '',
                        'CORDOBA': 'Córdoba',
                        'CORRIENTES': 'Corrientes',
                        'ENTRE RIOS': 'Entre.Ríos',
                        'FORMOSA': '',
                        'JUJUY': 'Jujuy',
                        'LA PAMPA': 'La.Pampa',
                        'LA RIOJA': '',
                        'MENDOZA': 'Mendoza',
                        'MISIONES': '',
                        'NEUQUEN': '',
                        'RIO NEGRO': '',
                        'SALTA': 'Salta',
                        'SAN JUAN': '',
                        'SAN LUIS': 'San.Luis',
                        'SANTA CRUZ': '',
                        'SANTA FE': 'Santa.Fe',
                        'SANTIAGO DEL ESTERO': 'Santiago.del.Estero',
                        'TIERRA DEL FUEGO': '',
                        'TUCUMAN': 'Tucumán',
                        'DESCONOCIDO': ''
                    }
                    
                    # Convertimos a mayúsculas para asegurar que las claves coincidan
                    df['Provincia'] = df['Provincia'].str.upper()
                    
                    # Reemplazamos los valores en la columna 'Provincia' basados en el diccionario
                    df['Provincia'] = df['Provincia'].replace(reemplazos)
                    #df['Provincia'] = df['Provincia'].astype(str)
                    df['Provincia'] = df['Provincia'].replace('nan', '')
                    
                    return df
                
                
                df = validar_provincia(df, conn)

                # TENENCIA #FUNCIONA
                def validar_tenencia(df,conn):
                    
                    tenencia_query = """SELECT valor FROM datcrea_tablas.variables_generales 
                                    WHERE variable = 'tenencia'"""
                    tenencia_df = pd.read_sql(tenencia_query,con=conn)
                    tenencia = tenencia_df['valor'].tolist()
                
                    df['Tenencia'] = df['Tenencia'].apply(lambda x: x if x in tenencia else '')
                
                    return df
                
                df = validar_tenencia(df,conn)

                
                # DESTINO #FUNCIONA
                def validar_destino(df,conn):
                    
                    destino_query = """SELECT valor FROM datcrea_tablas.variables_generales 
                                        WHERE variable = 'destino'"""
                    destino_df = pd.read_sql(destino_query,con=conn)
                    destino = destino_df['valor'].tolist()
                    df['DestinoProduccion'] = df['DestinoProduccion'].apply(lambda x: x if x in destino else '')
                    
                    return df
                
                df = validar_destino(df,conn)

                # CULTIVO ANTECESOR
                def validar_cultivo_antecesor(df,conn):
                    
                    # Lista de valores a reemplazar y sus correspondientes reemplazos
                    reemplazos = {
                        'acai - tardío': '',
                        'acai - de primera': '',
                        'acai - de segunda': '',
                        'achicoria - de primera': 'Achicoria',
                        'agave - tardío': '',
                        'agave - de segunda': '',
                        'agave - de primera': '',
                        'aguacate - de primera': '',
                        'ajo - de primera': 'Ajo',
                        'alfalfa - de segunda': 'Alfalfa',
                        'alfalfa - de primera': 'Alfalfa',
                        'alfalfa - tardío': 'Alfalfa',
                        'algodón - de segunda': 'Algodón',
                        'algodón - tardío': 'Algodón',
                        'algodón - de primera': 'Algodón',
                        'alpiste - de segunda': 'Alpiste',
                        'alpiste - de primera': 'Alpiste',
                        'anana - de primera': '',
                        'arroz - de primera': 'Arroz',
                        'arroz - tardío': 'Arroz',
                        'arroz - de segunda': 'Arroz',
                        'arveja - tardío': 'Arveja',
                        'arveja - de primera': 'Arveja',
                        'arveja - de segunda': 'Arveja',
                        'avena - de primera': 'Avena',
                        'avena - de segunda': 'Avena',
                        'avena - tardío': 'Avena',
                        'avena blanca - de segunda': 'Avena',
                        'avena blanca - de primera': 'Avena',
                        'avena strigosa - tardío': 'Avena',
                        'avena strigosa - de segunda': 'Avena',
                        'avena strigosa - de primera': 'Avena',
                        'banano - de primera': '',
                        'brócoli - de primera': '',
                        'café - safra': '',
                        'café - de segunda': '',
                        'café - de primera': '',
                        'camelina - de primera': 'Camelina',
                        'caña de azucar - de primera': 'Caña de Azucar',
                        'caña de azucar - tardío': 'Caña de Azucar',
                        'caña de azucar - de segunda': 'Caña de Azucar',
                        'canola - de primera': '',
                        'carinata - de primera': 'Carinata',
                        'carinata - de segunda': 'Carinata',
                        'cartamo - de primera': 'Cártamo',
                        'cebada - de primera': 'Cebada',
                        'cebada - tardío': 'Cebada',
                        'cebada - de segunda': 'Cebada',
                        'cebadilla - de primera': 'Cebadilla',
                        'centeno - de segunda': 'Centeno',
                        'centeno - de primera': 'Centeno',
                        'centeno - tardío': 'Centeno',
                        'chia - tardío': '',
                        'chia - de segunda': '',
                        'chia - de primera': '',
                        'cítricos - de segunda': '',
                        'cítricos - tardío': '',
                        'cítricos - de primera': '',
                        'colza - de primera': 'Colza',
                        'colza - de segunda': 'Colza',
                        'coriandro - de primera': 'Coriandro',
                        'crotalaria - de primera': '',
                        'cultivo de servicio - tardío': 'Cultivo de Servicio',
                        'cultivo de servicio - de primera': 'Cultivo de Servicio',
                        'cultivo de servicio - de segunda': 'Cultivo de Servicio',
                        'festuca - de primera': 'Pastura perenne',
                        'garbanzo - de segunda': 'Garbanzo',
                        'garbanzo - de primera': 'Garbanzo',
                        'girasol - tardío': 'Girasol',
                        'girasol - de primera': 'Girasol',
                        'girasol - de segunda': 'Girasol',
                        'kiwi - de primera': '',
                        'lechuga - de segunda': '',
                        'lenteja - de primera': 'Lenteja',
                        'lenteja - de segunda': 'Lenteja',
                        'lenteja - tardío': 'Lenteja',
                        'limon - de primera': '',
                        'lotus - de primera': 'Lotus',
                        'lupino - de primera': '',
                        'maiz' : 'Maíz',
                        'maíz - de primera': 'Maíz de 1° Temprano',
                        'maíz - tardío': 'Maíz de 1° Tardio',
                        'maíz - safrinha': '',
                        'maíz - de segunda': 'Maíz de 2°',
                        'mani - tardío': 'Maní',
                        'mani - de primera': 'Maní',
                        'mani - de segunda': 'Maní',
                        'melon - de primera': '',
                        'mijo - de segunda': 'Mijo',
                        'mijo - de primera': 'Mijo',
                        'mijo - tardío': 'Mijo',
                        'moha - de primera': 'Moha',
                        'moha - de segunda': 'Moha',
                        'moha - tardío': 'Moha',
                        'nabo - de segunda': 'Nabo',
                        'nabo - de primera': 'Nabo',
                        'olivo - tardío': '',
                        'papa - de segunda': 'Papa',
                        'papa - de primera': 'Papa',
                        'pasturas - tardío': 'Pastura perenne',
                        'pasturas - de primera': 'Pastura perenne',
                        'pasturas - de segunda': 'Pastura perenne',
                        'poroto - tardío': 'Poroto',
                        'poroto - de segunda': 'Poroto',
                        'poroto - de primera': 'Poroto',
                        'poroto mung - tardío': 'Poroto Mung',
                        'poroto mung - de segunda': 'Poroto Mung',
                        'poroto mung - de primera': 'Poroto Mung',
                        'raigrass - de segunda': 'Ryegrass',
                        'raigrass - de primera': 'Ryegrass',
                        'raigrass - tardío': 'Ryegrass',
                        'sandia - de primera': '',
                        'sesamo - de primera': '',
                        'soja' : 'Soja',
                        'soja - de segunda': 'Soja de 2°',
                        'soja - tardío': 'Soja de 1°',
                        'soja - primavera / verano': '',
                        'soja - safra': '',
                        'soja - semestre a': '',
                        'soja - de primera': 'Soja de 1°',
                        'sorgo - de primera': 'Sorgo',
                        'sorgo - tardío': 'Sorgo',
                        'sorgo - de segunda': 'Sorgo',
                        'tabaco - de primera': '',
                        'té - de primera': '',
                        'tomate - de segunda': '',
                        'tomate - de primera': '',
                        'trebol blanco - tardío': 'Trebol',
                        'trebol blanco - de primera': 'Trebol',
                        'trebol rojo - de primera': 'Trebol',
                        'trigo - de segunda': 'Trigo',
                        'trigo - de primera': 'Trigo',
                        'trigo - tardío': 'Trigo',
                        'triticale - de primera': 'Triticale',
                        'vicia - de segunda': 'Vicia',
                        'vicia - de primera': 'Vicia',
                        'vicia - tardío': 'Vicia',
                        'vid - de primera': 'Vid',
                        'vid - tardío': 'Vid',
                        'vid - de segunda': 'Vid',
                        'yerba mate - de primera': ''
                    }
                
                    # Convertir a minúsculas y reemplazar valores según el diccionario
                    df['CultivoAntecesor'] = df['CultivoAntecesor'].str.lower()
                    df['CultivoAntecesor'] = df['CultivoAntecesor'].replace(reemplazos)
                    
                    
                    return df
                    
                df = validar_cultivo_antecesor(df, conn)
                

                ## 6.7) Validación Calidad ambiente #FUNCIONA
                def validar_calidad_ambiente(df,conn):
                                
                    calidad_ambiente_query = """SELECT valor 
                                                FROM datcrea_tablas.variables_generales 
                                                WHERE variable = 'calidad_ambiente'"""
                    calidad_ambiente_df = pd.read_sql(calidad_ambiente_query,con=conn)
                    calidad_ambiente = calidad_ambiente_df['valor'].tolist()
                    
                    df['CalidadAmbiente'] = df['CalidadAmbiente'].apply(lambda x: x if x in calidad_ambiente else '')
                
                    return df
                    
                df = validar_calidad_ambiente(df, conn)
                

                def validar_subgrupo(df, conn): #REVISAR #FUNCIONA
                    
                    subgrupo_query = """SELECT subgrupo FROM datcrea_tablas.subgrupos_cultivos"""
                    subgrupo_df = pd.read_sql(subgrupo_query,con=conn)
                    subgrupo = subgrupo_df['subgrupo'].tolist()
                    df['SubgrupoCultivo'] = ['' if x not in subgrupo else x for x in df['SubgrupoCultivo']]
                
                
                    #Carga los subgrupos de cultivosvy los organiza en un diccionario
                    subgrupos_cultivos_query = """SELECT cultivo, subgrupo FROM datcrea_tablas.subgrupos_cultivos WHERE cultivo IS NOT NULL"""
                    subgrupos_cultivos = pd.read_sql(subgrupos_cultivos_query, con=conn)
                
                    # Crear un diccionario para almacenar los resultados
                    subgrupos_cultivos_dict = {}
                    for index, row in subgrupos_cultivos.iterrows():
                        cultivo = row['cultivo']
                        subgrupo = row['subgrupo']
                        if cultivo in subgrupos_cultivos_dict:
                            subgrupos_cultivos_dict[cultivo].append(subgrupo)
                        else:
                            subgrupos_cultivos_dict[cultivo] = [subgrupo]
                
                    def validar_subgrupo(row):
                        
                        valor_cultivo = row['Cultivo']
                        valor_subgrupo = row['SubgrupoCultivo']
                
                        if valor_cultivo in subgrupos_dict:
                            if valor_subgrupo in subgrupos_dict[valor_cultivo]:
                                return valor_subgrupo
                        return ''
                
                    return df
                
                df = validar_subgrupo(df, conn)
                

                # FERTILIZACIÓN
                def validar_fertilizacion(df,conn):
                    
                    fertilizacion_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'fertilizacion'"""
                    fertilizacion = pd.read_sql(fertilizacion_query, con=conn)
                    fertilizacion = fertilizacion['valor'].tolist()
                    df['Fertilizacion'] = ['' if x not in fertilizacion else x for x in df['Fertilizacion']]
                    return df
                
                df = validar_fertilizacion(df, conn)
                


                ## 6.12) Reemplazo de 0 por Na
                # REEMPLAZAR CEROS
                def replace_ceros(df):
                    df['SuperficieSembrada'] = df['SuperficieSembrada'].replace(0, pd.NA)
                    df['espaciamiento'] = df['espaciamiento'].replace(0, pd.NA)
                    df['Rendimiento'] = df['Rendimiento'].replace(0, pd.NA)
                    df['Superficie'] = df['Superficie'].replace(0, pd.NA)
                    return df
                
                df = replace_ceros(df)
                
                # 6.13) Validación de Forma
                
                # VALIDAR FORMA DE FERTILIZACIÓN
                def validar_forma_fertilizacion(df,conn):
                    forma_fertilizacion_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'forma_fertilizacion'"""
                    forma_fertilizacion = pd.read_sql(forma_fertilizacion_query, con=conn)
                    forma_fertilizacion = forma_fertilizacion['valor'].tolist()
                    df['1_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['1_Forma']]
                    df['2_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['2_Forma']]
                    df['3_Forma'] = ['' if x not in forma_fertilizacion else x for x in df['3_Forma']]
                    return df
                
                df = validar_forma_fertilizacion(df, conn)

                
                # Validación de Sistema de riego
                # SISTEMA RIEGO
                def validar_sistema_riego(df,conn):
                    sistema_riego_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'sistema_riego'"""
                    sistema_riego = pd.read_sql(sistema_riego_query, con=conn)
                    sistema_riego = sistema_riego['valor'].tolist()
                    df['Riego'] = ['' if x not in sistema_riego else x for x in df['Riego']]
                    return df
                
                df = validar_sistema_riego(df,conn)


                # INFLUENCIA NAPA
                def influencia_napa(df,conn):
                    
                    influencia_napa_query = """SELECT variable, valor FROM datcrea_tablas.variables_generales WHERE variable = 'influencia_napa'"""
                    influencia_napa = pd.read_sql(influencia_napa_query,con=conn)
                    influencia_napa = influencia_napa['valor'].tolist()
                    df['Napa'] = ['' if x not in influencia_napa else x for x in df['Napa']]
                    return df
                
                df = influencia_napa(df, conn)
                

                # Conexión a la base
                import pandas as pd
                from sqlalchemy import create_engine
                
                host="dw.crea.org.ar"
                port="54322"
                user="postgres"
                password="Tr0pema44cr34#"
                database="warehouse"
                
                conn2 = create_engine("postgresql+psycopg2://{user}:{password}@{host}:{port}/{database}".format(host=host,user=user,password=password,database=database,port=port))
                
                # Escribir en celdas excel
                region_grupo_query = """SELECT r.region_sigla, r.region_nombre,c.crea_numero,c.crea_nombre  
                FROM crm_crea.regiones AS r
                JOIN crm_crea.crea AS c 
                ON r.region_id = c.crea_region_id
                WHERE c.crea_baja IS NULL
                GROUP BY r.region_sigla, r.region_nombre, c.crea_numero,c.crea_nombre
                ORDER BY r.region_sigla, c.crea_nombre;"""
                region_grupo = pd.read_sql(region_grupo_query, con=conn2)
                
                #cuic = 'SSF145002M'
                if not pd.isnull(df['CUIC'][0]) and df['CUIC'][0] != 'Otro':
                    
                    cuic = df['CUIC'][0]
                    
                    region = region_grupo.loc[region_grupo['region_sigla'].str.startswith(cuic[:3]), 'region_nombre'].iloc[0]
                    #region = replace region por el formato dat que sale de una lista.
                    grupo = region_grupo.loc[region_grupo['crea_numero'].str.contains(cuic[3:6]), 'crea_nombre'].iloc[0]
                    
                    regiones_dat_dict = {
                        
                        'CENTRO':'CENTRO',
                        'CHACO SANTIAGUEÑO':'CHACO.SANTIAGUEÑO',
                        'CORDOBA NORTE':'CORDOBA.NORTE',
                        'ESTE':'ESTE',
                        'LITORAL NORTE':'LITORAL.NORTE',
                        'LITORAL SUR':'LITORAL.SUR',
                        'MAR Y SIERRAS':'MAR.Y.SIERRAS',
                        'NORTE DE BUENOS AIRES':'NORTE.BUENOS.AIRES',
                        'NOA':'NOA',
                        'NORTE DE SANTA FE':'NORTE.SANTA.FE',
                        'OESTE ARENOSO':'OESTE.ARENOSO',
                        'OESTE':'OESTE',
                        'PATAGONIA':'PATAGONIA',
                        'SEMIARIDA':'SEMIARIDA',
                        'SUDESTE':'SUDESTE',
                        'SANTA FE CENTRO':'SANTA.FE.CENTRO',
                        'SUR DE SANTA FE':'SUR.SANTA.FE',
                        'SUDOESTE':'SUDOESTE',
                        'VALLES CORDILLERANOS':'VALLES.CORDILLERANOS'
                        }
                    # Reemplazo directo sin función
                    for clave, valor in regiones_dat_dict.items():
                        region = region.replace(clave, valor)
                    
                    #print('REGIÓN: ', region)
                    
                    grupos_dat_dict = {'ALEJANDRO CHAJAN':'ALEJANDRO.CHAJAN',
                    'BUENA ESPERANZA':'BUENA.ESPERANZA',
                    'CANALS':'CANALS',
                    'CARNERILLO':'CARNERILLO',
                    'CAÑADA SECA':'CAÑADA.SECA',
                    'CTALAMOCHITA':'CTALAMOCHITA',
                    'HUINCA RENANCO':'HUINCA.RENANCO',
                    'LA CESIRA TAMBERO':'LA.CESIRA.TAMBERO',
                    'LA PORTADA':'LA.PORTADA',
                    'LABOULAYE - BOUCHARDO':'LABOULAYE.BUCHARDO',
                    'MELO SERRANO':'MELO.SERRANO',
                    'RANQUELES':'RANQUELES',
                    'RIO CUARTO':'RIO.CUARTO',
                    'RIO QUINTO':'RÍO.QUINTO',
                    'TAMBERO LABOULAYE':'TAMBERO.LABOULAYE',
                    'TAMBERO VILLA MARIA':'VILLA.MARIA',
                    'TEGUA':'TEGUA',
                    'VALLE DEL CONLARA':'VALLE.DEL.CONLARA',
                    'WASHINGTON MACKENNA':'WASHINGTON.MACKENNA',
                    'CAMPO GALLO MONTE QUEMADO':'CAMPO.GALLO.MONTE.QUEMADO',
                    'GUAYACAN':'GUAYACAN',
                    'IBARRETA':'IBARRETA',
                    'LOMITAS':'LOMITAS',
                    'PALMARES':'PALMARES',
                    'PAMPA DEL INFIERNO':'PAMPA.DEL.INFIERNO',
                    'QUIMILI':'QUIMILI',
                    'RENOVALES':'RENOVALES',
                    'SACHAYOJ':'SACHAYOJ',
                    'SANAVIRONES':'SANAVIRONES',
                    'SEMIARIDO NORTE':'SEMIARIDO.NORTE',
                    'SUDESTE SANTIAGUEÑO':'SES',
                    'TINTINA':'TINTINA',
                    'ARROYITO':'ARROYITO',
                    'BARRANCA YACO':'BARRANCA.YACO',
                    'CAROYA':'CAROYA',
                    'CAÑADA DE LUQUE SITON':'CAÑADA.DE.LUQUE.SITON',
                    'DEL ESTE':'DEL.ESTE',
                    'GANADERO DEL NOROESTE':'GANADERO.DEL.NOROESTE',
                    'JESUS MARIA':'JESUS.MARIA',
                    'LAGUNA LARGA':'LAGUNA.LARGA',
                    'LEOPOLDO LUGONES':'LEOPOLDO.LUGONES',
                    'MONTE CRISTO':'MONTE.CRISTO',
                    'RIO PRIMERO':'RIO.PRIMERO',
                    'SIERRAS CHICAS':'SIERRAS.CHICAS',
                    'TOTORAL':'TOTORAL',
                    'ABASTO':'Otro',
                    'CAÑUELAS':'CAÑUELAS',
                    'GELAS':'GELAS',
                    'LUJAN':'LUJAN',
                    'NAVARRO II':'NAVARROII',
                    'PIONEROS ABASTO':'Otro',
                    'AVATI - I - ARROCERO':'AVATI.I.ARROCERO',
                    'CURUZU CUATIA':'CURUZU.CUATIA',
                    'ESQUINA':'ESQUINA',
                    'MALEZALES':'Otro',
                    'MERCEDES':'MERCEDES',
                    'TIERRA COLORADA':'TIERRA.COLORADA',
                    'URUNDAY':'URUNDAY',
                    'ÑANDUBAY':'ÑANDUBAY',
                    'BOVRIL EL SOLAR':'BOVRIL.EL.SOLAR',
                    'CONCEPCION URUGUAY':'CONCEPCION.URUGUAY',
                    'CONCORDIA CHAJARI':'CONCORDIA.CHAJARI',
                    'GALARZA':'GALARZA',
                    'GUALEGUAYCHU':'GUALEGUAYCHU',
                    'ISLAS DEL IBICUY':'ISLAS.DEL.IBICUY',
                    'LA PAZ':'LA.PAZ',
                    'LARROQUE GUALEGUAY':'LARROQUE.GUALEGUAY',
                    'MANDISOVI CONCORDIA':'MANDISOVI.CONCORDIA',
                    'MONTOYA':'MONTOYA',
                    'SAN JAIME':'SAN.JAIME',
                    'VICTORIA':'VICTORIA',
                    'VILLAGUAY':'VILLAGUAY',
                    'ARROYO DE LOS HUESOS':'ARROYO.DE.LOS.HUESOS',
                    'AZUL CHILLAR':'AZUL.CHILLAR',
                    'CTE.N.OTAMENDI':'CTE.N.OTAMENDI',
                    'DEFFERRARI':'DEFFERRARI',
                    'FRONTERA':'FRONTERA',
                    'FULTON':'FULTON',
                    'LOBERIAS GRANDES':'LOBERIAS.GRANDES',
                    'NECOCHEA QUEQUEN':'NECOCHEA.QUEQUEN',
                    'QUEQUEN SALADO':'QUEQUEN.SALADO',
                    'SAN CAYETANO - TRES ARROYOS':'SAN.CAYETANO.TRES.ARROYOS',
                    'SAN FRANCISCO DE BELLOCQ':'SAN.FRANCISCO.DE.BELLOCQ',
                    'SAN MANUEL':'SAN.MANUEL',
                    'TAMBERO MAR Y SIERRA':'TAMBERO.MAR.Y.SIERRA',
                    'TANDIL':'TANDIL',
                    'TRES ARROYOS':'TRES.ARROYOS',
                    'ZONA 4 LECHERA':'ZONA.4.LECHERA',
                    'ALBERDI':'ALBERDI',
                    'ALBERTI- PLA':'ALBERTI.PLA',
                    'ARROYO DEL MEDIO':'ARROYO.DEL.MEDIO',
                    'BRAGADO':'BRAGADO',
                    'GIDAG':'GIDAG',
                    'PERGAMINO':'PERGAMINO',
                    'RAWSON TRES SARGENTOS':'RAWSON.TRES.SARGENTOS',
                    'SAN ANTONIO DE ARECO':'SAN.ANTONIO.DE.ARECO',
                    'SAN PEDRO VILLA LIA':'SAN.PEDRO.VILLA.LIA',
                    'SEGUI LA ORIENTAL':'SEGUI.LA.ORIENTAL',
                    'CAÑAVERALES DE TUCUMAN':'CAÑAVERALES.DE.TUCUMAN',
                    'EL PALOMAR':'EL.PALOMAR',
                    'EL RODEO':'EL.RODEO',
                    'LA COCHA':'LA.COCHA',
                    'LOS ALGARROBOS':'LOS.ALGARROBOS',
                    'METAN':'METAN',
                    'PALO SANTO':'Otro',
                    'SAN PATRICIO':'SAN.PATRICIO',
                    'SANTA ROSA CATAMARCA':'SANTA.ROSA.CATAMARCA',
                    'SURCOS':'Otro',
                    'SUYAY':'SUYAY',
                    'VALLES TEMPLADOS':'Otro',
                    'YUNGAS':'YUNGAS',
                    'CUÑA BOSCOSA':'CUÑA.BOSCOSA',
                    'MARGARITA CAMPO ALEMAN':'MARGARITA.CAMPO.ALEMAN',
                    'RAMAYON':'RAMAYON',
                    'SAN CRISTOBAL-LA LUCILA':'SAN.CRISTOBAL.LALUCILA',
                    'VILLA ANA-ARANDU':'VILLA.ANA.ARANDÚ',
                    'VILLA OCAMPO':'VILLA.OCAMPO',
                    'AMERICA':'AMERICA',
                    'AMERICA II':'AMERICA.II',
                    'AMERICA LECHERO':'AMERICA.LECHERO',
                    'ATREUCO':'ATREUCO',
                    'CORRALERO':'CORRALERO',
                    'PELLEGRINI - TRES LOMAS':'PELLEGRINI.TRES.LOMAS',
                    'PICO BARON':'PICO.BARON',
                    'PICO QUEMU':'PICO.QUEMU',
                    'QUEMU CATRILO':'QUEMU.CATRILO',
                    'TRENQUE LAUQUEN II':'TRENQUE.LAUQUEN.II',
                    'AGROGANADERO 9 DE JULIO':'AGROGANADERO.9.DE.JULIO',
                    'AMEGHINO':'AMEGHINO',
                    'BOLIVAR':'BOLIVAR',
                    'CASARES - 9 DE JULIO':'CASARES.9.DE.JULIO',
                    'GENERAL PINTO':'GENERAL.PINTO',
                    'GENERAL VILLEGAS':'GENERAL.VILLEGAS',
                    'GUANACO LAS TOSCAS':'GUANACO.LAS.TOSCAS',
                    'HENDERSON-DAIREAUX':'HENDERSON.DAIREAUX',
                    'HERRERA VEGAS - PEHUAJO':'HERRERA.VEGAS.PEHUAJO',
                    'INFOSURA':'INFOSURA',
                    'LA VIA':'LA.VIA',
                    'LINCOLN':'LINCOLN',
                    'MONES CAZON PEHUAJO':'MONES.CAZON.PEHUAJO',
                    'NUEVE DE JULIO':'NUEVE.DE.JULIO',
                    'PEHUAJO CASARES':'PEHUAJO.CASARES',
                    'PIROVANO LA LARGA':'PIROVANO.LA.LARGA',
                    'SALAZAR MONES CAZON':'SALAZAR.MONES.CAZON',
                    'TAMBERO AMEGHINO VILLEGAS':'TAMBERO.AMEGHINO.VILLEGAS',
                    'TEJEDOR':'TEJEDOR',
                    'TREINTA AGOSTO- MARI LAUQUEN':'TREINTA.AGOSTO.MARI.LAUQUEN',
                    'ALTO VALLE - VALLE MEDIO':'ALTO.VALLE.VALLE.MEDIO',
                    'CUENCA DEL AGRIO':'CUENCA.DEL.AGRIO',
                    'ESQUEL':'ESQUEL',
                    'LANIN':'LANIN',
                    'SANTA CRUZ':'SANTACRUZ',
                    'TIERRA DEL FUEGO':'TIERRADELFUEGO',
                    'VIEDMA':'VIEDMA',
                    'AOKEN AL':'AOKEN.AL',
                    'CALEUCHE':'CALEUCHE',
                    'GUATRACHE':'GUATRACHE',
                    'HOLISTICO':'HOLISTICO',
                    'PEHUENCHE':'PEHUENCHE',
                    'SOVEN':'SOVEN',
                    'TAMBERO GUATRACHE':'TAMBERO.GUATRACHE',
                    'UTRACAN':'UTRACAN',
                    '25 DE MAYO':'CREA.25.DE.MAYO',
                    'ARROYO DE LAS FLORES':'ARROYO.DE.LAS.FLORES',
                    'ARROYO LANGUEYU':'ARROYO.LANGUEYU',
                    'AYACUCHO':'AYACUCHO',
                    'CASTELLI - BELGRANO':'CASTELLI.BELGRANO',
                    'DEL TUYU':'DEL.TUYU',
                    'FORTIN MULITAS':'FORTIN.MULITAS',
                    'LEZAMA':'LEZAMA',
                    'MAIPU':'MAIPU',
                    'MAR CHIQUITA':'MAR.CHIQUITA',
                    'MONTE':'MONTE',
                    'PILA':'PILA',
                    'RAUCH - UDAQUIOLA':'RAUCH.UDAQUIOLA',
                    'RIO SALADO':'RIO.SALADO',
                    'ROQUE PEREZ SALADILLO':'ROQUE.PEREZ.SALADILLO',
                    'TAPALQUE II':'TAPALQUE.II',
                    'VALLIMANCA':'VALLIMANCA',
                    'CASTELAR':'CASTELAR',
                    'CENTRO OESTE SANTAFESINO':'C.O.S',
                    'CUENCA':'CUENCA',
                    'EL CEIBO':'EL.CEIBO',
                    'ESPERANZA':'ESPERANZA',
                    'GALVEZ':'GALVEZ',
                    'RAFAELA':'RAFAELA',
                    'SAN FRANCISCO':'SAN.FRANCISCO',
                    'SAN GUILLERMO':'Otro',
                    'SAN MARTIN DE LAS ESCOBAS - COLONIA BELGRANO':'S.M.E.C.B',
                    'SUNCHALES':'Otro',
                    'APLICADORES':'Otro',
                    'ARMSTRONG MONTES DE OCA':'ARMSTRONG.MONTES.DE.OCA',
                    'ASCENSION':'ASCENSION',
                    'COLONIA MEDICI':'COLONIA.MEDICI',
                    'COSTAS DEL CARCARAÑA':'COSTAS.DEL.CARCARAÑA',
                    'EL ABROJO':'EL.ABROJO',
                    'GENERAL ARENALES':'GENERAL.ARENALES',
                    'GENERAL BALDISSERA':'GENERAL.BALDISSERA',
                    'LA CALANDRIA':'LA.CALANDRIA',
                    'LA MAROMA':'LA.MAROMA',
                    'LAS PETACAS':'LAS.PETACAS',
                    'MARIA TERESA':'MARIA.TERESA',
                    'MONTE BUEY INRIVILLE':'MONTE.BUEY.INRIVILLE',
                    'MONTE MAIZ':'MONTE.MAIZ',
                    'POSTA ESPINILLOS':'POSTA.ESPINILLOS',
                    'ROSARIO':'ROSARIO',
                    'SAN JORGE LAS ROSAS':'SAN.JORGE.LAS.ROSAS',
                    'SANTA ISABEL':'SANTA.ISABEL',
                    'SANTA MARIA':'SANTA.MARIA',
                    'TEODELINA':'TEODELINA',
                    'BENITO JUAREZ':'BENITO.JUAREZ',
                    'CARHUE HUANGUELEN':'CARHUE.HUANGUELEN',
                    'CORONEL PRINGLES II':'CORONEL.PRINGLES.II',
                    'CORONEL SUAREZ':'CORONEL.SUAREZ',
                    'GENERAL LAMADRID':'GENERAL.LAMADRID',
                    'LAPRIDA':'LAPRIDA',
                    'NUESTRA SEÑORA DE LAS PAMPAS':'NUESTRA.SEÑORA.DE.LAS.PAMPAS',
                    'OLAVARRIA':'OLAVARRIA',
                    'PEDRO LURO':'PEDRO.LURO',
                    'SAN ELOY - PIÑEYRO':'SAN.ELOY.PIÑEYRO',
                    'VENTANIA':'VENTANIA',
                    'ACONCAGUA':'Otro',
                    'ARAUCO':'Otro',
                    'CALCHAQUI':'Otro',
                    'FRUTICOLA CUYO':'Otro',
                    'HUARPE':'Otro',
                    'LAS ACEQUIAS':'Otro',
                    'LOS ANDES':'Otro',
                    'NOGALERO DEL NORTE':'Otro',
                    'OLIVICOLA SAN JUAN':'Otro',
                    'VIGNERONS':'Otro'
                    }
                    
                    # Reemplazo directo sin función
                    for clave, valor in grupos_dat_dict.items():
                        grupo = grupo.replace(clave, valor)
                    
                    
                else: 
                    cuic = 'Otro'
                    region = ''
                    grupo = ''

                # Filtro el DataFrame
                # CULTIVO VERANO
                def filtrar_cultivos_invierno(df, conn):
                    query = """SELECT cultivo FROM datcrea_tablas.cultivos WHERE ciclo = 'i'"""
                    invierno_df = pd.read_sql(query, conn)
                    cultivos_invierno = invierno_df['cultivo'].tolist()
                    df = df[df['Cultivo'].isin(cultivos_invierno)]
                    return df
                
                invierno = filtrar_cultivos_invierno(df, conn)
                
                def exportar_sima_invierno(invierno):
                    seccion = 200 # Máx registros por planilla xls
                    registros_campana = len(invierno.index) # cuenta el total de registros
                    registros_wb = math.ceil(registros_campana / seccion) #cociente para saber cantidad de worksheets
                    
                    # Armo listado de subset de 200 (máx) registros cada uno
                    dataframes = []
                    
                    for i in range(registros_wb):
                        inicio = i * seccion
                        fin = (i + 1) * seccion
                        seccion_df = invierno.iloc[inicio:fin]
                        dataframes.append(seccion_df)
                    
                    # Crear una workbook para cada dataframe
                    
                    # Ruta del archivo original
                    ruta_original = 'C:/Users/EPolacco/Documents/9 - DAT/SIMA/DAT-Cultivos-de-invierno-2023-24.xlsx'
                    
                    # Cargar el archivo original
                    wb_original = load_workbook(filename=ruta_original)
                    
                    # Obtener la hoja original
                    planilla_original = wb_original['Planilla ']
                    
                    columns_to_exclude = {'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'U', 'W', 'Y', 'Z',
                                        'AF', 'AG', 'AK', 'AN', 'CB', 'CK', 'CN', 'CM', 'CO', 'CP', 'CQ', 
                                        'CR', 'DU', 'DV', 'DW', 'IA','IB','IC','ID','IE','IF','IG','IH',
                                        'II','IJ','IK','IL','IM','IN','IO' }
                    
                    
                    # Iterar sobre los df y guardar en archivos xls separados
                    for i, seccion_df in enumerate(dataframes, 1):
                        # Crear un workbook para cada sección
                        if i == 1:
                            wb = wb_original  # Utilizar el archivo original para el primer df
                            sheet = planilla_original
                        else:
                            # Crear una copia del archivo original para los df subsiguientes
                            ruta_copia = f'C:/Users/EPolacco/Documents/9 - DAT/SIMA/{cuic}_DAT-Cultivos-de-invierno-2023-24_SECCION_{i}.xlsx'
                            shutil.copy(ruta_original, ruta_copia)
                            
                            # Cargar la copia
                            wb = load_workbook(filename=ruta_copia)
                            sheet = wb['Planilla ']
                            sheet.title = f'Planilla '
                    
                        # Borrar datos existentes en la hoja, después de la fila 14 porque sino se pisan los datos
                        
                        for row in sheet.iter_rows(min_row=15, max_row=sheet.max_row, min_col=6, max_col=sheet.max_column):
                            for cell in row:
                                if cell.column_letter not in columns_to_exclude:
                                    cell.value = None
                    
                    
                        # Obtener las filas de datos del df
                        rows = dataframe_to_rows(seccion_df,index=False, header=False)
                        
                        # Copiar en worksheet a partir de la fila 15 y columna 7
                        for r_idx, row in enumerate(rows, 15):
                            for c_idx, value in enumerate(row, 6):
                                if not pd.isna(value):
                                    sheet.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Región
                        sheet['F3'] = region
                        # Grupo
                        sheet['F6'] = grupo
                        
                        # Guardar el workbook, si no es el primer df
                        if i != 1:
                            wb.save(ruta_copia)
                    
                    
                    # Guardar workbook con el primer df
                    ruta_alternativa = f'C:/Users/EPolacco/Documents/9 - DAT/SIMA/{cuic}_DAT-Cultivos-de-invierno-2023-24_SECCION_1.xlsx'
                    wb_original.save(ruta_alternativa)
                    
                
                exportar_sima_invierno(invierno) 


        convertir_a_sima(archivo_excel, estacion, campana_filtro)
        st.success('Proceso terminado')